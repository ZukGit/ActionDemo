#  encoding: utf-8

import tushare as ts
import os
import pandas as pd  # https://www.jianshu.com/p/5c0aa1fa19af
import openpyxl 
from openpyxl import load_workbook,Workbook
import time,datetime  # https://www.jb51.net/article/66019.htm
import re
import os
import tempfile

############################## 全局数据初始化块 数据初始化函数  End ##############################

##############  properities 函数 Begin ##############

class Properties:

    def __init__(self, file_name):
        self.file_name = file_name
        self.properties = {}
        try:
            fopen = open(self.file_name, 'r')
            for line in fopen:
                line = line.strip()
                if line.find('=') > 0 and not line.startswith('#'):
                    strs = line.split('=')
                    self.properties[strs[0].strip()] = strs[1].strip()
        except Exception :
            raise
        else:
            fopen.close()

    def has_key(self, key):
        return key in self.properties

    def get(self, key, default_value=''):
        if key in self.properties:
            return self.properties[key]
        return default_value

    def put(self, key, value):
        self.properties[key] = value
        replace_property(self.file_name, key + '=.*', key + '=' + value, True)


def replace_property(file_name, from_regex, to_str, append_on_not_exists=True):
    tmpfile = tempfile.TemporaryFile()

    if os.path.exists(file_name):
        r_open = open(file_name, 'r')
        pattern = re.compile(r'' + from_regex)
        found = None
        for line in r_open:
            if pattern.search(line) and not line.strip().startswith('#'):
                found = True
                line = re.sub(from_regex, to_str, line)
            tmpfile.write(line.encode())
        if not found and append_on_not_exists:
            tmpfile.write(('\n' + to_str).encode())
        r_open.close()
        tmpfile.seek(0)

        content = tmpfile.read()

        if os.path.exists(file_name):
            os.remove(file_name)

        w_open = open(file_name, 'wb')
        w_open.write(content)
        w_open.close()

        tmpfile.close()
    else:
        print ("file %s not found" % file_name)

##############  properities 函数 End ##############


##############  XLSX excel 函数 Begin ##############

def createexcel(filename):    ### 创建本地文件名称为 filename的文件
    wb = Workbook()
    curPath = os.path.abspath('.')
    cur_desktop_path=os.path.join(os.path.expanduser("~"), 'Desktop')
    curDir = cur_desktop_path+os.sep+"zbin"+os.sep+"J0_Data"+os.sep  ## 创建~/Desktop/zbin/J0_Data 这个 这个工作目录
    #curDir = curPath+os.sep
    if not os.path.exists(curDir):
        os.makedirs(curDir)
    #curPath = os.path.dirname(os.path.abspath('.'))
    #t = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    #suffix = '.xlsx' # 文件类型
    #newfile = t + suffix
    xlsxPath = curDir + filename
    print("curPath curDir=" + curPath)
    print("createexcel zbin_JO_Dir=" + curDir)
    print("createexcel path=" + xlsxPath)
    if not os.path.exists(xlsxPath):
        wb.save(xlsxPath)
        print(" wb.save(xlsxPath)  xlsxPath =" + xlsxPath)
        time.sleep(1)
    return xlsxPath

def getColumnIndex(table, columnName):   ## 返回 table 中 名称为  columnName 的 那列 的索引
    columnIndex = None
    for i in range(table.ncols):
        if(table.cell_value(0, i) == columnName):
            columnIndex = i
            break
    return columnIndex
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007

##############  XLSX excel 函数 End ##############



############################## Prop初始化 Begin ##############################

#
#Thu Aug 13 22:33:45 CST 2020
#rixianhangqing-time_record_date=20200707
#rixianhangqing-time_start_date=20010101

desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
zbin_path = str(desktop_path)+os.sep+"zbin"
j0_properties_path= str(zbin_path)+os.sep+"J0.properties"
print("desktop_path = "+ str(desktop_path))
print("zbin_path = "+ str(zbin_path))
print("j0_properties_path = "+ str(j0_properties_path))
J0_PROPS =  Properties(j0_properties_path)


############################## Prop初始化 End ##############################


############## tscode_股票列表的初始化  Begin ##############
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007
def init_tscode_data(book_name, sheet_name,ts_code_set,tscode_name_dict):
    # 读取excel
    wb = openpyxl.load_workbook(book_name)
    # 读取sheet表
    ws = wb[sheet_name]
    for i in range(ws.max_row):
         # 获取下拉框中的所有选择值
         if (i == 0 or i == 1):
             continue
         #print("i="+str(i)+" 总的列数:" + str(ws.max_row)+"  value:"+str(ws.cell(i,2).value))
         tscode_item = str(ws.cell(i,1).value)  ##  20201010--> 2:ts_code    4_name   ##  20201116--> 1:ts_code    3_name 
         tscode_name_item = str(ws.cell(i,3).value)
         #print("index="+str(i)+" tscode_item = "+ str(tscode_item) + "   tscode_name_item="+str(tscode_name_item))
         ts_code_set.add(str(ws.cell(i,1).value))
         tscode_name_dict[tscode_item]=tscode_name_item



tscode_set = set()    #### 股票代码的集合   000001.SZ   .... 999999.SH
tscode_name_dict = dict()  #### code-name 的 map的集合
init_tscode_data(zbin_path+os.sep+"J0_Python"+os.sep+"J0_股票列表.xlsx","股票列表",tscode_set,tscode_name_dict)
############## tscode_股票列表的初始化  End  ##############


############################## 运行属性 Begin ##############################
pd.set_option('display.max_rows', None)   ##  解决纵向出现...
#pd.set_option('display.width', 1000) 
pd.set_option('expand_frame_repr', False)  ##  解决横向出现...
Cur_Abs_Path=os.path.abspath('.')   # 表示当前所处的文件夹的绝对路径
print("当前绝对路径:"+Cur_Abs_Path)
Cur_Ref_Path=os.path.abspath('..')  # 表示当前所处的文件夹上一级文件夹的绝对路径
print("当前父目录绝对路径:"+Cur_Ref_Path)
############################## 运行属性 End ##############################

############################## 时间Date Begin ##############################

now_yyyymmdd=str(time.strftime('%Y%m%d', time.localtime(time.time())))
print("now_yyyymmdd = "+ str(now_yyyymmdd))

############################## 时间Date End ##############################





pro = ts.pro_api('43acb9a5ddc2cf73c6c4ea54796748f965457ed57daaa736bb778ea2')
# print(J0_PROPS.get(tree_node_name+"record_date"))           #根据key读取value
# J0_PROPS.put(tree_node_name+"record_date", now_yyyymmdd)       ###  覆盖原有的 key 和 value
#  zukgit
# rongzirongquanjiaoyimingxi_zukgit_website  =   https://tushare.pro/document/2?doc_id=59
tree_node_name="rongzirongquanjiaoyimingxi"+"_"
createexcel('margin_detail_2020.xlsx')
margin_detail_2020_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx')
margin_detail_2020_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx', engine='openpyxl')
margin_detail_2020_excel_writer.book = margin_detail_2020_book
margin_detail_2020_excel_writer.sheets = dict((ws.title, ws) for ws in margin_detail_2020_book.worksheets)
margin_detail_2020_1_xlsx_frame=pd.DataFrame()
if '1' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_1_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='1' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20210104")       ###  更新 记录日期
margin_detail_20210104 = pro.margin_detail(trade_date='20210104', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210104_tscode_list = list() 
margin_detail_20210104_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210104.columns.values.tolist():
    for ts_code_sh in margin_detail_20210104['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210104_tscode_list.append(stock_name)
    margin_detail_20210104_addname_dataframe['cname'] = margin_detail_20210104_tscode_list
for table_name in margin_detail_20210104.columns.values.tolist():
    margin_detail_20210104_addname_dataframe[table_name] = margin_detail_20210104[table_name]
print("融资融券交易明细  margin_detail_20210104 返回数据 row 行数 = "+str(margin_detail_20210104.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210104_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210105")       ###  更新 记录日期
margin_detail_20210105 = pro.margin_detail(trade_date='20210105', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210105_tscode_list = list() 
margin_detail_20210105_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210105.columns.values.tolist():
    for ts_code_sh in margin_detail_20210105['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210105_tscode_list.append(stock_name)
    margin_detail_20210105_addname_dataframe['cname'] = margin_detail_20210105_tscode_list
for table_name in margin_detail_20210105.columns.values.tolist():
    margin_detail_20210105_addname_dataframe[table_name] = margin_detail_20210105[table_name]
print("融资融券交易明细  margin_detail_20210105 返回数据 row 行数 = "+str(margin_detail_20210105.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210105_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210106")       ###  更新 记录日期
margin_detail_20210106 = pro.margin_detail(trade_date='20210106', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210106_tscode_list = list() 
margin_detail_20210106_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210106.columns.values.tolist():
    for ts_code_sh in margin_detail_20210106['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210106_tscode_list.append(stock_name)
    margin_detail_20210106_addname_dataframe['cname'] = margin_detail_20210106_tscode_list
for table_name in margin_detail_20210106.columns.values.tolist():
    margin_detail_20210106_addname_dataframe[table_name] = margin_detail_20210106[table_name]
print("融资融券交易明细  margin_detail_20210106 返回数据 row 行数 = "+str(margin_detail_20210106.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210106_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210107")       ###  更新 记录日期
margin_detail_20210107 = pro.margin_detail(trade_date='20210107', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210107_tscode_list = list() 
margin_detail_20210107_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210107.columns.values.tolist():
    for ts_code_sh in margin_detail_20210107['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210107_tscode_list.append(stock_name)
    margin_detail_20210107_addname_dataframe['cname'] = margin_detail_20210107_tscode_list
for table_name in margin_detail_20210107.columns.values.tolist():
    margin_detail_20210107_addname_dataframe[table_name] = margin_detail_20210107[table_name]
print("融资融券交易明细  margin_detail_20210107 返回数据 row 行数 = "+str(margin_detail_20210107.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210107_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210108")       ###  更新 记录日期
margin_detail_20210108 = pro.margin_detail(trade_date='20210108', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210108_tscode_list = list() 
margin_detail_20210108_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210108.columns.values.tolist():
    for ts_code_sh in margin_detail_20210108['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210108_tscode_list.append(stock_name)
    margin_detail_20210108_addname_dataframe['cname'] = margin_detail_20210108_tscode_list
for table_name in margin_detail_20210108.columns.values.tolist():
    margin_detail_20210108_addname_dataframe[table_name] = margin_detail_20210108[table_name]
print("融资融券交易明细  margin_detail_20210108 返回数据 row 行数 = "+str(margin_detail_20210108.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210108_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210111")       ###  更新 记录日期
margin_detail_20210111 = pro.margin_detail(trade_date='20210111', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210111_tscode_list = list() 
margin_detail_20210111_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210111.columns.values.tolist():
    for ts_code_sh in margin_detail_20210111['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210111_tscode_list.append(stock_name)
    margin_detail_20210111_addname_dataframe['cname'] = margin_detail_20210111_tscode_list
for table_name in margin_detail_20210111.columns.values.tolist():
    margin_detail_20210111_addname_dataframe[table_name] = margin_detail_20210111[table_name]
print("融资融券交易明细  margin_detail_20210111 返回数据 row 行数 = "+str(margin_detail_20210111.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210111_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210112")       ###  更新 记录日期
margin_detail_20210112 = pro.margin_detail(trade_date='20210112', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210112_tscode_list = list() 
margin_detail_20210112_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210112.columns.values.tolist():
    for ts_code_sh in margin_detail_20210112['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210112_tscode_list.append(stock_name)
    margin_detail_20210112_addname_dataframe['cname'] = margin_detail_20210112_tscode_list
for table_name in margin_detail_20210112.columns.values.tolist():
    margin_detail_20210112_addname_dataframe[table_name] = margin_detail_20210112[table_name]
print("融资融券交易明细  margin_detail_20210112 返回数据 row 行数 = "+str(margin_detail_20210112.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210112_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210113")       ###  更新 记录日期
margin_detail_20210113 = pro.margin_detail(trade_date='20210113', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210113_tscode_list = list() 
margin_detail_20210113_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210113.columns.values.tolist():
    for ts_code_sh in margin_detail_20210113['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210113_tscode_list.append(stock_name)
    margin_detail_20210113_addname_dataframe['cname'] = margin_detail_20210113_tscode_list
for table_name in margin_detail_20210113.columns.values.tolist():
    margin_detail_20210113_addname_dataframe[table_name] = margin_detail_20210113[table_name]
print("融资融券交易明细  margin_detail_20210113 返回数据 row 行数 = "+str(margin_detail_20210113.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210113_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210114")       ###  更新 记录日期
margin_detail_20210114 = pro.margin_detail(trade_date='20210114', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210114_tscode_list = list() 
margin_detail_20210114_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210114.columns.values.tolist():
    for ts_code_sh in margin_detail_20210114['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210114_tscode_list.append(stock_name)
    margin_detail_20210114_addname_dataframe['cname'] = margin_detail_20210114_tscode_list
for table_name in margin_detail_20210114.columns.values.tolist():
    margin_detail_20210114_addname_dataframe[table_name] = margin_detail_20210114[table_name]
print("融资融券交易明细  margin_detail_20210114 返回数据 row 行数 = "+str(margin_detail_20210114.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210114_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210115")       ###  更新 记录日期
margin_detail_20210115 = pro.margin_detail(trade_date='20210115', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210115_tscode_list = list() 
margin_detail_20210115_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210115.columns.values.tolist():
    for ts_code_sh in margin_detail_20210115['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210115_tscode_list.append(stock_name)
    margin_detail_20210115_addname_dataframe['cname'] = margin_detail_20210115_tscode_list
for table_name in margin_detail_20210115.columns.values.tolist():
    margin_detail_20210115_addname_dataframe[table_name] = margin_detail_20210115[table_name]
print("融资融券交易明细  margin_detail_20210115 返回数据 row 行数 = "+str(margin_detail_20210115.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210115_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210118")       ###  更新 记录日期
margin_detail_20210118 = pro.margin_detail(trade_date='20210118', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210118_tscode_list = list() 
margin_detail_20210118_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210118.columns.values.tolist():
    for ts_code_sh in margin_detail_20210118['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210118_tscode_list.append(stock_name)
    margin_detail_20210118_addname_dataframe['cname'] = margin_detail_20210118_tscode_list
for table_name in margin_detail_20210118.columns.values.tolist():
    margin_detail_20210118_addname_dataframe[table_name] = margin_detail_20210118[table_name]
print("融资融券交易明细  margin_detail_20210118 返回数据 row 行数 = "+str(margin_detail_20210118.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210118_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210119")       ###  更新 记录日期
margin_detail_20210119 = pro.margin_detail(trade_date='20210119', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210119_tscode_list = list() 
margin_detail_20210119_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210119.columns.values.tolist():
    for ts_code_sh in margin_detail_20210119['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210119_tscode_list.append(stock_name)
    margin_detail_20210119_addname_dataframe['cname'] = margin_detail_20210119_tscode_list
for table_name in margin_detail_20210119.columns.values.tolist():
    margin_detail_20210119_addname_dataframe[table_name] = margin_detail_20210119[table_name]
print("融资融券交易明细  margin_detail_20210119 返回数据 row 行数 = "+str(margin_detail_20210119.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210119_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210120")       ###  更新 记录日期
margin_detail_20210120 = pro.margin_detail(trade_date='20210120', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210120_tscode_list = list() 
margin_detail_20210120_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210120.columns.values.tolist():
    for ts_code_sh in margin_detail_20210120['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210120_tscode_list.append(stock_name)
    margin_detail_20210120_addname_dataframe['cname'] = margin_detail_20210120_tscode_list
for table_name in margin_detail_20210120.columns.values.tolist():
    margin_detail_20210120_addname_dataframe[table_name] = margin_detail_20210120[table_name]
print("融资融券交易明细  margin_detail_20210120 返回数据 row 行数 = "+str(margin_detail_20210120.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210120_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210121")       ###  更新 记录日期
margin_detail_20210121 = pro.margin_detail(trade_date='20210121', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210121_tscode_list = list() 
margin_detail_20210121_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210121.columns.values.tolist():
    for ts_code_sh in margin_detail_20210121['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210121_tscode_list.append(stock_name)
    margin_detail_20210121_addname_dataframe['cname'] = margin_detail_20210121_tscode_list
for table_name in margin_detail_20210121.columns.values.tolist():
    margin_detail_20210121_addname_dataframe[table_name] = margin_detail_20210121[table_name]
print("融资融券交易明细  margin_detail_20210121 返回数据 row 行数 = "+str(margin_detail_20210121.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210121_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210122")       ###  更新 记录日期
margin_detail_20210122 = pro.margin_detail(trade_date='20210122', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210122_tscode_list = list() 
margin_detail_20210122_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210122.columns.values.tolist():
    for ts_code_sh in margin_detail_20210122['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210122_tscode_list.append(stock_name)
    margin_detail_20210122_addname_dataframe['cname'] = margin_detail_20210122_tscode_list
for table_name in margin_detail_20210122.columns.values.tolist():
    margin_detail_20210122_addname_dataframe[table_name] = margin_detail_20210122[table_name]
print("融资融券交易明细  margin_detail_20210122 返回数据 row 行数 = "+str(margin_detail_20210122.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210122_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210125")       ###  更新 记录日期
margin_detail_20210125 = pro.margin_detail(trade_date='20210125', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210125_tscode_list = list() 
margin_detail_20210125_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210125.columns.values.tolist():
    for ts_code_sh in margin_detail_20210125['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210125_tscode_list.append(stock_name)
    margin_detail_20210125_addname_dataframe['cname'] = margin_detail_20210125_tscode_list
for table_name in margin_detail_20210125.columns.values.tolist():
    margin_detail_20210125_addname_dataframe[table_name] = margin_detail_20210125[table_name]
print("融资融券交易明细  margin_detail_20210125 返回数据 row 行数 = "+str(margin_detail_20210125.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210125_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210126")       ###  更新 记录日期
margin_detail_20210126 = pro.margin_detail(trade_date='20210126', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210126_tscode_list = list() 
margin_detail_20210126_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210126.columns.values.tolist():
    for ts_code_sh in margin_detail_20210126['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210126_tscode_list.append(stock_name)
    margin_detail_20210126_addname_dataframe['cname'] = margin_detail_20210126_tscode_list
for table_name in margin_detail_20210126.columns.values.tolist():
    margin_detail_20210126_addname_dataframe[table_name] = margin_detail_20210126[table_name]
print("融资融券交易明细  margin_detail_20210126 返回数据 row 行数 = "+str(margin_detail_20210126.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210126_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210127")       ###  更新 记录日期
margin_detail_20210127 = pro.margin_detail(trade_date='20210127', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210127_tscode_list = list() 
margin_detail_20210127_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210127.columns.values.tolist():
    for ts_code_sh in margin_detail_20210127['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210127_tscode_list.append(stock_name)
    margin_detail_20210127_addname_dataframe['cname'] = margin_detail_20210127_tscode_list
for table_name in margin_detail_20210127.columns.values.tolist():
    margin_detail_20210127_addname_dataframe[table_name] = margin_detail_20210127[table_name]
print("融资融券交易明细  margin_detail_20210127 返回数据 row 行数 = "+str(margin_detail_20210127.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210127_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210128")       ###  更新 记录日期
margin_detail_20210128 = pro.margin_detail(trade_date='20210128', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210128_tscode_list = list() 
margin_detail_20210128_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210128.columns.values.tolist():
    for ts_code_sh in margin_detail_20210128['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210128_tscode_list.append(stock_name)
    margin_detail_20210128_addname_dataframe['cname'] = margin_detail_20210128_tscode_list
for table_name in margin_detail_20210128.columns.values.tolist():
    margin_detail_20210128_addname_dataframe[table_name] = margin_detail_20210128[table_name]
print("融资融券交易明细  margin_detail_20210128 返回数据 row 行数 = "+str(margin_detail_20210128.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210128_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210129")       ###  更新 记录日期
margin_detail_20210129 = pro.margin_detail(trade_date='20210129', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210129_tscode_list = list() 
margin_detail_20210129_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210129.columns.values.tolist():
    for ts_code_sh in margin_detail_20210129['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210129_tscode_list.append(stock_name)
    margin_detail_20210129_addname_dataframe['cname'] = margin_detail_20210129_tscode_list
for table_name in margin_detail_20210129.columns.values.tolist():
    margin_detail_20210129_addname_dataframe[table_name] = margin_detail_20210129[table_name]
print("融资融券交易明细  margin_detail_20210129 返回数据 row 行数 = "+str(margin_detail_20210129.shape[0]))
margin_detail_2020_1_xlsx_frame=margin_detail_2020_1_xlsx_frame.append(margin_detail_20210129_addname_dataframe,ignore_index=True)
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_1_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'1',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_2_xlsx_frame=pd.DataFrame()
if '2' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_2_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='2' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20210201")       ###  更新 记录日期
margin_detail_20210201 = pro.margin_detail(trade_date='20210201', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210201_tscode_list = list() 
margin_detail_20210201_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210201.columns.values.tolist():
    for ts_code_sh in margin_detail_20210201['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210201_tscode_list.append(stock_name)
    margin_detail_20210201_addname_dataframe['cname'] = margin_detail_20210201_tscode_list
for table_name in margin_detail_20210201.columns.values.tolist():
    margin_detail_20210201_addname_dataframe[table_name] = margin_detail_20210201[table_name]
print("融资融券交易明细  margin_detail_20210201 返回数据 row 行数 = "+str(margin_detail_20210201.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210201_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210202")       ###  更新 记录日期
margin_detail_20210202 = pro.margin_detail(trade_date='20210202', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210202_tscode_list = list() 
margin_detail_20210202_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210202.columns.values.tolist():
    for ts_code_sh in margin_detail_20210202['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210202_tscode_list.append(stock_name)
    margin_detail_20210202_addname_dataframe['cname'] = margin_detail_20210202_tscode_list
for table_name in margin_detail_20210202.columns.values.tolist():
    margin_detail_20210202_addname_dataframe[table_name] = margin_detail_20210202[table_name]
print("融资融券交易明细  margin_detail_20210202 返回数据 row 行数 = "+str(margin_detail_20210202.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210202_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210203")       ###  更新 记录日期
margin_detail_20210203 = pro.margin_detail(trade_date='20210203', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210203_tscode_list = list() 
margin_detail_20210203_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210203.columns.values.tolist():
    for ts_code_sh in margin_detail_20210203['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210203_tscode_list.append(stock_name)
    margin_detail_20210203_addname_dataframe['cname'] = margin_detail_20210203_tscode_list
for table_name in margin_detail_20210203.columns.values.tolist():
    margin_detail_20210203_addname_dataframe[table_name] = margin_detail_20210203[table_name]
print("融资融券交易明细  margin_detail_20210203 返回数据 row 行数 = "+str(margin_detail_20210203.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210203_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210204")       ###  更新 记录日期
margin_detail_20210204 = pro.margin_detail(trade_date='20210204', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210204_tscode_list = list() 
margin_detail_20210204_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210204.columns.values.tolist():
    for ts_code_sh in margin_detail_20210204['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210204_tscode_list.append(stock_name)
    margin_detail_20210204_addname_dataframe['cname'] = margin_detail_20210204_tscode_list
for table_name in margin_detail_20210204.columns.values.tolist():
    margin_detail_20210204_addname_dataframe[table_name] = margin_detail_20210204[table_name]
print("融资融券交易明细  margin_detail_20210204 返回数据 row 行数 = "+str(margin_detail_20210204.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210204_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210205")       ###  更新 记录日期
margin_detail_20210205 = pro.margin_detail(trade_date='20210205', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210205_tscode_list = list() 
margin_detail_20210205_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210205.columns.values.tolist():
    for ts_code_sh in margin_detail_20210205['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210205_tscode_list.append(stock_name)
    margin_detail_20210205_addname_dataframe['cname'] = margin_detail_20210205_tscode_list
for table_name in margin_detail_20210205.columns.values.tolist():
    margin_detail_20210205_addname_dataframe[table_name] = margin_detail_20210205[table_name]
print("融资融券交易明细  margin_detail_20210205 返回数据 row 行数 = "+str(margin_detail_20210205.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210205_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210208")       ###  更新 记录日期
margin_detail_20210208 = pro.margin_detail(trade_date='20210208', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210208_tscode_list = list() 
margin_detail_20210208_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210208.columns.values.tolist():
    for ts_code_sh in margin_detail_20210208['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210208_tscode_list.append(stock_name)
    margin_detail_20210208_addname_dataframe['cname'] = margin_detail_20210208_tscode_list
for table_name in margin_detail_20210208.columns.values.tolist():
    margin_detail_20210208_addname_dataframe[table_name] = margin_detail_20210208[table_name]
print("融资融券交易明细  margin_detail_20210208 返回数据 row 行数 = "+str(margin_detail_20210208.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210208_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210209")       ###  更新 记录日期
margin_detail_20210209 = pro.margin_detail(trade_date='20210209', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210209_tscode_list = list() 
margin_detail_20210209_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210209.columns.values.tolist():
    for ts_code_sh in margin_detail_20210209['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210209_tscode_list.append(stock_name)
    margin_detail_20210209_addname_dataframe['cname'] = margin_detail_20210209_tscode_list
for table_name in margin_detail_20210209.columns.values.tolist():
    margin_detail_20210209_addname_dataframe[table_name] = margin_detail_20210209[table_name]
print("融资融券交易明细  margin_detail_20210209 返回数据 row 行数 = "+str(margin_detail_20210209.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210209_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210210")       ###  更新 记录日期
margin_detail_20210210 = pro.margin_detail(trade_date='20210210', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210210_tscode_list = list() 
margin_detail_20210210_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210210.columns.values.tolist():
    for ts_code_sh in margin_detail_20210210['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210210_tscode_list.append(stock_name)
    margin_detail_20210210_addname_dataframe['cname'] = margin_detail_20210210_tscode_list
for table_name in margin_detail_20210210.columns.values.tolist():
    margin_detail_20210210_addname_dataframe[table_name] = margin_detail_20210210[table_name]
print("融资融券交易明细  margin_detail_20210210 返回数据 row 行数 = "+str(margin_detail_20210210.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210210_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210218")       ###  更新 记录日期
margin_detail_20210218 = pro.margin_detail(trade_date='20210218', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210218_tscode_list = list() 
margin_detail_20210218_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210218.columns.values.tolist():
    for ts_code_sh in margin_detail_20210218['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210218_tscode_list.append(stock_name)
    margin_detail_20210218_addname_dataframe['cname'] = margin_detail_20210218_tscode_list
for table_name in margin_detail_20210218.columns.values.tolist():
    margin_detail_20210218_addname_dataframe[table_name] = margin_detail_20210218[table_name]
print("融资融券交易明细  margin_detail_20210218 返回数据 row 行数 = "+str(margin_detail_20210218.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210218_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210219")       ###  更新 记录日期
margin_detail_20210219 = pro.margin_detail(trade_date='20210219', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210219_tscode_list = list() 
margin_detail_20210219_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210219.columns.values.tolist():
    for ts_code_sh in margin_detail_20210219['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210219_tscode_list.append(stock_name)
    margin_detail_20210219_addname_dataframe['cname'] = margin_detail_20210219_tscode_list
for table_name in margin_detail_20210219.columns.values.tolist():
    margin_detail_20210219_addname_dataframe[table_name] = margin_detail_20210219[table_name]
print("融资融券交易明细  margin_detail_20210219 返回数据 row 行数 = "+str(margin_detail_20210219.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210219_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210222")       ###  更新 记录日期
margin_detail_20210222 = pro.margin_detail(trade_date='20210222', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210222_tscode_list = list() 
margin_detail_20210222_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210222.columns.values.tolist():
    for ts_code_sh in margin_detail_20210222['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210222_tscode_list.append(stock_name)
    margin_detail_20210222_addname_dataframe['cname'] = margin_detail_20210222_tscode_list
for table_name in margin_detail_20210222.columns.values.tolist():
    margin_detail_20210222_addname_dataframe[table_name] = margin_detail_20210222[table_name]
print("融资融券交易明细  margin_detail_20210222 返回数据 row 行数 = "+str(margin_detail_20210222.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210222_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210223")       ###  更新 记录日期
margin_detail_20210223 = pro.margin_detail(trade_date='20210223', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210223_tscode_list = list() 
margin_detail_20210223_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210223.columns.values.tolist():
    for ts_code_sh in margin_detail_20210223['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210223_tscode_list.append(stock_name)
    margin_detail_20210223_addname_dataframe['cname'] = margin_detail_20210223_tscode_list
for table_name in margin_detail_20210223.columns.values.tolist():
    margin_detail_20210223_addname_dataframe[table_name] = margin_detail_20210223[table_name]
print("融资融券交易明细  margin_detail_20210223 返回数据 row 行数 = "+str(margin_detail_20210223.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210223_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210224")       ###  更新 记录日期
margin_detail_20210224 = pro.margin_detail(trade_date='20210224', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210224_tscode_list = list() 
margin_detail_20210224_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210224.columns.values.tolist():
    for ts_code_sh in margin_detail_20210224['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210224_tscode_list.append(stock_name)
    margin_detail_20210224_addname_dataframe['cname'] = margin_detail_20210224_tscode_list
for table_name in margin_detail_20210224.columns.values.tolist():
    margin_detail_20210224_addname_dataframe[table_name] = margin_detail_20210224[table_name]
print("融资融券交易明细  margin_detail_20210224 返回数据 row 行数 = "+str(margin_detail_20210224.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210224_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210225")       ###  更新 记录日期
margin_detail_20210225 = pro.margin_detail(trade_date='20210225', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210225_tscode_list = list() 
margin_detail_20210225_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210225.columns.values.tolist():
    for ts_code_sh in margin_detail_20210225['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210225_tscode_list.append(stock_name)
    margin_detail_20210225_addname_dataframe['cname'] = margin_detail_20210225_tscode_list
for table_name in margin_detail_20210225.columns.values.tolist():
    margin_detail_20210225_addname_dataframe[table_name] = margin_detail_20210225[table_name]
print("融资融券交易明细  margin_detail_20210225 返回数据 row 行数 = "+str(margin_detail_20210225.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210225_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210226")       ###  更新 记录日期
margin_detail_20210226 = pro.margin_detail(trade_date='20210226', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210226_tscode_list = list() 
margin_detail_20210226_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210226.columns.values.tolist():
    for ts_code_sh in margin_detail_20210226['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210226_tscode_list.append(stock_name)
    margin_detail_20210226_addname_dataframe['cname'] = margin_detail_20210226_tscode_list
for table_name in margin_detail_20210226.columns.values.tolist():
    margin_detail_20210226_addname_dataframe[table_name] = margin_detail_20210226[table_name]
print("融资融券交易明细  margin_detail_20210226 返回数据 row 行数 = "+str(margin_detail_20210226.shape[0]))
margin_detail_2020_2_xlsx_frame=margin_detail_2020_2_xlsx_frame.append(margin_detail_20210226_addname_dataframe,ignore_index=True)
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_2_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'2',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_3_xlsx_frame=pd.DataFrame()
if '3' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_3_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='3' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20210301")       ###  更新 记录日期
margin_detail_20210301 = pro.margin_detail(trade_date='20210301', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210301_tscode_list = list() 
margin_detail_20210301_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210301.columns.values.tolist():
    for ts_code_sh in margin_detail_20210301['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210301_tscode_list.append(stock_name)
    margin_detail_20210301_addname_dataframe['cname'] = margin_detail_20210301_tscode_list
for table_name in margin_detail_20210301.columns.values.tolist():
    margin_detail_20210301_addname_dataframe[table_name] = margin_detail_20210301[table_name]
print("融资融券交易明细  margin_detail_20210301 返回数据 row 行数 = "+str(margin_detail_20210301.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210301_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210302")       ###  更新 记录日期
margin_detail_20210302 = pro.margin_detail(trade_date='20210302', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210302_tscode_list = list() 
margin_detail_20210302_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210302.columns.values.tolist():
    for ts_code_sh in margin_detail_20210302['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210302_tscode_list.append(stock_name)
    margin_detail_20210302_addname_dataframe['cname'] = margin_detail_20210302_tscode_list
for table_name in margin_detail_20210302.columns.values.tolist():
    margin_detail_20210302_addname_dataframe[table_name] = margin_detail_20210302[table_name]
print("融资融券交易明细  margin_detail_20210302 返回数据 row 行数 = "+str(margin_detail_20210302.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210302_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210303")       ###  更新 记录日期
margin_detail_20210303 = pro.margin_detail(trade_date='20210303', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210303_tscode_list = list() 
margin_detail_20210303_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210303.columns.values.tolist():
    for ts_code_sh in margin_detail_20210303['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210303_tscode_list.append(stock_name)
    margin_detail_20210303_addname_dataframe['cname'] = margin_detail_20210303_tscode_list
for table_name in margin_detail_20210303.columns.values.tolist():
    margin_detail_20210303_addname_dataframe[table_name] = margin_detail_20210303[table_name]
print("融资融券交易明细  margin_detail_20210303 返回数据 row 行数 = "+str(margin_detail_20210303.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210303_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210304")       ###  更新 记录日期
margin_detail_20210304 = pro.margin_detail(trade_date='20210304', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210304_tscode_list = list() 
margin_detail_20210304_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210304.columns.values.tolist():
    for ts_code_sh in margin_detail_20210304['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210304_tscode_list.append(stock_name)
    margin_detail_20210304_addname_dataframe['cname'] = margin_detail_20210304_tscode_list
for table_name in margin_detail_20210304.columns.values.tolist():
    margin_detail_20210304_addname_dataframe[table_name] = margin_detail_20210304[table_name]
print("融资融券交易明细  margin_detail_20210304 返回数据 row 行数 = "+str(margin_detail_20210304.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210304_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210305")       ###  更新 记录日期
margin_detail_20210305 = pro.margin_detail(trade_date='20210305', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210305_tscode_list = list() 
margin_detail_20210305_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210305.columns.values.tolist():
    for ts_code_sh in margin_detail_20210305['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210305_tscode_list.append(stock_name)
    margin_detail_20210305_addname_dataframe['cname'] = margin_detail_20210305_tscode_list
for table_name in margin_detail_20210305.columns.values.tolist():
    margin_detail_20210305_addname_dataframe[table_name] = margin_detail_20210305[table_name]
print("融资融券交易明细  margin_detail_20210305 返回数据 row 行数 = "+str(margin_detail_20210305.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210305_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210308")       ###  更新 记录日期
margin_detail_20210308 = pro.margin_detail(trade_date='20210308', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210308_tscode_list = list() 
margin_detail_20210308_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210308.columns.values.tolist():
    for ts_code_sh in margin_detail_20210308['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210308_tscode_list.append(stock_name)
    margin_detail_20210308_addname_dataframe['cname'] = margin_detail_20210308_tscode_list
for table_name in margin_detail_20210308.columns.values.tolist():
    margin_detail_20210308_addname_dataframe[table_name] = margin_detail_20210308[table_name]
print("融资融券交易明细  margin_detail_20210308 返回数据 row 行数 = "+str(margin_detail_20210308.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210308_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210309")       ###  更新 记录日期
margin_detail_20210309 = pro.margin_detail(trade_date='20210309', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210309_tscode_list = list() 
margin_detail_20210309_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210309.columns.values.tolist():
    for ts_code_sh in margin_detail_20210309['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210309_tscode_list.append(stock_name)
    margin_detail_20210309_addname_dataframe['cname'] = margin_detail_20210309_tscode_list
for table_name in margin_detail_20210309.columns.values.tolist():
    margin_detail_20210309_addname_dataframe[table_name] = margin_detail_20210309[table_name]
print("融资融券交易明细  margin_detail_20210309 返回数据 row 行数 = "+str(margin_detail_20210309.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210309_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210310")       ###  更新 记录日期
margin_detail_20210310 = pro.margin_detail(trade_date='20210310', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210310_tscode_list = list() 
margin_detail_20210310_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210310.columns.values.tolist():
    for ts_code_sh in margin_detail_20210310['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210310_tscode_list.append(stock_name)
    margin_detail_20210310_addname_dataframe['cname'] = margin_detail_20210310_tscode_list
for table_name in margin_detail_20210310.columns.values.tolist():
    margin_detail_20210310_addname_dataframe[table_name] = margin_detail_20210310[table_name]
print("融资融券交易明细  margin_detail_20210310 返回数据 row 行数 = "+str(margin_detail_20210310.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210310_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210311")       ###  更新 记录日期
margin_detail_20210311 = pro.margin_detail(trade_date='20210311', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210311_tscode_list = list() 
margin_detail_20210311_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210311.columns.values.tolist():
    for ts_code_sh in margin_detail_20210311['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210311_tscode_list.append(stock_name)
    margin_detail_20210311_addname_dataframe['cname'] = margin_detail_20210311_tscode_list
for table_name in margin_detail_20210311.columns.values.tolist():
    margin_detail_20210311_addname_dataframe[table_name] = margin_detail_20210311[table_name]
print("融资融券交易明细  margin_detail_20210311 返回数据 row 行数 = "+str(margin_detail_20210311.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210311_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210312")       ###  更新 记录日期
margin_detail_20210312 = pro.margin_detail(trade_date='20210312', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210312_tscode_list = list() 
margin_detail_20210312_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210312.columns.values.tolist():
    for ts_code_sh in margin_detail_20210312['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210312_tscode_list.append(stock_name)
    margin_detail_20210312_addname_dataframe['cname'] = margin_detail_20210312_tscode_list
for table_name in margin_detail_20210312.columns.values.tolist():
    margin_detail_20210312_addname_dataframe[table_name] = margin_detail_20210312[table_name]
print("融资融券交易明细  margin_detail_20210312 返回数据 row 行数 = "+str(margin_detail_20210312.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210312_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210315")       ###  更新 记录日期
margin_detail_20210315 = pro.margin_detail(trade_date='20210315', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210315_tscode_list = list() 
margin_detail_20210315_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210315.columns.values.tolist():
    for ts_code_sh in margin_detail_20210315['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210315_tscode_list.append(stock_name)
    margin_detail_20210315_addname_dataframe['cname'] = margin_detail_20210315_tscode_list
for table_name in margin_detail_20210315.columns.values.tolist():
    margin_detail_20210315_addname_dataframe[table_name] = margin_detail_20210315[table_name]
print("融资融券交易明细  margin_detail_20210315 返回数据 row 行数 = "+str(margin_detail_20210315.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210315_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210316")       ###  更新 记录日期
margin_detail_20210316 = pro.margin_detail(trade_date='20210316', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210316_tscode_list = list() 
margin_detail_20210316_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210316.columns.values.tolist():
    for ts_code_sh in margin_detail_20210316['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210316_tscode_list.append(stock_name)
    margin_detail_20210316_addname_dataframe['cname'] = margin_detail_20210316_tscode_list
for table_name in margin_detail_20210316.columns.values.tolist():
    margin_detail_20210316_addname_dataframe[table_name] = margin_detail_20210316[table_name]
print("融资融券交易明细  margin_detail_20210316 返回数据 row 行数 = "+str(margin_detail_20210316.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210316_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210317")       ###  更新 记录日期
margin_detail_20210317 = pro.margin_detail(trade_date='20210317', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210317_tscode_list = list() 
margin_detail_20210317_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210317.columns.values.tolist():
    for ts_code_sh in margin_detail_20210317['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210317_tscode_list.append(stock_name)
    margin_detail_20210317_addname_dataframe['cname'] = margin_detail_20210317_tscode_list
for table_name in margin_detail_20210317.columns.values.tolist():
    margin_detail_20210317_addname_dataframe[table_name] = margin_detail_20210317[table_name]
print("融资融券交易明细  margin_detail_20210317 返回数据 row 行数 = "+str(margin_detail_20210317.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210317_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210318")       ###  更新 记录日期
margin_detail_20210318 = pro.margin_detail(trade_date='20210318', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210318_tscode_list = list() 
margin_detail_20210318_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210318.columns.values.tolist():
    for ts_code_sh in margin_detail_20210318['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210318_tscode_list.append(stock_name)
    margin_detail_20210318_addname_dataframe['cname'] = margin_detail_20210318_tscode_list
for table_name in margin_detail_20210318.columns.values.tolist():
    margin_detail_20210318_addname_dataframe[table_name] = margin_detail_20210318[table_name]
print("融资融券交易明细  margin_detail_20210318 返回数据 row 行数 = "+str(margin_detail_20210318.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210318_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210319")       ###  更新 记录日期
margin_detail_20210319 = pro.margin_detail(trade_date='20210319', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210319_tscode_list = list() 
margin_detail_20210319_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210319.columns.values.tolist():
    for ts_code_sh in margin_detail_20210319['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210319_tscode_list.append(stock_name)
    margin_detail_20210319_addname_dataframe['cname'] = margin_detail_20210319_tscode_list
for table_name in margin_detail_20210319.columns.values.tolist():
    margin_detail_20210319_addname_dataframe[table_name] = margin_detail_20210319[table_name]
print("融资融券交易明细  margin_detail_20210319 返回数据 row 行数 = "+str(margin_detail_20210319.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210319_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210322")       ###  更新 记录日期
margin_detail_20210322 = pro.margin_detail(trade_date='20210322', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210322_tscode_list = list() 
margin_detail_20210322_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210322.columns.values.tolist():
    for ts_code_sh in margin_detail_20210322['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210322_tscode_list.append(stock_name)
    margin_detail_20210322_addname_dataframe['cname'] = margin_detail_20210322_tscode_list
for table_name in margin_detail_20210322.columns.values.tolist():
    margin_detail_20210322_addname_dataframe[table_name] = margin_detail_20210322[table_name]
print("融资融券交易明细  margin_detail_20210322 返回数据 row 行数 = "+str(margin_detail_20210322.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210322_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210323")       ###  更新 记录日期
margin_detail_20210323 = pro.margin_detail(trade_date='20210323', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210323_tscode_list = list() 
margin_detail_20210323_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210323.columns.values.tolist():
    for ts_code_sh in margin_detail_20210323['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210323_tscode_list.append(stock_name)
    margin_detail_20210323_addname_dataframe['cname'] = margin_detail_20210323_tscode_list
for table_name in margin_detail_20210323.columns.values.tolist():
    margin_detail_20210323_addname_dataframe[table_name] = margin_detail_20210323[table_name]
print("融资融券交易明细  margin_detail_20210323 返回数据 row 行数 = "+str(margin_detail_20210323.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210323_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210324")       ###  更新 记录日期
margin_detail_20210324 = pro.margin_detail(trade_date='20210324', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210324_tscode_list = list() 
margin_detail_20210324_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210324.columns.values.tolist():
    for ts_code_sh in margin_detail_20210324['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210324_tscode_list.append(stock_name)
    margin_detail_20210324_addname_dataframe['cname'] = margin_detail_20210324_tscode_list
for table_name in margin_detail_20210324.columns.values.tolist():
    margin_detail_20210324_addname_dataframe[table_name] = margin_detail_20210324[table_name]
print("融资融券交易明细  margin_detail_20210324 返回数据 row 行数 = "+str(margin_detail_20210324.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210324_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210325")       ###  更新 记录日期
margin_detail_20210325 = pro.margin_detail(trade_date='20210325', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210325_tscode_list = list() 
margin_detail_20210325_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210325.columns.values.tolist():
    for ts_code_sh in margin_detail_20210325['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210325_tscode_list.append(stock_name)
    margin_detail_20210325_addname_dataframe['cname'] = margin_detail_20210325_tscode_list
for table_name in margin_detail_20210325.columns.values.tolist():
    margin_detail_20210325_addname_dataframe[table_name] = margin_detail_20210325[table_name]
print("融资融券交易明细  margin_detail_20210325 返回数据 row 行数 = "+str(margin_detail_20210325.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210325_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210326")       ###  更新 记录日期
margin_detail_20210326 = pro.margin_detail(trade_date='20210326', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210326_tscode_list = list() 
margin_detail_20210326_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210326.columns.values.tolist():
    for ts_code_sh in margin_detail_20210326['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210326_tscode_list.append(stock_name)
    margin_detail_20210326_addname_dataframe['cname'] = margin_detail_20210326_tscode_list
for table_name in margin_detail_20210326.columns.values.tolist():
    margin_detail_20210326_addname_dataframe[table_name] = margin_detail_20210326[table_name]
print("融资融券交易明细  margin_detail_20210326 返回数据 row 行数 = "+str(margin_detail_20210326.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210326_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210329")       ###  更新 记录日期
margin_detail_20210329 = pro.margin_detail(trade_date='20210329', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210329_tscode_list = list() 
margin_detail_20210329_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210329.columns.values.tolist():
    for ts_code_sh in margin_detail_20210329['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210329_tscode_list.append(stock_name)
    margin_detail_20210329_addname_dataframe['cname'] = margin_detail_20210329_tscode_list
for table_name in margin_detail_20210329.columns.values.tolist():
    margin_detail_20210329_addname_dataframe[table_name] = margin_detail_20210329[table_name]
print("融资融券交易明细  margin_detail_20210329 返回数据 row 行数 = "+str(margin_detail_20210329.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210329_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210330")       ###  更新 记录日期
margin_detail_20210330 = pro.margin_detail(trade_date='20210330', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210330_tscode_list = list() 
margin_detail_20210330_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210330.columns.values.tolist():
    for ts_code_sh in margin_detail_20210330['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210330_tscode_list.append(stock_name)
    margin_detail_20210330_addname_dataframe['cname'] = margin_detail_20210330_tscode_list
for table_name in margin_detail_20210330.columns.values.tolist():
    margin_detail_20210330_addname_dataframe[table_name] = margin_detail_20210330[table_name]
print("融资融券交易明细  margin_detail_20210330 返回数据 row 行数 = "+str(margin_detail_20210330.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210330_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210331")       ###  更新 记录日期
margin_detail_20210331 = pro.margin_detail(trade_date='20210331', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210331_tscode_list = list() 
margin_detail_20210331_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210331.columns.values.tolist():
    for ts_code_sh in margin_detail_20210331['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210331_tscode_list.append(stock_name)
    margin_detail_20210331_addname_dataframe['cname'] = margin_detail_20210331_tscode_list
for table_name in margin_detail_20210331.columns.values.tolist():
    margin_detail_20210331_addname_dataframe[table_name] = margin_detail_20210331[table_name]
print("融资融券交易明细  margin_detail_20210331 返回数据 row 行数 = "+str(margin_detail_20210331.shape[0]))
margin_detail_2020_3_xlsx_frame=margin_detail_2020_3_xlsx_frame.append(margin_detail_20210331_addname_dataframe,ignore_index=True)
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_3_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'3',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_4_xlsx_frame=pd.DataFrame()
if '4' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_4_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='4' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20210401")       ###  更新 记录日期
margin_detail_20210401 = pro.margin_detail(trade_date='20210401', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210401_tscode_list = list() 
margin_detail_20210401_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210401.columns.values.tolist():
    for ts_code_sh in margin_detail_20210401['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210401_tscode_list.append(stock_name)
    margin_detail_20210401_addname_dataframe['cname'] = margin_detail_20210401_tscode_list
for table_name in margin_detail_20210401.columns.values.tolist():
    margin_detail_20210401_addname_dataframe[table_name] = margin_detail_20210401[table_name]
print("融资融券交易明细  margin_detail_20210401 返回数据 row 行数 = "+str(margin_detail_20210401.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210401_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210402")       ###  更新 记录日期
margin_detail_20210402 = pro.margin_detail(trade_date='20210402', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210402_tscode_list = list() 
margin_detail_20210402_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210402.columns.values.tolist():
    for ts_code_sh in margin_detail_20210402['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210402_tscode_list.append(stock_name)
    margin_detail_20210402_addname_dataframe['cname'] = margin_detail_20210402_tscode_list
for table_name in margin_detail_20210402.columns.values.tolist():
    margin_detail_20210402_addname_dataframe[table_name] = margin_detail_20210402[table_name]
print("融资融券交易明细  margin_detail_20210402 返回数据 row 行数 = "+str(margin_detail_20210402.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210402_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210406")       ###  更新 记录日期
margin_detail_20210406 = pro.margin_detail(trade_date='20210406', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210406_tscode_list = list() 
margin_detail_20210406_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210406.columns.values.tolist():
    for ts_code_sh in margin_detail_20210406['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210406_tscode_list.append(stock_name)
    margin_detail_20210406_addname_dataframe['cname'] = margin_detail_20210406_tscode_list
for table_name in margin_detail_20210406.columns.values.tolist():
    margin_detail_20210406_addname_dataframe[table_name] = margin_detail_20210406[table_name]
print("融资融券交易明细  margin_detail_20210406 返回数据 row 行数 = "+str(margin_detail_20210406.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210406_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210407")       ###  更新 记录日期
margin_detail_20210407 = pro.margin_detail(trade_date='20210407', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210407_tscode_list = list() 
margin_detail_20210407_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210407.columns.values.tolist():
    for ts_code_sh in margin_detail_20210407['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210407_tscode_list.append(stock_name)
    margin_detail_20210407_addname_dataframe['cname'] = margin_detail_20210407_tscode_list
for table_name in margin_detail_20210407.columns.values.tolist():
    margin_detail_20210407_addname_dataframe[table_name] = margin_detail_20210407[table_name]
print("融资融券交易明细  margin_detail_20210407 返回数据 row 行数 = "+str(margin_detail_20210407.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210407_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210408")       ###  更新 记录日期
margin_detail_20210408 = pro.margin_detail(trade_date='20210408', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210408_tscode_list = list() 
margin_detail_20210408_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210408.columns.values.tolist():
    for ts_code_sh in margin_detail_20210408['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210408_tscode_list.append(stock_name)
    margin_detail_20210408_addname_dataframe['cname'] = margin_detail_20210408_tscode_list
for table_name in margin_detail_20210408.columns.values.tolist():
    margin_detail_20210408_addname_dataframe[table_name] = margin_detail_20210408[table_name]
print("融资融券交易明细  margin_detail_20210408 返回数据 row 行数 = "+str(margin_detail_20210408.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210408_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210409")       ###  更新 记录日期
margin_detail_20210409 = pro.margin_detail(trade_date='20210409', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210409_tscode_list = list() 
margin_detail_20210409_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210409.columns.values.tolist():
    for ts_code_sh in margin_detail_20210409['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210409_tscode_list.append(stock_name)
    margin_detail_20210409_addname_dataframe['cname'] = margin_detail_20210409_tscode_list
for table_name in margin_detail_20210409.columns.values.tolist():
    margin_detail_20210409_addname_dataframe[table_name] = margin_detail_20210409[table_name]
print("融资融券交易明细  margin_detail_20210409 返回数据 row 行数 = "+str(margin_detail_20210409.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210409_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210412")       ###  更新 记录日期
margin_detail_20210412 = pro.margin_detail(trade_date='20210412', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210412_tscode_list = list() 
margin_detail_20210412_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210412.columns.values.tolist():
    for ts_code_sh in margin_detail_20210412['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210412_tscode_list.append(stock_name)
    margin_detail_20210412_addname_dataframe['cname'] = margin_detail_20210412_tscode_list
for table_name in margin_detail_20210412.columns.values.tolist():
    margin_detail_20210412_addname_dataframe[table_name] = margin_detail_20210412[table_name]
print("融资融券交易明细  margin_detail_20210412 返回数据 row 行数 = "+str(margin_detail_20210412.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210412_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210413")       ###  更新 记录日期
margin_detail_20210413 = pro.margin_detail(trade_date='20210413', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210413_tscode_list = list() 
margin_detail_20210413_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210413.columns.values.tolist():
    for ts_code_sh in margin_detail_20210413['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210413_tscode_list.append(stock_name)
    margin_detail_20210413_addname_dataframe['cname'] = margin_detail_20210413_tscode_list
for table_name in margin_detail_20210413.columns.values.tolist():
    margin_detail_20210413_addname_dataframe[table_name] = margin_detail_20210413[table_name]
print("融资融券交易明细  margin_detail_20210413 返回数据 row 行数 = "+str(margin_detail_20210413.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210413_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210414")       ###  更新 记录日期
margin_detail_20210414 = pro.margin_detail(trade_date='20210414', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210414_tscode_list = list() 
margin_detail_20210414_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210414.columns.values.tolist():
    for ts_code_sh in margin_detail_20210414['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210414_tscode_list.append(stock_name)
    margin_detail_20210414_addname_dataframe['cname'] = margin_detail_20210414_tscode_list
for table_name in margin_detail_20210414.columns.values.tolist():
    margin_detail_20210414_addname_dataframe[table_name] = margin_detail_20210414[table_name]
print("融资融券交易明细  margin_detail_20210414 返回数据 row 行数 = "+str(margin_detail_20210414.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210414_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210415")       ###  更新 记录日期
margin_detail_20210415 = pro.margin_detail(trade_date='20210415', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210415_tscode_list = list() 
margin_detail_20210415_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210415.columns.values.tolist():
    for ts_code_sh in margin_detail_20210415['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210415_tscode_list.append(stock_name)
    margin_detail_20210415_addname_dataframe['cname'] = margin_detail_20210415_tscode_list
for table_name in margin_detail_20210415.columns.values.tolist():
    margin_detail_20210415_addname_dataframe[table_name] = margin_detail_20210415[table_name]
print("融资融券交易明细  margin_detail_20210415 返回数据 row 行数 = "+str(margin_detail_20210415.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210415_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210416")       ###  更新 记录日期
margin_detail_20210416 = pro.margin_detail(trade_date='20210416', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210416_tscode_list = list() 
margin_detail_20210416_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210416.columns.values.tolist():
    for ts_code_sh in margin_detail_20210416['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210416_tscode_list.append(stock_name)
    margin_detail_20210416_addname_dataframe['cname'] = margin_detail_20210416_tscode_list
for table_name in margin_detail_20210416.columns.values.tolist():
    margin_detail_20210416_addname_dataframe[table_name] = margin_detail_20210416[table_name]
print("融资融券交易明细  margin_detail_20210416 返回数据 row 行数 = "+str(margin_detail_20210416.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210416_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210419")       ###  更新 记录日期
margin_detail_20210419 = pro.margin_detail(trade_date='20210419', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210419_tscode_list = list() 
margin_detail_20210419_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210419.columns.values.tolist():
    for ts_code_sh in margin_detail_20210419['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210419_tscode_list.append(stock_name)
    margin_detail_20210419_addname_dataframe['cname'] = margin_detail_20210419_tscode_list
for table_name in margin_detail_20210419.columns.values.tolist():
    margin_detail_20210419_addname_dataframe[table_name] = margin_detail_20210419[table_name]
print("融资融券交易明细  margin_detail_20210419 返回数据 row 行数 = "+str(margin_detail_20210419.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210419_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210420")       ###  更新 记录日期
margin_detail_20210420 = pro.margin_detail(trade_date='20210420', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210420_tscode_list = list() 
margin_detail_20210420_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210420.columns.values.tolist():
    for ts_code_sh in margin_detail_20210420['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210420_tscode_list.append(stock_name)
    margin_detail_20210420_addname_dataframe['cname'] = margin_detail_20210420_tscode_list
for table_name in margin_detail_20210420.columns.values.tolist():
    margin_detail_20210420_addname_dataframe[table_name] = margin_detail_20210420[table_name]
print("融资融券交易明细  margin_detail_20210420 返回数据 row 行数 = "+str(margin_detail_20210420.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210420_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210421")       ###  更新 记录日期
margin_detail_20210421 = pro.margin_detail(trade_date='20210421', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210421_tscode_list = list() 
margin_detail_20210421_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210421.columns.values.tolist():
    for ts_code_sh in margin_detail_20210421['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210421_tscode_list.append(stock_name)
    margin_detail_20210421_addname_dataframe['cname'] = margin_detail_20210421_tscode_list
for table_name in margin_detail_20210421.columns.values.tolist():
    margin_detail_20210421_addname_dataframe[table_name] = margin_detail_20210421[table_name]
print("融资融券交易明细  margin_detail_20210421 返回数据 row 行数 = "+str(margin_detail_20210421.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210421_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210422")       ###  更新 记录日期
margin_detail_20210422 = pro.margin_detail(trade_date='20210422', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210422_tscode_list = list() 
margin_detail_20210422_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210422.columns.values.tolist():
    for ts_code_sh in margin_detail_20210422['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210422_tscode_list.append(stock_name)
    margin_detail_20210422_addname_dataframe['cname'] = margin_detail_20210422_tscode_list
for table_name in margin_detail_20210422.columns.values.tolist():
    margin_detail_20210422_addname_dataframe[table_name] = margin_detail_20210422[table_name]
print("融资融券交易明细  margin_detail_20210422 返回数据 row 行数 = "+str(margin_detail_20210422.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210422_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210423")       ###  更新 记录日期
margin_detail_20210423 = pro.margin_detail(trade_date='20210423', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210423_tscode_list = list() 
margin_detail_20210423_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210423.columns.values.tolist():
    for ts_code_sh in margin_detail_20210423['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210423_tscode_list.append(stock_name)
    margin_detail_20210423_addname_dataframe['cname'] = margin_detail_20210423_tscode_list
for table_name in margin_detail_20210423.columns.values.tolist():
    margin_detail_20210423_addname_dataframe[table_name] = margin_detail_20210423[table_name]
print("融资融券交易明细  margin_detail_20210423 返回数据 row 行数 = "+str(margin_detail_20210423.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210423_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210426")       ###  更新 记录日期
margin_detail_20210426 = pro.margin_detail(trade_date='20210426', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210426_tscode_list = list() 
margin_detail_20210426_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210426.columns.values.tolist():
    for ts_code_sh in margin_detail_20210426['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210426_tscode_list.append(stock_name)
    margin_detail_20210426_addname_dataframe['cname'] = margin_detail_20210426_tscode_list
for table_name in margin_detail_20210426.columns.values.tolist():
    margin_detail_20210426_addname_dataframe[table_name] = margin_detail_20210426[table_name]
print("融资融券交易明细  margin_detail_20210426 返回数据 row 行数 = "+str(margin_detail_20210426.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210426_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210427")       ###  更新 记录日期
margin_detail_20210427 = pro.margin_detail(trade_date='20210427', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210427_tscode_list = list() 
margin_detail_20210427_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210427.columns.values.tolist():
    for ts_code_sh in margin_detail_20210427['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210427_tscode_list.append(stock_name)
    margin_detail_20210427_addname_dataframe['cname'] = margin_detail_20210427_tscode_list
for table_name in margin_detail_20210427.columns.values.tolist():
    margin_detail_20210427_addname_dataframe[table_name] = margin_detail_20210427[table_name]
print("融资融券交易明细  margin_detail_20210427 返回数据 row 行数 = "+str(margin_detail_20210427.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210427_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210428")       ###  更新 记录日期
margin_detail_20210428 = pro.margin_detail(trade_date='20210428', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210428_tscode_list = list() 
margin_detail_20210428_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210428.columns.values.tolist():
    for ts_code_sh in margin_detail_20210428['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210428_tscode_list.append(stock_name)
    margin_detail_20210428_addname_dataframe['cname'] = margin_detail_20210428_tscode_list
for table_name in margin_detail_20210428.columns.values.tolist():
    margin_detail_20210428_addname_dataframe[table_name] = margin_detail_20210428[table_name]
print("融资融券交易明细  margin_detail_20210428 返回数据 row 行数 = "+str(margin_detail_20210428.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210428_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210429")       ###  更新 记录日期
margin_detail_20210429 = pro.margin_detail(trade_date='20210429', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210429_tscode_list = list() 
margin_detail_20210429_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210429.columns.values.tolist():
    for ts_code_sh in margin_detail_20210429['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210429_tscode_list.append(stock_name)
    margin_detail_20210429_addname_dataframe['cname'] = margin_detail_20210429_tscode_list
for table_name in margin_detail_20210429.columns.values.tolist():
    margin_detail_20210429_addname_dataframe[table_name] = margin_detail_20210429[table_name]
print("融资融券交易明细  margin_detail_20210429 返回数据 row 行数 = "+str(margin_detail_20210429.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210429_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210430")       ###  更新 记录日期
margin_detail_20210430 = pro.margin_detail(trade_date='20210430', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210430_tscode_list = list() 
margin_detail_20210430_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210430.columns.values.tolist():
    for ts_code_sh in margin_detail_20210430['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210430_tscode_list.append(stock_name)
    margin_detail_20210430_addname_dataframe['cname'] = margin_detail_20210430_tscode_list
for table_name in margin_detail_20210430.columns.values.tolist():
    margin_detail_20210430_addname_dataframe[table_name] = margin_detail_20210430[table_name]
print("融资融券交易明细  margin_detail_20210430 返回数据 row 行数 = "+str(margin_detail_20210430.shape[0]))
margin_detail_2020_4_xlsx_frame=margin_detail_2020_4_xlsx_frame.append(margin_detail_20210430_addname_dataframe,ignore_index=True)
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_4_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'4',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_5_xlsx_frame=pd.DataFrame()
if '5' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_5_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='5' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20210506")       ###  更新 记录日期
margin_detail_20210506 = pro.margin_detail(trade_date='20210506', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210506_tscode_list = list() 
margin_detail_20210506_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210506.columns.values.tolist():
    for ts_code_sh in margin_detail_20210506['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210506_tscode_list.append(stock_name)
    margin_detail_20210506_addname_dataframe['cname'] = margin_detail_20210506_tscode_list
for table_name in margin_detail_20210506.columns.values.tolist():
    margin_detail_20210506_addname_dataframe[table_name] = margin_detail_20210506[table_name]
print("融资融券交易明细  margin_detail_20210506 返回数据 row 行数 = "+str(margin_detail_20210506.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210506_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210507")       ###  更新 记录日期
margin_detail_20210507 = pro.margin_detail(trade_date='20210507', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210507_tscode_list = list() 
margin_detail_20210507_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210507.columns.values.tolist():
    for ts_code_sh in margin_detail_20210507['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210507_tscode_list.append(stock_name)
    margin_detail_20210507_addname_dataframe['cname'] = margin_detail_20210507_tscode_list
for table_name in margin_detail_20210507.columns.values.tolist():
    margin_detail_20210507_addname_dataframe[table_name] = margin_detail_20210507[table_name]
print("融资融券交易明细  margin_detail_20210507 返回数据 row 行数 = "+str(margin_detail_20210507.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210507_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210510")       ###  更新 记录日期
margin_detail_20210510 = pro.margin_detail(trade_date='20210510', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210510_tscode_list = list() 
margin_detail_20210510_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210510.columns.values.tolist():
    for ts_code_sh in margin_detail_20210510['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210510_tscode_list.append(stock_name)
    margin_detail_20210510_addname_dataframe['cname'] = margin_detail_20210510_tscode_list
for table_name in margin_detail_20210510.columns.values.tolist():
    margin_detail_20210510_addname_dataframe[table_name] = margin_detail_20210510[table_name]
print("融资融券交易明细  margin_detail_20210510 返回数据 row 行数 = "+str(margin_detail_20210510.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210510_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210511")       ###  更新 记录日期
margin_detail_20210511 = pro.margin_detail(trade_date='20210511', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210511_tscode_list = list() 
margin_detail_20210511_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210511.columns.values.tolist():
    for ts_code_sh in margin_detail_20210511['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210511_tscode_list.append(stock_name)
    margin_detail_20210511_addname_dataframe['cname'] = margin_detail_20210511_tscode_list
for table_name in margin_detail_20210511.columns.values.tolist():
    margin_detail_20210511_addname_dataframe[table_name] = margin_detail_20210511[table_name]
print("融资融券交易明细  margin_detail_20210511 返回数据 row 行数 = "+str(margin_detail_20210511.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210511_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210512")       ###  更新 记录日期
margin_detail_20210512 = pro.margin_detail(trade_date='20210512', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210512_tscode_list = list() 
margin_detail_20210512_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210512.columns.values.tolist():
    for ts_code_sh in margin_detail_20210512['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210512_tscode_list.append(stock_name)
    margin_detail_20210512_addname_dataframe['cname'] = margin_detail_20210512_tscode_list
for table_name in margin_detail_20210512.columns.values.tolist():
    margin_detail_20210512_addname_dataframe[table_name] = margin_detail_20210512[table_name]
print("融资融券交易明细  margin_detail_20210512 返回数据 row 行数 = "+str(margin_detail_20210512.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210512_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210513")       ###  更新 记录日期
margin_detail_20210513 = pro.margin_detail(trade_date='20210513', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210513_tscode_list = list() 
margin_detail_20210513_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210513.columns.values.tolist():
    for ts_code_sh in margin_detail_20210513['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210513_tscode_list.append(stock_name)
    margin_detail_20210513_addname_dataframe['cname'] = margin_detail_20210513_tscode_list
for table_name in margin_detail_20210513.columns.values.tolist():
    margin_detail_20210513_addname_dataframe[table_name] = margin_detail_20210513[table_name]
print("融资融券交易明细  margin_detail_20210513 返回数据 row 行数 = "+str(margin_detail_20210513.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210513_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210514")       ###  更新 记录日期
margin_detail_20210514 = pro.margin_detail(trade_date='20210514', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210514_tscode_list = list() 
margin_detail_20210514_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210514.columns.values.tolist():
    for ts_code_sh in margin_detail_20210514['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210514_tscode_list.append(stock_name)
    margin_detail_20210514_addname_dataframe['cname'] = margin_detail_20210514_tscode_list
for table_name in margin_detail_20210514.columns.values.tolist():
    margin_detail_20210514_addname_dataframe[table_name] = margin_detail_20210514[table_name]
print("融资融券交易明细  margin_detail_20210514 返回数据 row 行数 = "+str(margin_detail_20210514.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210514_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210517")       ###  更新 记录日期
margin_detail_20210517 = pro.margin_detail(trade_date='20210517', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210517_tscode_list = list() 
margin_detail_20210517_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210517.columns.values.tolist():
    for ts_code_sh in margin_detail_20210517['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210517_tscode_list.append(stock_name)
    margin_detail_20210517_addname_dataframe['cname'] = margin_detail_20210517_tscode_list
for table_name in margin_detail_20210517.columns.values.tolist():
    margin_detail_20210517_addname_dataframe[table_name] = margin_detail_20210517[table_name]
print("融资融券交易明细  margin_detail_20210517 返回数据 row 行数 = "+str(margin_detail_20210517.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210517_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210518")       ###  更新 记录日期
margin_detail_20210518 = pro.margin_detail(trade_date='20210518', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210518_tscode_list = list() 
margin_detail_20210518_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210518.columns.values.tolist():
    for ts_code_sh in margin_detail_20210518['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210518_tscode_list.append(stock_name)
    margin_detail_20210518_addname_dataframe['cname'] = margin_detail_20210518_tscode_list
for table_name in margin_detail_20210518.columns.values.tolist():
    margin_detail_20210518_addname_dataframe[table_name] = margin_detail_20210518[table_name]
print("融资融券交易明细  margin_detail_20210518 返回数据 row 行数 = "+str(margin_detail_20210518.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210518_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210519")       ###  更新 记录日期
margin_detail_20210519 = pro.margin_detail(trade_date='20210519', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210519_tscode_list = list() 
margin_detail_20210519_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210519.columns.values.tolist():
    for ts_code_sh in margin_detail_20210519['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210519_tscode_list.append(stock_name)
    margin_detail_20210519_addname_dataframe['cname'] = margin_detail_20210519_tscode_list
for table_name in margin_detail_20210519.columns.values.tolist():
    margin_detail_20210519_addname_dataframe[table_name] = margin_detail_20210519[table_name]
print("融资融券交易明细  margin_detail_20210519 返回数据 row 行数 = "+str(margin_detail_20210519.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210519_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210520")       ###  更新 记录日期
margin_detail_20210520 = pro.margin_detail(trade_date='20210520', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210520_tscode_list = list() 
margin_detail_20210520_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210520.columns.values.tolist():
    for ts_code_sh in margin_detail_20210520['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210520_tscode_list.append(stock_name)
    margin_detail_20210520_addname_dataframe['cname'] = margin_detail_20210520_tscode_list
for table_name in margin_detail_20210520.columns.values.tolist():
    margin_detail_20210520_addname_dataframe[table_name] = margin_detail_20210520[table_name]
print("融资融券交易明细  margin_detail_20210520 返回数据 row 行数 = "+str(margin_detail_20210520.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210520_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210521")       ###  更新 记录日期
margin_detail_20210521 = pro.margin_detail(trade_date='20210521', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210521_tscode_list = list() 
margin_detail_20210521_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210521.columns.values.tolist():
    for ts_code_sh in margin_detail_20210521['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210521_tscode_list.append(stock_name)
    margin_detail_20210521_addname_dataframe['cname'] = margin_detail_20210521_tscode_list
for table_name in margin_detail_20210521.columns.values.tolist():
    margin_detail_20210521_addname_dataframe[table_name] = margin_detail_20210521[table_name]
print("融资融券交易明细  margin_detail_20210521 返回数据 row 行数 = "+str(margin_detail_20210521.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210521_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210524")       ###  更新 记录日期
margin_detail_20210524 = pro.margin_detail(trade_date='20210524', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210524_tscode_list = list() 
margin_detail_20210524_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210524.columns.values.tolist():
    for ts_code_sh in margin_detail_20210524['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210524_tscode_list.append(stock_name)
    margin_detail_20210524_addname_dataframe['cname'] = margin_detail_20210524_tscode_list
for table_name in margin_detail_20210524.columns.values.tolist():
    margin_detail_20210524_addname_dataframe[table_name] = margin_detail_20210524[table_name]
print("融资融券交易明细  margin_detail_20210524 返回数据 row 行数 = "+str(margin_detail_20210524.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210524_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210525")       ###  更新 记录日期
margin_detail_20210525 = pro.margin_detail(trade_date='20210525', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210525_tscode_list = list() 
margin_detail_20210525_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210525.columns.values.tolist():
    for ts_code_sh in margin_detail_20210525['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210525_tscode_list.append(stock_name)
    margin_detail_20210525_addname_dataframe['cname'] = margin_detail_20210525_tscode_list
for table_name in margin_detail_20210525.columns.values.tolist():
    margin_detail_20210525_addname_dataframe[table_name] = margin_detail_20210525[table_name]
print("融资融券交易明细  margin_detail_20210525 返回数据 row 行数 = "+str(margin_detail_20210525.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210525_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210526")       ###  更新 记录日期
margin_detail_20210526 = pro.margin_detail(trade_date='20210526', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210526_tscode_list = list() 
margin_detail_20210526_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210526.columns.values.tolist():
    for ts_code_sh in margin_detail_20210526['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210526_tscode_list.append(stock_name)
    margin_detail_20210526_addname_dataframe['cname'] = margin_detail_20210526_tscode_list
for table_name in margin_detail_20210526.columns.values.tolist():
    margin_detail_20210526_addname_dataframe[table_name] = margin_detail_20210526[table_name]
print("融资融券交易明细  margin_detail_20210526 返回数据 row 行数 = "+str(margin_detail_20210526.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210526_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210527")       ###  更新 记录日期
margin_detail_20210527 = pro.margin_detail(trade_date='20210527', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210527_tscode_list = list() 
margin_detail_20210527_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210527.columns.values.tolist():
    for ts_code_sh in margin_detail_20210527['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210527_tscode_list.append(stock_name)
    margin_detail_20210527_addname_dataframe['cname'] = margin_detail_20210527_tscode_list
for table_name in margin_detail_20210527.columns.values.tolist():
    margin_detail_20210527_addname_dataframe[table_name] = margin_detail_20210527[table_name]
print("融资融券交易明细  margin_detail_20210527 返回数据 row 行数 = "+str(margin_detail_20210527.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210527_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210528")       ###  更新 记录日期
margin_detail_20210528 = pro.margin_detail(trade_date='20210528', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210528_tscode_list = list() 
margin_detail_20210528_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210528.columns.values.tolist():
    for ts_code_sh in margin_detail_20210528['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210528_tscode_list.append(stock_name)
    margin_detail_20210528_addname_dataframe['cname'] = margin_detail_20210528_tscode_list
for table_name in margin_detail_20210528.columns.values.tolist():
    margin_detail_20210528_addname_dataframe[table_name] = margin_detail_20210528[table_name]
print("融资融券交易明细  margin_detail_20210528 返回数据 row 行数 = "+str(margin_detail_20210528.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210528_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210531")       ###  更新 记录日期
margin_detail_20210531 = pro.margin_detail(trade_date='20210531', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210531_tscode_list = list() 
margin_detail_20210531_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210531.columns.values.tolist():
    for ts_code_sh in margin_detail_20210531['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210531_tscode_list.append(stock_name)
    margin_detail_20210531_addname_dataframe['cname'] = margin_detail_20210531_tscode_list
for table_name in margin_detail_20210531.columns.values.tolist():
    margin_detail_20210531_addname_dataframe[table_name] = margin_detail_20210531[table_name]
print("融资融券交易明细  margin_detail_20210531 返回数据 row 行数 = "+str(margin_detail_20210531.shape[0]))
margin_detail_2020_5_xlsx_frame=margin_detail_2020_5_xlsx_frame.append(margin_detail_20210531_addname_dataframe,ignore_index=True)
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_5_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'5',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_6_xlsx_frame=pd.DataFrame()
if '6' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_6_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='6' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20210601")       ###  更新 记录日期
margin_detail_20210601 = pro.margin_detail(trade_date='20210601', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210601_tscode_list = list() 
margin_detail_20210601_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210601.columns.values.tolist():
    for ts_code_sh in margin_detail_20210601['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210601_tscode_list.append(stock_name)
    margin_detail_20210601_addname_dataframe['cname'] = margin_detail_20210601_tscode_list
for table_name in margin_detail_20210601.columns.values.tolist():
    margin_detail_20210601_addname_dataframe[table_name] = margin_detail_20210601[table_name]
print("融资融券交易明细  margin_detail_20210601 返回数据 row 行数 = "+str(margin_detail_20210601.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210601_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210602")       ###  更新 记录日期
margin_detail_20210602 = pro.margin_detail(trade_date='20210602', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210602_tscode_list = list() 
margin_detail_20210602_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210602.columns.values.tolist():
    for ts_code_sh in margin_detail_20210602['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210602_tscode_list.append(stock_name)
    margin_detail_20210602_addname_dataframe['cname'] = margin_detail_20210602_tscode_list
for table_name in margin_detail_20210602.columns.values.tolist():
    margin_detail_20210602_addname_dataframe[table_name] = margin_detail_20210602[table_name]
print("融资融券交易明细  margin_detail_20210602 返回数据 row 行数 = "+str(margin_detail_20210602.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210602_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210603")       ###  更新 记录日期
margin_detail_20210603 = pro.margin_detail(trade_date='20210603', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210603_tscode_list = list() 
margin_detail_20210603_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210603.columns.values.tolist():
    for ts_code_sh in margin_detail_20210603['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210603_tscode_list.append(stock_name)
    margin_detail_20210603_addname_dataframe['cname'] = margin_detail_20210603_tscode_list
for table_name in margin_detail_20210603.columns.values.tolist():
    margin_detail_20210603_addname_dataframe[table_name] = margin_detail_20210603[table_name]
print("融资融券交易明细  margin_detail_20210603 返回数据 row 行数 = "+str(margin_detail_20210603.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210603_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210604")       ###  更新 记录日期
margin_detail_20210604 = pro.margin_detail(trade_date='20210604', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210604_tscode_list = list() 
margin_detail_20210604_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210604.columns.values.tolist():
    for ts_code_sh in margin_detail_20210604['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210604_tscode_list.append(stock_name)
    margin_detail_20210604_addname_dataframe['cname'] = margin_detail_20210604_tscode_list
for table_name in margin_detail_20210604.columns.values.tolist():
    margin_detail_20210604_addname_dataframe[table_name] = margin_detail_20210604[table_name]
print("融资融券交易明细  margin_detail_20210604 返回数据 row 行数 = "+str(margin_detail_20210604.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210604_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210607")       ###  更新 记录日期
margin_detail_20210607 = pro.margin_detail(trade_date='20210607', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210607_tscode_list = list() 
margin_detail_20210607_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210607.columns.values.tolist():
    for ts_code_sh in margin_detail_20210607['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210607_tscode_list.append(stock_name)
    margin_detail_20210607_addname_dataframe['cname'] = margin_detail_20210607_tscode_list
for table_name in margin_detail_20210607.columns.values.tolist():
    margin_detail_20210607_addname_dataframe[table_name] = margin_detail_20210607[table_name]
print("融资融券交易明细  margin_detail_20210607 返回数据 row 行数 = "+str(margin_detail_20210607.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210607_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210608")       ###  更新 记录日期
margin_detail_20210608 = pro.margin_detail(trade_date='20210608', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210608_tscode_list = list() 
margin_detail_20210608_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210608.columns.values.tolist():
    for ts_code_sh in margin_detail_20210608['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210608_tscode_list.append(stock_name)
    margin_detail_20210608_addname_dataframe['cname'] = margin_detail_20210608_tscode_list
for table_name in margin_detail_20210608.columns.values.tolist():
    margin_detail_20210608_addname_dataframe[table_name] = margin_detail_20210608[table_name]
print("融资融券交易明细  margin_detail_20210608 返回数据 row 行数 = "+str(margin_detail_20210608.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210608_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210609")       ###  更新 记录日期
margin_detail_20210609 = pro.margin_detail(trade_date='20210609', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210609_tscode_list = list() 
margin_detail_20210609_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210609.columns.values.tolist():
    for ts_code_sh in margin_detail_20210609['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210609_tscode_list.append(stock_name)
    margin_detail_20210609_addname_dataframe['cname'] = margin_detail_20210609_tscode_list
for table_name in margin_detail_20210609.columns.values.tolist():
    margin_detail_20210609_addname_dataframe[table_name] = margin_detail_20210609[table_name]
print("融资融券交易明细  margin_detail_20210609 返回数据 row 行数 = "+str(margin_detail_20210609.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210609_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210610")       ###  更新 记录日期
margin_detail_20210610 = pro.margin_detail(trade_date='20210610', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210610_tscode_list = list() 
margin_detail_20210610_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210610.columns.values.tolist():
    for ts_code_sh in margin_detail_20210610['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210610_tscode_list.append(stock_name)
    margin_detail_20210610_addname_dataframe['cname'] = margin_detail_20210610_tscode_list
for table_name in margin_detail_20210610.columns.values.tolist():
    margin_detail_20210610_addname_dataframe[table_name] = margin_detail_20210610[table_name]
print("融资融券交易明细  margin_detail_20210610 返回数据 row 行数 = "+str(margin_detail_20210610.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210610_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210611")       ###  更新 记录日期
margin_detail_20210611 = pro.margin_detail(trade_date='20210611', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210611_tscode_list = list() 
margin_detail_20210611_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210611.columns.values.tolist():
    for ts_code_sh in margin_detail_20210611['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210611_tscode_list.append(stock_name)
    margin_detail_20210611_addname_dataframe['cname'] = margin_detail_20210611_tscode_list
for table_name in margin_detail_20210611.columns.values.tolist():
    margin_detail_20210611_addname_dataframe[table_name] = margin_detail_20210611[table_name]
print("融资融券交易明细  margin_detail_20210611 返回数据 row 行数 = "+str(margin_detail_20210611.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210611_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210615")       ###  更新 记录日期
margin_detail_20210615 = pro.margin_detail(trade_date='20210615', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210615_tscode_list = list() 
margin_detail_20210615_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210615.columns.values.tolist():
    for ts_code_sh in margin_detail_20210615['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210615_tscode_list.append(stock_name)
    margin_detail_20210615_addname_dataframe['cname'] = margin_detail_20210615_tscode_list
for table_name in margin_detail_20210615.columns.values.tolist():
    margin_detail_20210615_addname_dataframe[table_name] = margin_detail_20210615[table_name]
print("融资融券交易明细  margin_detail_20210615 返回数据 row 行数 = "+str(margin_detail_20210615.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210615_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210616")       ###  更新 记录日期
margin_detail_20210616 = pro.margin_detail(trade_date='20210616', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210616_tscode_list = list() 
margin_detail_20210616_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210616.columns.values.tolist():
    for ts_code_sh in margin_detail_20210616['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210616_tscode_list.append(stock_name)
    margin_detail_20210616_addname_dataframe['cname'] = margin_detail_20210616_tscode_list
for table_name in margin_detail_20210616.columns.values.tolist():
    margin_detail_20210616_addname_dataframe[table_name] = margin_detail_20210616[table_name]
print("融资融券交易明细  margin_detail_20210616 返回数据 row 行数 = "+str(margin_detail_20210616.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210616_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210617")       ###  更新 记录日期
margin_detail_20210617 = pro.margin_detail(trade_date='20210617', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210617_tscode_list = list() 
margin_detail_20210617_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210617.columns.values.tolist():
    for ts_code_sh in margin_detail_20210617['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210617_tscode_list.append(stock_name)
    margin_detail_20210617_addname_dataframe['cname'] = margin_detail_20210617_tscode_list
for table_name in margin_detail_20210617.columns.values.tolist():
    margin_detail_20210617_addname_dataframe[table_name] = margin_detail_20210617[table_name]
print("融资融券交易明细  margin_detail_20210617 返回数据 row 行数 = "+str(margin_detail_20210617.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210617_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210618")       ###  更新 记录日期
margin_detail_20210618 = pro.margin_detail(trade_date='20210618', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210618_tscode_list = list() 
margin_detail_20210618_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210618.columns.values.tolist():
    for ts_code_sh in margin_detail_20210618['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210618_tscode_list.append(stock_name)
    margin_detail_20210618_addname_dataframe['cname'] = margin_detail_20210618_tscode_list
for table_name in margin_detail_20210618.columns.values.tolist():
    margin_detail_20210618_addname_dataframe[table_name] = margin_detail_20210618[table_name]
print("融资融券交易明细  margin_detail_20210618 返回数据 row 行数 = "+str(margin_detail_20210618.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210618_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210621")       ###  更新 记录日期
margin_detail_20210621 = pro.margin_detail(trade_date='20210621', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210621_tscode_list = list() 
margin_detail_20210621_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210621.columns.values.tolist():
    for ts_code_sh in margin_detail_20210621['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210621_tscode_list.append(stock_name)
    margin_detail_20210621_addname_dataframe['cname'] = margin_detail_20210621_tscode_list
for table_name in margin_detail_20210621.columns.values.tolist():
    margin_detail_20210621_addname_dataframe[table_name] = margin_detail_20210621[table_name]
print("融资融券交易明细  margin_detail_20210621 返回数据 row 行数 = "+str(margin_detail_20210621.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210621_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210622")       ###  更新 记录日期
margin_detail_20210622 = pro.margin_detail(trade_date='20210622', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210622_tscode_list = list() 
margin_detail_20210622_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210622.columns.values.tolist():
    for ts_code_sh in margin_detail_20210622['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210622_tscode_list.append(stock_name)
    margin_detail_20210622_addname_dataframe['cname'] = margin_detail_20210622_tscode_list
for table_name in margin_detail_20210622.columns.values.tolist():
    margin_detail_20210622_addname_dataframe[table_name] = margin_detail_20210622[table_name]
print("融资融券交易明细  margin_detail_20210622 返回数据 row 行数 = "+str(margin_detail_20210622.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210622_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210623")       ###  更新 记录日期
margin_detail_20210623 = pro.margin_detail(trade_date='20210623', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210623_tscode_list = list() 
margin_detail_20210623_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210623.columns.values.tolist():
    for ts_code_sh in margin_detail_20210623['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210623_tscode_list.append(stock_name)
    margin_detail_20210623_addname_dataframe['cname'] = margin_detail_20210623_tscode_list
for table_name in margin_detail_20210623.columns.values.tolist():
    margin_detail_20210623_addname_dataframe[table_name] = margin_detail_20210623[table_name]
print("融资融券交易明细  margin_detail_20210623 返回数据 row 行数 = "+str(margin_detail_20210623.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210623_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210624")       ###  更新 记录日期
margin_detail_20210624 = pro.margin_detail(trade_date='20210624', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210624_tscode_list = list() 
margin_detail_20210624_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210624.columns.values.tolist():
    for ts_code_sh in margin_detail_20210624['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210624_tscode_list.append(stock_name)
    margin_detail_20210624_addname_dataframe['cname'] = margin_detail_20210624_tscode_list
for table_name in margin_detail_20210624.columns.values.tolist():
    margin_detail_20210624_addname_dataframe[table_name] = margin_detail_20210624[table_name]
print("融资融券交易明细  margin_detail_20210624 返回数据 row 行数 = "+str(margin_detail_20210624.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210624_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210625")       ###  更新 记录日期
margin_detail_20210625 = pro.margin_detail(trade_date='20210625', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210625_tscode_list = list() 
margin_detail_20210625_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210625.columns.values.tolist():
    for ts_code_sh in margin_detail_20210625['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210625_tscode_list.append(stock_name)
    margin_detail_20210625_addname_dataframe['cname'] = margin_detail_20210625_tscode_list
for table_name in margin_detail_20210625.columns.values.tolist():
    margin_detail_20210625_addname_dataframe[table_name] = margin_detail_20210625[table_name]
print("融资融券交易明细  margin_detail_20210625 返回数据 row 行数 = "+str(margin_detail_20210625.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210625_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210628")       ###  更新 记录日期
margin_detail_20210628 = pro.margin_detail(trade_date='20210628', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210628_tscode_list = list() 
margin_detail_20210628_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210628.columns.values.tolist():
    for ts_code_sh in margin_detail_20210628['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210628_tscode_list.append(stock_name)
    margin_detail_20210628_addname_dataframe['cname'] = margin_detail_20210628_tscode_list
for table_name in margin_detail_20210628.columns.values.tolist():
    margin_detail_20210628_addname_dataframe[table_name] = margin_detail_20210628[table_name]
print("融资融券交易明细  margin_detail_20210628 返回数据 row 行数 = "+str(margin_detail_20210628.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210628_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210629")       ###  更新 记录日期
margin_detail_20210629 = pro.margin_detail(trade_date='20210629', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210629_tscode_list = list() 
margin_detail_20210629_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210629.columns.values.tolist():
    for ts_code_sh in margin_detail_20210629['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210629_tscode_list.append(stock_name)
    margin_detail_20210629_addname_dataframe['cname'] = margin_detail_20210629_tscode_list
for table_name in margin_detail_20210629.columns.values.tolist():
    margin_detail_20210629_addname_dataframe[table_name] = margin_detail_20210629[table_name]
print("融资融券交易明细  margin_detail_20210629 返回数据 row 行数 = "+str(margin_detail_20210629.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210629_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210630")       ###  更新 记录日期
margin_detail_20210630 = pro.margin_detail(trade_date='20210630', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210630_tscode_list = list() 
margin_detail_20210630_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210630.columns.values.tolist():
    for ts_code_sh in margin_detail_20210630['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210630_tscode_list.append(stock_name)
    margin_detail_20210630_addname_dataframe['cname'] = margin_detail_20210630_tscode_list
for table_name in margin_detail_20210630.columns.values.tolist():
    margin_detail_20210630_addname_dataframe[table_name] = margin_detail_20210630[table_name]
print("融资融券交易明细  margin_detail_20210630 返回数据 row 行数 = "+str(margin_detail_20210630.shape[0]))
margin_detail_2020_6_xlsx_frame=margin_detail_2020_6_xlsx_frame.append(margin_detail_20210630_addname_dataframe,ignore_index=True)
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_6_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'6',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_7_xlsx_frame=pd.DataFrame()
if '7' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_7_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='7' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20210701")       ###  更新 记录日期
margin_detail_20210701 = pro.margin_detail(trade_date='20210701', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210701_tscode_list = list() 
margin_detail_20210701_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210701.columns.values.tolist():
    for ts_code_sh in margin_detail_20210701['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210701_tscode_list.append(stock_name)
    margin_detail_20210701_addname_dataframe['cname'] = margin_detail_20210701_tscode_list
for table_name in margin_detail_20210701.columns.values.tolist():
    margin_detail_20210701_addname_dataframe[table_name] = margin_detail_20210701[table_name]
print("融资融券交易明细  margin_detail_20210701 返回数据 row 行数 = "+str(margin_detail_20210701.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210701_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210702")       ###  更新 记录日期
margin_detail_20210702 = pro.margin_detail(trade_date='20210702', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210702_tscode_list = list() 
margin_detail_20210702_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210702.columns.values.tolist():
    for ts_code_sh in margin_detail_20210702['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210702_tscode_list.append(stock_name)
    margin_detail_20210702_addname_dataframe['cname'] = margin_detail_20210702_tscode_list
for table_name in margin_detail_20210702.columns.values.tolist():
    margin_detail_20210702_addname_dataframe[table_name] = margin_detail_20210702[table_name]
print("融资融券交易明细  margin_detail_20210702 返回数据 row 行数 = "+str(margin_detail_20210702.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210702_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210705")       ###  更新 记录日期
margin_detail_20210705 = pro.margin_detail(trade_date='20210705', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210705_tscode_list = list() 
margin_detail_20210705_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210705.columns.values.tolist():
    for ts_code_sh in margin_detail_20210705['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210705_tscode_list.append(stock_name)
    margin_detail_20210705_addname_dataframe['cname'] = margin_detail_20210705_tscode_list
for table_name in margin_detail_20210705.columns.values.tolist():
    margin_detail_20210705_addname_dataframe[table_name] = margin_detail_20210705[table_name]
print("融资融券交易明细  margin_detail_20210705 返回数据 row 行数 = "+str(margin_detail_20210705.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210705_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210706")       ###  更新 记录日期
margin_detail_20210706 = pro.margin_detail(trade_date='20210706', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210706_tscode_list = list() 
margin_detail_20210706_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210706.columns.values.tolist():
    for ts_code_sh in margin_detail_20210706['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210706_tscode_list.append(stock_name)
    margin_detail_20210706_addname_dataframe['cname'] = margin_detail_20210706_tscode_list
for table_name in margin_detail_20210706.columns.values.tolist():
    margin_detail_20210706_addname_dataframe[table_name] = margin_detail_20210706[table_name]
print("融资融券交易明细  margin_detail_20210706 返回数据 row 行数 = "+str(margin_detail_20210706.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210706_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210707")       ###  更新 记录日期
margin_detail_20210707 = pro.margin_detail(trade_date='20210707', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210707_tscode_list = list() 
margin_detail_20210707_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210707.columns.values.tolist():
    for ts_code_sh in margin_detail_20210707['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210707_tscode_list.append(stock_name)
    margin_detail_20210707_addname_dataframe['cname'] = margin_detail_20210707_tscode_list
for table_name in margin_detail_20210707.columns.values.tolist():
    margin_detail_20210707_addname_dataframe[table_name] = margin_detail_20210707[table_name]
print("融资融券交易明细  margin_detail_20210707 返回数据 row 行数 = "+str(margin_detail_20210707.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210707_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210708")       ###  更新 记录日期
margin_detail_20210708 = pro.margin_detail(trade_date='20210708', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210708_tscode_list = list() 
margin_detail_20210708_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210708.columns.values.tolist():
    for ts_code_sh in margin_detail_20210708['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210708_tscode_list.append(stock_name)
    margin_detail_20210708_addname_dataframe['cname'] = margin_detail_20210708_tscode_list
for table_name in margin_detail_20210708.columns.values.tolist():
    margin_detail_20210708_addname_dataframe[table_name] = margin_detail_20210708[table_name]
print("融资融券交易明细  margin_detail_20210708 返回数据 row 行数 = "+str(margin_detail_20210708.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210708_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210709")       ###  更新 记录日期
margin_detail_20210709 = pro.margin_detail(trade_date='20210709', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210709_tscode_list = list() 
margin_detail_20210709_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210709.columns.values.tolist():
    for ts_code_sh in margin_detail_20210709['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210709_tscode_list.append(stock_name)
    margin_detail_20210709_addname_dataframe['cname'] = margin_detail_20210709_tscode_list
for table_name in margin_detail_20210709.columns.values.tolist():
    margin_detail_20210709_addname_dataframe[table_name] = margin_detail_20210709[table_name]
print("融资融券交易明细  margin_detail_20210709 返回数据 row 行数 = "+str(margin_detail_20210709.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210709_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210712")       ###  更新 记录日期
margin_detail_20210712 = pro.margin_detail(trade_date='20210712', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210712_tscode_list = list() 
margin_detail_20210712_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210712.columns.values.tolist():
    for ts_code_sh in margin_detail_20210712['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210712_tscode_list.append(stock_name)
    margin_detail_20210712_addname_dataframe['cname'] = margin_detail_20210712_tscode_list
for table_name in margin_detail_20210712.columns.values.tolist():
    margin_detail_20210712_addname_dataframe[table_name] = margin_detail_20210712[table_name]
print("融资融券交易明细  margin_detail_20210712 返回数据 row 行数 = "+str(margin_detail_20210712.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210712_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210713")       ###  更新 记录日期
margin_detail_20210713 = pro.margin_detail(trade_date='20210713', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210713_tscode_list = list() 
margin_detail_20210713_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210713.columns.values.tolist():
    for ts_code_sh in margin_detail_20210713['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210713_tscode_list.append(stock_name)
    margin_detail_20210713_addname_dataframe['cname'] = margin_detail_20210713_tscode_list
for table_name in margin_detail_20210713.columns.values.tolist():
    margin_detail_20210713_addname_dataframe[table_name] = margin_detail_20210713[table_name]
print("融资融券交易明细  margin_detail_20210713 返回数据 row 行数 = "+str(margin_detail_20210713.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210713_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210714")       ###  更新 记录日期
margin_detail_20210714 = pro.margin_detail(trade_date='20210714', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210714_tscode_list = list() 
margin_detail_20210714_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210714.columns.values.tolist():
    for ts_code_sh in margin_detail_20210714['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210714_tscode_list.append(stock_name)
    margin_detail_20210714_addname_dataframe['cname'] = margin_detail_20210714_tscode_list
for table_name in margin_detail_20210714.columns.values.tolist():
    margin_detail_20210714_addname_dataframe[table_name] = margin_detail_20210714[table_name]
print("融资融券交易明细  margin_detail_20210714 返回数据 row 行数 = "+str(margin_detail_20210714.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210714_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210715")       ###  更新 记录日期
margin_detail_20210715 = pro.margin_detail(trade_date='20210715', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210715_tscode_list = list() 
margin_detail_20210715_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210715.columns.values.tolist():
    for ts_code_sh in margin_detail_20210715['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210715_tscode_list.append(stock_name)
    margin_detail_20210715_addname_dataframe['cname'] = margin_detail_20210715_tscode_list
for table_name in margin_detail_20210715.columns.values.tolist():
    margin_detail_20210715_addname_dataframe[table_name] = margin_detail_20210715[table_name]
print("融资融券交易明细  margin_detail_20210715 返回数据 row 行数 = "+str(margin_detail_20210715.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210715_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210716")       ###  更新 记录日期
margin_detail_20210716 = pro.margin_detail(trade_date='20210716', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210716_tscode_list = list() 
margin_detail_20210716_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210716.columns.values.tolist():
    for ts_code_sh in margin_detail_20210716['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210716_tscode_list.append(stock_name)
    margin_detail_20210716_addname_dataframe['cname'] = margin_detail_20210716_tscode_list
for table_name in margin_detail_20210716.columns.values.tolist():
    margin_detail_20210716_addname_dataframe[table_name] = margin_detail_20210716[table_name]
print("融资融券交易明细  margin_detail_20210716 返回数据 row 行数 = "+str(margin_detail_20210716.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210716_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210719")       ###  更新 记录日期
margin_detail_20210719 = pro.margin_detail(trade_date='20210719', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210719_tscode_list = list() 
margin_detail_20210719_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210719.columns.values.tolist():
    for ts_code_sh in margin_detail_20210719['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210719_tscode_list.append(stock_name)
    margin_detail_20210719_addname_dataframe['cname'] = margin_detail_20210719_tscode_list
for table_name in margin_detail_20210719.columns.values.tolist():
    margin_detail_20210719_addname_dataframe[table_name] = margin_detail_20210719[table_name]
print("融资融券交易明细  margin_detail_20210719 返回数据 row 行数 = "+str(margin_detail_20210719.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210719_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210720")       ###  更新 记录日期
margin_detail_20210720 = pro.margin_detail(trade_date='20210720', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210720_tscode_list = list() 
margin_detail_20210720_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210720.columns.values.tolist():
    for ts_code_sh in margin_detail_20210720['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210720_tscode_list.append(stock_name)
    margin_detail_20210720_addname_dataframe['cname'] = margin_detail_20210720_tscode_list
for table_name in margin_detail_20210720.columns.values.tolist():
    margin_detail_20210720_addname_dataframe[table_name] = margin_detail_20210720[table_name]
print("融资融券交易明细  margin_detail_20210720 返回数据 row 行数 = "+str(margin_detail_20210720.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210720_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210721")       ###  更新 记录日期
margin_detail_20210721 = pro.margin_detail(trade_date='20210721', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210721_tscode_list = list() 
margin_detail_20210721_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210721.columns.values.tolist():
    for ts_code_sh in margin_detail_20210721['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210721_tscode_list.append(stock_name)
    margin_detail_20210721_addname_dataframe['cname'] = margin_detail_20210721_tscode_list
for table_name in margin_detail_20210721.columns.values.tolist():
    margin_detail_20210721_addname_dataframe[table_name] = margin_detail_20210721[table_name]
print("融资融券交易明细  margin_detail_20210721 返回数据 row 行数 = "+str(margin_detail_20210721.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210721_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210722")       ###  更新 记录日期
margin_detail_20210722 = pro.margin_detail(trade_date='20210722', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210722_tscode_list = list() 
margin_detail_20210722_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210722.columns.values.tolist():
    for ts_code_sh in margin_detail_20210722['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210722_tscode_list.append(stock_name)
    margin_detail_20210722_addname_dataframe['cname'] = margin_detail_20210722_tscode_list
for table_name in margin_detail_20210722.columns.values.tolist():
    margin_detail_20210722_addname_dataframe[table_name] = margin_detail_20210722[table_name]
print("融资融券交易明细  margin_detail_20210722 返回数据 row 行数 = "+str(margin_detail_20210722.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210722_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210723")       ###  更新 记录日期
margin_detail_20210723 = pro.margin_detail(trade_date='20210723', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210723_tscode_list = list() 
margin_detail_20210723_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210723.columns.values.tolist():
    for ts_code_sh in margin_detail_20210723['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210723_tscode_list.append(stock_name)
    margin_detail_20210723_addname_dataframe['cname'] = margin_detail_20210723_tscode_list
for table_name in margin_detail_20210723.columns.values.tolist():
    margin_detail_20210723_addname_dataframe[table_name] = margin_detail_20210723[table_name]
print("融资融券交易明细  margin_detail_20210723 返回数据 row 行数 = "+str(margin_detail_20210723.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210723_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210726")       ###  更新 记录日期
margin_detail_20210726 = pro.margin_detail(trade_date='20210726', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210726_tscode_list = list() 
margin_detail_20210726_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210726.columns.values.tolist():
    for ts_code_sh in margin_detail_20210726['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210726_tscode_list.append(stock_name)
    margin_detail_20210726_addname_dataframe['cname'] = margin_detail_20210726_tscode_list
for table_name in margin_detail_20210726.columns.values.tolist():
    margin_detail_20210726_addname_dataframe[table_name] = margin_detail_20210726[table_name]
print("融资融券交易明细  margin_detail_20210726 返回数据 row 行数 = "+str(margin_detail_20210726.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210726_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210727")       ###  更新 记录日期
margin_detail_20210727 = pro.margin_detail(trade_date='20210727', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210727_tscode_list = list() 
margin_detail_20210727_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210727.columns.values.tolist():
    for ts_code_sh in margin_detail_20210727['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210727_tscode_list.append(stock_name)
    margin_detail_20210727_addname_dataframe['cname'] = margin_detail_20210727_tscode_list
for table_name in margin_detail_20210727.columns.values.tolist():
    margin_detail_20210727_addname_dataframe[table_name] = margin_detail_20210727[table_name]
print("融资融券交易明细  margin_detail_20210727 返回数据 row 行数 = "+str(margin_detail_20210727.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210727_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210728")       ###  更新 记录日期
margin_detail_20210728 = pro.margin_detail(trade_date='20210728', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210728_tscode_list = list() 
margin_detail_20210728_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210728.columns.values.tolist():
    for ts_code_sh in margin_detail_20210728['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210728_tscode_list.append(stock_name)
    margin_detail_20210728_addname_dataframe['cname'] = margin_detail_20210728_tscode_list
for table_name in margin_detail_20210728.columns.values.tolist():
    margin_detail_20210728_addname_dataframe[table_name] = margin_detail_20210728[table_name]
print("融资融券交易明细  margin_detail_20210728 返回数据 row 行数 = "+str(margin_detail_20210728.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210728_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210729")       ###  更新 记录日期
margin_detail_20210729 = pro.margin_detail(trade_date='20210729', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210729_tscode_list = list() 
margin_detail_20210729_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210729.columns.values.tolist():
    for ts_code_sh in margin_detail_20210729['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210729_tscode_list.append(stock_name)
    margin_detail_20210729_addname_dataframe['cname'] = margin_detail_20210729_tscode_list
for table_name in margin_detail_20210729.columns.values.tolist():
    margin_detail_20210729_addname_dataframe[table_name] = margin_detail_20210729[table_name]
print("融资融券交易明细  margin_detail_20210729 返回数据 row 行数 = "+str(margin_detail_20210729.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210729_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210730")       ###  更新 记录日期
margin_detail_20210730 = pro.margin_detail(trade_date='20210730', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210730_tscode_list = list() 
margin_detail_20210730_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210730.columns.values.tolist():
    for ts_code_sh in margin_detail_20210730['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210730_tscode_list.append(stock_name)
    margin_detail_20210730_addname_dataframe['cname'] = margin_detail_20210730_tscode_list
for table_name in margin_detail_20210730.columns.values.tolist():
    margin_detail_20210730_addname_dataframe[table_name] = margin_detail_20210730[table_name]
print("融资融券交易明细  margin_detail_20210730 返回数据 row 行数 = "+str(margin_detail_20210730.shape[0]))
margin_detail_2020_7_xlsx_frame=margin_detail_2020_7_xlsx_frame.append(margin_detail_20210730_addname_dataframe,ignore_index=True)
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_7_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'7',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_8_xlsx_frame=pd.DataFrame()
if '8' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_8_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='8' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20210802")       ###  更新 记录日期
margin_detail_20210802 = pro.margin_detail(trade_date='20210802', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210802_tscode_list = list() 
margin_detail_20210802_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210802.columns.values.tolist():
    for ts_code_sh in margin_detail_20210802['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210802_tscode_list.append(stock_name)
    margin_detail_20210802_addname_dataframe['cname'] = margin_detail_20210802_tscode_list
for table_name in margin_detail_20210802.columns.values.tolist():
    margin_detail_20210802_addname_dataframe[table_name] = margin_detail_20210802[table_name]
print("融资融券交易明细  margin_detail_20210802 返回数据 row 行数 = "+str(margin_detail_20210802.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210802_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210803")       ###  更新 记录日期
margin_detail_20210803 = pro.margin_detail(trade_date='20210803', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210803_tscode_list = list() 
margin_detail_20210803_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210803.columns.values.tolist():
    for ts_code_sh in margin_detail_20210803['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210803_tscode_list.append(stock_name)
    margin_detail_20210803_addname_dataframe['cname'] = margin_detail_20210803_tscode_list
for table_name in margin_detail_20210803.columns.values.tolist():
    margin_detail_20210803_addname_dataframe[table_name] = margin_detail_20210803[table_name]
print("融资融券交易明细  margin_detail_20210803 返回数据 row 行数 = "+str(margin_detail_20210803.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210803_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210804")       ###  更新 记录日期
margin_detail_20210804 = pro.margin_detail(trade_date='20210804', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210804_tscode_list = list() 
margin_detail_20210804_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210804.columns.values.tolist():
    for ts_code_sh in margin_detail_20210804['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210804_tscode_list.append(stock_name)
    margin_detail_20210804_addname_dataframe['cname'] = margin_detail_20210804_tscode_list
for table_name in margin_detail_20210804.columns.values.tolist():
    margin_detail_20210804_addname_dataframe[table_name] = margin_detail_20210804[table_name]
print("融资融券交易明细  margin_detail_20210804 返回数据 row 行数 = "+str(margin_detail_20210804.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210804_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210805")       ###  更新 记录日期
margin_detail_20210805 = pro.margin_detail(trade_date='20210805', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210805_tscode_list = list() 
margin_detail_20210805_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210805.columns.values.tolist():
    for ts_code_sh in margin_detail_20210805['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210805_tscode_list.append(stock_name)
    margin_detail_20210805_addname_dataframe['cname'] = margin_detail_20210805_tscode_list
for table_name in margin_detail_20210805.columns.values.tolist():
    margin_detail_20210805_addname_dataframe[table_name] = margin_detail_20210805[table_name]
print("融资融券交易明细  margin_detail_20210805 返回数据 row 行数 = "+str(margin_detail_20210805.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210805_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210806")       ###  更新 记录日期
margin_detail_20210806 = pro.margin_detail(trade_date='20210806', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210806_tscode_list = list() 
margin_detail_20210806_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210806.columns.values.tolist():
    for ts_code_sh in margin_detail_20210806['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210806_tscode_list.append(stock_name)
    margin_detail_20210806_addname_dataframe['cname'] = margin_detail_20210806_tscode_list
for table_name in margin_detail_20210806.columns.values.tolist():
    margin_detail_20210806_addname_dataframe[table_name] = margin_detail_20210806[table_name]
print("融资融券交易明细  margin_detail_20210806 返回数据 row 行数 = "+str(margin_detail_20210806.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210806_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210809")       ###  更新 记录日期
margin_detail_20210809 = pro.margin_detail(trade_date='20210809', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210809_tscode_list = list() 
margin_detail_20210809_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210809.columns.values.tolist():
    for ts_code_sh in margin_detail_20210809['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210809_tscode_list.append(stock_name)
    margin_detail_20210809_addname_dataframe['cname'] = margin_detail_20210809_tscode_list
for table_name in margin_detail_20210809.columns.values.tolist():
    margin_detail_20210809_addname_dataframe[table_name] = margin_detail_20210809[table_name]
print("融资融券交易明细  margin_detail_20210809 返回数据 row 行数 = "+str(margin_detail_20210809.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210809_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210810")       ###  更新 记录日期
margin_detail_20210810 = pro.margin_detail(trade_date='20210810', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210810_tscode_list = list() 
margin_detail_20210810_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210810.columns.values.tolist():
    for ts_code_sh in margin_detail_20210810['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210810_tscode_list.append(stock_name)
    margin_detail_20210810_addname_dataframe['cname'] = margin_detail_20210810_tscode_list
for table_name in margin_detail_20210810.columns.values.tolist():
    margin_detail_20210810_addname_dataframe[table_name] = margin_detail_20210810[table_name]
print("融资融券交易明细  margin_detail_20210810 返回数据 row 行数 = "+str(margin_detail_20210810.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210810_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210811")       ###  更新 记录日期
margin_detail_20210811 = pro.margin_detail(trade_date='20210811', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210811_tscode_list = list() 
margin_detail_20210811_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210811.columns.values.tolist():
    for ts_code_sh in margin_detail_20210811['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210811_tscode_list.append(stock_name)
    margin_detail_20210811_addname_dataframe['cname'] = margin_detail_20210811_tscode_list
for table_name in margin_detail_20210811.columns.values.tolist():
    margin_detail_20210811_addname_dataframe[table_name] = margin_detail_20210811[table_name]
print("融资融券交易明细  margin_detail_20210811 返回数据 row 行数 = "+str(margin_detail_20210811.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210811_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210812")       ###  更新 记录日期
margin_detail_20210812 = pro.margin_detail(trade_date='20210812', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210812_tscode_list = list() 
margin_detail_20210812_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210812.columns.values.tolist():
    for ts_code_sh in margin_detail_20210812['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210812_tscode_list.append(stock_name)
    margin_detail_20210812_addname_dataframe['cname'] = margin_detail_20210812_tscode_list
for table_name in margin_detail_20210812.columns.values.tolist():
    margin_detail_20210812_addname_dataframe[table_name] = margin_detail_20210812[table_name]
print("融资融券交易明细  margin_detail_20210812 返回数据 row 行数 = "+str(margin_detail_20210812.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210812_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210813")       ###  更新 记录日期
margin_detail_20210813 = pro.margin_detail(trade_date='20210813', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210813_tscode_list = list() 
margin_detail_20210813_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210813.columns.values.tolist():
    for ts_code_sh in margin_detail_20210813['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210813_tscode_list.append(stock_name)
    margin_detail_20210813_addname_dataframe['cname'] = margin_detail_20210813_tscode_list
for table_name in margin_detail_20210813.columns.values.tolist():
    margin_detail_20210813_addname_dataframe[table_name] = margin_detail_20210813[table_name]
print("融资融券交易明细  margin_detail_20210813 返回数据 row 行数 = "+str(margin_detail_20210813.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210813_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210816")       ###  更新 记录日期
margin_detail_20210816 = pro.margin_detail(trade_date='20210816', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210816_tscode_list = list() 
margin_detail_20210816_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210816.columns.values.tolist():
    for ts_code_sh in margin_detail_20210816['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210816_tscode_list.append(stock_name)
    margin_detail_20210816_addname_dataframe['cname'] = margin_detail_20210816_tscode_list
for table_name in margin_detail_20210816.columns.values.tolist():
    margin_detail_20210816_addname_dataframe[table_name] = margin_detail_20210816[table_name]
print("融资融券交易明细  margin_detail_20210816 返回数据 row 行数 = "+str(margin_detail_20210816.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210816_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210817")       ###  更新 记录日期
margin_detail_20210817 = pro.margin_detail(trade_date='20210817', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210817_tscode_list = list() 
margin_detail_20210817_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210817.columns.values.tolist():
    for ts_code_sh in margin_detail_20210817['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210817_tscode_list.append(stock_name)
    margin_detail_20210817_addname_dataframe['cname'] = margin_detail_20210817_tscode_list
for table_name in margin_detail_20210817.columns.values.tolist():
    margin_detail_20210817_addname_dataframe[table_name] = margin_detail_20210817[table_name]
print("融资融券交易明细  margin_detail_20210817 返回数据 row 行数 = "+str(margin_detail_20210817.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210817_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210818")       ###  更新 记录日期
margin_detail_20210818 = pro.margin_detail(trade_date='20210818', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210818_tscode_list = list() 
margin_detail_20210818_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210818.columns.values.tolist():
    for ts_code_sh in margin_detail_20210818['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210818_tscode_list.append(stock_name)
    margin_detail_20210818_addname_dataframe['cname'] = margin_detail_20210818_tscode_list
for table_name in margin_detail_20210818.columns.values.tolist():
    margin_detail_20210818_addname_dataframe[table_name] = margin_detail_20210818[table_name]
print("融资融券交易明细  margin_detail_20210818 返回数据 row 行数 = "+str(margin_detail_20210818.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210818_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210819")       ###  更新 记录日期
margin_detail_20210819 = pro.margin_detail(trade_date='20210819', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210819_tscode_list = list() 
margin_detail_20210819_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210819.columns.values.tolist():
    for ts_code_sh in margin_detail_20210819['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210819_tscode_list.append(stock_name)
    margin_detail_20210819_addname_dataframe['cname'] = margin_detail_20210819_tscode_list
for table_name in margin_detail_20210819.columns.values.tolist():
    margin_detail_20210819_addname_dataframe[table_name] = margin_detail_20210819[table_name]
print("融资融券交易明细  margin_detail_20210819 返回数据 row 行数 = "+str(margin_detail_20210819.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210819_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210820")       ###  更新 记录日期
margin_detail_20210820 = pro.margin_detail(trade_date='20210820', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210820_tscode_list = list() 
margin_detail_20210820_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210820.columns.values.tolist():
    for ts_code_sh in margin_detail_20210820['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210820_tscode_list.append(stock_name)
    margin_detail_20210820_addname_dataframe['cname'] = margin_detail_20210820_tscode_list
for table_name in margin_detail_20210820.columns.values.tolist():
    margin_detail_20210820_addname_dataframe[table_name] = margin_detail_20210820[table_name]
print("融资融券交易明细  margin_detail_20210820 返回数据 row 行数 = "+str(margin_detail_20210820.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210820_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210823")       ###  更新 记录日期
margin_detail_20210823 = pro.margin_detail(trade_date='20210823', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210823_tscode_list = list() 
margin_detail_20210823_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210823.columns.values.tolist():
    for ts_code_sh in margin_detail_20210823['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210823_tscode_list.append(stock_name)
    margin_detail_20210823_addname_dataframe['cname'] = margin_detail_20210823_tscode_list
for table_name in margin_detail_20210823.columns.values.tolist():
    margin_detail_20210823_addname_dataframe[table_name] = margin_detail_20210823[table_name]
print("融资融券交易明细  margin_detail_20210823 返回数据 row 行数 = "+str(margin_detail_20210823.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210823_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210824")       ###  更新 记录日期
margin_detail_20210824 = pro.margin_detail(trade_date='20210824', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210824_tscode_list = list() 
margin_detail_20210824_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210824.columns.values.tolist():
    for ts_code_sh in margin_detail_20210824['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210824_tscode_list.append(stock_name)
    margin_detail_20210824_addname_dataframe['cname'] = margin_detail_20210824_tscode_list
for table_name in margin_detail_20210824.columns.values.tolist():
    margin_detail_20210824_addname_dataframe[table_name] = margin_detail_20210824[table_name]
print("融资融券交易明细  margin_detail_20210824 返回数据 row 行数 = "+str(margin_detail_20210824.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210824_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210825")       ###  更新 记录日期
margin_detail_20210825 = pro.margin_detail(trade_date='20210825', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210825_tscode_list = list() 
margin_detail_20210825_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210825.columns.values.tolist():
    for ts_code_sh in margin_detail_20210825['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210825_tscode_list.append(stock_name)
    margin_detail_20210825_addname_dataframe['cname'] = margin_detail_20210825_tscode_list
for table_name in margin_detail_20210825.columns.values.tolist():
    margin_detail_20210825_addname_dataframe[table_name] = margin_detail_20210825[table_name]
print("融资融券交易明细  margin_detail_20210825 返回数据 row 行数 = "+str(margin_detail_20210825.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210825_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210826")       ###  更新 记录日期
margin_detail_20210826 = pro.margin_detail(trade_date='20210826', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210826_tscode_list = list() 
margin_detail_20210826_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210826.columns.values.tolist():
    for ts_code_sh in margin_detail_20210826['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210826_tscode_list.append(stock_name)
    margin_detail_20210826_addname_dataframe['cname'] = margin_detail_20210826_tscode_list
for table_name in margin_detail_20210826.columns.values.tolist():
    margin_detail_20210826_addname_dataframe[table_name] = margin_detail_20210826[table_name]
print("融资融券交易明细  margin_detail_20210826 返回数据 row 行数 = "+str(margin_detail_20210826.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210826_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210827")       ###  更新 记录日期
margin_detail_20210827 = pro.margin_detail(trade_date='20210827', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210827_tscode_list = list() 
margin_detail_20210827_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210827.columns.values.tolist():
    for ts_code_sh in margin_detail_20210827['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210827_tscode_list.append(stock_name)
    margin_detail_20210827_addname_dataframe['cname'] = margin_detail_20210827_tscode_list
for table_name in margin_detail_20210827.columns.values.tolist():
    margin_detail_20210827_addname_dataframe[table_name] = margin_detail_20210827[table_name]
print("融资融券交易明细  margin_detail_20210827 返回数据 row 行数 = "+str(margin_detail_20210827.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210827_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210830")       ###  更新 记录日期
margin_detail_20210830 = pro.margin_detail(trade_date='20210830', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210830_tscode_list = list() 
margin_detail_20210830_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210830.columns.values.tolist():
    for ts_code_sh in margin_detail_20210830['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210830_tscode_list.append(stock_name)
    margin_detail_20210830_addname_dataframe['cname'] = margin_detail_20210830_tscode_list
for table_name in margin_detail_20210830.columns.values.tolist():
    margin_detail_20210830_addname_dataframe[table_name] = margin_detail_20210830[table_name]
print("融资融券交易明细  margin_detail_20210830 返回数据 row 行数 = "+str(margin_detail_20210830.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210830_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210831")       ###  更新 记录日期
margin_detail_20210831 = pro.margin_detail(trade_date='20210831', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210831_tscode_list = list() 
margin_detail_20210831_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210831.columns.values.tolist():
    for ts_code_sh in margin_detail_20210831['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210831_tscode_list.append(stock_name)
    margin_detail_20210831_addname_dataframe['cname'] = margin_detail_20210831_tscode_list
for table_name in margin_detail_20210831.columns.values.tolist():
    margin_detail_20210831_addname_dataframe[table_name] = margin_detail_20210831[table_name]
print("融资融券交易明细  margin_detail_20210831 返回数据 row 行数 = "+str(margin_detail_20210831.shape[0]))
margin_detail_2020_8_xlsx_frame=margin_detail_2020_8_xlsx_frame.append(margin_detail_20210831_addname_dataframe,ignore_index=True)
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_8_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'8',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_9_xlsx_frame=pd.DataFrame()
if '9' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_9_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='9' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20210901")       ###  更新 记录日期
margin_detail_20210901 = pro.margin_detail(trade_date='20210901', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210901_tscode_list = list() 
margin_detail_20210901_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210901.columns.values.tolist():
    for ts_code_sh in margin_detail_20210901['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210901_tscode_list.append(stock_name)
    margin_detail_20210901_addname_dataframe['cname'] = margin_detail_20210901_tscode_list
for table_name in margin_detail_20210901.columns.values.tolist():
    margin_detail_20210901_addname_dataframe[table_name] = margin_detail_20210901[table_name]
print("融资融券交易明细  margin_detail_20210901 返回数据 row 行数 = "+str(margin_detail_20210901.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210901_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210902")       ###  更新 记录日期
margin_detail_20210902 = pro.margin_detail(trade_date='20210902', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210902_tscode_list = list() 
margin_detail_20210902_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210902.columns.values.tolist():
    for ts_code_sh in margin_detail_20210902['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210902_tscode_list.append(stock_name)
    margin_detail_20210902_addname_dataframe['cname'] = margin_detail_20210902_tscode_list
for table_name in margin_detail_20210902.columns.values.tolist():
    margin_detail_20210902_addname_dataframe[table_name] = margin_detail_20210902[table_name]
print("融资融券交易明细  margin_detail_20210902 返回数据 row 行数 = "+str(margin_detail_20210902.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210902_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210903")       ###  更新 记录日期
margin_detail_20210903 = pro.margin_detail(trade_date='20210903', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210903_tscode_list = list() 
margin_detail_20210903_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210903.columns.values.tolist():
    for ts_code_sh in margin_detail_20210903['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210903_tscode_list.append(stock_name)
    margin_detail_20210903_addname_dataframe['cname'] = margin_detail_20210903_tscode_list
for table_name in margin_detail_20210903.columns.values.tolist():
    margin_detail_20210903_addname_dataframe[table_name] = margin_detail_20210903[table_name]
print("融资融券交易明细  margin_detail_20210903 返回数据 row 行数 = "+str(margin_detail_20210903.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210903_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210906")       ###  更新 记录日期
margin_detail_20210906 = pro.margin_detail(trade_date='20210906', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210906_tscode_list = list() 
margin_detail_20210906_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210906.columns.values.tolist():
    for ts_code_sh in margin_detail_20210906['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210906_tscode_list.append(stock_name)
    margin_detail_20210906_addname_dataframe['cname'] = margin_detail_20210906_tscode_list
for table_name in margin_detail_20210906.columns.values.tolist():
    margin_detail_20210906_addname_dataframe[table_name] = margin_detail_20210906[table_name]
print("融资融券交易明细  margin_detail_20210906 返回数据 row 行数 = "+str(margin_detail_20210906.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210906_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210907")       ###  更新 记录日期
margin_detail_20210907 = pro.margin_detail(trade_date='20210907', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210907_tscode_list = list() 
margin_detail_20210907_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210907.columns.values.tolist():
    for ts_code_sh in margin_detail_20210907['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210907_tscode_list.append(stock_name)
    margin_detail_20210907_addname_dataframe['cname'] = margin_detail_20210907_tscode_list
for table_name in margin_detail_20210907.columns.values.tolist():
    margin_detail_20210907_addname_dataframe[table_name] = margin_detail_20210907[table_name]
print("融资融券交易明细  margin_detail_20210907 返回数据 row 行数 = "+str(margin_detail_20210907.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210907_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210908")       ###  更新 记录日期
margin_detail_20210908 = pro.margin_detail(trade_date='20210908', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210908_tscode_list = list() 
margin_detail_20210908_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210908.columns.values.tolist():
    for ts_code_sh in margin_detail_20210908['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210908_tscode_list.append(stock_name)
    margin_detail_20210908_addname_dataframe['cname'] = margin_detail_20210908_tscode_list
for table_name in margin_detail_20210908.columns.values.tolist():
    margin_detail_20210908_addname_dataframe[table_name] = margin_detail_20210908[table_name]
print("融资融券交易明细  margin_detail_20210908 返回数据 row 行数 = "+str(margin_detail_20210908.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210908_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210909")       ###  更新 记录日期
margin_detail_20210909 = pro.margin_detail(trade_date='20210909', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210909_tscode_list = list() 
margin_detail_20210909_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210909.columns.values.tolist():
    for ts_code_sh in margin_detail_20210909['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210909_tscode_list.append(stock_name)
    margin_detail_20210909_addname_dataframe['cname'] = margin_detail_20210909_tscode_list
for table_name in margin_detail_20210909.columns.values.tolist():
    margin_detail_20210909_addname_dataframe[table_name] = margin_detail_20210909[table_name]
print("融资融券交易明细  margin_detail_20210909 返回数据 row 行数 = "+str(margin_detail_20210909.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210909_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210910")       ###  更新 记录日期
margin_detail_20210910 = pro.margin_detail(trade_date='20210910', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210910_tscode_list = list() 
margin_detail_20210910_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210910.columns.values.tolist():
    for ts_code_sh in margin_detail_20210910['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210910_tscode_list.append(stock_name)
    margin_detail_20210910_addname_dataframe['cname'] = margin_detail_20210910_tscode_list
for table_name in margin_detail_20210910.columns.values.tolist():
    margin_detail_20210910_addname_dataframe[table_name] = margin_detail_20210910[table_name]
print("融资融券交易明细  margin_detail_20210910 返回数据 row 行数 = "+str(margin_detail_20210910.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210910_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210913")       ###  更新 记录日期
margin_detail_20210913 = pro.margin_detail(trade_date='20210913', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210913_tscode_list = list() 
margin_detail_20210913_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210913.columns.values.tolist():
    for ts_code_sh in margin_detail_20210913['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210913_tscode_list.append(stock_name)
    margin_detail_20210913_addname_dataframe['cname'] = margin_detail_20210913_tscode_list
for table_name in margin_detail_20210913.columns.values.tolist():
    margin_detail_20210913_addname_dataframe[table_name] = margin_detail_20210913[table_name]
print("融资融券交易明细  margin_detail_20210913 返回数据 row 行数 = "+str(margin_detail_20210913.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210913_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210914")       ###  更新 记录日期
margin_detail_20210914 = pro.margin_detail(trade_date='20210914', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210914_tscode_list = list() 
margin_detail_20210914_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210914.columns.values.tolist():
    for ts_code_sh in margin_detail_20210914['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210914_tscode_list.append(stock_name)
    margin_detail_20210914_addname_dataframe['cname'] = margin_detail_20210914_tscode_list
for table_name in margin_detail_20210914.columns.values.tolist():
    margin_detail_20210914_addname_dataframe[table_name] = margin_detail_20210914[table_name]
print("融资融券交易明细  margin_detail_20210914 返回数据 row 行数 = "+str(margin_detail_20210914.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210914_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210915")       ###  更新 记录日期
margin_detail_20210915 = pro.margin_detail(trade_date='20210915', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210915_tscode_list = list() 
margin_detail_20210915_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210915.columns.values.tolist():
    for ts_code_sh in margin_detail_20210915['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210915_tscode_list.append(stock_name)
    margin_detail_20210915_addname_dataframe['cname'] = margin_detail_20210915_tscode_list
for table_name in margin_detail_20210915.columns.values.tolist():
    margin_detail_20210915_addname_dataframe[table_name] = margin_detail_20210915[table_name]
print("融资融券交易明细  margin_detail_20210915 返回数据 row 行数 = "+str(margin_detail_20210915.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210915_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210916")       ###  更新 记录日期
margin_detail_20210916 = pro.margin_detail(trade_date='20210916', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210916_tscode_list = list() 
margin_detail_20210916_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210916.columns.values.tolist():
    for ts_code_sh in margin_detail_20210916['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210916_tscode_list.append(stock_name)
    margin_detail_20210916_addname_dataframe['cname'] = margin_detail_20210916_tscode_list
for table_name in margin_detail_20210916.columns.values.tolist():
    margin_detail_20210916_addname_dataframe[table_name] = margin_detail_20210916[table_name]
print("融资融券交易明细  margin_detail_20210916 返回数据 row 行数 = "+str(margin_detail_20210916.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210916_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210917")       ###  更新 记录日期
margin_detail_20210917 = pro.margin_detail(trade_date='20210917', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210917_tscode_list = list() 
margin_detail_20210917_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210917.columns.values.tolist():
    for ts_code_sh in margin_detail_20210917['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210917_tscode_list.append(stock_name)
    margin_detail_20210917_addname_dataframe['cname'] = margin_detail_20210917_tscode_list
for table_name in margin_detail_20210917.columns.values.tolist():
    margin_detail_20210917_addname_dataframe[table_name] = margin_detail_20210917[table_name]
print("融资融券交易明细  margin_detail_20210917 返回数据 row 行数 = "+str(margin_detail_20210917.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210917_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210922")       ###  更新 记录日期
margin_detail_20210922 = pro.margin_detail(trade_date='20210922', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210922_tscode_list = list() 
margin_detail_20210922_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210922.columns.values.tolist():
    for ts_code_sh in margin_detail_20210922['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210922_tscode_list.append(stock_name)
    margin_detail_20210922_addname_dataframe['cname'] = margin_detail_20210922_tscode_list
for table_name in margin_detail_20210922.columns.values.tolist():
    margin_detail_20210922_addname_dataframe[table_name] = margin_detail_20210922[table_name]
print("融资融券交易明细  margin_detail_20210922 返回数据 row 行数 = "+str(margin_detail_20210922.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210922_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210923")       ###  更新 记录日期
margin_detail_20210923 = pro.margin_detail(trade_date='20210923', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210923_tscode_list = list() 
margin_detail_20210923_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210923.columns.values.tolist():
    for ts_code_sh in margin_detail_20210923['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210923_tscode_list.append(stock_name)
    margin_detail_20210923_addname_dataframe['cname'] = margin_detail_20210923_tscode_list
for table_name in margin_detail_20210923.columns.values.tolist():
    margin_detail_20210923_addname_dataframe[table_name] = margin_detail_20210923[table_name]
print("融资融券交易明细  margin_detail_20210923 返回数据 row 行数 = "+str(margin_detail_20210923.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210923_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210924")       ###  更新 记录日期
margin_detail_20210924 = pro.margin_detail(trade_date='20210924', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210924_tscode_list = list() 
margin_detail_20210924_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210924.columns.values.tolist():
    for ts_code_sh in margin_detail_20210924['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210924_tscode_list.append(stock_name)
    margin_detail_20210924_addname_dataframe['cname'] = margin_detail_20210924_tscode_list
for table_name in margin_detail_20210924.columns.values.tolist():
    margin_detail_20210924_addname_dataframe[table_name] = margin_detail_20210924[table_name]
print("融资融券交易明细  margin_detail_20210924 返回数据 row 行数 = "+str(margin_detail_20210924.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210924_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210927")       ###  更新 记录日期
margin_detail_20210927 = pro.margin_detail(trade_date='20210927', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210927_tscode_list = list() 
margin_detail_20210927_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210927.columns.values.tolist():
    for ts_code_sh in margin_detail_20210927['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210927_tscode_list.append(stock_name)
    margin_detail_20210927_addname_dataframe['cname'] = margin_detail_20210927_tscode_list
for table_name in margin_detail_20210927.columns.values.tolist():
    margin_detail_20210927_addname_dataframe[table_name] = margin_detail_20210927[table_name]
print("融资融券交易明细  margin_detail_20210927 返回数据 row 行数 = "+str(margin_detail_20210927.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210927_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210928")       ###  更新 记录日期
margin_detail_20210928 = pro.margin_detail(trade_date='20210928', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210928_tscode_list = list() 
margin_detail_20210928_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210928.columns.values.tolist():
    for ts_code_sh in margin_detail_20210928['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210928_tscode_list.append(stock_name)
    margin_detail_20210928_addname_dataframe['cname'] = margin_detail_20210928_tscode_list
for table_name in margin_detail_20210928.columns.values.tolist():
    margin_detail_20210928_addname_dataframe[table_name] = margin_detail_20210928[table_name]
print("融资融券交易明细  margin_detail_20210928 返回数据 row 行数 = "+str(margin_detail_20210928.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210928_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210929")       ###  更新 记录日期
margin_detail_20210929 = pro.margin_detail(trade_date='20210929', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210929_tscode_list = list() 
margin_detail_20210929_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210929.columns.values.tolist():
    for ts_code_sh in margin_detail_20210929['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210929_tscode_list.append(stock_name)
    margin_detail_20210929_addname_dataframe['cname'] = margin_detail_20210929_tscode_list
for table_name in margin_detail_20210929.columns.values.tolist():
    margin_detail_20210929_addname_dataframe[table_name] = margin_detail_20210929[table_name]
print("融资融券交易明细  margin_detail_20210929 返回数据 row 行数 = "+str(margin_detail_20210929.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210929_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20210930")       ###  更新 记录日期
margin_detail_20210930 = pro.margin_detail(trade_date='20210930', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20210930_tscode_list = list() 
margin_detail_20210930_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20210930.columns.values.tolist():
    for ts_code_sh in margin_detail_20210930['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20210930_tscode_list.append(stock_name)
    margin_detail_20210930_addname_dataframe['cname'] = margin_detail_20210930_tscode_list
for table_name in margin_detail_20210930.columns.values.tolist():
    margin_detail_20210930_addname_dataframe[table_name] = margin_detail_20210930[table_name]
print("融资融券交易明细  margin_detail_20210930 返回数据 row 行数 = "+str(margin_detail_20210930.shape[0]))
margin_detail_2020_9_xlsx_frame=margin_detail_2020_9_xlsx_frame.append(margin_detail_20210930_addname_dataframe,ignore_index=True)
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_9_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'9',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_10_xlsx_frame=pd.DataFrame()
if '10' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_10_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='10' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20211008")       ###  更新 记录日期
margin_detail_20211008 = pro.margin_detail(trade_date='20211008', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211008_tscode_list = list() 
margin_detail_20211008_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211008.columns.values.tolist():
    for ts_code_sh in margin_detail_20211008['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211008_tscode_list.append(stock_name)
    margin_detail_20211008_addname_dataframe['cname'] = margin_detail_20211008_tscode_list
for table_name in margin_detail_20211008.columns.values.tolist():
    margin_detail_20211008_addname_dataframe[table_name] = margin_detail_20211008[table_name]
print("融资融券交易明细  margin_detail_20211008 返回数据 row 行数 = "+str(margin_detail_20211008.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211008_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211011")       ###  更新 记录日期
margin_detail_20211011 = pro.margin_detail(trade_date='20211011', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211011_tscode_list = list() 
margin_detail_20211011_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211011.columns.values.tolist():
    for ts_code_sh in margin_detail_20211011['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211011_tscode_list.append(stock_name)
    margin_detail_20211011_addname_dataframe['cname'] = margin_detail_20211011_tscode_list
for table_name in margin_detail_20211011.columns.values.tolist():
    margin_detail_20211011_addname_dataframe[table_name] = margin_detail_20211011[table_name]
print("融资融券交易明细  margin_detail_20211011 返回数据 row 行数 = "+str(margin_detail_20211011.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211011_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211012")       ###  更新 记录日期
margin_detail_20211012 = pro.margin_detail(trade_date='20211012', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211012_tscode_list = list() 
margin_detail_20211012_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211012.columns.values.tolist():
    for ts_code_sh in margin_detail_20211012['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211012_tscode_list.append(stock_name)
    margin_detail_20211012_addname_dataframe['cname'] = margin_detail_20211012_tscode_list
for table_name in margin_detail_20211012.columns.values.tolist():
    margin_detail_20211012_addname_dataframe[table_name] = margin_detail_20211012[table_name]
print("融资融券交易明细  margin_detail_20211012 返回数据 row 行数 = "+str(margin_detail_20211012.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211012_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211013")       ###  更新 记录日期
margin_detail_20211013 = pro.margin_detail(trade_date='20211013', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211013_tscode_list = list() 
margin_detail_20211013_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211013.columns.values.tolist():
    for ts_code_sh in margin_detail_20211013['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211013_tscode_list.append(stock_name)
    margin_detail_20211013_addname_dataframe['cname'] = margin_detail_20211013_tscode_list
for table_name in margin_detail_20211013.columns.values.tolist():
    margin_detail_20211013_addname_dataframe[table_name] = margin_detail_20211013[table_name]
print("融资融券交易明细  margin_detail_20211013 返回数据 row 行数 = "+str(margin_detail_20211013.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211013_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211014")       ###  更新 记录日期
margin_detail_20211014 = pro.margin_detail(trade_date='20211014', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211014_tscode_list = list() 
margin_detail_20211014_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211014.columns.values.tolist():
    for ts_code_sh in margin_detail_20211014['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211014_tscode_list.append(stock_name)
    margin_detail_20211014_addname_dataframe['cname'] = margin_detail_20211014_tscode_list
for table_name in margin_detail_20211014.columns.values.tolist():
    margin_detail_20211014_addname_dataframe[table_name] = margin_detail_20211014[table_name]
print("融资融券交易明细  margin_detail_20211014 返回数据 row 行数 = "+str(margin_detail_20211014.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211014_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211015")       ###  更新 记录日期
margin_detail_20211015 = pro.margin_detail(trade_date='20211015', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211015_tscode_list = list() 
margin_detail_20211015_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211015.columns.values.tolist():
    for ts_code_sh in margin_detail_20211015['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211015_tscode_list.append(stock_name)
    margin_detail_20211015_addname_dataframe['cname'] = margin_detail_20211015_tscode_list
for table_name in margin_detail_20211015.columns.values.tolist():
    margin_detail_20211015_addname_dataframe[table_name] = margin_detail_20211015[table_name]
print("融资融券交易明细  margin_detail_20211015 返回数据 row 行数 = "+str(margin_detail_20211015.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211015_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211018")       ###  更新 记录日期
margin_detail_20211018 = pro.margin_detail(trade_date='20211018', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211018_tscode_list = list() 
margin_detail_20211018_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211018.columns.values.tolist():
    for ts_code_sh in margin_detail_20211018['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211018_tscode_list.append(stock_name)
    margin_detail_20211018_addname_dataframe['cname'] = margin_detail_20211018_tscode_list
for table_name in margin_detail_20211018.columns.values.tolist():
    margin_detail_20211018_addname_dataframe[table_name] = margin_detail_20211018[table_name]
print("融资融券交易明细  margin_detail_20211018 返回数据 row 行数 = "+str(margin_detail_20211018.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211018_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211019")       ###  更新 记录日期
margin_detail_20211019 = pro.margin_detail(trade_date='20211019', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211019_tscode_list = list() 
margin_detail_20211019_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211019.columns.values.tolist():
    for ts_code_sh in margin_detail_20211019['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211019_tscode_list.append(stock_name)
    margin_detail_20211019_addname_dataframe['cname'] = margin_detail_20211019_tscode_list
for table_name in margin_detail_20211019.columns.values.tolist():
    margin_detail_20211019_addname_dataframe[table_name] = margin_detail_20211019[table_name]
print("融资融券交易明细  margin_detail_20211019 返回数据 row 行数 = "+str(margin_detail_20211019.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211019_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211020")       ###  更新 记录日期
margin_detail_20211020 = pro.margin_detail(trade_date='20211020', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211020_tscode_list = list() 
margin_detail_20211020_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211020.columns.values.tolist():
    for ts_code_sh in margin_detail_20211020['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211020_tscode_list.append(stock_name)
    margin_detail_20211020_addname_dataframe['cname'] = margin_detail_20211020_tscode_list
for table_name in margin_detail_20211020.columns.values.tolist():
    margin_detail_20211020_addname_dataframe[table_name] = margin_detail_20211020[table_name]
print("融资融券交易明细  margin_detail_20211020 返回数据 row 行数 = "+str(margin_detail_20211020.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211020_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211021")       ###  更新 记录日期
margin_detail_20211021 = pro.margin_detail(trade_date='20211021', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211021_tscode_list = list() 
margin_detail_20211021_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211021.columns.values.tolist():
    for ts_code_sh in margin_detail_20211021['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211021_tscode_list.append(stock_name)
    margin_detail_20211021_addname_dataframe['cname'] = margin_detail_20211021_tscode_list
for table_name in margin_detail_20211021.columns.values.tolist():
    margin_detail_20211021_addname_dataframe[table_name] = margin_detail_20211021[table_name]
print("融资融券交易明细  margin_detail_20211021 返回数据 row 行数 = "+str(margin_detail_20211021.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211021_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211022")       ###  更新 记录日期
margin_detail_20211022 = pro.margin_detail(trade_date='20211022', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211022_tscode_list = list() 
margin_detail_20211022_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211022.columns.values.tolist():
    for ts_code_sh in margin_detail_20211022['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211022_tscode_list.append(stock_name)
    margin_detail_20211022_addname_dataframe['cname'] = margin_detail_20211022_tscode_list
for table_name in margin_detail_20211022.columns.values.tolist():
    margin_detail_20211022_addname_dataframe[table_name] = margin_detail_20211022[table_name]
print("融资融券交易明细  margin_detail_20211022 返回数据 row 行数 = "+str(margin_detail_20211022.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211022_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211025")       ###  更新 记录日期
margin_detail_20211025 = pro.margin_detail(trade_date='20211025', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211025_tscode_list = list() 
margin_detail_20211025_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211025.columns.values.tolist():
    for ts_code_sh in margin_detail_20211025['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211025_tscode_list.append(stock_name)
    margin_detail_20211025_addname_dataframe['cname'] = margin_detail_20211025_tscode_list
for table_name in margin_detail_20211025.columns.values.tolist():
    margin_detail_20211025_addname_dataframe[table_name] = margin_detail_20211025[table_name]
print("融资融券交易明细  margin_detail_20211025 返回数据 row 行数 = "+str(margin_detail_20211025.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211025_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211026")       ###  更新 记录日期
margin_detail_20211026 = pro.margin_detail(trade_date='20211026', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211026_tscode_list = list() 
margin_detail_20211026_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211026.columns.values.tolist():
    for ts_code_sh in margin_detail_20211026['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211026_tscode_list.append(stock_name)
    margin_detail_20211026_addname_dataframe['cname'] = margin_detail_20211026_tscode_list
for table_name in margin_detail_20211026.columns.values.tolist():
    margin_detail_20211026_addname_dataframe[table_name] = margin_detail_20211026[table_name]
print("融资融券交易明细  margin_detail_20211026 返回数据 row 行数 = "+str(margin_detail_20211026.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211026_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211027")       ###  更新 记录日期
margin_detail_20211027 = pro.margin_detail(trade_date='20211027', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211027_tscode_list = list() 
margin_detail_20211027_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211027.columns.values.tolist():
    for ts_code_sh in margin_detail_20211027['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211027_tscode_list.append(stock_name)
    margin_detail_20211027_addname_dataframe['cname'] = margin_detail_20211027_tscode_list
for table_name in margin_detail_20211027.columns.values.tolist():
    margin_detail_20211027_addname_dataframe[table_name] = margin_detail_20211027[table_name]
print("融资融券交易明细  margin_detail_20211027 返回数据 row 行数 = "+str(margin_detail_20211027.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211027_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211028")       ###  更新 记录日期
margin_detail_20211028 = pro.margin_detail(trade_date='20211028', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211028_tscode_list = list() 
margin_detail_20211028_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211028.columns.values.tolist():
    for ts_code_sh in margin_detail_20211028['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211028_tscode_list.append(stock_name)
    margin_detail_20211028_addname_dataframe['cname'] = margin_detail_20211028_tscode_list
for table_name in margin_detail_20211028.columns.values.tolist():
    margin_detail_20211028_addname_dataframe[table_name] = margin_detail_20211028[table_name]
print("融资融券交易明细  margin_detail_20211028 返回数据 row 行数 = "+str(margin_detail_20211028.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211028_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211029")       ###  更新 记录日期
margin_detail_20211029 = pro.margin_detail(trade_date='20211029', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211029_tscode_list = list() 
margin_detail_20211029_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211029.columns.values.tolist():
    for ts_code_sh in margin_detail_20211029['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211029_tscode_list.append(stock_name)
    margin_detail_20211029_addname_dataframe['cname'] = margin_detail_20211029_tscode_list
for table_name in margin_detail_20211029.columns.values.tolist():
    margin_detail_20211029_addname_dataframe[table_name] = margin_detail_20211029[table_name]
print("融资融券交易明细  margin_detail_20211029 返回数据 row 行数 = "+str(margin_detail_20211029.shape[0]))
margin_detail_2020_10_xlsx_frame=margin_detail_2020_10_xlsx_frame.append(margin_detail_20211029_addname_dataframe,ignore_index=True)
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_10_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'10',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_11_xlsx_frame=pd.DataFrame()
if '11' in margin_detail_2020_excel_writer.sheets:
    margin_detail_2020_11_xlsx_frame=pd.read_excel('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\margin_detail_2020.xlsx',sheet_name ='11' , index=False)
J0_PROPS.put(tree_node_name+"record_date", "20211101")       ###  更新 记录日期
margin_detail_20211101 = pro.margin_detail(trade_date='20211101', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211101_tscode_list = list() 
margin_detail_20211101_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211101.columns.values.tolist():
    for ts_code_sh in margin_detail_20211101['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211101_tscode_list.append(stock_name)
    margin_detail_20211101_addname_dataframe['cname'] = margin_detail_20211101_tscode_list
for table_name in margin_detail_20211101.columns.values.tolist():
    margin_detail_20211101_addname_dataframe[table_name] = margin_detail_20211101[table_name]
print("融资融券交易明细  margin_detail_20211101 返回数据 row 行数 = "+str(margin_detail_20211101.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211101_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211102")       ###  更新 记录日期
margin_detail_20211102 = pro.margin_detail(trade_date='20211102', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211102_tscode_list = list() 
margin_detail_20211102_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211102.columns.values.tolist():
    for ts_code_sh in margin_detail_20211102['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211102_tscode_list.append(stock_name)
    margin_detail_20211102_addname_dataframe['cname'] = margin_detail_20211102_tscode_list
for table_name in margin_detail_20211102.columns.values.tolist():
    margin_detail_20211102_addname_dataframe[table_name] = margin_detail_20211102[table_name]
print("融资融券交易明细  margin_detail_20211102 返回数据 row 行数 = "+str(margin_detail_20211102.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211102_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211103")       ###  更新 记录日期
margin_detail_20211103 = pro.margin_detail(trade_date='20211103', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211103_tscode_list = list() 
margin_detail_20211103_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211103.columns.values.tolist():
    for ts_code_sh in margin_detail_20211103['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211103_tscode_list.append(stock_name)
    margin_detail_20211103_addname_dataframe['cname'] = margin_detail_20211103_tscode_list
for table_name in margin_detail_20211103.columns.values.tolist():
    margin_detail_20211103_addname_dataframe[table_name] = margin_detail_20211103[table_name]
print("融资融券交易明细  margin_detail_20211103 返回数据 row 行数 = "+str(margin_detail_20211103.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211103_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211104")       ###  更新 记录日期
margin_detail_20211104 = pro.margin_detail(trade_date='20211104', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211104_tscode_list = list() 
margin_detail_20211104_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211104.columns.values.tolist():
    for ts_code_sh in margin_detail_20211104['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211104_tscode_list.append(stock_name)
    margin_detail_20211104_addname_dataframe['cname'] = margin_detail_20211104_tscode_list
for table_name in margin_detail_20211104.columns.values.tolist():
    margin_detail_20211104_addname_dataframe[table_name] = margin_detail_20211104[table_name]
print("融资融券交易明细  margin_detail_20211104 返回数据 row 行数 = "+str(margin_detail_20211104.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211104_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211105")       ###  更新 记录日期
margin_detail_20211105 = pro.margin_detail(trade_date='20211105', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211105_tscode_list = list() 
margin_detail_20211105_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211105.columns.values.tolist():
    for ts_code_sh in margin_detail_20211105['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211105_tscode_list.append(stock_name)
    margin_detail_20211105_addname_dataframe['cname'] = margin_detail_20211105_tscode_list
for table_name in margin_detail_20211105.columns.values.tolist():
    margin_detail_20211105_addname_dataframe[table_name] = margin_detail_20211105[table_name]
print("融资融券交易明细  margin_detail_20211105 返回数据 row 行数 = "+str(margin_detail_20211105.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211105_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211108")       ###  更新 记录日期
margin_detail_20211108 = pro.margin_detail(trade_date='20211108', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211108_tscode_list = list() 
margin_detail_20211108_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211108.columns.values.tolist():
    for ts_code_sh in margin_detail_20211108['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211108_tscode_list.append(stock_name)
    margin_detail_20211108_addname_dataframe['cname'] = margin_detail_20211108_tscode_list
for table_name in margin_detail_20211108.columns.values.tolist():
    margin_detail_20211108_addname_dataframe[table_name] = margin_detail_20211108[table_name]
print("融资融券交易明细  margin_detail_20211108 返回数据 row 行数 = "+str(margin_detail_20211108.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211108_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211109")       ###  更新 记录日期
margin_detail_20211109 = pro.margin_detail(trade_date='20211109', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211109_tscode_list = list() 
margin_detail_20211109_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211109.columns.values.tolist():
    for ts_code_sh in margin_detail_20211109['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211109_tscode_list.append(stock_name)
    margin_detail_20211109_addname_dataframe['cname'] = margin_detail_20211109_tscode_list
for table_name in margin_detail_20211109.columns.values.tolist():
    margin_detail_20211109_addname_dataframe[table_name] = margin_detail_20211109[table_name]
print("融资融券交易明细  margin_detail_20211109 返回数据 row 行数 = "+str(margin_detail_20211109.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211109_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211110")       ###  更新 记录日期
margin_detail_20211110 = pro.margin_detail(trade_date='20211110', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211110_tscode_list = list() 
margin_detail_20211110_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211110.columns.values.tolist():
    for ts_code_sh in margin_detail_20211110['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211110_tscode_list.append(stock_name)
    margin_detail_20211110_addname_dataframe['cname'] = margin_detail_20211110_tscode_list
for table_name in margin_detail_20211110.columns.values.tolist():
    margin_detail_20211110_addname_dataframe[table_name] = margin_detail_20211110[table_name]
print("融资融券交易明细  margin_detail_20211110 返回数据 row 行数 = "+str(margin_detail_20211110.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211110_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211111")       ###  更新 记录日期
margin_detail_20211111 = pro.margin_detail(trade_date='20211111', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211111_tscode_list = list() 
margin_detail_20211111_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211111.columns.values.tolist():
    for ts_code_sh in margin_detail_20211111['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211111_tscode_list.append(stock_name)
    margin_detail_20211111_addname_dataframe['cname'] = margin_detail_20211111_tscode_list
for table_name in margin_detail_20211111.columns.values.tolist():
    margin_detail_20211111_addname_dataframe[table_name] = margin_detail_20211111[table_name]
print("融资融券交易明细  margin_detail_20211111 返回数据 row 行数 = "+str(margin_detail_20211111.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211111_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211112")       ###  更新 记录日期
margin_detail_20211112 = pro.margin_detail(trade_date='20211112', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211112_tscode_list = list() 
margin_detail_20211112_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211112.columns.values.tolist():
    for ts_code_sh in margin_detail_20211112['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211112_tscode_list.append(stock_name)
    margin_detail_20211112_addname_dataframe['cname'] = margin_detail_20211112_tscode_list
for table_name in margin_detail_20211112.columns.values.tolist():
    margin_detail_20211112_addname_dataframe[table_name] = margin_detail_20211112[table_name]
print("融资融券交易明细  margin_detail_20211112 返回数据 row 行数 = "+str(margin_detail_20211112.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211112_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211115")       ###  更新 记录日期
margin_detail_20211115 = pro.margin_detail(trade_date='20211115', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211115_tscode_list = list() 
margin_detail_20211115_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211115.columns.values.tolist():
    for ts_code_sh in margin_detail_20211115['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211115_tscode_list.append(stock_name)
    margin_detail_20211115_addname_dataframe['cname'] = margin_detail_20211115_tscode_list
for table_name in margin_detail_20211115.columns.values.tolist():
    margin_detail_20211115_addname_dataframe[table_name] = margin_detail_20211115[table_name]
print("融资融券交易明细  margin_detail_20211115 返回数据 row 行数 = "+str(margin_detail_20211115.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211115_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211116")       ###  更新 记录日期
margin_detail_20211116 = pro.margin_detail(trade_date='20211116', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211116_tscode_list = list() 
margin_detail_20211116_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211116.columns.values.tolist():
    for ts_code_sh in margin_detail_20211116['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211116_tscode_list.append(stock_name)
    margin_detail_20211116_addname_dataframe['cname'] = margin_detail_20211116_tscode_list
for table_name in margin_detail_20211116.columns.values.tolist():
    margin_detail_20211116_addname_dataframe[table_name] = margin_detail_20211116[table_name]
print("融资融券交易明细  margin_detail_20211116 返回数据 row 行数 = "+str(margin_detail_20211116.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211116_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211117")       ###  更新 记录日期
margin_detail_20211117 = pro.margin_detail(trade_date='20211117', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211117_tscode_list = list() 
margin_detail_20211117_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211117.columns.values.tolist():
    for ts_code_sh in margin_detail_20211117['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211117_tscode_list.append(stock_name)
    margin_detail_20211117_addname_dataframe['cname'] = margin_detail_20211117_tscode_list
for table_name in margin_detail_20211117.columns.values.tolist():
    margin_detail_20211117_addname_dataframe[table_name] = margin_detail_20211117[table_name]
print("融资融券交易明细  margin_detail_20211117 返回数据 row 行数 = "+str(margin_detail_20211117.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211117_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211118")       ###  更新 记录日期
margin_detail_20211118 = pro.margin_detail(trade_date='20211118', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211118_tscode_list = list() 
margin_detail_20211118_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211118.columns.values.tolist():
    for ts_code_sh in margin_detail_20211118['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211118_tscode_list.append(stock_name)
    margin_detail_20211118_addname_dataframe['cname'] = margin_detail_20211118_tscode_list
for table_name in margin_detail_20211118.columns.values.tolist():
    margin_detail_20211118_addname_dataframe[table_name] = margin_detail_20211118[table_name]
print("融资融券交易明细  margin_detail_20211118 返回数据 row 行数 = "+str(margin_detail_20211118.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211118_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211119")       ###  更新 记录日期
margin_detail_20211119 = pro.margin_detail(trade_date='20211119', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211119_tscode_list = list() 
margin_detail_20211119_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211119.columns.values.tolist():
    for ts_code_sh in margin_detail_20211119['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211119_tscode_list.append(stock_name)
    margin_detail_20211119_addname_dataframe['cname'] = margin_detail_20211119_tscode_list
for table_name in margin_detail_20211119.columns.values.tolist():
    margin_detail_20211119_addname_dataframe[table_name] = margin_detail_20211119[table_name]
print("融资融券交易明细  margin_detail_20211119 返回数据 row 行数 = "+str(margin_detail_20211119.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211119_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211122")       ###  更新 记录日期
margin_detail_20211122 = pro.margin_detail(trade_date='20211122', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211122_tscode_list = list() 
margin_detail_20211122_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211122.columns.values.tolist():
    for ts_code_sh in margin_detail_20211122['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211122_tscode_list.append(stock_name)
    margin_detail_20211122_addname_dataframe['cname'] = margin_detail_20211122_tscode_list
for table_name in margin_detail_20211122.columns.values.tolist():
    margin_detail_20211122_addname_dataframe[table_name] = margin_detail_20211122[table_name]
print("融资融券交易明细  margin_detail_20211122 返回数据 row 行数 = "+str(margin_detail_20211122.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211122_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211123")       ###  更新 记录日期
margin_detail_20211123 = pro.margin_detail(trade_date='20211123', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211123_tscode_list = list() 
margin_detail_20211123_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211123.columns.values.tolist():
    for ts_code_sh in margin_detail_20211123['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211123_tscode_list.append(stock_name)
    margin_detail_20211123_addname_dataframe['cname'] = margin_detail_20211123_tscode_list
for table_name in margin_detail_20211123.columns.values.tolist():
    margin_detail_20211123_addname_dataframe[table_name] = margin_detail_20211123[table_name]
print("融资融券交易明细  margin_detail_20211123 返回数据 row 行数 = "+str(margin_detail_20211123.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211123_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211124")       ###  更新 记录日期
margin_detail_20211124 = pro.margin_detail(trade_date='20211124', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211124_tscode_list = list() 
margin_detail_20211124_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211124.columns.values.tolist():
    for ts_code_sh in margin_detail_20211124['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211124_tscode_list.append(stock_name)
    margin_detail_20211124_addname_dataframe['cname'] = margin_detail_20211124_tscode_list
for table_name in margin_detail_20211124.columns.values.tolist():
    margin_detail_20211124_addname_dataframe[table_name] = margin_detail_20211124[table_name]
print("融资融券交易明细  margin_detail_20211124 返回数据 row 行数 = "+str(margin_detail_20211124.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211124_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211125")       ###  更新 记录日期
margin_detail_20211125 = pro.margin_detail(trade_date='20211125', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211125_tscode_list = list() 
margin_detail_20211125_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211125.columns.values.tolist():
    for ts_code_sh in margin_detail_20211125['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211125_tscode_list.append(stock_name)
    margin_detail_20211125_addname_dataframe['cname'] = margin_detail_20211125_tscode_list
for table_name in margin_detail_20211125.columns.values.tolist():
    margin_detail_20211125_addname_dataframe[table_name] = margin_detail_20211125[table_name]
print("融资融券交易明细  margin_detail_20211125 返回数据 row 行数 = "+str(margin_detail_20211125.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211125_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211126")       ###  更新 记录日期
margin_detail_20211126 = pro.margin_detail(trade_date='20211126', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211126_tscode_list = list() 
margin_detail_20211126_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211126.columns.values.tolist():
    for ts_code_sh in margin_detail_20211126['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211126_tscode_list.append(stock_name)
    margin_detail_20211126_addname_dataframe['cname'] = margin_detail_20211126_tscode_list
for table_name in margin_detail_20211126.columns.values.tolist():
    margin_detail_20211126_addname_dataframe[table_name] = margin_detail_20211126[table_name]
print("融资融券交易明细  margin_detail_20211126 返回数据 row 行数 = "+str(margin_detail_20211126.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211126_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211129")       ###  更新 记录日期
margin_detail_20211129 = pro.margin_detail(trade_date='20211129', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211129_tscode_list = list() 
margin_detail_20211129_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211129.columns.values.tolist():
    for ts_code_sh in margin_detail_20211129['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211129_tscode_list.append(stock_name)
    margin_detail_20211129_addname_dataframe['cname'] = margin_detail_20211129_tscode_list
for table_name in margin_detail_20211129.columns.values.tolist():
    margin_detail_20211129_addname_dataframe[table_name] = margin_detail_20211129[table_name]
print("融资融券交易明细  margin_detail_20211129 返回数据 row 行数 = "+str(margin_detail_20211129.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211129_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211130")       ###  更新 记录日期
margin_detail_20211130 = pro.margin_detail(trade_date='20211130', fields='trade_date,ts_code,name,rzye,rqye,rzmre,rqyl,rzche,rqchl,rqmcl,rzrqye')
margin_detail_20211130_tscode_list = list() 
margin_detail_20211130_addname_dataframe=pd.DataFrame()
if 'ts_code'  in margin_detail_20211130.columns.values.tolist():
    for ts_code_sh in margin_detail_20211130['ts_code']:
        stock_name = tscode_name_dict.get(ts_code_sh)
        if stock_name is None:
            stock_name = 'null'
        margin_detail_20211130_tscode_list.append(stock_name)
    margin_detail_20211130_addname_dataframe['cname'] = margin_detail_20211130_tscode_list
for table_name in margin_detail_20211130.columns.values.tolist():
    margin_detail_20211130_addname_dataframe[table_name] = margin_detail_20211130[table_name]
print("融资融券交易明细  margin_detail_20211130 返回数据 row 行数 = "+str(margin_detail_20211130.shape[0]))
margin_detail_2020_11_xlsx_frame=margin_detail_2020_11_xlsx_frame.append(margin_detail_20211130_addname_dataframe,ignore_index=True)
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
margin_detail_2020_11_xlsx_frame.to_excel(margin_detail_2020_excel_writer,'11',index=False)
margin_detail_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20211201")       ###  更新 记录日期
