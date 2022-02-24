#  encoding: utf-8

import tushare as ts
import os
import pandas as pd  # https://www.jianshu.com/p/5c0aa1fa19af
import openpyxl 
from openpyxl import load_workbook,Workbook
import time,datetime  # https://www.jb51.net/article/66019.htm
import re
import os
import tempfile

############################## 全局数据初始化块 数据初始化函数  End ##############################

##############  properities 函数 Begin ##############

class Properties:

    def __init__(self, file_name):
        self.file_name = file_name
        self.properties = {}
        try:
            fopen = open(self.file_name, 'r')
            for line in fopen:
                line = line.strip()
                if line.find('=') > 0 and not line.startswith('#'):
                    strs = line.split('=')
                    self.properties[strs[0].strip()] = strs[1].strip()
        except Exception :
            raise
        else:
            fopen.close()

    def has_key(self, key):
        return key in self.properties

    def get(self, key, default_value=''):
        if key in self.properties:
            return self.properties[key]
        return default_value

    def put(self, key, value):
        self.properties[key] = value
        replace_property(self.file_name, key + '=.*', key + '=' + value, True)


def replace_property(file_name, from_regex, to_str, append_on_not_exists=True):
    tmpfile = tempfile.TemporaryFile()

    if os.path.exists(file_name):
        r_open = open(file_name, 'r')
        pattern = re.compile(r'' + from_regex)
        found = None
        for line in r_open:
            if pattern.search(line) and not line.strip().startswith('#'):
                found = True
                line = re.sub(from_regex, to_str, line)
            tmpfile.write(line.encode())
        if not found and append_on_not_exists:
            tmpfile.write(('\n' + to_str).encode())
        r_open.close()
        tmpfile.seek(0)

        content = tmpfile.read()

        if os.path.exists(file_name):
            os.remove(file_name)

        w_open = open(file_name, 'wb')
        w_open.write(content)
        w_open.close()

        tmpfile.close()
    else:
        print ("file %s not found" % file_name)

##############  properities 函数 End ##############


##############  XLSX excel 函数 Begin ##############

def createexcel(filename):    ### 创建本地文件名称为 filename的文件
    wb = Workbook()
    curPath = os.path.abspath('.')
    cur_desktop_path=os.path.join(os.path.expanduser("~"), 'Desktop')
    curDir = cur_desktop_path+os.sep+"zbin"+os.sep+"J0_Data"+os.sep  ## 创建~/Desktop/zbin/J0_Data 这个 这个工作目录
    #curDir = curPath+os.sep
    if not os.path.exists(curDir):
        os.makedirs(curDir)
    #curPath = os.path.dirname(os.path.abspath('.'))
    #t = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    #suffix = '.xlsx' # 文件类型
    #newfile = t + suffix
    xlsxPath = curDir + filename
    print("curPath curDir=" + curPath)
    print("createexcel zbin_JO_Dir=" + curDir)
    print("createexcel path=" + xlsxPath)
    if not os.path.exists(xlsxPath):
        wb.save(xlsxPath)
        print(" wb.save(xlsxPath)  xlsxPath =" + xlsxPath)
        time.sleep(1)
    return xlsxPath

def getColumnIndex(table, columnName):   ## 返回 table 中 名称为  columnName 的 那列 的索引
    columnIndex = None
    for i in range(table.ncols):
        if(table.cell_value(0, i) == columnName):
            columnIndex = i
            break
    return columnIndex
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007

##############  XLSX excel 函数 End ##############



############################## Prop初始化 Begin ##############################

#
#Thu Aug 13 22:33:45 CST 2020
#rixianhangqing-time_record_date=20200707
#rixianhangqing-time_start_date=20010101

desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
zbin_path = str(desktop_path)+os.sep+"zbin"
j0_properties_path= str(zbin_path)+os.sep+"J0.properties"
print("desktop_path = "+ str(desktop_path))
print("zbin_path = "+ str(zbin_path))
print("j0_properties_path = "+ str(j0_properties_path))
J0_PROPS =  Properties(j0_properties_path)


############################## Prop初始化 End ##############################


############## tscode_股票列表的初始化  Begin ##############
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007
def init_tscode_data(book_name, sheet_name,ts_code_set,tscode_name_dict):
    # 读取excel
    wb = openpyxl.load_workbook(book_name)
    # 读取sheet表
    ws = wb[sheet_name]
    for i in range(ws.max_row):
         # 获取下拉框中的所有选择值
         if (i == 0 or i == 1):
             continue
         #print("i="+str(i)+" 总的列数:" + str(ws.max_row)+"  value:"+str(ws.cell(i,2).value))
         tscode_item = str(ws.cell(i,1).value)  ##  20201010--> 2:ts_code    4_name   ##  20201116--> 1:ts_code    3_name 
         tscode_name_item = str(ws.cell(i,3).value)
         #print("index="+str(i)+" tscode_item = "+ str(tscode_item) + "   tscode_name_item="+str(tscode_name_item))
         ts_code_set.add(str(ws.cell(i,1).value))
         tscode_name_dict[tscode_item]=tscode_name_item



tscode_set = set()    #### 股票代码的集合   000001.SZ   .... 999999.SH
tscode_name_dict = dict()  #### code-name 的 map的集合
init_tscode_data(zbin_path+os.sep+"J0_Python"+os.sep+"J0_股票列表.xlsx","股票列表",tscode_set,tscode_name_dict)
############## tscode_股票列表的初始化  End  ##############


############################## 运行属性 Begin ##############################
pd.set_option('display.max_rows', None)   ##  解决纵向出现...
#pd.set_option('display.width', 1000) 
pd.set_option('expand_frame_repr', False)  ##  解决横向出现...
Cur_Abs_Path=os.path.abspath('.')   # 表示当前所处的文件夹的绝对路径
print("当前绝对路径:"+Cur_Abs_Path)
Cur_Ref_Path=os.path.abspath('..')  # 表示当前所处的文件夹上一级文件夹的绝对路径
print("当前父目录绝对路径:"+Cur_Ref_Path)
############################## 运行属性 End ##############################

############################## 时间Date Begin ##############################

now_yyyymmdd=str(time.strftime('%Y%m%d', time.localtime(time.time())))
print("now_yyyymmdd = "+ str(now_yyyymmdd))

############################## 时间Date End ##############################





pro = ts.pro_api('43acb9a5ddc2cf73c6c4ea54796748f965457ed57daaa736bb778ea2')
# print(J0_PROPS.get(tree_node_name+"record_date"))           #根据key读取value
# J0_PROPS.put(tree_node_name+"record_date", now_yyyymmdd)       ###  覆盖原有的 key 和 value
#  zukgit
# shangshigongsiguanliceng_zukgit_website  =   https://tushare.pro/document/2?doc_id=193
createexcel('上市公司管理层.xlsx')
stk_managers_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\上市公司管理层.xlsx')
stk_managers_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\上市公司管理层.xlsx', engine='openpyxl')
stk_managers_excel_writer.book = stk_managers_book
stk_managers_excel_writer.sheets = dict((ws.title, ws) for ws in stk_managers_book.worksheets)
stk_managers_data_List = pd.DataFrame()
stk_managers_item_0 = pro.stk_managers(ts_code='000001.SZ,000002.SZ,000003.SZ,000004.SZ,000005.SZ,000006.SZ,000007.SZ,000008.SZ,000009.SZ,000010.SZ,000011.SZ,000012.SZ,000013.SZ,000014.SZ,000015.SZ,000016.SZ,000017.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_0 返回数据 row 行数 = "+str(stk_managers_item_0.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_0)


stk_managers_item_1 = pro.stk_managers(ts_code='000018.SZ,000019.SZ,000020.SZ,000021.SZ,000023.SZ,000024.SZ,000025.SZ,000026.SZ,000027.SZ,000028.SZ,000029.SZ,000030.SZ,000031.SZ,000032.SZ,000033.SZ,000034.SZ,000035.SZ,000036.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_1 返回数据 row 行数 = "+str(stk_managers_item_1.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_1)


stk_managers_item_2 = pro.stk_managers(ts_code='000037.SZ,000038.SZ,000039.SZ,000040.SZ,000042.SZ,000045.SZ,000046.SZ,000047.SZ,000048.SZ,000049.SZ,000050.SZ,000055.SZ,000056.SZ,000058.SZ,000059.SZ,000060.SZ,000061.SZ,000062.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_2 返回数据 row 行数 = "+str(stk_managers_item_2.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_2)


stk_managers_item_3 = pro.stk_managers(ts_code='000063.SZ,000065.SZ,000066.SZ,000068.SZ,000069.SZ,000070.SZ,000078.SZ,000088.SZ,000089.SZ,000090.SZ,000096.SZ,000099.SZ,000100.SZ,000150.SZ,000151.SZ,000153.SZ,000155.SZ,000156.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_3 返回数据 row 行数 = "+str(stk_managers_item_3.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_3)


stk_managers_item_4 = pro.stk_managers(ts_code='000157.SZ,000158.SZ,000159.SZ,000166.SZ,000301.SZ,000333.SZ,000338.SZ,000400.SZ,000401.SZ,000402.SZ,000403.SZ,000404.SZ,000405.SZ,000406.SZ,000407.SZ,000408.SZ,000409.SZ,000410.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_4 返回数据 row 行数 = "+str(stk_managers_item_4.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_4)


stk_managers_item_5 = pro.stk_managers(ts_code='000411.SZ,000412.SZ,000413.SZ,000415.SZ,000416.SZ,000417.SZ,000418.SZ,000419.SZ,000420.SZ,000421.SZ,000422.SZ,000423.SZ,000425.SZ,000426.SZ,000428.SZ,000429.SZ,000430.SZ,000488.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_5 返回数据 row 行数 = "+str(stk_managers_item_5.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_5)


stk_managers_item_6 = pro.stk_managers(ts_code='000498.SZ,000501.SZ,000502.SZ,000503.SZ,000504.SZ,000505.SZ,000506.SZ,000507.SZ,000508.SZ,000509.SZ,000510.SZ,000511.SZ,000513.SZ,000514.SZ,000515.SZ,000516.SZ,000517.SZ,000518.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_6 返回数据 row 行数 = "+str(stk_managers_item_6.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_6)


stk_managers_item_7 = pro.stk_managers(ts_code='000519.SZ,000520.SZ,000521.SZ,000522.SZ,000523.SZ,000524.SZ,000525.SZ,000526.SZ,000527.SZ,000528.SZ,000529.SZ,000530.SZ,000531.SZ,000532.SZ,000533.SZ,000534.SZ,000535.SZ,000536.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_7 返回数据 row 行数 = "+str(stk_managers_item_7.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_7)


stk_managers_item_8 = pro.stk_managers(ts_code='000537.SZ,000538.SZ,000539.SZ,000540.SZ,000541.SZ,000542.SZ,000543.SZ,000544.SZ,000545.SZ,000546.SZ,000547.SZ,000548.SZ,000549.SZ,000550.SZ,000551.SZ,000552.SZ,000553.SZ,000554.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_8 返回数据 row 行数 = "+str(stk_managers_item_8.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_8)


stk_managers_item_9 = pro.stk_managers(ts_code='000555.SZ,000556.SZ,000557.SZ,000558.SZ,000559.SZ,000560.SZ,000561.SZ,000562.SZ,000563.SZ,000564.SZ,000565.SZ,000566.SZ,000567.SZ,000568.SZ,000569.SZ,000570.SZ,000571.SZ,000572.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_9 返回数据 row 行数 = "+str(stk_managers_item_9.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_9)


stk_managers_item_10 = pro.stk_managers(ts_code='000573.SZ,000576.SZ,000578.SZ,000581.SZ,000582.SZ,000583.SZ,000584.SZ,000585.SZ,000586.SZ,000587.SZ,000588.SZ,000589.SZ,000590.SZ,000591.SZ,000592.SZ,000593.SZ,000594.SZ,000595.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_10 返回数据 row 行数 = "+str(stk_managers_item_10.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_10)


stk_managers_item_11 = pro.stk_managers(ts_code='000596.SZ,000597.SZ,000598.SZ,000599.SZ,000600.SZ,000601.SZ,000602.SZ,000603.SZ,000605.SZ,000606.SZ,000607.SZ,000608.SZ,000609.SZ,000610.SZ,000611.SZ,000612.SZ,000613.SZ,000615.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_11 返回数据 row 行数 = "+str(stk_managers_item_11.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_11)


stk_managers_item_12 = pro.stk_managers(ts_code='000616.SZ,000617.SZ,000618.SZ,000619.SZ,000620.SZ,000621.SZ,000622.SZ,000623.SZ,000625.SZ,000626.SZ,000627.SZ,000628.SZ,000629.SZ,000630.SZ,000631.SZ,000632.SZ,000633.SZ,000635.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_12 返回数据 row 行数 = "+str(stk_managers_item_12.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_12)


stk_managers_item_13 = pro.stk_managers(ts_code='000636.SZ,000637.SZ,000638.SZ,000639.SZ,000650.SZ,000651.SZ,000652.SZ,000653.SZ,000655.SZ,000656.SZ,000657.SZ,000658.SZ,000659.SZ,000660.SZ,000661.SZ,000662.SZ,000663.SZ,000665.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_13 返回数据 row 行数 = "+str(stk_managers_item_13.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_13)


stk_managers_item_14 = pro.stk_managers(ts_code='000666.SZ,000667.SZ,000668.SZ,000669.SZ,000670.SZ,000671.SZ,000672.SZ,000673.SZ,000675.SZ,000676.SZ,000677.SZ,000678.SZ,000679.SZ,000680.SZ,000681.SZ,000682.SZ,000683.SZ,000685.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_14 返回数据 row 行数 = "+str(stk_managers_item_14.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_14)


stk_managers_item_15 = pro.stk_managers(ts_code='000686.SZ,000687.SZ,000688.SZ,000689.SZ,000690.SZ,000691.SZ,000692.SZ,000693.SZ,000695.SZ,000697.SZ,000698.SZ,000699.SZ,000700.SZ,000701.SZ,000702.SZ,000703.SZ,000705.SZ,000707.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_15 返回数据 row 行数 = "+str(stk_managers_item_15.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_15)


stk_managers_item_16 = pro.stk_managers(ts_code='000708.SZ,000709.SZ,000710.SZ,000711.SZ,000712.SZ,000713.SZ,000715.SZ,000716.SZ,000717.SZ,000718.SZ,000719.SZ,000720.SZ,000721.SZ,000722.SZ,000723.SZ,000725.SZ,000726.SZ,000727.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_16 返回数据 row 行数 = "+str(stk_managers_item_16.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_16)


stk_managers_item_17 = pro.stk_managers(ts_code='000728.SZ,000729.SZ,000730.SZ,000731.SZ,000732.SZ,000733.SZ,000735.SZ,000736.SZ,000737.SZ,000738.SZ,000739.SZ,000748.SZ,000750.SZ,000751.SZ,000752.SZ,000753.SZ,000755.SZ,000756.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_17 返回数据 row 行数 = "+str(stk_managers_item_17.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_17)


stk_managers_item_18 = pro.stk_managers(ts_code='000757.SZ,000758.SZ,000759.SZ,000760.SZ,000761.SZ,000762.SZ,000763.SZ,000765.SZ,000766.SZ,000767.SZ,000768.SZ,000769.SZ,000776.SZ,000777.SZ,000778.SZ,000779.SZ,000780.SZ,000782.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_18 返回数据 row 行数 = "+str(stk_managers_item_18.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_18)


stk_managers_item_19 = pro.stk_managers(ts_code='000783.SZ,000785.SZ,000786.SZ,000787.SZ,000788.SZ,000789.SZ,000790.SZ,000791.SZ,000792.SZ,000793.SZ,000795.SZ,000796.SZ,000797.SZ,000798.SZ,000799.SZ,000800.SZ,000801.SZ,000802.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_19 返回数据 row 行数 = "+str(stk_managers_item_19.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_19)


stk_managers_item_20 = pro.stk_managers(ts_code='000803.SZ,000805.SZ,000806.SZ,000807.SZ,000809.SZ,000810.SZ,000811.SZ,000812.SZ,000813.SZ,000815.SZ,000816.SZ,000817.SZ,000818.SZ,000819.SZ,000820.SZ,000821.SZ,000822.SZ,000823.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_20 返回数据 row 行数 = "+str(stk_managers_item_20.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_20)


stk_managers_item_21 = pro.stk_managers(ts_code='000825.SZ,000826.SZ,000827.SZ,000828.SZ,000829.SZ,000830.SZ,000831.SZ,000832.SZ,000833.SZ,000835.SZ,000836.SZ,000837.SZ,000838.SZ,000839.SZ,000848.SZ,000850.SZ,000851.SZ,000852.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_21 返回数据 row 行数 = "+str(stk_managers_item_21.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_21)


stk_managers_item_22 = pro.stk_managers(ts_code='000856.SZ,000858.SZ,000859.SZ,000860.SZ,000861.SZ,000862.SZ,000863.SZ,000866.SZ,000868.SZ,000869.SZ,000875.SZ,000876.SZ,000877.SZ,000878.SZ,000880.SZ,000881.SZ,000882.SZ,000883.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_22 返回数据 row 行数 = "+str(stk_managers_item_22.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_22)


stk_managers_item_23 = pro.stk_managers(ts_code='000885.SZ,000886.SZ,000887.SZ,000888.SZ,000889.SZ,000890.SZ,000892.SZ,000893.SZ,000895.SZ,000897.SZ,000898.SZ,000899.SZ,000900.SZ,000901.SZ,000902.SZ,000903.SZ,000905.SZ,000906.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_23 返回数据 row 行数 = "+str(stk_managers_item_23.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_23)


stk_managers_item_24 = pro.stk_managers(ts_code='000908.SZ,000909.SZ,000910.SZ,000911.SZ,000912.SZ,000913.SZ,000915.SZ,000916.SZ,000917.SZ,000918.SZ,000919.SZ,000920.SZ,000921.SZ,000922.SZ,000923.SZ,000925.SZ,000926.SZ,000927.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_24 返回数据 row 行数 = "+str(stk_managers_item_24.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_24)


stk_managers_item_25 = pro.stk_managers(ts_code='000928.SZ,000929.SZ,000930.SZ,000931.SZ,000932.SZ,000933.SZ,000935.SZ,000936.SZ,000937.SZ,000938.SZ,000939.SZ,000948.SZ,000949.SZ,000950.SZ,000951.SZ,000952.SZ,000953.SZ,000955.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_25 返回数据 row 行数 = "+str(stk_managers_item_25.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_25)


stk_managers_item_26 = pro.stk_managers(ts_code='000956.SZ,000957.SZ,000958.SZ,000959.SZ,000960.SZ,000961.SZ,000962.SZ,000963.SZ,000965.SZ,000966.SZ,000967.SZ,000968.SZ,000969.SZ,000970.SZ,000971.SZ,000972.SZ,000973.SZ,000975.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_26 返回数据 row 行数 = "+str(stk_managers_item_26.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_26)


stk_managers_item_27 = pro.stk_managers(ts_code='000976.SZ,000977.SZ,000978.SZ,000979.SZ,000980.SZ,000981.SZ,000982.SZ,000983.SZ,000985.SZ,000987.SZ,000988.SZ,000989.SZ,000990.SZ,000993.SZ,000995.SZ,000996.SZ,000997.SZ,000998.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_27 返回数据 row 行数 = "+str(stk_managers_item_27.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_27)


stk_managers_item_28 = pro.stk_managers(ts_code='000999.SZ,001201.SZ,001202.SZ,001203.SZ,001205.SZ,001206.SZ,001207.SZ,001208.SZ,001696.SZ,001872.SZ,001896.SZ,001914.SZ,001965.SZ,001979.SZ,002001.SZ,002002.SZ,002003.SZ,002004.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_28 返回数据 row 行数 = "+str(stk_managers_item_28.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_28)


stk_managers_item_29 = pro.stk_managers(ts_code='002005.SZ,002006.SZ,002007.SZ,002008.SZ,002009.SZ,002010.SZ,002011.SZ,002012.SZ,002013.SZ,002014.SZ,002015.SZ,002016.SZ,002017.SZ,002018.SZ,002019.SZ,002020.SZ,002021.SZ,002022.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_29 返回数据 row 行数 = "+str(stk_managers_item_29.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_29)


stk_managers_item_30 = pro.stk_managers(ts_code='002023.SZ,002024.SZ,002025.SZ,002026.SZ,002027.SZ,002028.SZ,002029.SZ,002030.SZ,002031.SZ,002032.SZ,002033.SZ,002034.SZ,002035.SZ,002036.SZ,002037.SZ,002038.SZ,002039.SZ,002040.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_30 返回数据 row 行数 = "+str(stk_managers_item_30.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_30)


stk_managers_item_31 = pro.stk_managers(ts_code='002041.SZ,002042.SZ,002043.SZ,002044.SZ,002045.SZ,002046.SZ,002047.SZ,002048.SZ,002049.SZ,002050.SZ,002051.SZ,002052.SZ,002053.SZ,002054.SZ,002055.SZ,002056.SZ,002057.SZ,002058.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_31 返回数据 row 行数 = "+str(stk_managers_item_31.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_31)


stk_managers_item_32 = pro.stk_managers(ts_code='002059.SZ,002060.SZ,002061.SZ,002062.SZ,002063.SZ,002064.SZ,002065.SZ,002066.SZ,002067.SZ,002068.SZ,002069.SZ,002070.SZ,002071.SZ,002072.SZ,002073.SZ,002074.SZ,002075.SZ,002076.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_32 返回数据 row 行数 = "+str(stk_managers_item_32.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_32)


stk_managers_item_33 = pro.stk_managers(ts_code='002077.SZ,002078.SZ,002079.SZ,002080.SZ,002081.SZ,002082.SZ,002083.SZ,002084.SZ,002085.SZ,002086.SZ,002087.SZ,002088.SZ,002089.SZ,002090.SZ,002091.SZ,002092.SZ,002093.SZ,002094.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_33 返回数据 row 行数 = "+str(stk_managers_item_33.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_33)


stk_managers_item_34 = pro.stk_managers(ts_code='002095.SZ,002096.SZ,002097.SZ,002098.SZ,002099.SZ,002100.SZ,002101.SZ,002102.SZ,002103.SZ,002104.SZ,002105.SZ,002106.SZ,002107.SZ,002108.SZ,002109.SZ,002110.SZ,002111.SZ,002112.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_34 返回数据 row 行数 = "+str(stk_managers_item_34.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_34)


stk_managers_item_35 = pro.stk_managers(ts_code='002113.SZ,002114.SZ,002115.SZ,002116.SZ,002117.SZ,002118.SZ,002119.SZ,002120.SZ,002121.SZ,002122.SZ,002123.SZ,002124.SZ,002125.SZ,002126.SZ,002127.SZ,002128.SZ,002129.SZ,002130.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_35 返回数据 row 行数 = "+str(stk_managers_item_35.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_35)


stk_managers_item_36 = pro.stk_managers(ts_code='002131.SZ,002132.SZ,002133.SZ,002134.SZ,002135.SZ,002136.SZ,002137.SZ,002138.SZ,002139.SZ,002140.SZ,002141.SZ,002142.SZ,002143.SZ,002144.SZ,002145.SZ,002146.SZ,002147.SZ,002148.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_36 返回数据 row 行数 = "+str(stk_managers_item_36.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_36)


stk_managers_item_37 = pro.stk_managers(ts_code='002149.SZ,002150.SZ,002151.SZ,002152.SZ,002153.SZ,002154.SZ,002155.SZ,002156.SZ,002157.SZ,002158.SZ,002159.SZ,002160.SZ,002161.SZ,002162.SZ,002163.SZ,002164.SZ,002165.SZ,002166.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_37 返回数据 row 行数 = "+str(stk_managers_item_37.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_37)


stk_managers_item_38 = pro.stk_managers(ts_code='002167.SZ,002168.SZ,002169.SZ,002170.SZ,002171.SZ,002172.SZ,002173.SZ,002174.SZ,002175.SZ,002176.SZ,002177.SZ,002178.SZ,002179.SZ,002180.SZ,002181.SZ,002182.SZ,002183.SZ,002184.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_38 返回数据 row 行数 = "+str(stk_managers_item_38.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_38)


stk_managers_item_39 = pro.stk_managers(ts_code='002185.SZ,002186.SZ,002187.SZ,002188.SZ,002189.SZ,002190.SZ,002191.SZ,002192.SZ,002193.SZ,002194.SZ,002195.SZ,002196.SZ,002197.SZ,002198.SZ,002199.SZ,002200.SZ,002201.SZ,002202.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_39 返回数据 row 行数 = "+str(stk_managers_item_39.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_39)


stk_managers_item_40 = pro.stk_managers(ts_code='002203.SZ,002204.SZ,002205.SZ,002206.SZ,002207.SZ,002208.SZ,002209.SZ,002210.SZ,002211.SZ,002212.SZ,002213.SZ,002214.SZ,002215.SZ,002216.SZ,002217.SZ,002218.SZ,002219.SZ,002220.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_40 返回数据 row 行数 = "+str(stk_managers_item_40.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_40)


stk_managers_item_41 = pro.stk_managers(ts_code='002221.SZ,002222.SZ,002223.SZ,002224.SZ,002225.SZ,002226.SZ,002227.SZ,002228.SZ,002229.SZ,002230.SZ,002231.SZ,002232.SZ,002233.SZ,002234.SZ,002235.SZ,002236.SZ,002237.SZ,002238.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_41 返回数据 row 行数 = "+str(stk_managers_item_41.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_41)


stk_managers_item_42 = pro.stk_managers(ts_code='002239.SZ,002240.SZ,002241.SZ,002242.SZ,002243.SZ,002244.SZ,002245.SZ,002246.SZ,002247.SZ,002248.SZ,002249.SZ,002250.SZ,002251.SZ,002252.SZ,002253.SZ,002254.SZ,002255.SZ,002256.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_42 返回数据 row 行数 = "+str(stk_managers_item_42.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_42)


stk_managers_item_43 = pro.stk_managers(ts_code='002258.SZ,002259.SZ,002260.SZ,002261.SZ,002262.SZ,002263.SZ,002264.SZ,002265.SZ,002266.SZ,002267.SZ,002268.SZ,002269.SZ,002270.SZ,002271.SZ,002272.SZ,002273.SZ,002274.SZ,002275.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_43 返回数据 row 行数 = "+str(stk_managers_item_43.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_43)


stk_managers_item_44 = pro.stk_managers(ts_code='002276.SZ,002277.SZ,002278.SZ,002279.SZ,002280.SZ,002281.SZ,002282.SZ,002283.SZ,002284.SZ,002285.SZ,002286.SZ,002287.SZ,002288.SZ,002289.SZ,002290.SZ,002291.SZ,002292.SZ,002293.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_44 返回数据 row 行数 = "+str(stk_managers_item_44.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_44)


stk_managers_item_45 = pro.stk_managers(ts_code='002294.SZ,002295.SZ,002296.SZ,002297.SZ,002298.SZ,002299.SZ,002300.SZ,002301.SZ,002302.SZ,002303.SZ,002304.SZ,002305.SZ,002306.SZ,002307.SZ,002308.SZ,002309.SZ,002310.SZ,002311.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_45 返回数据 row 行数 = "+str(stk_managers_item_45.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_45)


stk_managers_item_46 = pro.stk_managers(ts_code='002312.SZ,002313.SZ,002314.SZ,002315.SZ,002316.SZ,002317.SZ,002318.SZ,002319.SZ,002320.SZ,002321.SZ,002322.SZ,002323.SZ,002324.SZ,002325.SZ,002326.SZ,002327.SZ,002328.SZ,002329.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_46 返回数据 row 行数 = "+str(stk_managers_item_46.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_46)


stk_managers_item_47 = pro.stk_managers(ts_code='002330.SZ,002331.SZ,002332.SZ,002333.SZ,002334.SZ,002335.SZ,002336.SZ,002337.SZ,002338.SZ,002339.SZ,002340.SZ,002341.SZ,002342.SZ,002343.SZ,002344.SZ,002345.SZ,002346.SZ,002347.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_47 返回数据 row 行数 = "+str(stk_managers_item_47.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_47)


stk_managers_item_48 = pro.stk_managers(ts_code='002348.SZ,002349.SZ,002350.SZ,002351.SZ,002352.SZ,002353.SZ,002354.SZ,002355.SZ,002356.SZ,002357.SZ,002358.SZ,002359.SZ,002360.SZ,002361.SZ,002362.SZ,002363.SZ,002364.SZ,002365.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_48 返回数据 row 行数 = "+str(stk_managers_item_48.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_48)


stk_managers_item_49 = pro.stk_managers(ts_code='002366.SZ,002367.SZ,002368.SZ,002369.SZ,002370.SZ,002371.SZ,002372.SZ,002373.SZ,002374.SZ,002375.SZ,002376.SZ,002377.SZ,002378.SZ,002379.SZ,002380.SZ,002381.SZ,002382.SZ,002383.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_49 返回数据 row 行数 = "+str(stk_managers_item_49.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_49)


stk_managers_item_50 = pro.stk_managers(ts_code='002384.SZ,002385.SZ,002386.SZ,002387.SZ,002388.SZ,002389.SZ,002390.SZ,002391.SZ,002392.SZ,002393.SZ,002394.SZ,002395.SZ,002396.SZ,002397.SZ,002398.SZ,002399.SZ,002400.SZ,002401.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_50 返回数据 row 行数 = "+str(stk_managers_item_50.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_50)


stk_managers_item_51 = pro.stk_managers(ts_code='002402.SZ,002403.SZ,002404.SZ,002405.SZ,002406.SZ,002407.SZ,002408.SZ,002409.SZ,002410.SZ,002411.SZ,002412.SZ,002413.SZ,002414.SZ,002415.SZ,002416.SZ,002417.SZ,002418.SZ,002419.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_51 返回数据 row 行数 = "+str(stk_managers_item_51.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_51)


stk_managers_item_52 = pro.stk_managers(ts_code='002420.SZ,002421.SZ,002422.SZ,002423.SZ,002424.SZ,002425.SZ,002426.SZ,002427.SZ,002428.SZ,002429.SZ,002430.SZ,002431.SZ,002432.SZ,002433.SZ,002434.SZ,002435.SZ,002436.SZ,002437.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_52 返回数据 row 行数 = "+str(stk_managers_item_52.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_52)


stk_managers_item_53 = pro.stk_managers(ts_code='002438.SZ,002439.SZ,002440.SZ,002441.SZ,002442.SZ,002443.SZ,002444.SZ,002445.SZ,002446.SZ,002447.SZ,002448.SZ,002449.SZ,002450.SZ,002451.SZ,002452.SZ,002453.SZ,002454.SZ,002455.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_53 返回数据 row 行数 = "+str(stk_managers_item_53.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_53)


stk_managers_item_54 = pro.stk_managers(ts_code='002456.SZ,002457.SZ,002458.SZ,002459.SZ,002460.SZ,002461.SZ,002462.SZ,002463.SZ,002464.SZ,002465.SZ,002466.SZ,002467.SZ,002468.SZ,002469.SZ,002470.SZ,002471.SZ,002472.SZ,002473.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_54 返回数据 row 行数 = "+str(stk_managers_item_54.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_54)


stk_managers_item_55 = pro.stk_managers(ts_code='002474.SZ,002475.SZ,002476.SZ,002477.SZ,002478.SZ,002479.SZ,002480.SZ,002481.SZ,002482.SZ,002483.SZ,002484.SZ,002485.SZ,002486.SZ,002487.SZ,002488.SZ,002489.SZ,002490.SZ,002491.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_55 返回数据 row 行数 = "+str(stk_managers_item_55.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_55)


stk_managers_item_56 = pro.stk_managers(ts_code='002492.SZ,002493.SZ,002494.SZ,002495.SZ,002496.SZ,002497.SZ,002498.SZ,002499.SZ,002500.SZ,002501.SZ,002502.SZ,002503.SZ,002504.SZ,002505.SZ,002506.SZ,002507.SZ,002508.SZ,002509.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_56 返回数据 row 行数 = "+str(stk_managers_item_56.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_56)


stk_managers_item_57 = pro.stk_managers(ts_code='002510.SZ,002511.SZ,002512.SZ,002513.SZ,002514.SZ,002515.SZ,002516.SZ,002517.SZ,002518.SZ,002519.SZ,002520.SZ,002521.SZ,002522.SZ,002523.SZ,002524.SZ,002526.SZ,002527.SZ,002528.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_57 返回数据 row 行数 = "+str(stk_managers_item_57.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_57)


stk_managers_item_58 = pro.stk_managers(ts_code='002529.SZ,002530.SZ,002531.SZ,002532.SZ,002533.SZ,002534.SZ,002535.SZ,002536.SZ,002537.SZ,002538.SZ,002539.SZ,002540.SZ,002541.SZ,002542.SZ,002543.SZ,002544.SZ,002545.SZ,002546.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_58 返回数据 row 行数 = "+str(stk_managers_item_58.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_58)


stk_managers_item_59 = pro.stk_managers(ts_code='002547.SZ,002548.SZ,002549.SZ,002550.SZ,002551.SZ,002552.SZ,002553.SZ,002554.SZ,002555.SZ,002556.SZ,002557.SZ,002558.SZ,002559.SZ,002560.SZ,002561.SZ,002562.SZ,002563.SZ,002564.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_59 返回数据 row 行数 = "+str(stk_managers_item_59.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_59)


stk_managers_item_60 = pro.stk_managers(ts_code='002565.SZ,002566.SZ,002567.SZ,002568.SZ,002569.SZ,002570.SZ,002571.SZ,002572.SZ,002573.SZ,002574.SZ,002575.SZ,002576.SZ,002577.SZ,002578.SZ,002579.SZ,002580.SZ,002581.SZ,002582.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_60 返回数据 row 行数 = "+str(stk_managers_item_60.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_60)


stk_managers_item_61 = pro.stk_managers(ts_code='002583.SZ,002584.SZ,002585.SZ,002586.SZ,002587.SZ,002588.SZ,002589.SZ,002590.SZ,002591.SZ,002592.SZ,002593.SZ,002594.SZ,002595.SZ,002596.SZ,002597.SZ,002598.SZ,002599.SZ,002600.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_61 返回数据 row 行数 = "+str(stk_managers_item_61.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_61)


stk_managers_item_62 = pro.stk_managers(ts_code='002601.SZ,002602.SZ,002603.SZ,002604.SZ,002605.SZ,002606.SZ,002607.SZ,002608.SZ,002609.SZ,002610.SZ,002611.SZ,002612.SZ,002613.SZ,002614.SZ,002615.SZ,002616.SZ,002617.SZ,002618.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_62 返回数据 row 行数 = "+str(stk_managers_item_62.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_62)


stk_managers_item_63 = pro.stk_managers(ts_code='002619.SZ,002620.SZ,002621.SZ,002622.SZ,002623.SZ,002624.SZ,002625.SZ,002626.SZ,002627.SZ,002628.SZ,002629.SZ,002630.SZ,002631.SZ,002632.SZ,002633.SZ,002634.SZ,002635.SZ,002636.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_63 返回数据 row 行数 = "+str(stk_managers_item_63.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_63)


stk_managers_item_64 = pro.stk_managers(ts_code='002637.SZ,002638.SZ,002639.SZ,002640.SZ,002641.SZ,002642.SZ,002643.SZ,002644.SZ,002645.SZ,002646.SZ,002647.SZ,002648.SZ,002649.SZ,002650.SZ,002651.SZ,002652.SZ,002653.SZ,002654.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_64 返回数据 row 行数 = "+str(stk_managers_item_64.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_64)


stk_managers_item_65 = pro.stk_managers(ts_code='002655.SZ,002656.SZ,002657.SZ,002658.SZ,002659.SZ,002660.SZ,002661.SZ,002662.SZ,002663.SZ,002664.SZ,002665.SZ,002666.SZ,002667.SZ,002668.SZ,002669.SZ,002670.SZ,002671.SZ,002672.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_65 返回数据 row 行数 = "+str(stk_managers_item_65.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_65)


stk_managers_item_66 = pro.stk_managers(ts_code='002673.SZ,002674.SZ,002675.SZ,002676.SZ,002677.SZ,002678.SZ,002679.SZ,002680.SZ,002681.SZ,002682.SZ,002683.SZ,002684.SZ,002685.SZ,002686.SZ,002687.SZ,002688.SZ,002689.SZ,002690.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_66 返回数据 row 行数 = "+str(stk_managers_item_66.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_66)


stk_managers_item_67 = pro.stk_managers(ts_code='002691.SZ,002692.SZ,002693.SZ,002694.SZ,002695.SZ,002696.SZ,002697.SZ,002698.SZ,002699.SZ,002700.SZ,002701.SZ,002702.SZ,002703.SZ,002705.SZ,002706.SZ,002707.SZ,002708.SZ,002709.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_67 返回数据 row 行数 = "+str(stk_managers_item_67.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_67)


stk_managers_item_68 = pro.stk_managers(ts_code='002711.SZ,002712.SZ,002713.SZ,002714.SZ,002715.SZ,002716.SZ,002717.SZ,002718.SZ,002719.SZ,002721.SZ,002722.SZ,002723.SZ,002724.SZ,002725.SZ,002726.SZ,002727.SZ,002728.SZ,002729.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_68 返回数据 row 行数 = "+str(stk_managers_item_68.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_68)


stk_managers_item_69 = pro.stk_managers(ts_code='002730.SZ,002731.SZ,002732.SZ,002733.SZ,002734.SZ,002735.SZ,002736.SZ,002737.SZ,002738.SZ,002739.SZ,002740.SZ,002741.SZ,002742.SZ,002743.SZ,002745.SZ,002746.SZ,002747.SZ,002748.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_69 返回数据 row 行数 = "+str(stk_managers_item_69.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_69)


stk_managers_item_70 = pro.stk_managers(ts_code='002749.SZ,002750.SZ,002751.SZ,002752.SZ,002753.SZ,002755.SZ,002756.SZ,002757.SZ,002758.SZ,002759.SZ,002760.SZ,002761.SZ,002762.SZ,002763.SZ,002765.SZ,002766.SZ,002767.SZ,002768.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_70 返回数据 row 行数 = "+str(stk_managers_item_70.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_70)


stk_managers_item_71 = pro.stk_managers(ts_code='002769.SZ,002770.SZ,002771.SZ,002772.SZ,002773.SZ,002774.SZ,002775.SZ,002776.SZ,002777.SZ,002778.SZ,002779.SZ,002780.SZ,002781.SZ,002782.SZ,002783.SZ,002785.SZ,002786.SZ,002787.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_71 返回数据 row 行数 = "+str(stk_managers_item_71.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_71)


stk_managers_item_72 = pro.stk_managers(ts_code='002788.SZ,002789.SZ,002790.SZ,002791.SZ,002792.SZ,002793.SZ,002795.SZ,002796.SZ,002797.SZ,002798.SZ,002799.SZ,002800.SZ,002801.SZ,002802.SZ,002803.SZ,002805.SZ,002806.SZ,002807.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_72 返回数据 row 行数 = "+str(stk_managers_item_72.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_72)


stk_managers_item_73 = pro.stk_managers(ts_code='002808.SZ,002809.SZ,002810.SZ,002811.SZ,002812.SZ,002813.SZ,002815.SZ,002816.SZ,002817.SZ,002818.SZ,002819.SZ,002820.SZ,002821.SZ,002822.SZ,002823.SZ,002824.SZ,002825.SZ,002826.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_73 返回数据 row 行数 = "+str(stk_managers_item_73.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_73)


stk_managers_item_74 = pro.stk_managers(ts_code='002827.SZ,002828.SZ,002829.SZ,002830.SZ,002831.SZ,002832.SZ,002833.SZ,002835.SZ,002836.SZ,002837.SZ,002838.SZ,002839.SZ,002840.SZ,002841.SZ,002842.SZ,002843.SZ,002845.SZ,002846.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_74 返回数据 row 行数 = "+str(stk_managers_item_74.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_74)


stk_managers_item_75 = pro.stk_managers(ts_code='002847.SZ,002848.SZ,002849.SZ,002850.SZ,002851.SZ,002852.SZ,002853.SZ,002855.SZ,002856.SZ,002857.SZ,002858.SZ,002859.SZ,002860.SZ,002861.SZ,002862.SZ,002863.SZ,002864.SZ,002865.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_75 返回数据 row 行数 = "+str(stk_managers_item_75.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_75)


stk_managers_item_76 = pro.stk_managers(ts_code='002866.SZ,002867.SZ,002868.SZ,002869.SZ,002870.SZ,002871.SZ,002872.SZ,002873.SZ,002875.SZ,002876.SZ,002877.SZ,002878.SZ,002879.SZ,002880.SZ,002881.SZ,002882.SZ,002883.SZ,002884.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_76 返回数据 row 行数 = "+str(stk_managers_item_76.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_76)


stk_managers_item_77 = pro.stk_managers(ts_code='002885.SZ,002886.SZ,002887.SZ,002888.SZ,002889.SZ,002890.SZ,002891.SZ,002892.SZ,002893.SZ,002895.SZ,002896.SZ,002897.SZ,002898.SZ,002899.SZ,002900.SZ,002901.SZ,002902.SZ,002903.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_77 返回数据 row 行数 = "+str(stk_managers_item_77.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_77)


stk_managers_item_78 = pro.stk_managers(ts_code='002905.SZ,002906.SZ,002907.SZ,002908.SZ,002909.SZ,002910.SZ,002911.SZ,002912.SZ,002913.SZ,002915.SZ,002916.SZ,002917.SZ,002918.SZ,002919.SZ,002920.SZ,002921.SZ,002922.SZ,002923.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_78 返回数据 row 行数 = "+str(stk_managers_item_78.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_78)


stk_managers_item_79 = pro.stk_managers(ts_code='002925.SZ,002926.SZ,002927.SZ,002928.SZ,002929.SZ,002930.SZ,002931.SZ,002932.SZ,002933.SZ,002935.SZ,002936.SZ,002937.SZ,002938.SZ,002939.SZ,002940.SZ,002941.SZ,002942.SZ,002943.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_79 返回数据 row 行数 = "+str(stk_managers_item_79.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_79)


stk_managers_item_80 = pro.stk_managers(ts_code='002945.SZ,002946.SZ,002947.SZ,002948.SZ,002949.SZ,002950.SZ,002951.SZ,002952.SZ,002953.SZ,002955.SZ,002956.SZ,002957.SZ,002958.SZ,002959.SZ,002960.SZ,002961.SZ,002962.SZ,002963.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_80 返回数据 row 行数 = "+str(stk_managers_item_80.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_80)


stk_managers_item_81 = pro.stk_managers(ts_code='002965.SZ,002966.SZ,002967.SZ,002968.SZ,002969.SZ,002970.SZ,002971.SZ,002972.SZ,002973.SZ,002975.SZ,002976.SZ,002977.SZ,002978.SZ,002979.SZ,002980.SZ,002981.SZ,002982.SZ,002983.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_81 返回数据 row 行数 = "+str(stk_managers_item_81.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_81)


stk_managers_item_82 = pro.stk_managers(ts_code='002984.SZ,002985.SZ,002986.SZ,002987.SZ,002988.SZ,002989.SZ,002990.SZ,002991.SZ,002992.SZ,002993.SZ,002995.SZ,002996.SZ,002997.SZ,002998.SZ,002999.SZ,003000.SZ,003001.SZ,003002.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_82 返回数据 row 行数 = "+str(stk_managers_item_82.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_82)


stk_managers_item_83 = pro.stk_managers(ts_code='003003.SZ,003004.SZ,003005.SZ,003006.SZ,003007.SZ,003008.SZ,003009.SZ,003010.SZ,003011.SZ,003012.SZ,003013.SZ,003015.SZ,003016.SZ,003017.SZ,003018.SZ,003019.SZ,003020.SZ,003021.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_83 返回数据 row 行数 = "+str(stk_managers_item_83.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_83)


stk_managers_item_84 = pro.stk_managers(ts_code='003022.SZ,003023.SZ,003025.SZ,003026.SZ,003027.SZ,003028.SZ,003029.SZ,003030.SZ,003031.SZ,003032.SZ,003033.SZ,003035.SZ,003036.SZ,003037.SZ,003038.SZ,003039.SZ,003040.SZ,003041.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_84 返回数据 row 行数 = "+str(stk_managers_item_84.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_84)


stk_managers_item_85 = pro.stk_managers(ts_code='003042.SZ,003043.SZ,003816.SZ,300001.SZ,300002.SZ,300003.SZ,300004.SZ,300005.SZ,300006.SZ,300007.SZ,300008.SZ,300009.SZ,300010.SZ,300011.SZ,300012.SZ,300013.SZ,300014.SZ,300015.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_85 返回数据 row 行数 = "+str(stk_managers_item_85.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_85)


stk_managers_item_86 = pro.stk_managers(ts_code='300016.SZ,300017.SZ,300018.SZ,300019.SZ,300020.SZ,300021.SZ,300022.SZ,300023.SZ,300024.SZ,300025.SZ,300026.SZ,300027.SZ,300028.SZ,300029.SZ,300030.SZ,300031.SZ,300032.SZ,300033.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_86 返回数据 row 行数 = "+str(stk_managers_item_86.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_86)


stk_managers_item_87 = pro.stk_managers(ts_code='300034.SZ,300035.SZ,300036.SZ,300037.SZ,300038.SZ,300039.SZ,300040.SZ,300041.SZ,300042.SZ,300043.SZ,300044.SZ,300045.SZ,300046.SZ,300047.SZ,300048.SZ,300049.SZ,300050.SZ,300051.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_87 返回数据 row 行数 = "+str(stk_managers_item_87.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_87)


stk_managers_item_88 = pro.stk_managers(ts_code='300052.SZ,300053.SZ,300054.SZ,300055.SZ,300056.SZ,300057.SZ,300058.SZ,300059.SZ,300061.SZ,300062.SZ,300063.SZ,300064.SZ,300065.SZ,300066.SZ,300067.SZ,300068.SZ,300069.SZ,300070.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_88 返回数据 row 行数 = "+str(stk_managers_item_88.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_88)


stk_managers_item_89 = pro.stk_managers(ts_code='300071.SZ,300072.SZ,300073.SZ,300074.SZ,300075.SZ,300076.SZ,300077.SZ,300078.SZ,300079.SZ,300080.SZ,300081.SZ,300082.SZ,300083.SZ,300084.SZ,300085.SZ,300086.SZ,300087.SZ,300088.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_89 返回数据 row 行数 = "+str(stk_managers_item_89.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_89)


stk_managers_item_90 = pro.stk_managers(ts_code='300089.SZ,300090.SZ,300091.SZ,300092.SZ,300093.SZ,300094.SZ,300095.SZ,300096.SZ,300097.SZ,300098.SZ,300099.SZ,300100.SZ,300101.SZ,300102.SZ,300103.SZ,300104.SZ,300105.SZ,300106.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_90 返回数据 row 行数 = "+str(stk_managers_item_90.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_90)


stk_managers_item_91 = pro.stk_managers(ts_code='300107.SZ,300108.SZ,300109.SZ,300110.SZ,300111.SZ,300112.SZ,300113.SZ,300114.SZ,300115.SZ,300116.SZ,300117.SZ,300118.SZ,300119.SZ,300120.SZ,300121.SZ,300122.SZ,300123.SZ,300124.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_91 返回数据 row 行数 = "+str(stk_managers_item_91.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_91)


stk_managers_item_92 = pro.stk_managers(ts_code='300125.SZ,300126.SZ,300127.SZ,300128.SZ,300129.SZ,300130.SZ,300131.SZ,300132.SZ,300133.SZ,300134.SZ,300135.SZ,300136.SZ,300137.SZ,300138.SZ,300139.SZ,300140.SZ,300141.SZ,300142.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_92 返回数据 row 行数 = "+str(stk_managers_item_92.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_92)


stk_managers_item_93 = pro.stk_managers(ts_code='300143.SZ,300144.SZ,300145.SZ,300146.SZ,300147.SZ,300148.SZ,300149.SZ,300150.SZ,300151.SZ,300152.SZ,300153.SZ,300154.SZ,300155.SZ,300156.SZ,300157.SZ,300158.SZ,300159.SZ,300160.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_93 返回数据 row 行数 = "+str(stk_managers_item_93.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_93)


stk_managers_item_94 = pro.stk_managers(ts_code='300161.SZ,300162.SZ,300163.SZ,300164.SZ,300165.SZ,300166.SZ,300167.SZ,300168.SZ,300169.SZ,300170.SZ,300171.SZ,300172.SZ,300173.SZ,300174.SZ,300175.SZ,300176.SZ,300177.SZ,300178.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_94 返回数据 row 行数 = "+str(stk_managers_item_94.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_94)


stk_managers_item_95 = pro.stk_managers(ts_code='300179.SZ,300180.SZ,300181.SZ,300182.SZ,300183.SZ,300184.SZ,300185.SZ,300186.SZ,300187.SZ,300188.SZ,300189.SZ,300190.SZ,300191.SZ,300192.SZ,300193.SZ,300194.SZ,300195.SZ,300196.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_95 返回数据 row 行数 = "+str(stk_managers_item_95.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_95)


stk_managers_item_96 = pro.stk_managers(ts_code='300197.SZ,300198.SZ,300199.SZ,300200.SZ,300201.SZ,300202.SZ,300203.SZ,300204.SZ,300205.SZ,300206.SZ,300207.SZ,300208.SZ,300209.SZ,300210.SZ,300211.SZ,300212.SZ,300213.SZ,300214.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_96 返回数据 row 行数 = "+str(stk_managers_item_96.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_96)


stk_managers_item_97 = pro.stk_managers(ts_code='300215.SZ,300216.SZ,300217.SZ,300218.SZ,300219.SZ,300220.SZ,300221.SZ,300222.SZ,300223.SZ,300224.SZ,300225.SZ,300226.SZ,300227.SZ,300228.SZ,300229.SZ,300230.SZ,300231.SZ,300232.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_97 返回数据 row 行数 = "+str(stk_managers_item_97.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_97)


stk_managers_item_98 = pro.stk_managers(ts_code='300233.SZ,300234.SZ,300235.SZ,300236.SZ,300237.SZ,300238.SZ,300239.SZ,300240.SZ,300241.SZ,300242.SZ,300243.SZ,300244.SZ,300245.SZ,300246.SZ,300247.SZ,300248.SZ,300249.SZ,300250.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_98 返回数据 row 行数 = "+str(stk_managers_item_98.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_98)


stk_managers_item_99 = pro.stk_managers(ts_code='300251.SZ,300252.SZ,300253.SZ,300254.SZ,300255.SZ,300256.SZ,300257.SZ,300258.SZ,300259.SZ,300260.SZ,300261.SZ,300262.SZ,300263.SZ,300264.SZ,300265.SZ,300266.SZ,300267.SZ,300268.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_99 返回数据 row 行数 = "+str(stk_managers_item_99.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_99)


stk_managers_item_100 = pro.stk_managers(ts_code='300269.SZ,300270.SZ,300271.SZ,300272.SZ,300273.SZ,300274.SZ,300275.SZ,300276.SZ,300277.SZ,300278.SZ,300279.SZ,300280.SZ,300281.SZ,300282.SZ,300283.SZ,300284.SZ,300285.SZ,300286.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_100 返回数据 row 行数 = "+str(stk_managers_item_100.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_100)


stk_managers_item_101 = pro.stk_managers(ts_code='300287.SZ,300288.SZ,300289.SZ,300290.SZ,300291.SZ,300292.SZ,300293.SZ,300294.SZ,300295.SZ,300296.SZ,300297.SZ,300298.SZ,300299.SZ,300300.SZ,300301.SZ,300302.SZ,300303.SZ,300304.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_101 返回数据 row 行数 = "+str(stk_managers_item_101.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_101)


stk_managers_item_102 = pro.stk_managers(ts_code='300305.SZ,300306.SZ,300307.SZ,300308.SZ,300309.SZ,300310.SZ,300311.SZ,300312.SZ,300313.SZ,300314.SZ,300315.SZ,300316.SZ,300317.SZ,300318.SZ,300319.SZ,300320.SZ,300321.SZ,300322.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_102 返回数据 row 行数 = "+str(stk_managers_item_102.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_102)


stk_managers_item_103 = pro.stk_managers(ts_code='300323.SZ,300324.SZ,300325.SZ,300326.SZ,300327.SZ,300328.SZ,300329.SZ,300330.SZ,300331.SZ,300332.SZ,300333.SZ,300334.SZ,300335.SZ,300336.SZ,300337.SZ,300338.SZ,300339.SZ,300340.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_103 返回数据 row 行数 = "+str(stk_managers_item_103.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_103)


stk_managers_item_104 = pro.stk_managers(ts_code='300341.SZ,300342.SZ,300343.SZ,300344.SZ,300345.SZ,300346.SZ,300347.SZ,300348.SZ,300349.SZ,300350.SZ,300351.SZ,300352.SZ,300353.SZ,300354.SZ,300355.SZ,300356.SZ,300357.SZ,300358.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_104 返回数据 row 行数 = "+str(stk_managers_item_104.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_104)


stk_managers_item_105 = pro.stk_managers(ts_code='300359.SZ,300360.SZ,300362.SZ,300363.SZ,300364.SZ,300365.SZ,300366.SZ,300367.SZ,300368.SZ,300369.SZ,300370.SZ,300371.SZ,300372.SZ,300373.SZ,300374.SZ,300375.SZ,300376.SZ,300377.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_105 返回数据 row 行数 = "+str(stk_managers_item_105.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_105)


stk_managers_item_106 = pro.stk_managers(ts_code='300378.SZ,300379.SZ,300380.SZ,300381.SZ,300382.SZ,300383.SZ,300384.SZ,300385.SZ,300386.SZ,300387.SZ,300388.SZ,300389.SZ,300390.SZ,300391.SZ,300392.SZ,300393.SZ,300394.SZ,300395.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_106 返回数据 row 行数 = "+str(stk_managers_item_106.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_106)


stk_managers_item_107 = pro.stk_managers(ts_code='300396.SZ,300397.SZ,300398.SZ,300399.SZ,300400.SZ,300401.SZ,300402.SZ,300403.SZ,300404.SZ,300405.SZ,300406.SZ,300407.SZ,300408.SZ,300409.SZ,300410.SZ,300411.SZ,300412.SZ,300413.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_107 返回数据 row 行数 = "+str(stk_managers_item_107.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_107)


stk_managers_item_108 = pro.stk_managers(ts_code='300414.SZ,300415.SZ,300416.SZ,300417.SZ,300418.SZ,300419.SZ,300420.SZ,300421.SZ,300422.SZ,300423.SZ,300424.SZ,300425.SZ,300426.SZ,300427.SZ,300428.SZ,300429.SZ,300430.SZ,300431.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_108 返回数据 row 行数 = "+str(stk_managers_item_108.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_108)


stk_managers_item_109 = pro.stk_managers(ts_code='300432.SZ,300433.SZ,300434.SZ,300435.SZ,300436.SZ,300437.SZ,300438.SZ,300439.SZ,300440.SZ,300441.SZ,300442.SZ,300443.SZ,300444.SZ,300445.SZ,300446.SZ,300447.SZ,300448.SZ,300449.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_109 返回数据 row 行数 = "+str(stk_managers_item_109.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_109)


stk_managers_item_110 = pro.stk_managers(ts_code='300450.SZ,300451.SZ,300452.SZ,300453.SZ,300454.SZ,300455.SZ,300456.SZ,300457.SZ,300458.SZ,300459.SZ,300460.SZ,300461.SZ,300462.SZ,300463.SZ,300464.SZ,300465.SZ,300466.SZ,300467.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_110 返回数据 row 行数 = "+str(stk_managers_item_110.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_110)


stk_managers_item_111 = pro.stk_managers(ts_code='300468.SZ,300469.SZ,300470.SZ,300471.SZ,300472.SZ,300473.SZ,300474.SZ,300475.SZ,300476.SZ,300477.SZ,300478.SZ,300479.SZ,300480.SZ,300481.SZ,300482.SZ,300483.SZ,300484.SZ,300485.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_111 返回数据 row 行数 = "+str(stk_managers_item_111.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_111)


stk_managers_item_112 = pro.stk_managers(ts_code='300486.SZ,300487.SZ,300488.SZ,300489.SZ,300490.SZ,300491.SZ,300492.SZ,300493.SZ,300494.SZ,300495.SZ,300496.SZ,300497.SZ,300498.SZ,300499.SZ,300500.SZ,300501.SZ,300502.SZ,300503.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_112 返回数据 row 行数 = "+str(stk_managers_item_112.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_112)


stk_managers_item_113 = pro.stk_managers(ts_code='300504.SZ,300505.SZ,300506.SZ,300507.SZ,300508.SZ,300509.SZ,300510.SZ,300511.SZ,300512.SZ,300513.SZ,300514.SZ,300515.SZ,300516.SZ,300517.SZ,300518.SZ,300519.SZ,300520.SZ,300521.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_113 返回数据 row 行数 = "+str(stk_managers_item_113.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_113)


stk_managers_item_114 = pro.stk_managers(ts_code='300522.SZ,300523.SZ,300525.SZ,300526.SZ,300527.SZ,300528.SZ,300529.SZ,300530.SZ,300531.SZ,300532.SZ,300533.SZ,300534.SZ,300535.SZ,300536.SZ,300537.SZ,300538.SZ,300539.SZ,300540.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_114 返回数据 row 行数 = "+str(stk_managers_item_114.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_114)


stk_managers_item_115 = pro.stk_managers(ts_code='300541.SZ,300542.SZ,300543.SZ,300545.SZ,300546.SZ,300547.SZ,300548.SZ,300549.SZ,300550.SZ,300551.SZ,300552.SZ,300553.SZ,300554.SZ,300555.SZ,300556.SZ,300557.SZ,300558.SZ,300559.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_115 返回数据 row 行数 = "+str(stk_managers_item_115.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_115)


stk_managers_item_116 = pro.stk_managers(ts_code='300560.SZ,300561.SZ,300562.SZ,300563.SZ,300564.SZ,300565.SZ,300566.SZ,300567.SZ,300568.SZ,300569.SZ,300570.SZ,300571.SZ,300572.SZ,300573.SZ,300575.SZ,300576.SZ,300577.SZ,300578.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_116 返回数据 row 行数 = "+str(stk_managers_item_116.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_116)


stk_managers_item_117 = pro.stk_managers(ts_code='300579.SZ,300580.SZ,300581.SZ,300582.SZ,300583.SZ,300584.SZ,300585.SZ,300586.SZ,300587.SZ,300588.SZ,300589.SZ,300590.SZ,300591.SZ,300592.SZ,300593.SZ,300594.SZ,300595.SZ,300596.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_117 返回数据 row 行数 = "+str(stk_managers_item_117.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_117)


stk_managers_item_118 = pro.stk_managers(ts_code='300597.SZ,300598.SZ,300599.SZ,300600.SZ,300601.SZ,300602.SZ,300603.SZ,300604.SZ,300605.SZ,300606.SZ,300607.SZ,300608.SZ,300609.SZ,300610.SZ,300611.SZ,300612.SZ,300613.SZ,300614.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_118 返回数据 row 行数 = "+str(stk_managers_item_118.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_118)


stk_managers_item_119 = pro.stk_managers(ts_code='300615.SZ,300616.SZ,300617.SZ,300618.SZ,300619.SZ,300620.SZ,300621.SZ,300622.SZ,300623.SZ,300624.SZ,300625.SZ,300626.SZ,300627.SZ,300628.SZ,300629.SZ,300630.SZ,300631.SZ,300632.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_119 返回数据 row 行数 = "+str(stk_managers_item_119.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_119)


stk_managers_item_120 = pro.stk_managers(ts_code='300633.SZ,300634.SZ,300635.SZ,300636.SZ,300637.SZ,300638.SZ,300639.SZ,300640.SZ,300641.SZ,300642.SZ,300643.SZ,300644.SZ,300645.SZ,300647.SZ,300648.SZ,300649.SZ,300650.SZ,300651.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_120 返回数据 row 行数 = "+str(stk_managers_item_120.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_120)


stk_managers_item_121 = pro.stk_managers(ts_code='300652.SZ,300653.SZ,300654.SZ,300655.SZ,300656.SZ,300657.SZ,300658.SZ,300659.SZ,300660.SZ,300661.SZ,300662.SZ,300663.SZ,300664.SZ,300665.SZ,300666.SZ,300667.SZ,300668.SZ,300669.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_121 返回数据 row 行数 = "+str(stk_managers_item_121.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_121)


stk_managers_item_122 = pro.stk_managers(ts_code='300670.SZ,300671.SZ,300672.SZ,300673.SZ,300674.SZ,300675.SZ,300676.SZ,300677.SZ,300678.SZ,300679.SZ,300680.SZ,300681.SZ,300682.SZ,300683.SZ,300684.SZ,300685.SZ,300686.SZ,300687.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_122 返回数据 row 行数 = "+str(stk_managers_item_122.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_122)


stk_managers_item_123 = pro.stk_managers(ts_code='300688.SZ,300689.SZ,300690.SZ,300691.SZ,300692.SZ,300693.SZ,300694.SZ,300695.SZ,300696.SZ,300697.SZ,300698.SZ,300699.SZ,300700.SZ,300701.SZ,300702.SZ,300703.SZ,300705.SZ,300706.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_123 返回数据 row 行数 = "+str(stk_managers_item_123.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_123)


stk_managers_item_124 = pro.stk_managers(ts_code='300707.SZ,300708.SZ,300709.SZ,300710.SZ,300711.SZ,300712.SZ,300713.SZ,300715.SZ,300716.SZ,300717.SZ,300718.SZ,300719.SZ,300720.SZ,300721.SZ,300722.SZ,300723.SZ,300724.SZ,300725.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_124 返回数据 row 行数 = "+str(stk_managers_item_124.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_124)


stk_managers_item_125 = pro.stk_managers(ts_code='300726.SZ,300727.SZ,300729.SZ,300730.SZ,300731.SZ,300732.SZ,300733.SZ,300735.SZ,300736.SZ,300737.SZ,300738.SZ,300739.SZ,300740.SZ,300741.SZ,300742.SZ,300743.SZ,300745.SZ,300746.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_125 返回数据 row 行数 = "+str(stk_managers_item_125.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_125)


stk_managers_item_126 = pro.stk_managers(ts_code='300747.SZ,300748.SZ,300749.SZ,300750.SZ,300751.SZ,300752.SZ,300753.SZ,300755.SZ,300756.SZ,300757.SZ,300758.SZ,300759.SZ,300760.SZ,300761.SZ,300762.SZ,300763.SZ,300765.SZ,300766.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_126 返回数据 row 行数 = "+str(stk_managers_item_126.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_126)


stk_managers_item_127 = pro.stk_managers(ts_code='300767.SZ,300768.SZ,300769.SZ,300770.SZ,300771.SZ,300772.SZ,300773.SZ,300775.SZ,300776.SZ,300777.SZ,300778.SZ,300779.SZ,300780.SZ,300781.SZ,300782.SZ,300783.SZ,300785.SZ,300786.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_127 返回数据 row 行数 = "+str(stk_managers_item_127.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_127)


stk_managers_item_128 = pro.stk_managers(ts_code='300787.SZ,300788.SZ,300789.SZ,300790.SZ,300791.SZ,300792.SZ,300793.SZ,300795.SZ,300796.SZ,300797.SZ,300798.SZ,300799.SZ,300800.SZ,300801.SZ,300802.SZ,300803.SZ,300805.SZ,300806.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_128 返回数据 row 行数 = "+str(stk_managers_item_128.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_128)


stk_managers_item_129 = pro.stk_managers(ts_code='300807.SZ,300808.SZ,300809.SZ,300810.SZ,300811.SZ,300812.SZ,300813.SZ,300815.SZ,300816.SZ,300817.SZ,300818.SZ,300819.SZ,300820.SZ,300821.SZ,300822.SZ,300823.SZ,300824.SZ,300825.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_129 返回数据 row 行数 = "+str(stk_managers_item_129.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_129)


stk_managers_item_130 = pro.stk_managers(ts_code='300826.SZ,300827.SZ,300828.SZ,300829.SZ,300830.SZ,300831.SZ,300832.SZ,300833.SZ,300835.SZ,300836.SZ,300837.SZ,300838.SZ,300839.SZ,300840.SZ,300841.SZ,300842.SZ,300843.SZ,300845.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_130 返回数据 row 行数 = "+str(stk_managers_item_130.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_130)


stk_managers_item_131 = pro.stk_managers(ts_code='300846.SZ,300847.SZ,300848.SZ,300849.SZ,300850.SZ,300851.SZ,300852.SZ,300853.SZ,300855.SZ,300856.SZ,300857.SZ,300858.SZ,300859.SZ,300860.SZ,300861.SZ,300862.SZ,300863.SZ,300864.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_131 返回数据 row 行数 = "+str(stk_managers_item_131.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_131)


stk_managers_item_132 = pro.stk_managers(ts_code='300865.SZ,300866.SZ,300867.SZ,300868.SZ,300869.SZ,300870.SZ,300871.SZ,300872.SZ,300873.SZ,300875.SZ,300876.SZ,300877.SZ,300878.SZ,300879.SZ,300880.SZ,300881.SZ,300882.SZ,300883.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_132 返回数据 row 行数 = "+str(stk_managers_item_132.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_132)


stk_managers_item_133 = pro.stk_managers(ts_code='300884.SZ,300885.SZ,300886.SZ,300887.SZ,300888.SZ,300889.SZ,300890.SZ,300891.SZ,300892.SZ,300893.SZ,300894.SZ,300895.SZ,300896.SZ,300897.SZ,300898.SZ,300899.SZ,300900.SZ,300901.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_133 返回数据 row 行数 = "+str(stk_managers_item_133.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_133)


stk_managers_item_134 = pro.stk_managers(ts_code='300902.SZ,300903.SZ,300905.SZ,300906.SZ,300907.SZ,300908.SZ,300909.SZ,300910.SZ,300911.SZ,300912.SZ,300913.SZ,300915.SZ,300916.SZ,300917.SZ,300918.SZ,300919.SZ,300920.SZ,300921.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_134 返回数据 row 行数 = "+str(stk_managers_item_134.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_134)


stk_managers_item_135 = pro.stk_managers(ts_code='300922.SZ,300923.SZ,300925.SZ,300926.SZ,300927.SZ,300928.SZ,300929.SZ,300930.SZ,300931.SZ,300932.SZ,300933.SZ,300935.SZ,300936.SZ,300937.SZ,300938.SZ,300939.SZ,300940.SZ,300941.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_135 返回数据 row 行数 = "+str(stk_managers_item_135.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_135)


stk_managers_item_136 = pro.stk_managers(ts_code='300942.SZ,300943.SZ,300945.SZ,300946.SZ,300947.SZ,300948.SZ,300949.SZ,300950.SZ,300951.SZ,300952.SZ,300953.SZ,300955.SZ,300956.SZ,300957.SZ,300958.SZ,300959.SZ,300960.SZ,300961.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_136 返回数据 row 行数 = "+str(stk_managers_item_136.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_136)


stk_managers_item_137 = pro.stk_managers(ts_code='300962.SZ,300963.SZ,300965.SZ,300966.SZ,300967.SZ,300968.SZ,300969.SZ,300970.SZ,300971.SZ,300972.SZ,300973.SZ,300975.SZ,300976.SZ,300977.SZ,300978.SZ,300979.SZ,300980.SZ,300981.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_137 返回数据 row 行数 = "+str(stk_managers_item_137.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_137)


stk_managers_item_138 = pro.stk_managers(ts_code='300982.SZ,300983.SZ,300984.SZ,300985.SZ,300986.SZ,300987.SZ,300988.SZ,300989.SZ,300990.SZ,300991.SZ,300992.SZ,300993.SZ,300995.SZ,300996.SZ,300997.SZ,300998.SZ,300999.SZ,301000.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_138 返回数据 row 行数 = "+str(stk_managers_item_138.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_138)


stk_managers_item_139 = pro.stk_managers(ts_code='301001.SZ,301002.SZ,301003.SZ,301004.SZ,301005.SZ,301006.SZ,301007.SZ,301008.SZ,301009.SZ,301010.SZ,301011.SZ,301012.SZ,301013.SZ,301015.SZ,301016.SZ,301017.SZ,301018.SZ,301019.SZ', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_139 返回数据 row 行数 = "+str(stk_managers_item_139.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_139)


stk_managers_item_140 = pro.stk_managers(ts_code='301020.SZ,301021.SZ,301022.SZ,301023.SZ,600000.SH,600001.SH,600002.SH,600003.SH,600004.SH,600005.SH,600006.SH,600007.SH,600008.SH,600009.SH,600010.SH,600011.SH,600012.SH,600015.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_140 返回数据 row 行数 = "+str(stk_managers_item_140.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_140)


stk_managers_item_141 = pro.stk_managers(ts_code='600016.SH,600017.SH,600018.SH,600019.SH,600020.SH,600021.SH,600022.SH,600023.SH,600025.SH,600026.SH,600027.SH,600028.SH,600029.SH,600030.SH,600031.SH,600032.SH,600033.SH,600035.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_141 返回数据 row 行数 = "+str(stk_managers_item_141.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_141)


stk_managers_item_142 = pro.stk_managers(ts_code='600036.SH,600037.SH,600038.SH,600039.SH,600048.SH,600050.SH,600051.SH,600052.SH,600053.SH,600054.SH,600055.SH,600056.SH,600057.SH,600058.SH,600059.SH,600060.SH,600061.SH,600062.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_142 返回数据 row 行数 = "+str(stk_managers_item_142.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_142)


stk_managers_item_143 = pro.stk_managers(ts_code='600063.SH,600064.SH,600065.SH,600066.SH,600067.SH,600068.SH,600069.SH,600070.SH,600071.SH,600072.SH,600073.SH,600074.SH,600075.SH,600076.SH,600077.SH,600078.SH,600079.SH,600080.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_143 返回数据 row 行数 = "+str(stk_managers_item_143.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_143)


stk_managers_item_144 = pro.stk_managers(ts_code='600081.SH,600082.SH,600083.SH,600084.SH,600085.SH,600086.SH,600087.SH,600088.SH,600089.SH,600090.SH,600091.SH,600092.SH,600093.SH,600094.SH,600095.SH,600096.SH,600097.SH,600098.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_144 返回数据 row 行数 = "+str(stk_managers_item_144.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_144)


stk_managers_item_145 = pro.stk_managers(ts_code='600099.SH,600100.SH,600101.SH,600102.SH,600103.SH,600104.SH,600105.SH,600106.SH,600107.SH,600108.SH,600109.SH,600110.SH,600111.SH,600112.SH,600113.SH,600114.SH,600115.SH,600116.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_145 返回数据 row 行数 = "+str(stk_managers_item_145.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_145)


stk_managers_item_146 = pro.stk_managers(ts_code='600117.SH,600118.SH,600119.SH,600120.SH,600121.SH,600122.SH,600123.SH,600125.SH,600126.SH,600127.SH,600128.SH,600129.SH,600130.SH,600131.SH,600132.SH,600133.SH,600135.SH,600136.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_146 返回数据 row 行数 = "+str(stk_managers_item_146.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_146)


stk_managers_item_147 = pro.stk_managers(ts_code='600137.SH,600138.SH,600139.SH,600141.SH,600143.SH,600145.SH,600146.SH,600148.SH,600149.SH,600150.SH,600151.SH,600152.SH,600153.SH,600155.SH,600156.SH,600157.SH,600158.SH,600159.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_147 返回数据 row 行数 = "+str(stk_managers_item_147.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_147)


stk_managers_item_148 = pro.stk_managers(ts_code='600160.SH,600161.SH,600162.SH,600163.SH,600165.SH,600166.SH,600167.SH,600168.SH,600169.SH,600170.SH,600171.SH,600172.SH,600173.SH,600175.SH,600176.SH,600177.SH,600178.SH,600179.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_148 返回数据 row 行数 = "+str(stk_managers_item_148.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_148)


stk_managers_item_149 = pro.stk_managers(ts_code='600180.SH,600181.SH,600182.SH,600183.SH,600184.SH,600185.SH,600186.SH,600187.SH,600188.SH,600189.SH,600190.SH,600191.SH,600192.SH,600193.SH,600195.SH,600196.SH,600197.SH,600198.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_149 返回数据 row 行数 = "+str(stk_managers_item_149.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_149)


stk_managers_item_150 = pro.stk_managers(ts_code='600199.SH,600200.SH,600201.SH,600202.SH,600203.SH,600205.SH,600206.SH,600207.SH,600208.SH,600209.SH,600210.SH,600211.SH,600212.SH,600213.SH,600215.SH,600216.SH,600217.SH,600218.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_150 返回数据 row 行数 = "+str(stk_managers_item_150.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_150)


stk_managers_item_151 = pro.stk_managers(ts_code='600219.SH,600220.SH,600221.SH,600222.SH,600223.SH,600225.SH,600226.SH,600227.SH,600228.SH,600229.SH,600230.SH,600231.SH,600232.SH,600233.SH,600234.SH,600235.SH,600236.SH,600237.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_151 返回数据 row 行数 = "+str(stk_managers_item_151.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_151)


stk_managers_item_152 = pro.stk_managers(ts_code='600238.SH,600239.SH,600240.SH,600241.SH,600242.SH,600243.SH,600246.SH,600247.SH,600248.SH,600249.SH,600250.SH,600251.SH,600252.SH,600253.SH,600255.SH,600256.SH,600257.SH,600258.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_152 返回数据 row 行数 = "+str(stk_managers_item_152.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_152)


stk_managers_item_153 = pro.stk_managers(ts_code='600259.SH,600260.SH,600261.SH,600262.SH,600263.SH,600265.SH,600266.SH,600267.SH,600268.SH,600269.SH,600270.SH,600271.SH,600272.SH,600273.SH,600275.SH,600276.SH,600277.SH,600278.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_153 返回数据 row 行数 = "+str(stk_managers_item_153.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_153)


stk_managers_item_154 = pro.stk_managers(ts_code='600279.SH,600280.SH,600281.SH,600282.SH,600283.SH,600284.SH,600285.SH,600286.SH,600287.SH,600288.SH,600289.SH,600290.SH,600291.SH,600292.SH,600293.SH,600295.SH,600296.SH,600297.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_154 返回数据 row 行数 = "+str(stk_managers_item_154.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_154)


stk_managers_item_155 = pro.stk_managers(ts_code='600298.SH,600299.SH,600300.SH,600301.SH,600302.SH,600303.SH,600305.SH,600306.SH,600307.SH,600308.SH,600309.SH,600310.SH,600311.SH,600312.SH,600313.SH,600315.SH,600316.SH,600317.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_155 返回数据 row 行数 = "+str(stk_managers_item_155.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_155)


stk_managers_item_156 = pro.stk_managers(ts_code='600318.SH,600319.SH,600320.SH,600321.SH,600322.SH,600323.SH,600325.SH,600326.SH,600327.SH,600328.SH,600329.SH,600330.SH,600331.SH,600332.SH,600333.SH,600335.SH,600336.SH,600337.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_156 返回数据 row 行数 = "+str(stk_managers_item_156.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_156)


stk_managers_item_157 = pro.stk_managers(ts_code='600338.SH,600339.SH,600340.SH,600343.SH,600345.SH,600346.SH,600348.SH,600350.SH,600351.SH,600352.SH,600353.SH,600354.SH,600355.SH,600356.SH,600357.SH,600358.SH,600359.SH,600360.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_157 返回数据 row 行数 = "+str(stk_managers_item_157.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_157)


stk_managers_item_158 = pro.stk_managers(ts_code='600361.SH,600362.SH,600363.SH,600365.SH,600366.SH,600367.SH,600368.SH,600369.SH,600370.SH,600371.SH,600372.SH,600373.SH,600375.SH,600376.SH,600377.SH,600378.SH,600379.SH,600380.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_158 返回数据 row 行数 = "+str(stk_managers_item_158.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_158)


stk_managers_item_159 = pro.stk_managers(ts_code='600381.SH,600382.SH,600383.SH,600385.SH,600386.SH,600387.SH,600388.SH,600389.SH,600390.SH,600391.SH,600392.SH,600393.SH,600395.SH,600396.SH,600397.SH,600398.SH,600399.SH,600400.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_159 返回数据 row 行数 = "+str(stk_managers_item_159.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_159)


stk_managers_item_160 = pro.stk_managers(ts_code='600401.SH,600403.SH,600405.SH,600406.SH,600408.SH,600409.SH,600410.SH,600415.SH,600416.SH,600418.SH,600419.SH,600420.SH,600421.SH,600422.SH,600423.SH,600425.SH,600426.SH,600428.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_160 返回数据 row 行数 = "+str(stk_managers_item_160.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_160)


stk_managers_item_161 = pro.stk_managers(ts_code='600429.SH,600432.SH,600433.SH,600435.SH,600436.SH,600438.SH,600439.SH,600444.SH,600446.SH,600448.SH,600449.SH,600452.SH,600455.SH,600456.SH,600458.SH,600459.SH,600460.SH,600461.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_161 返回数据 row 行数 = "+str(stk_managers_item_161.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_161)


stk_managers_item_162 = pro.stk_managers(ts_code='600462.SH,600463.SH,600466.SH,600467.SH,600468.SH,600469.SH,600470.SH,600472.SH,600475.SH,600476.SH,600477.SH,600478.SH,600479.SH,600480.SH,600481.SH,600482.SH,600483.SH,600485.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_162 返回数据 row 行数 = "+str(stk_managers_item_162.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_162)


stk_managers_item_163 = pro.stk_managers(ts_code='600486.SH,600487.SH,600488.SH,600489.SH,600490.SH,600491.SH,600493.SH,600495.SH,600496.SH,600497.SH,600498.SH,600499.SH,600500.SH,600501.SH,600502.SH,600503.SH,600505.SH,600506.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_163 返回数据 row 行数 = "+str(stk_managers_item_163.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_163)


stk_managers_item_164 = pro.stk_managers(ts_code='600507.SH,600508.SH,600509.SH,600510.SH,600511.SH,600512.SH,600513.SH,600515.SH,600516.SH,600517.SH,600518.SH,600519.SH,600520.SH,600521.SH,600522.SH,600523.SH,600525.SH,600526.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_164 返回数据 row 行数 = "+str(stk_managers_item_164.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_164)


stk_managers_item_165 = pro.stk_managers(ts_code='600527.SH,600528.SH,600529.SH,600530.SH,600531.SH,600532.SH,600533.SH,600535.SH,600536.SH,600537.SH,600538.SH,600539.SH,600540.SH,600543.SH,600545.SH,600546.SH,600547.SH,600548.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_165 返回数据 row 行数 = "+str(stk_managers_item_165.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_165)


stk_managers_item_166 = pro.stk_managers(ts_code='600549.SH,600550.SH,600551.SH,600552.SH,600553.SH,600555.SH,600556.SH,600557.SH,600558.SH,600559.SH,600560.SH,600561.SH,600562.SH,600563.SH,600565.SH,600566.SH,600567.SH,600568.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_166 返回数据 row 行数 = "+str(stk_managers_item_166.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_166)


stk_managers_item_167 = pro.stk_managers(ts_code='600569.SH,600570.SH,600571.SH,600572.SH,600573.SH,600575.SH,600576.SH,600577.SH,600578.SH,600579.SH,600580.SH,600581.SH,600582.SH,600583.SH,600584.SH,600585.SH,600586.SH,600587.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_167 返回数据 row 行数 = "+str(stk_managers_item_167.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_167)


stk_managers_item_168 = pro.stk_managers(ts_code='600588.SH,600589.SH,600590.SH,600591.SH,600592.SH,600593.SH,600594.SH,600595.SH,600596.SH,600597.SH,600598.SH,600599.SH,600600.SH,600601.SH,600602.SH,600603.SH,600604.SH,600605.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_168 返回数据 row 行数 = "+str(stk_managers_item_168.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_168)


stk_managers_item_169 = pro.stk_managers(ts_code='600606.SH,600607.SH,600608.SH,600609.SH,600610.SH,600611.SH,600612.SH,600613.SH,600614.SH,600615.SH,600616.SH,600617.SH,600618.SH,600619.SH,600620.SH,600621.SH,600622.SH,600623.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_169 返回数据 row 行数 = "+str(stk_managers_item_169.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_169)


stk_managers_item_170 = pro.stk_managers(ts_code='600624.SH,600625.SH,600626.SH,600627.SH,600628.SH,600629.SH,600630.SH,600631.SH,600632.SH,600633.SH,600634.SH,600635.SH,600636.SH,600637.SH,600638.SH,600639.SH,600640.SH,600641.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_170 返回数据 row 行数 = "+str(stk_managers_item_170.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_170)


stk_managers_item_171 = pro.stk_managers(ts_code='600642.SH,600643.SH,600644.SH,600645.SH,600646.SH,600647.SH,600648.SH,600649.SH,600650.SH,600651.SH,600652.SH,600653.SH,600654.SH,600655.SH,600656.SH,600657.SH,600658.SH,600659.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_171 返回数据 row 行数 = "+str(stk_managers_item_171.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_171)


stk_managers_item_172 = pro.stk_managers(ts_code='600660.SH,600661.SH,600662.SH,600663.SH,600664.SH,600665.SH,600666.SH,600667.SH,600668.SH,600669.SH,600670.SH,600671.SH,600672.SH,600673.SH,600674.SH,600675.SH,600676.SH,600677.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_172 返回数据 row 行数 = "+str(stk_managers_item_172.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_172)


stk_managers_item_173 = pro.stk_managers(ts_code='600678.SH,600679.SH,600680.SH,600681.SH,600682.SH,600683.SH,600684.SH,600685.SH,600686.SH,600687.SH,600688.SH,600689.SH,600690.SH,600691.SH,600692.SH,600693.SH,600694.SH,600695.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_173 返回数据 row 行数 = "+str(stk_managers_item_173.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_173)


stk_managers_item_174 = pro.stk_managers(ts_code='600696.SH,600697.SH,600698.SH,600699.SH,600700.SH,600701.SH,600702.SH,600703.SH,600704.SH,600705.SH,600706.SH,600707.SH,600708.SH,600709.SH,600710.SH,600711.SH,600712.SH,600713.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_174 返回数据 row 行数 = "+str(stk_managers_item_174.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_174)


stk_managers_item_175 = pro.stk_managers(ts_code='600714.SH,600715.SH,600716.SH,600717.SH,600718.SH,600719.SH,600720.SH,600721.SH,600722.SH,600723.SH,600724.SH,600725.SH,600726.SH,600727.SH,600728.SH,600729.SH,600730.SH,600731.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_175 返回数据 row 行数 = "+str(stk_managers_item_175.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_175)


stk_managers_item_176 = pro.stk_managers(ts_code='600732.SH,600733.SH,600734.SH,600735.SH,600736.SH,600737.SH,600738.SH,600739.SH,600740.SH,600741.SH,600742.SH,600743.SH,600744.SH,600745.SH,600746.SH,600747.SH,600748.SH,600749.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_176 返回数据 row 行数 = "+str(stk_managers_item_176.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_176)


stk_managers_item_177 = pro.stk_managers(ts_code='600750.SH,600751.SH,600752.SH,600753.SH,600754.SH,600755.SH,600756.SH,600757.SH,600758.SH,600759.SH,600760.SH,600761.SH,600762.SH,600763.SH,600764.SH,600765.SH,600766.SH,600767.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_177 返回数据 row 行数 = "+str(stk_managers_item_177.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_177)


stk_managers_item_178 = pro.stk_managers(ts_code='600768.SH,600769.SH,600770.SH,600771.SH,600772.SH,600773.SH,600774.SH,600775.SH,600776.SH,600777.SH,600778.SH,600779.SH,600780.SH,600781.SH,600782.SH,600783.SH,600784.SH,600785.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_178 返回数据 row 行数 = "+str(stk_managers_item_178.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_178)


stk_managers_item_179 = pro.stk_managers(ts_code='600786.SH,600787.SH,600788.SH,600789.SH,600790.SH,600791.SH,600792.SH,600793.SH,600794.SH,600795.SH,600796.SH,600797.SH,600798.SH,600799.SH,600800.SH,600801.SH,600802.SH,600803.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_179 返回数据 row 行数 = "+str(stk_managers_item_179.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_179)


stk_managers_item_180 = pro.stk_managers(ts_code='600804.SH,600805.SH,600806.SH,600807.SH,600808.SH,600809.SH,600810.SH,600811.SH,600812.SH,600813.SH,600814.SH,600815.SH,600816.SH,600817.SH,600818.SH,600819.SH,600820.SH,600821.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_180 返回数据 row 行数 = "+str(stk_managers_item_180.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_180)


stk_managers_item_181 = pro.stk_managers(ts_code='600822.SH,600823.SH,600824.SH,600825.SH,600826.SH,600827.SH,600828.SH,600829.SH,600830.SH,600831.SH,600832.SH,600833.SH,600834.SH,600835.SH,600836.SH,600837.SH,600838.SH,600839.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_181 返回数据 row 行数 = "+str(stk_managers_item_181.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_181)


stk_managers_item_182 = pro.stk_managers(ts_code='600840.SH,600841.SH,600842.SH,600843.SH,600844.SH,600845.SH,600846.SH,600847.SH,600848.SH,600850.SH,600851.SH,600852.SH,600853.SH,600854.SH,600855.SH,600856.SH,600857.SH,600858.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_182 返回数据 row 行数 = "+str(stk_managers_item_182.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_182)


stk_managers_item_183 = pro.stk_managers(ts_code='600859.SH,600860.SH,600861.SH,600862.SH,600863.SH,600864.SH,600865.SH,600866.SH,600867.SH,600868.SH,600869.SH,600870.SH,600871.SH,600872.SH,600873.SH,600874.SH,600875.SH,600876.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_183 返回数据 row 行数 = "+str(stk_managers_item_183.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_183)


stk_managers_item_184 = pro.stk_managers(ts_code='600877.SH,600878.SH,600879.SH,600880.SH,600881.SH,600882.SH,600883.SH,600884.SH,600885.SH,600886.SH,600887.SH,600888.SH,600889.SH,600890.SH,600891.SH,600892.SH,600893.SH,600894.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_184 返回数据 row 行数 = "+str(stk_managers_item_184.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_184)


stk_managers_item_185 = pro.stk_managers(ts_code='600895.SH,600896.SH,600897.SH,600898.SH,600899.SH,600900.SH,600901.SH,600903.SH,600905.SH,600906.SH,600908.SH,600909.SH,600916.SH,600917.SH,600918.SH,600919.SH,600926.SH,600928.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_185 返回数据 row 行数 = "+str(stk_managers_item_185.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_185)


stk_managers_item_186 = pro.stk_managers(ts_code='600929.SH,600933.SH,600936.SH,600939.SH,600956.SH,600958.SH,600959.SH,600960.SH,600961.SH,600962.SH,600963.SH,600965.SH,600966.SH,600967.SH,600968.SH,600969.SH,600970.SH,600971.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_186 返回数据 row 行数 = "+str(stk_managers_item_186.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_186)


stk_managers_item_187 = pro.stk_managers(ts_code='600973.SH,600975.SH,600976.SH,600977.SH,600978.SH,600979.SH,600980.SH,600981.SH,600982.SH,600983.SH,600984.SH,600985.SH,600986.SH,600987.SH,600988.SH,600989.SH,600990.SH,600991.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_187 返回数据 row 行数 = "+str(stk_managers_item_187.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_187)


stk_managers_item_188 = pro.stk_managers(ts_code='600992.SH,600993.SH,600995.SH,600996.SH,600997.SH,600998.SH,600999.SH,601000.SH,601001.SH,601002.SH,601003.SH,601005.SH,601006.SH,601007.SH,601008.SH,601009.SH,601010.SH,601011.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_188 返回数据 row 行数 = "+str(stk_managers_item_188.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_188)


stk_managers_item_189 = pro.stk_managers(ts_code='601012.SH,601015.SH,601016.SH,601018.SH,601019.SH,601020.SH,601021.SH,601028.SH,601038.SH,601058.SH,601066.SH,601068.SH,601069.SH,601077.SH,601086.SH,601088.SH,601098.SH,601099.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_189 返回数据 row 行数 = "+str(stk_managers_item_189.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_189)


stk_managers_item_190 = pro.stk_managers(ts_code='601100.SH,601101.SH,601106.SH,601107.SH,601108.SH,601111.SH,601113.SH,601116.SH,601117.SH,601118.SH,601126.SH,601127.SH,601128.SH,601137.SH,601138.SH,601139.SH,601155.SH,601156.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_190 返回数据 row 行数 = "+str(stk_managers_item_190.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_190)


stk_managers_item_191 = pro.stk_managers(ts_code='601158.SH,601162.SH,601163.SH,601166.SH,601168.SH,601169.SH,601177.SH,601179.SH,601186.SH,601187.SH,601188.SH,601198.SH,601199.SH,601200.SH,601208.SH,601211.SH,601212.SH,601216.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_191 返回数据 row 行数 = "+str(stk_managers_item_191.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_191)


stk_managers_item_192 = pro.stk_managers(ts_code='601218.SH,601222.SH,601225.SH,601226.SH,601228.SH,601229.SH,601231.SH,601233.SH,601236.SH,601238.SH,601258.SH,601268.SH,601279.SH,601288.SH,601298.SH,601299.SH,601311.SH,601318.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_192 返回数据 row 行数 = "+str(stk_managers_item_192.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_192)


stk_managers_item_193 = pro.stk_managers(ts_code='601319.SH,601326.SH,601328.SH,601330.SH,601333.SH,601336.SH,601339.SH,601360.SH,601366.SH,601368.SH,601369.SH,601375.SH,601377.SH,601388.SH,601390.SH,601398.SH,601399.SH,601456.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_193 返回数据 row 行数 = "+str(stk_managers_item_193.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_193)


stk_managers_item_194 = pro.stk_managers(ts_code='601500.SH,601512.SH,601515.SH,601518.SH,601519.SH,601528.SH,601555.SH,601558.SH,601566.SH,601567.SH,601568.SH,601577.SH,601579.SH,601588.SH,601595.SH,601598.SH,601599.SH,601600.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_194 返回数据 row 行数 = "+str(stk_managers_item_194.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_194)


stk_managers_item_195 = pro.stk_managers(ts_code='601601.SH,601606.SH,601607.SH,601608.SH,601609.SH,601611.SH,601615.SH,601616.SH,601618.SH,601619.SH,601628.SH,601633.SH,601636.SH,601658.SH,601665.SH,601666.SH,601668.SH,601669.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_195 返回数据 row 行数 = "+str(stk_managers_item_195.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_195)


stk_managers_item_196 = pro.stk_managers(ts_code='601677.SH,601678.SH,601686.SH,601688.SH,601689.SH,601696.SH,601698.SH,601699.SH,601700.SH,601702.SH,601717.SH,601718.SH,601727.SH,601766.SH,601777.SH,601778.SH,601788.SH,601789.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_196 返回数据 row 行数 = "+str(stk_managers_item_196.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_196)


stk_managers_item_197 = pro.stk_managers(ts_code='601798.SH,601799.SH,601800.SH,601801.SH,601808.SH,601811.SH,601816.SH,601818.SH,601827.SH,601828.SH,601838.SH,601857.SH,601858.SH,601860.SH,601865.SH,601866.SH,601869.SH,601872.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_197 返回数据 row 行数 = "+str(stk_managers_item_197.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_197)


stk_managers_item_198 = pro.stk_managers(ts_code='601877.SH,601878.SH,601880.SH,601881.SH,601882.SH,601886.SH,601888.SH,601890.SH,601898.SH,601899.SH,601900.SH,601901.SH,601908.SH,601916.SH,601918.SH,601919.SH,601928.SH,601929.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_198 返回数据 row 行数 = "+str(stk_managers_item_198.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_198)


stk_managers_item_199 = pro.stk_managers(ts_code='601933.SH,601939.SH,601949.SH,601952.SH,601956.SH,601958.SH,601963.SH,601965.SH,601966.SH,601968.SH,601969.SH,601975.SH,601985.SH,601988.SH,601989.SH,601990.SH,601991.SH,601992.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_199 返回数据 row 行数 = "+str(stk_managers_item_199.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_199)


stk_managers_item_200 = pro.stk_managers(ts_code='601995.SH,601996.SH,601997.SH,601998.SH,601999.SH,603000.SH,603001.SH,603002.SH,603003.SH,603005.SH,603006.SH,603007.SH,603008.SH,603009.SH,603010.SH,603011.SH,603012.SH,603013.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_200 返回数据 row 行数 = "+str(stk_managers_item_200.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_200)


stk_managers_item_201 = pro.stk_managers(ts_code='603015.SH,603016.SH,603017.SH,603018.SH,603019.SH,603020.SH,603021.SH,603022.SH,603023.SH,603025.SH,603026.SH,603027.SH,603028.SH,603029.SH,603030.SH,603031.SH,603032.SH,603033.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_201 返回数据 row 行数 = "+str(stk_managers_item_201.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_201)


stk_managers_item_202 = pro.stk_managers(ts_code='603035.SH,603036.SH,603037.SH,603038.SH,603039.SH,603040.SH,603041.SH,603042.SH,603043.SH,603045.SH,603050.SH,603053.SH,603055.SH,603056.SH,603058.SH,603059.SH,603060.SH,603063.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_202 返回数据 row 行数 = "+str(stk_managers_item_202.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_202)


stk_managers_item_203 = pro.stk_managers(ts_code='603066.SH,603067.SH,603068.SH,603069.SH,603076.SH,603077.SH,603078.SH,603079.SH,603080.SH,603081.SH,603083.SH,603085.SH,603086.SH,603087.SH,603088.SH,603089.SH,603090.SH,603093.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_203 返回数据 row 行数 = "+str(stk_managers_item_203.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_203)


stk_managers_item_204 = pro.stk_managers(ts_code='603095.SH,603096.SH,603098.SH,603099.SH,603100.SH,603101.SH,603103.SH,603105.SH,603106.SH,603108.SH,603109.SH,603110.SH,603111.SH,603112.SH,603113.SH,603115.SH,603116.SH,603117.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_204 返回数据 row 行数 = "+str(stk_managers_item_204.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_204)


stk_managers_item_205 = pro.stk_managers(ts_code='603118.SH,603121.SH,603123.SH,603126.SH,603127.SH,603128.SH,603129.SH,603131.SH,603133.SH,603136.SH,603138.SH,603139.SH,603155.SH,603156.SH,603157.SH,603158.SH,603159.SH,603160.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_205 返回数据 row 行数 = "+str(stk_managers_item_205.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_205)


stk_managers_item_206 = pro.stk_managers(ts_code='603161.SH,603165.SH,603166.SH,603167.SH,603168.SH,603169.SH,603171.SH,603177.SH,603178.SH,603179.SH,603180.SH,603181.SH,603183.SH,603185.SH,603186.SH,603187.SH,603188.SH,603189.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_206 返回数据 row 行数 = "+str(stk_managers_item_206.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_206)


stk_managers_item_207 = pro.stk_managers(ts_code='603192.SH,603195.SH,603196.SH,603197.SH,603198.SH,603199.SH,603200.SH,603203.SH,603208.SH,603212.SH,603214.SH,603217.SH,603218.SH,603220.SH,603221.SH,603222.SH,603223.SH,603225.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_207 返回数据 row 行数 = "+str(stk_managers_item_207.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_207)


stk_managers_item_208 = pro.stk_managers(ts_code='603226.SH,603227.SH,603228.SH,603229.SH,603232.SH,603233.SH,603236.SH,603238.SH,603239.SH,603256.SH,603258.SH,603259.SH,603260.SH,603266.SH,603267.SH,603268.SH,603269.SH,603277.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_208 返回数据 row 行数 = "+str(stk_managers_item_208.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_208)


stk_managers_item_209 = pro.stk_managers(ts_code='603278.SH,603279.SH,603283.SH,603286.SH,603288.SH,603289.SH,603290.SH,603297.SH,603298.SH,603299.SH,603300.SH,603301.SH,603303.SH,603305.SH,603306.SH,603308.SH,603309.SH,603311.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_209 返回数据 row 行数 = "+str(stk_managers_item_209.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_209)


stk_managers_item_210 = pro.stk_managers(ts_code='603313.SH,603315.SH,603316.SH,603317.SH,603318.SH,603319.SH,603320.SH,603321.SH,603322.SH,603323.SH,603324.SH,603326.SH,603327.SH,603328.SH,603329.SH,603330.SH,603331.SH,603332.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_210 返回数据 row 行数 = "+str(stk_managers_item_210.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_210)


stk_managers_item_211 = pro.stk_managers(ts_code='603333.SH,603335.SH,603336.SH,603337.SH,603338.SH,603339.SH,603345.SH,603348.SH,603351.SH,603353.SH,603355.SH,603356.SH,603357.SH,603358.SH,603359.SH,603360.SH,603363.SH,603365.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_211 返回数据 row 行数 = "+str(stk_managers_item_211.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_211)


stk_managers_item_212 = pro.stk_managers(ts_code='603366.SH,603367.SH,603368.SH,603369.SH,603377.SH,603378.SH,603379.SH,603380.SH,603383.SH,603385.SH,603386.SH,603387.SH,603388.SH,603389.SH,603390.SH,603392.SH,603393.SH,603396.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_212 返回数据 row 行数 = "+str(stk_managers_item_212.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_212)


stk_managers_item_213 = pro.stk_managers(ts_code='603398.SH,603399.SH,603408.SH,603416.SH,603421.SH,603429.SH,603439.SH,603444.SH,603456.SH,603458.SH,603466.SH,603477.SH,603486.SH,603488.SH,603489.SH,603496.SH,603499.SH,603500.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_213 返回数据 row 行数 = "+str(stk_managers_item_213.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_213)


stk_managers_item_214 = pro.stk_managers(ts_code='603501.SH,603505.SH,603506.SH,603507.SH,603508.SH,603511.SH,603515.SH,603516.SH,603517.SH,603518.SH,603519.SH,603520.SH,603527.SH,603528.SH,603529.SH,603530.SH,603533.SH,603535.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_214 返回数据 row 行数 = "+str(stk_managers_item_214.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_214)


stk_managers_item_215 = pro.stk_managers(ts_code='603536.SH,603538.SH,603551.SH,603555.SH,603556.SH,603557.SH,603558.SH,603559.SH,603565.SH,603566.SH,603567.SH,603568.SH,603569.SH,603577.SH,603578.SH,603579.SH,603580.SH,603583.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_215 返回数据 row 行数 = "+str(stk_managers_item_215.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_215)


stk_managers_item_216 = pro.stk_managers(ts_code='603585.SH,603586.SH,603587.SH,603588.SH,603589.SH,603590.SH,603595.SH,603596.SH,603598.SH,603599.SH,603600.SH,603601.SH,603602.SH,603603.SH,603605.SH,603606.SH,603607.SH,603608.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_216 返回数据 row 行数 = "+str(stk_managers_item_216.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_216)


stk_managers_item_217 = pro.stk_managers(ts_code='603609.SH,603610.SH,603611.SH,603612.SH,603613.SH,603615.SH,603616.SH,603617.SH,603618.SH,603619.SH,603626.SH,603628.SH,603629.SH,603630.SH,603633.SH,603636.SH,603637.SH,603638.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_217 返回数据 row 行数 = "+str(stk_managers_item_217.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_217)


stk_managers_item_218 = pro.stk_managers(ts_code='603639.SH,603648.SH,603650.SH,603655.SH,603656.SH,603657.SH,603658.SH,603659.SH,603660.SH,603661.SH,603662.SH,603663.SH,603665.SH,603666.SH,603667.SH,603668.SH,603669.SH,603676.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_218 返回数据 row 行数 = "+str(stk_managers_item_218.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_218)


stk_managers_item_219 = pro.stk_managers(ts_code='603677.SH,603678.SH,603679.SH,603680.SH,603681.SH,603682.SH,603683.SH,603685.SH,603686.SH,603687.SH,603688.SH,603689.SH,603690.SH,603693.SH,603696.SH,603697.SH,603698.SH,603699.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_219 返回数据 row 行数 = "+str(stk_managers_item_219.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_219)


stk_managers_item_220 = pro.stk_managers(ts_code='603700.SH,603701.SH,603703.SH,603706.SH,603707.SH,603708.SH,603709.SH,603711.SH,603712.SH,603713.SH,603716.SH,603717.SH,603718.SH,603719.SH,603721.SH,603722.SH,603725.SH,603726.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_220 返回数据 row 行数 = "+str(stk_managers_item_220.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_220)


stk_managers_item_221 = pro.stk_managers(ts_code='603727.SH,603728.SH,603729.SH,603730.SH,603733.SH,603737.SH,603738.SH,603739.SH,603755.SH,603757.SH,603758.SH,603759.SH,603766.SH,603767.SH,603768.SH,603773.SH,603776.SH,603777.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_221 返回数据 row 行数 = "+str(stk_managers_item_221.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_221)


stk_managers_item_222 = pro.stk_managers(ts_code='603778.SH,603779.SH,603786.SH,603787.SH,603788.SH,603789.SH,603790.SH,603797.SH,603798.SH,603799.SH,603800.SH,603801.SH,603803.SH,603806.SH,603808.SH,603809.SH,603810.SH,603811.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_222 返回数据 row 行数 = "+str(stk_managers_item_222.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_222)


stk_managers_item_223 = pro.stk_managers(ts_code='603813.SH,603815.SH,603816.SH,603817.SH,603818.SH,603819.SH,603822.SH,603823.SH,603825.SH,603826.SH,603828.SH,603829.SH,603833.SH,603836.SH,603838.SH,603839.SH,603843.SH,603848.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_223 返回数据 row 行数 = "+str(stk_managers_item_223.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_223)


stk_managers_item_224 = pro.stk_managers(ts_code='603855.SH,603856.SH,603858.SH,603859.SH,603860.SH,603861.SH,603863.SH,603866.SH,603867.SH,603868.SH,603869.SH,603871.SH,603876.SH,603877.SH,603878.SH,603879.SH,603880.SH,603881.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_224 返回数据 row 行数 = "+str(stk_managers_item_224.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_224)


stk_managers_item_225 = pro.stk_managers(ts_code='603882.SH,603883.SH,603885.SH,603886.SH,603887.SH,603888.SH,603889.SH,603890.SH,603893.SH,603895.SH,603896.SH,603897.SH,603898.SH,603899.SH,603900.SH,603901.SH,603903.SH,603906.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_225 返回数据 row 行数 = "+str(stk_managers_item_225.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_225)


stk_managers_item_226 = pro.stk_managers(ts_code='603908.SH,603909.SH,603912.SH,603915.SH,603916.SH,603917.SH,603918.SH,603919.SH,603920.SH,603922.SH,603926.SH,603927.SH,603928.SH,603929.SH,603931.SH,603933.SH,603936.SH,603937.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_226 返回数据 row 行数 = "+str(stk_managers_item_226.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_226)


stk_managers_item_227 = pro.stk_managers(ts_code='603938.SH,603939.SH,603948.SH,603949.SH,603950.SH,603955.SH,603956.SH,603958.SH,603959.SH,603960.SH,603963.SH,603966.SH,603967.SH,603968.SH,603969.SH,603970.SH,603976.SH,603977.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_227 返回数据 row 行数 = "+str(stk_managers_item_227.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_227)


stk_managers_item_228 = pro.stk_managers(ts_code='603978.SH,603979.SH,603980.SH,603982.SH,603983.SH,603985.SH,603986.SH,603987.SH,603988.SH,603989.SH,603990.SH,603991.SH,603992.SH,603993.SH,603995.SH,603996.SH,603997.SH,603998.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_228 返回数据 row 行数 = "+str(stk_managers_item_228.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_228)


stk_managers_item_229 = pro.stk_managers(ts_code='603999.SH,605001.SH,605003.SH,605005.SH,605006.SH,605007.SH,605008.SH,605009.SH,605011.SH,605016.SH,605018.SH,605050.SH,605055.SH,605058.SH,605060.SH,605066.SH,605068.SH,605077.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_229 返回数据 row 行数 = "+str(stk_managers_item_229.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_229)


stk_managers_item_230 = pro.stk_managers(ts_code='605080.SH,605081.SH,605086.SH,605088.SH,605089.SH,605090.SH,605098.SH,605099.SH,605100.SH,605108.SH,605111.SH,605116.SH,605117.SH,605118.SH,605122.SH,605123.SH,605128.SH,605133.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_230 返回数据 row 行数 = "+str(stk_managers_item_230.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_230)


stk_managers_item_231 = pro.stk_managers(ts_code='605136.SH,605151.SH,605155.SH,605158.SH,605162.SH,605166.SH,605168.SH,605169.SH,605177.SH,605178.SH,605179.SH,605180.SH,605183.SH,605186.SH,605188.SH,605189.SH,605196.SH,605198.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_231 返回数据 row 行数 = "+str(stk_managers_item_231.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_231)


stk_managers_item_232 = pro.stk_managers(ts_code='605199.SH,605208.SH,605218.SH,605222.SH,605228.SH,605255.SH,605258.SH,605259.SH,605266.SH,605268.SH,605277.SH,605286.SH,605287.SH,605288.SH,605289.SH,605296.SH,605298.SH,605299.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_232 返回数据 row 行数 = "+str(stk_managers_item_232.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_232)


stk_managers_item_233 = pro.stk_managers(ts_code='605300.SH,605303.SH,605305.SH,605318.SH,605319.SH,605333.SH,605336.SH,605337.SH,605338.SH,605339.SH,605358.SH,605366.SH,605368.SH,605369.SH,605376.SH,605377.SH,605378.SH,605388.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_233 返回数据 row 行数 = "+str(stk_managers_item_233.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_233)


stk_managers_item_234 = pro.stk_managers(ts_code='605389.SH,605398.SH,605399.SH,605488.SH,605499.SH,605500.SH,688001.SH,688002.SH,688003.SH,688004.SH,688005.SH,688006.SH,688007.SH,688008.SH,688009.SH,688010.SH,688011.SH,688012.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_234 返回数据 row 行数 = "+str(stk_managers_item_234.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_234)


stk_managers_item_235 = pro.stk_managers(ts_code='688013.SH,688015.SH,688016.SH,688017.SH,688018.SH,688019.SH,688020.SH,688021.SH,688022.SH,688023.SH,688025.SH,688026.SH,688027.SH,688028.SH,688029.SH,688030.SH,688033.SH,688036.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_235 返回数据 row 行数 = "+str(stk_managers_item_235.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_235)


stk_managers_item_236 = pro.stk_managers(ts_code='688037.SH,688039.SH,688050.SH,688051.SH,688055.SH,688056.SH,688057.SH,688058.SH,688059.SH,688060.SH,688063.SH,688065.SH,688066.SH,688067.SH,688068.SH,688069.SH,688070.SH,688076.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_236 返回数据 row 行数 = "+str(stk_managers_item_236.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_236)


stk_managers_item_237 = pro.stk_managers(ts_code='688077.SH,688078.SH,688079.SH,688080.SH,688081.SH,688083.SH,688085.SH,688086.SH,688088.SH,688089.SH,688090.SH,688092.SH,688093.SH,688095.SH,688096.SH,688097.SH,688098.SH,688099.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_237 返回数据 row 行数 = "+str(stk_managers_item_237.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_237)


stk_managers_item_238 = pro.stk_managers(ts_code='688100.SH,688101.SH,688106.SH,688108.SH,688109.SH,688111.SH,688113.SH,688116.SH,688117.SH,688118.SH,688122.SH,688123.SH,688126.SH,688127.SH,688128.SH,688129.SH,688131.SH,688133.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_238 返回数据 row 行数 = "+str(stk_managers_item_238.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_238)


stk_managers_item_239 = pro.stk_managers(ts_code='688135.SH,688136.SH,688138.SH,688139.SH,688155.SH,688156.SH,688157.SH,688158.SH,688159.SH,688160.SH,688161.SH,688165.SH,688166.SH,688168.SH,688169.SH,688177.SH,688178.SH,688179.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_239 返回数据 row 行数 = "+str(stk_managers_item_239.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_239)


stk_managers_item_240 = pro.stk_managers(ts_code='688180.SH,688181.SH,688183.SH,688185.SH,688186.SH,688188.SH,688189.SH,688191.SH,688195.SH,688196.SH,688198.SH,688199.SH,688200.SH,688201.SH,688202.SH,688208.SH,688215.SH,688216.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_240 返回数据 row 行数 = "+str(stk_managers_item_240.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_240)


stk_managers_item_241 = pro.stk_managers(ts_code='688217.SH,688218.SH,688219.SH,688221.SH,688222.SH,688226.SH,688228.SH,688229.SH,688233.SH,688239.SH,688256.SH,688258.SH,688260.SH,688266.SH,688268.SH,688269.SH,688276.SH,688277.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_241 返回数据 row 行数 = "+str(stk_managers_item_241.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_241)


stk_managers_item_242 = pro.stk_managers(ts_code='688278.SH,688286.SH,688288.SH,688289.SH,688298.SH,688299.SH,688300.SH,688301.SH,688308.SH,688309.SH,688310.SH,688311.SH,688312.SH,688313.SH,688314.SH,688315.SH,688316.SH,688317.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_242 返回数据 row 行数 = "+str(stk_managers_item_242.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_242)


stk_managers_item_243 = pro.stk_managers(ts_code='688318.SH,688319.SH,688321.SH,688323.SH,688328.SH,688329.SH,688330.SH,688333.SH,688335.SH,688336.SH,688338.SH,688339.SH,688345.SH,688350.SH,688355.SH,688356.SH,688357.SH,688358.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_243 返回数据 row 行数 = "+str(stk_managers_item_243.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_243)


stk_managers_item_244 = pro.stk_managers(ts_code='688359.SH,688360.SH,688363.SH,688365.SH,688366.SH,688367.SH,688368.SH,688369.SH,688377.SH,688378.SH,688379.SH,688383.SH,688386.SH,688388.SH,688389.SH,688390.SH,688393.SH,688395.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_244 返回数据 row 行数 = "+str(stk_managers_item_244.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_244)


stk_managers_item_245 = pro.stk_managers(ts_code='688396.SH,688398.SH,688399.SH,688408.SH,688418.SH,688425.SH,688456.SH,688466.SH,688468.SH,688488.SH,688499.SH,688500.SH,688505.SH,688508.SH,688510.SH,688513.SH,688516.SH,688517.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_245 返回数据 row 行数 = "+str(stk_managers_item_245.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_245)


stk_managers_item_246 = pro.stk_managers(ts_code='688518.SH,688519.SH,688520.SH,688521.SH,688526.SH,688528.SH,688529.SH,688533.SH,688536.SH,688538.SH,688550.SH,688551.SH,688555.SH,688556.SH,688557.SH,688558.SH,688559.SH,688560.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_246 返回数据 row 行数 = "+str(stk_managers_item_246.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_246)


stk_managers_item_247 = pro.stk_managers(ts_code='688561.SH,688565.SH,688566.SH,688567.SH,688568.SH,688569.SH,688571.SH,688575.SH,688577.SH,688578.SH,688579.SH,688580.SH,688585.SH,688586.SH,688588.SH,688589.SH,688590.SH,688595.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_247 返回数据 row 行数 = "+str(stk_managers_item_247.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_247)


stk_managers_item_248 = pro.stk_managers(ts_code='688596.SH,688597.SH,688598.SH,688599.SH,688600.SH,688601.SH,688606.SH,688607.SH,688608.SH,688609.SH,688611.SH,688613.SH,688616.SH,688617.SH,688618.SH,688619.SH,688621.SH,688625.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_248 返回数据 row 行数 = "+str(stk_managers_item_248.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_248)


stk_managers_item_249 = pro.stk_managers(ts_code='688626.SH,688628.SH,688630.SH,688633.SH,688636.SH,688639.SH,688655.SH,688656.SH,688658.SH,688659.SH,688660.SH,688661.SH,688662.SH,688663.SH,688665.SH,688667.SH,688668.SH,688669.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_249 返回数据 row 行数 = "+str(stk_managers_item_249.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_249)


stk_managers_item_250 = pro.stk_managers(ts_code='688676.SH,688677.SH,688678.SH,688679.SH,688680.SH,688681.SH,688682.SH,688683.SH,688685.SH,688686.SH,688687.SH,688689.SH,688690.SH,688696.SH,688698.SH,688699.SH,688700.SH,688777.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_250 返回数据 row 行数 = "+str(stk_managers_item_250.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_250)


stk_managers_item_251 = pro.stk_managers(ts_code='688788.SH,688819.SH,688981.SH,T00018.SH', fields='ts_code,ann_date,name,gender,lev,title,edu,national,birthday,begin_date,end_date,resume')
print(" 上市公司管理层 stk_managers_item_251 返回数据 row 行数 = "+str(stk_managers_item_251.shape[0]))
stk_managers_data_List = stk_managers_data_List.append(stk_managers_item_251)


print("stk_managers_data_List.__len__() = "+str(stk_managers_data_List.__len__()))
stk_managers_data_List.to_excel(stk_managers_excel_writer,'上市公司管理层',index=False)
stk_managers_excel_writer.save()
