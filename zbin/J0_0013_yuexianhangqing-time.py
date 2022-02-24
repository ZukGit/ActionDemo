#  encoding: utf-8

import tushare as ts
import os
import pandas as pd  # https://www.jianshu.com/p/5c0aa1fa19af
import openpyxl 
from openpyxl import load_workbook,Workbook
import time,datetime  # https://www.jb51.net/article/66019.htm
import re
import os
import tempfile

############################## 全局数据初始化块 数据初始化函数  End ##############################

##############  properities 函数 Begin ##############

class Properties:

    def __init__(self, file_name):
        self.file_name = file_name
        self.properties = {}
        try:
            fopen = open(self.file_name, 'r')
            for line in fopen:
                line = line.strip()
                if line.find('=') > 0 and not line.startswith('#'):
                    strs = line.split('=')
                    self.properties[strs[0].strip()] = strs[1].strip()
        except Exception :
            raise
        else:
            fopen.close()

    def has_key(self, key):
        return key in self.properties

    def get(self, key, default_value=''):
        if key in self.properties:
            return self.properties[key]
        return default_value

    def put(self, key, value):
        self.properties[key] = value
        replace_property(self.file_name, key + '=.*', key + '=' + value, True)


def replace_property(file_name, from_regex, to_str, append_on_not_exists=True):
    tmpfile = tempfile.TemporaryFile()

    if os.path.exists(file_name):
        r_open = open(file_name, 'r')
        pattern = re.compile(r'' + from_regex)
        found = None
        for line in r_open:
            if pattern.search(line) and not line.strip().startswith('#'):
                found = True
                line = re.sub(from_regex, to_str, line)
            tmpfile.write(line.encode())
        if not found and append_on_not_exists:
            tmpfile.write(('\n' + to_str).encode())
        r_open.close()
        tmpfile.seek(0)

        content = tmpfile.read()

        if os.path.exists(file_name):
            os.remove(file_name)

        w_open = open(file_name, 'wb')
        w_open.write(content)
        w_open.close()

        tmpfile.close()
    else:
        print ("file %s not found" % file_name)

##############  properities 函数 End ##############


##############  XLSX excel 函数 Begin ##############

def createexcel(filename):    ### 创建本地文件名称为 filename的文件
    wb = Workbook()
    curPath = os.path.abspath('.')
    cur_desktop_path=os.path.join(os.path.expanduser("~"), 'Desktop')
    curDir = cur_desktop_path+os.sep+"zbin"+os.sep+"J0_Data"+os.sep  ## 创建~/Desktop/zbin/J0_Data 这个 这个工作目录
    #curDir = curPath+os.sep
    if not os.path.exists(curDir):
        os.makedirs(curDir)
    #curPath = os.path.dirname(os.path.abspath('.'))
    #t = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    #suffix = '.xlsx' # 文件类型
    #newfile = t + suffix
    xlsxPath = curDir + filename
    print("curPath curDir=" + curPath)
    print("createexcel zbin_JO_Dir=" + curDir)
    print("createexcel path=" + xlsxPath)
    if not os.path.exists(xlsxPath):
        wb.save(xlsxPath)
        print(" wb.save(xlsxPath)  xlsxPath =" + xlsxPath)
        time.sleep(1)
    return xlsxPath

def getColumnIndex(table, columnName):   ## 返回 table 中 名称为  columnName 的 那列 的索引
    columnIndex = None
    for i in range(table.ncols):
        if(table.cell_value(0, i) == columnName):
            columnIndex = i
            break
    return columnIndex
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007

##############  XLSX excel 函数 End ##############



############################## Prop初始化 Begin ##############################

#
#Thu Aug 13 22:33:45 CST 2020
#rixianhangqing-time_record_date=20200707
#rixianhangqing-time_start_date=20010101

desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
zbin_path = str(desktop_path)+os.sep+"zbin"
j0_properties_path= str(zbin_path)+os.sep+"J0.properties"
print("desktop_path = "+ str(desktop_path))
print("zbin_path = "+ str(zbin_path))
print("j0_properties_path = "+ str(j0_properties_path))
J0_PROPS =  Properties(j0_properties_path)


############################## Prop初始化 End ##############################


############## tscode_股票列表的初始化  Begin ##############
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007
def init_tscode_data(book_name, sheet_name,ts_code_set,tscode_name_dict):
    # 读取excel
    wb = openpyxl.load_workbook(book_name)
    # 读取sheet表
    ws = wb[sheet_name]
    for i in range(ws.max_row):
         # 获取下拉框中的所有选择值
         if (i == 0 or i == 1):
             continue
         #print("i="+str(i)+" 总的列数:" + str(ws.max_row)+"  value:"+str(ws.cell(i,2).value))
         tscode_item = str(ws.cell(i,1).value)  ##  20201010--> 2:ts_code    4_name   ##  20201116--> 1:ts_code    3_name 
         tscode_name_item = str(ws.cell(i,3).value)
         #print("index="+str(i)+" tscode_item = "+ str(tscode_item) + "   tscode_name_item="+str(tscode_name_item))
         ts_code_set.add(str(ws.cell(i,1).value))
         tscode_name_dict[tscode_item]=tscode_name_item



tscode_set = set()    #### 股票代码的集合   000001.SZ   .... 999999.SH
tscode_name_dict = dict()  #### code-name 的 map的集合
init_tscode_data(zbin_path+os.sep+"J0_Python"+os.sep+"J0_股票列表.xlsx","股票列表",tscode_set,tscode_name_dict)
############## tscode_股票列表的初始化  End  ##############


############################## 运行属性 Begin ##############################
pd.set_option('display.max_rows', None)   ##  解决纵向出现...
#pd.set_option('display.width', 1000) 
pd.set_option('expand_frame_repr', False)  ##  解决横向出现...
Cur_Abs_Path=os.path.abspath('.')   # 表示当前所处的文件夹的绝对路径
print("当前绝对路径:"+Cur_Abs_Path)
Cur_Ref_Path=os.path.abspath('..')  # 表示当前所处的文件夹上一级文件夹的绝对路径
print("当前父目录绝对路径:"+Cur_Ref_Path)
############################## 运行属性 End ##############################

############################## 时间Date Begin ##############################

now_yyyymmdd=str(time.strftime('%Y%m%d', time.localtime(time.time())))
print("now_yyyymmdd = "+ str(now_yyyymmdd))

############################## 时间Date End ##############################





pro = ts.pro_api('43acb9a5ddc2cf73c6c4ea54796748f965457ed57daaa736bb778ea2')
# print(J0_PROPS.get(tree_node_name+"record_date"))           #根据key读取value
# J0_PROPS.put(tree_node_name+"record_date", now_yyyymmdd)       ###  覆盖原有的 key 和 value
#  zukgit
# yuexianhangqing-time_zukgit_website  =   https://tushare.pro/document/2?doc_id=145
tree_node_name="yuexianhangqing-time"+"_"
createexcel('monthly_2010.xlsx')
monthly_2010_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2010.xlsx')
monthly_2010_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2010.xlsx', engine='openpyxl')
monthly_2010_excel_writer.book = monthly_2010_book
monthly_2010_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2010_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20100101")       ###  更新 周线记录日期
monthly_20100101 = pro.monthly(trade_date='20100129', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20100101_tscode_list = list() 
for ts_code_sh in monthly_20100101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20100101_tscode_list.append(stock_name)
monthly_20100101_addname_dataframe=pd.DataFrame()
monthly_20100101_addname_dataframe['cname'] = monthly_20100101_tscode_list
for table_name in monthly_20100101.columns.values.tolist():
    monthly_20100101_addname_dataframe[table_name] = monthly_20100101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20100101 20100129 返回数据 row 行数 = "+str(monthly_20100101.shape[0]))
monthly_20100101_addname_dataframe.to_excel(monthly_2010_excel_writer,'1',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100201")       ###  更新 周线记录日期
monthly_20100201 = pro.monthly(trade_date='20100226', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20100201_tscode_list = list() 
for ts_code_sh in monthly_20100201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20100201_tscode_list.append(stock_name)
monthly_20100201_addname_dataframe=pd.DataFrame()
monthly_20100201_addname_dataframe['cname'] = monthly_20100201_tscode_list
for table_name in monthly_20100201.columns.values.tolist():
    monthly_20100201_addname_dataframe[table_name] = monthly_20100201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20100201 20100226 返回数据 row 行数 = "+str(monthly_20100201.shape[0]))
monthly_20100201_addname_dataframe.to_excel(monthly_2010_excel_writer,'2',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100301")       ###  更新 周线记录日期
monthly_20100301 = pro.monthly(trade_date='20100331', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20100301_tscode_list = list() 
for ts_code_sh in monthly_20100301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20100301_tscode_list.append(stock_name)
monthly_20100301_addname_dataframe=pd.DataFrame()
monthly_20100301_addname_dataframe['cname'] = monthly_20100301_tscode_list
for table_name in monthly_20100301.columns.values.tolist():
    monthly_20100301_addname_dataframe[table_name] = monthly_20100301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20100301 20100331 返回数据 row 行数 = "+str(monthly_20100301.shape[0]))
monthly_20100301_addname_dataframe.to_excel(monthly_2010_excel_writer,'3',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100401")       ###  更新 周线记录日期
monthly_20100401 = pro.monthly(trade_date='20100430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20100401_tscode_list = list() 
for ts_code_sh in monthly_20100401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20100401_tscode_list.append(stock_name)
monthly_20100401_addname_dataframe=pd.DataFrame()
monthly_20100401_addname_dataframe['cname'] = monthly_20100401_tscode_list
for table_name in monthly_20100401.columns.values.tolist():
    monthly_20100401_addname_dataframe[table_name] = monthly_20100401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20100401 20100430 返回数据 row 行数 = "+str(monthly_20100401.shape[0]))
monthly_20100401_addname_dataframe.to_excel(monthly_2010_excel_writer,'4',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100501")       ###  更新 周线记录日期
monthly_20100501 = pro.monthly(trade_date='20100531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20100501_tscode_list = list() 
for ts_code_sh in monthly_20100501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20100501_tscode_list.append(stock_name)
monthly_20100501_addname_dataframe=pd.DataFrame()
monthly_20100501_addname_dataframe['cname'] = monthly_20100501_tscode_list
for table_name in monthly_20100501.columns.values.tolist():
    monthly_20100501_addname_dataframe[table_name] = monthly_20100501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20100501 20100531 返回数据 row 行数 = "+str(monthly_20100501.shape[0]))
monthly_20100501_addname_dataframe.to_excel(monthly_2010_excel_writer,'5',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100601")       ###  更新 周线记录日期
monthly_20100601 = pro.monthly(trade_date='20100630', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20100601_tscode_list = list() 
for ts_code_sh in monthly_20100601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20100601_tscode_list.append(stock_name)
monthly_20100601_addname_dataframe=pd.DataFrame()
monthly_20100601_addname_dataframe['cname'] = monthly_20100601_tscode_list
for table_name in monthly_20100601.columns.values.tolist():
    monthly_20100601_addname_dataframe[table_name] = monthly_20100601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20100601 20100630 返回数据 row 行数 = "+str(monthly_20100601.shape[0]))
monthly_20100601_addname_dataframe.to_excel(monthly_2010_excel_writer,'6',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100701")       ###  更新 周线记录日期
monthly_20100701 = pro.monthly(trade_date='20100730', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20100701_tscode_list = list() 
for ts_code_sh in monthly_20100701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20100701_tscode_list.append(stock_name)
monthly_20100701_addname_dataframe=pd.DataFrame()
monthly_20100701_addname_dataframe['cname'] = monthly_20100701_tscode_list
for table_name in monthly_20100701.columns.values.tolist():
    monthly_20100701_addname_dataframe[table_name] = monthly_20100701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20100701 20100730 返回数据 row 行数 = "+str(monthly_20100701.shape[0]))
monthly_20100701_addname_dataframe.to_excel(monthly_2010_excel_writer,'7',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100801")       ###  更新 周线记录日期
monthly_20100801 = pro.monthly(trade_date='20100831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20100801_tscode_list = list() 
for ts_code_sh in monthly_20100801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20100801_tscode_list.append(stock_name)
monthly_20100801_addname_dataframe=pd.DataFrame()
monthly_20100801_addname_dataframe['cname'] = monthly_20100801_tscode_list
for table_name in monthly_20100801.columns.values.tolist():
    monthly_20100801_addname_dataframe[table_name] = monthly_20100801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20100801 20100831 返回数据 row 行数 = "+str(monthly_20100801.shape[0]))
monthly_20100801_addname_dataframe.to_excel(monthly_2010_excel_writer,'8',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100901")       ###  更新 周线记录日期
monthly_20100901 = pro.monthly(trade_date='20100930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20100901_tscode_list = list() 
for ts_code_sh in monthly_20100901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20100901_tscode_list.append(stock_name)
monthly_20100901_addname_dataframe=pd.DataFrame()
monthly_20100901_addname_dataframe['cname'] = monthly_20100901_tscode_list
for table_name in monthly_20100901.columns.values.tolist():
    monthly_20100901_addname_dataframe[table_name] = monthly_20100901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20100901 20100930 返回数据 row 行数 = "+str(monthly_20100901.shape[0]))
monthly_20100901_addname_dataframe.to_excel(monthly_2010_excel_writer,'9',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101001")       ###  更新 周线记录日期
monthly_20101001 = pro.monthly(trade_date='20101029', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20101001_tscode_list = list() 
for ts_code_sh in monthly_20101001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20101001_tscode_list.append(stock_name)
monthly_20101001_addname_dataframe=pd.DataFrame()
monthly_20101001_addname_dataframe['cname'] = monthly_20101001_tscode_list
for table_name in monthly_20101001.columns.values.tolist():
    monthly_20101001_addname_dataframe[table_name] = monthly_20101001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20101001 20101029 返回数据 row 行数 = "+str(monthly_20101001.shape[0]))
monthly_20101001_addname_dataframe.to_excel(monthly_2010_excel_writer,'10',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101101")       ###  更新 周线记录日期
monthly_20101101 = pro.monthly(trade_date='20101130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20101101_tscode_list = list() 
for ts_code_sh in monthly_20101101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20101101_tscode_list.append(stock_name)
monthly_20101101_addname_dataframe=pd.DataFrame()
monthly_20101101_addname_dataframe['cname'] = monthly_20101101_tscode_list
for table_name in monthly_20101101.columns.values.tolist():
    monthly_20101101_addname_dataframe[table_name] = monthly_20101101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20101101 20101130 返回数据 row 行数 = "+str(monthly_20101101.shape[0]))
monthly_20101101_addname_dataframe.to_excel(monthly_2010_excel_writer,'11',index=False)
monthly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101201")       ###  更新 周线记录日期
monthly_20101201 = pro.monthly(trade_date='20101231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20101201_tscode_list = list() 
for ts_code_sh in monthly_20101201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20101201_tscode_list.append(stock_name)
monthly_20101201_addname_dataframe=pd.DataFrame()
monthly_20101201_addname_dataframe['cname'] = monthly_20101201_tscode_list
for table_name in monthly_20101201.columns.values.tolist():
    monthly_20101201_addname_dataframe[table_name] = monthly_20101201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20101201 20101231 返回数据 row 行数 = "+str(monthly_20101201.shape[0]))
monthly_20101201_addname_dataframe.to_excel(monthly_2010_excel_writer,'12',index=False)
monthly_2010_excel_writer.save()
createexcel('monthly_2011.xlsx')
monthly_2011_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2011.xlsx')
monthly_2011_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2011.xlsx', engine='openpyxl')
monthly_2011_excel_writer.book = monthly_2011_book
monthly_2011_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2011_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20110101")       ###  更新 周线记录日期
monthly_20110101 = pro.monthly(trade_date='20110131', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20110101_tscode_list = list() 
for ts_code_sh in monthly_20110101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20110101_tscode_list.append(stock_name)
monthly_20110101_addname_dataframe=pd.DataFrame()
monthly_20110101_addname_dataframe['cname'] = monthly_20110101_tscode_list
for table_name in monthly_20110101.columns.values.tolist():
    monthly_20110101_addname_dataframe[table_name] = monthly_20110101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20110101 20110131 返回数据 row 行数 = "+str(monthly_20110101.shape[0]))
monthly_20110101_addname_dataframe.to_excel(monthly_2011_excel_writer,'1',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110201")       ###  更新 周线记录日期
monthly_20110201 = pro.monthly(trade_date='20110228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20110201_tscode_list = list() 
for ts_code_sh in monthly_20110201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20110201_tscode_list.append(stock_name)
monthly_20110201_addname_dataframe=pd.DataFrame()
monthly_20110201_addname_dataframe['cname'] = monthly_20110201_tscode_list
for table_name in monthly_20110201.columns.values.tolist():
    monthly_20110201_addname_dataframe[table_name] = monthly_20110201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20110201 20110228 返回数据 row 行数 = "+str(monthly_20110201.shape[0]))
monthly_20110201_addname_dataframe.to_excel(monthly_2011_excel_writer,'2',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110301")       ###  更新 周线记录日期
monthly_20110301 = pro.monthly(trade_date='20110331', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20110301_tscode_list = list() 
for ts_code_sh in monthly_20110301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20110301_tscode_list.append(stock_name)
monthly_20110301_addname_dataframe=pd.DataFrame()
monthly_20110301_addname_dataframe['cname'] = monthly_20110301_tscode_list
for table_name in monthly_20110301.columns.values.tolist():
    monthly_20110301_addname_dataframe[table_name] = monthly_20110301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20110301 20110331 返回数据 row 行数 = "+str(monthly_20110301.shape[0]))
monthly_20110301_addname_dataframe.to_excel(monthly_2011_excel_writer,'3',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110401")       ###  更新 周线记录日期
monthly_20110401 = pro.monthly(trade_date='20110429', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20110401_tscode_list = list() 
for ts_code_sh in monthly_20110401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20110401_tscode_list.append(stock_name)
monthly_20110401_addname_dataframe=pd.DataFrame()
monthly_20110401_addname_dataframe['cname'] = monthly_20110401_tscode_list
for table_name in monthly_20110401.columns.values.tolist():
    monthly_20110401_addname_dataframe[table_name] = monthly_20110401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20110401 20110429 返回数据 row 行数 = "+str(monthly_20110401.shape[0]))
monthly_20110401_addname_dataframe.to_excel(monthly_2011_excel_writer,'4',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110501")       ###  更新 周线记录日期
monthly_20110501 = pro.monthly(trade_date='20110531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20110501_tscode_list = list() 
for ts_code_sh in monthly_20110501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20110501_tscode_list.append(stock_name)
monthly_20110501_addname_dataframe=pd.DataFrame()
monthly_20110501_addname_dataframe['cname'] = monthly_20110501_tscode_list
for table_name in monthly_20110501.columns.values.tolist():
    monthly_20110501_addname_dataframe[table_name] = monthly_20110501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20110501 20110531 返回数据 row 行数 = "+str(monthly_20110501.shape[0]))
monthly_20110501_addname_dataframe.to_excel(monthly_2011_excel_writer,'5',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110601")       ###  更新 周线记录日期
monthly_20110601 = pro.monthly(trade_date='20110630', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20110601_tscode_list = list() 
for ts_code_sh in monthly_20110601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20110601_tscode_list.append(stock_name)
monthly_20110601_addname_dataframe=pd.DataFrame()
monthly_20110601_addname_dataframe['cname'] = monthly_20110601_tscode_list
for table_name in monthly_20110601.columns.values.tolist():
    monthly_20110601_addname_dataframe[table_name] = monthly_20110601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20110601 20110630 返回数据 row 行数 = "+str(monthly_20110601.shape[0]))
monthly_20110601_addname_dataframe.to_excel(monthly_2011_excel_writer,'6',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110701")       ###  更新 周线记录日期
monthly_20110701 = pro.monthly(trade_date='20110729', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20110701_tscode_list = list() 
for ts_code_sh in monthly_20110701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20110701_tscode_list.append(stock_name)
monthly_20110701_addname_dataframe=pd.DataFrame()
monthly_20110701_addname_dataframe['cname'] = monthly_20110701_tscode_list
for table_name in monthly_20110701.columns.values.tolist():
    monthly_20110701_addname_dataframe[table_name] = monthly_20110701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20110701 20110729 返回数据 row 行数 = "+str(monthly_20110701.shape[0]))
monthly_20110701_addname_dataframe.to_excel(monthly_2011_excel_writer,'7',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110801")       ###  更新 周线记录日期
monthly_20110801 = pro.monthly(trade_date='20110831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20110801_tscode_list = list() 
for ts_code_sh in monthly_20110801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20110801_tscode_list.append(stock_name)
monthly_20110801_addname_dataframe=pd.DataFrame()
monthly_20110801_addname_dataframe['cname'] = monthly_20110801_tscode_list
for table_name in monthly_20110801.columns.values.tolist():
    monthly_20110801_addname_dataframe[table_name] = monthly_20110801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20110801 20110831 返回数据 row 行数 = "+str(monthly_20110801.shape[0]))
monthly_20110801_addname_dataframe.to_excel(monthly_2011_excel_writer,'8',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110901")       ###  更新 周线记录日期
monthly_20110901 = pro.monthly(trade_date='20110930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20110901_tscode_list = list() 
for ts_code_sh in monthly_20110901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20110901_tscode_list.append(stock_name)
monthly_20110901_addname_dataframe=pd.DataFrame()
monthly_20110901_addname_dataframe['cname'] = monthly_20110901_tscode_list
for table_name in monthly_20110901.columns.values.tolist():
    monthly_20110901_addname_dataframe[table_name] = monthly_20110901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20110901 20110930 返回数据 row 行数 = "+str(monthly_20110901.shape[0]))
monthly_20110901_addname_dataframe.to_excel(monthly_2011_excel_writer,'9',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111001")       ###  更新 周线记录日期
monthly_20111001 = pro.monthly(trade_date='20111031', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20111001_tscode_list = list() 
for ts_code_sh in monthly_20111001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20111001_tscode_list.append(stock_name)
monthly_20111001_addname_dataframe=pd.DataFrame()
monthly_20111001_addname_dataframe['cname'] = monthly_20111001_tscode_list
for table_name in monthly_20111001.columns.values.tolist():
    monthly_20111001_addname_dataframe[table_name] = monthly_20111001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20111001 20111031 返回数据 row 行数 = "+str(monthly_20111001.shape[0]))
monthly_20111001_addname_dataframe.to_excel(monthly_2011_excel_writer,'10',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111101")       ###  更新 周线记录日期
monthly_20111101 = pro.monthly(trade_date='20111130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20111101_tscode_list = list() 
for ts_code_sh in monthly_20111101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20111101_tscode_list.append(stock_name)
monthly_20111101_addname_dataframe=pd.DataFrame()
monthly_20111101_addname_dataframe['cname'] = monthly_20111101_tscode_list
for table_name in monthly_20111101.columns.values.tolist():
    monthly_20111101_addname_dataframe[table_name] = monthly_20111101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20111101 20111130 返回数据 row 行数 = "+str(monthly_20111101.shape[0]))
monthly_20111101_addname_dataframe.to_excel(monthly_2011_excel_writer,'11',index=False)
monthly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111201")       ###  更新 周线记录日期
monthly_20111201 = pro.monthly(trade_date='20111230', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20111201_tscode_list = list() 
for ts_code_sh in monthly_20111201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20111201_tscode_list.append(stock_name)
monthly_20111201_addname_dataframe=pd.DataFrame()
monthly_20111201_addname_dataframe['cname'] = monthly_20111201_tscode_list
for table_name in monthly_20111201.columns.values.tolist():
    monthly_20111201_addname_dataframe[table_name] = monthly_20111201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20111201 20111230 返回数据 row 行数 = "+str(monthly_20111201.shape[0]))
monthly_20111201_addname_dataframe.to_excel(monthly_2011_excel_writer,'12',index=False)
monthly_2011_excel_writer.save()
createexcel('monthly_2012.xlsx')
monthly_2012_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2012.xlsx')
monthly_2012_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2012.xlsx', engine='openpyxl')
monthly_2012_excel_writer.book = monthly_2012_book
monthly_2012_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2012_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20120101")       ###  更新 周线记录日期
monthly_20120101 = pro.monthly(trade_date='20120131', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20120101_tscode_list = list() 
for ts_code_sh in monthly_20120101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20120101_tscode_list.append(stock_name)
monthly_20120101_addname_dataframe=pd.DataFrame()
monthly_20120101_addname_dataframe['cname'] = monthly_20120101_tscode_list
for table_name in monthly_20120101.columns.values.tolist():
    monthly_20120101_addname_dataframe[table_name] = monthly_20120101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20120101 20120131 返回数据 row 行数 = "+str(monthly_20120101.shape[0]))
monthly_20120101_addname_dataframe.to_excel(monthly_2012_excel_writer,'1',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120201")       ###  更新 周线记录日期
monthly_20120201 = pro.monthly(trade_date='20120229', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20120201_tscode_list = list() 
for ts_code_sh in monthly_20120201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20120201_tscode_list.append(stock_name)
monthly_20120201_addname_dataframe=pd.DataFrame()
monthly_20120201_addname_dataframe['cname'] = monthly_20120201_tscode_list
for table_name in monthly_20120201.columns.values.tolist():
    monthly_20120201_addname_dataframe[table_name] = monthly_20120201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20120201 20120229 返回数据 row 行数 = "+str(monthly_20120201.shape[0]))
monthly_20120201_addname_dataframe.to_excel(monthly_2012_excel_writer,'2',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120301")       ###  更新 周线记录日期
monthly_20120301 = pro.monthly(trade_date='20120330', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20120301_tscode_list = list() 
for ts_code_sh in monthly_20120301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20120301_tscode_list.append(stock_name)
monthly_20120301_addname_dataframe=pd.DataFrame()
monthly_20120301_addname_dataframe['cname'] = monthly_20120301_tscode_list
for table_name in monthly_20120301.columns.values.tolist():
    monthly_20120301_addname_dataframe[table_name] = monthly_20120301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20120301 20120330 返回数据 row 行数 = "+str(monthly_20120301.shape[0]))
monthly_20120301_addname_dataframe.to_excel(monthly_2012_excel_writer,'3',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120401")       ###  更新 周线记录日期
monthly_20120401 = pro.monthly(trade_date='20120427', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20120401_tscode_list = list() 
for ts_code_sh in monthly_20120401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20120401_tscode_list.append(stock_name)
monthly_20120401_addname_dataframe=pd.DataFrame()
monthly_20120401_addname_dataframe['cname'] = monthly_20120401_tscode_list
for table_name in monthly_20120401.columns.values.tolist():
    monthly_20120401_addname_dataframe[table_name] = monthly_20120401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20120401 20120427 返回数据 row 行数 = "+str(monthly_20120401.shape[0]))
monthly_20120401_addname_dataframe.to_excel(monthly_2012_excel_writer,'4',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120501")       ###  更新 周线记录日期
monthly_20120501 = pro.monthly(trade_date='20120531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20120501_tscode_list = list() 
for ts_code_sh in monthly_20120501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20120501_tscode_list.append(stock_name)
monthly_20120501_addname_dataframe=pd.DataFrame()
monthly_20120501_addname_dataframe['cname'] = monthly_20120501_tscode_list
for table_name in monthly_20120501.columns.values.tolist():
    monthly_20120501_addname_dataframe[table_name] = monthly_20120501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20120501 20120531 返回数据 row 行数 = "+str(monthly_20120501.shape[0]))
monthly_20120501_addname_dataframe.to_excel(monthly_2012_excel_writer,'5',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120601")       ###  更新 周线记录日期
monthly_20120601 = pro.monthly(trade_date='20120629', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20120601_tscode_list = list() 
for ts_code_sh in monthly_20120601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20120601_tscode_list.append(stock_name)
monthly_20120601_addname_dataframe=pd.DataFrame()
monthly_20120601_addname_dataframe['cname'] = monthly_20120601_tscode_list
for table_name in monthly_20120601.columns.values.tolist():
    monthly_20120601_addname_dataframe[table_name] = monthly_20120601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20120601 20120629 返回数据 row 行数 = "+str(monthly_20120601.shape[0]))
monthly_20120601_addname_dataframe.to_excel(monthly_2012_excel_writer,'6',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120701")       ###  更新 周线记录日期
monthly_20120701 = pro.monthly(trade_date='20120731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20120701_tscode_list = list() 
for ts_code_sh in monthly_20120701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20120701_tscode_list.append(stock_name)
monthly_20120701_addname_dataframe=pd.DataFrame()
monthly_20120701_addname_dataframe['cname'] = monthly_20120701_tscode_list
for table_name in monthly_20120701.columns.values.tolist():
    monthly_20120701_addname_dataframe[table_name] = monthly_20120701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20120701 20120731 返回数据 row 行数 = "+str(monthly_20120701.shape[0]))
monthly_20120701_addname_dataframe.to_excel(monthly_2012_excel_writer,'7',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120801")       ###  更新 周线记录日期
monthly_20120801 = pro.monthly(trade_date='20120831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20120801_tscode_list = list() 
for ts_code_sh in monthly_20120801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20120801_tscode_list.append(stock_name)
monthly_20120801_addname_dataframe=pd.DataFrame()
monthly_20120801_addname_dataframe['cname'] = monthly_20120801_tscode_list
for table_name in monthly_20120801.columns.values.tolist():
    monthly_20120801_addname_dataframe[table_name] = monthly_20120801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20120801 20120831 返回数据 row 行数 = "+str(monthly_20120801.shape[0]))
monthly_20120801_addname_dataframe.to_excel(monthly_2012_excel_writer,'8',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120901")       ###  更新 周线记录日期
monthly_20120901 = pro.monthly(trade_date='20120928', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20120901_tscode_list = list() 
for ts_code_sh in monthly_20120901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20120901_tscode_list.append(stock_name)
monthly_20120901_addname_dataframe=pd.DataFrame()
monthly_20120901_addname_dataframe['cname'] = monthly_20120901_tscode_list
for table_name in monthly_20120901.columns.values.tolist():
    monthly_20120901_addname_dataframe[table_name] = monthly_20120901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20120901 20120928 返回数据 row 行数 = "+str(monthly_20120901.shape[0]))
monthly_20120901_addname_dataframe.to_excel(monthly_2012_excel_writer,'9',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121001")       ###  更新 周线记录日期
monthly_20121001 = pro.monthly(trade_date='20121031', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20121001_tscode_list = list() 
for ts_code_sh in monthly_20121001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20121001_tscode_list.append(stock_name)
monthly_20121001_addname_dataframe=pd.DataFrame()
monthly_20121001_addname_dataframe['cname'] = monthly_20121001_tscode_list
for table_name in monthly_20121001.columns.values.tolist():
    monthly_20121001_addname_dataframe[table_name] = monthly_20121001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20121001 20121031 返回数据 row 行数 = "+str(monthly_20121001.shape[0]))
monthly_20121001_addname_dataframe.to_excel(monthly_2012_excel_writer,'10',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121101")       ###  更新 周线记录日期
monthly_20121101 = pro.monthly(trade_date='20121130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20121101_tscode_list = list() 
for ts_code_sh in monthly_20121101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20121101_tscode_list.append(stock_name)
monthly_20121101_addname_dataframe=pd.DataFrame()
monthly_20121101_addname_dataframe['cname'] = monthly_20121101_tscode_list
for table_name in monthly_20121101.columns.values.tolist():
    monthly_20121101_addname_dataframe[table_name] = monthly_20121101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20121101 20121130 返回数据 row 行数 = "+str(monthly_20121101.shape[0]))
monthly_20121101_addname_dataframe.to_excel(monthly_2012_excel_writer,'11',index=False)
monthly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121201")       ###  更新 周线记录日期
monthly_20121201 = pro.monthly(trade_date='20121231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20121201_tscode_list = list() 
for ts_code_sh in monthly_20121201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20121201_tscode_list.append(stock_name)
monthly_20121201_addname_dataframe=pd.DataFrame()
monthly_20121201_addname_dataframe['cname'] = monthly_20121201_tscode_list
for table_name in monthly_20121201.columns.values.tolist():
    monthly_20121201_addname_dataframe[table_name] = monthly_20121201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20121201 20121231 返回数据 row 行数 = "+str(monthly_20121201.shape[0]))
monthly_20121201_addname_dataframe.to_excel(monthly_2012_excel_writer,'12',index=False)
monthly_2012_excel_writer.save()
createexcel('monthly_2013.xlsx')
monthly_2013_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2013.xlsx')
monthly_2013_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2013.xlsx', engine='openpyxl')
monthly_2013_excel_writer.book = monthly_2013_book
monthly_2013_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2013_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20130101")       ###  更新 周线记录日期
monthly_20130101 = pro.monthly(trade_date='20130131', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20130101_tscode_list = list() 
for ts_code_sh in monthly_20130101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20130101_tscode_list.append(stock_name)
monthly_20130101_addname_dataframe=pd.DataFrame()
monthly_20130101_addname_dataframe['cname'] = monthly_20130101_tscode_list
for table_name in monthly_20130101.columns.values.tolist():
    monthly_20130101_addname_dataframe[table_name] = monthly_20130101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20130101 20130131 返回数据 row 行数 = "+str(monthly_20130101.shape[0]))
monthly_20130101_addname_dataframe.to_excel(monthly_2013_excel_writer,'1',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130201")       ###  更新 周线记录日期
monthly_20130201 = pro.monthly(trade_date='20130228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20130201_tscode_list = list() 
for ts_code_sh in monthly_20130201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20130201_tscode_list.append(stock_name)
monthly_20130201_addname_dataframe=pd.DataFrame()
monthly_20130201_addname_dataframe['cname'] = monthly_20130201_tscode_list
for table_name in monthly_20130201.columns.values.tolist():
    monthly_20130201_addname_dataframe[table_name] = monthly_20130201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20130201 20130228 返回数据 row 行数 = "+str(monthly_20130201.shape[0]))
monthly_20130201_addname_dataframe.to_excel(monthly_2013_excel_writer,'2',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130301")       ###  更新 周线记录日期
monthly_20130301 = pro.monthly(trade_date='20130329', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20130301_tscode_list = list() 
for ts_code_sh in monthly_20130301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20130301_tscode_list.append(stock_name)
monthly_20130301_addname_dataframe=pd.DataFrame()
monthly_20130301_addname_dataframe['cname'] = monthly_20130301_tscode_list
for table_name in monthly_20130301.columns.values.tolist():
    monthly_20130301_addname_dataframe[table_name] = monthly_20130301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20130301 20130329 返回数据 row 行数 = "+str(monthly_20130301.shape[0]))
monthly_20130301_addname_dataframe.to_excel(monthly_2013_excel_writer,'3',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130401")       ###  更新 周线记录日期
monthly_20130401 = pro.monthly(trade_date='20130426', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20130401_tscode_list = list() 
for ts_code_sh in monthly_20130401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20130401_tscode_list.append(stock_name)
monthly_20130401_addname_dataframe=pd.DataFrame()
monthly_20130401_addname_dataframe['cname'] = monthly_20130401_tscode_list
for table_name in monthly_20130401.columns.values.tolist():
    monthly_20130401_addname_dataframe[table_name] = monthly_20130401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20130401 20130426 返回数据 row 行数 = "+str(monthly_20130401.shape[0]))
monthly_20130401_addname_dataframe.to_excel(monthly_2013_excel_writer,'4',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130501")       ###  更新 周线记录日期
monthly_20130501 = pro.monthly(trade_date='20130531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20130501_tscode_list = list() 
for ts_code_sh in monthly_20130501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20130501_tscode_list.append(stock_name)
monthly_20130501_addname_dataframe=pd.DataFrame()
monthly_20130501_addname_dataframe['cname'] = monthly_20130501_tscode_list
for table_name in monthly_20130501.columns.values.tolist():
    monthly_20130501_addname_dataframe[table_name] = monthly_20130501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20130501 20130531 返回数据 row 行数 = "+str(monthly_20130501.shape[0]))
monthly_20130501_addname_dataframe.to_excel(monthly_2013_excel_writer,'5',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130601")       ###  更新 周线记录日期
monthly_20130601 = pro.monthly(trade_date='20130628', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20130601_tscode_list = list() 
for ts_code_sh in monthly_20130601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20130601_tscode_list.append(stock_name)
monthly_20130601_addname_dataframe=pd.DataFrame()
monthly_20130601_addname_dataframe['cname'] = monthly_20130601_tscode_list
for table_name in monthly_20130601.columns.values.tolist():
    monthly_20130601_addname_dataframe[table_name] = monthly_20130601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20130601 20130628 返回数据 row 行数 = "+str(monthly_20130601.shape[0]))
monthly_20130601_addname_dataframe.to_excel(monthly_2013_excel_writer,'6',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130701")       ###  更新 周线记录日期
monthly_20130701 = pro.monthly(trade_date='20130731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20130701_tscode_list = list() 
for ts_code_sh in monthly_20130701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20130701_tscode_list.append(stock_name)
monthly_20130701_addname_dataframe=pd.DataFrame()
monthly_20130701_addname_dataframe['cname'] = monthly_20130701_tscode_list
for table_name in monthly_20130701.columns.values.tolist():
    monthly_20130701_addname_dataframe[table_name] = monthly_20130701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20130701 20130731 返回数据 row 行数 = "+str(monthly_20130701.shape[0]))
monthly_20130701_addname_dataframe.to_excel(monthly_2013_excel_writer,'7',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130801")       ###  更新 周线记录日期
monthly_20130801 = pro.monthly(trade_date='20130830', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20130801_tscode_list = list() 
for ts_code_sh in monthly_20130801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20130801_tscode_list.append(stock_name)
monthly_20130801_addname_dataframe=pd.DataFrame()
monthly_20130801_addname_dataframe['cname'] = monthly_20130801_tscode_list
for table_name in monthly_20130801.columns.values.tolist():
    monthly_20130801_addname_dataframe[table_name] = monthly_20130801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20130801 20130830 返回数据 row 行数 = "+str(monthly_20130801.shape[0]))
monthly_20130801_addname_dataframe.to_excel(monthly_2013_excel_writer,'8',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130901")       ###  更新 周线记录日期
monthly_20130901 = pro.monthly(trade_date='20130930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20130901_tscode_list = list() 
for ts_code_sh in monthly_20130901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20130901_tscode_list.append(stock_name)
monthly_20130901_addname_dataframe=pd.DataFrame()
monthly_20130901_addname_dataframe['cname'] = monthly_20130901_tscode_list
for table_name in monthly_20130901.columns.values.tolist():
    monthly_20130901_addname_dataframe[table_name] = monthly_20130901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20130901 20130930 返回数据 row 行数 = "+str(monthly_20130901.shape[0]))
monthly_20130901_addname_dataframe.to_excel(monthly_2013_excel_writer,'9',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131001")       ###  更新 周线记录日期
monthly_20131001 = pro.monthly(trade_date='20131031', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20131001_tscode_list = list() 
for ts_code_sh in monthly_20131001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20131001_tscode_list.append(stock_name)
monthly_20131001_addname_dataframe=pd.DataFrame()
monthly_20131001_addname_dataframe['cname'] = monthly_20131001_tscode_list
for table_name in monthly_20131001.columns.values.tolist():
    monthly_20131001_addname_dataframe[table_name] = monthly_20131001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20131001 20131031 返回数据 row 行数 = "+str(monthly_20131001.shape[0]))
monthly_20131001_addname_dataframe.to_excel(monthly_2013_excel_writer,'10',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131101")       ###  更新 周线记录日期
monthly_20131101 = pro.monthly(trade_date='20131129', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20131101_tscode_list = list() 
for ts_code_sh in monthly_20131101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20131101_tscode_list.append(stock_name)
monthly_20131101_addname_dataframe=pd.DataFrame()
monthly_20131101_addname_dataframe['cname'] = monthly_20131101_tscode_list
for table_name in monthly_20131101.columns.values.tolist():
    monthly_20131101_addname_dataframe[table_name] = monthly_20131101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20131101 20131129 返回数据 row 行数 = "+str(monthly_20131101.shape[0]))
monthly_20131101_addname_dataframe.to_excel(monthly_2013_excel_writer,'11',index=False)
monthly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131201")       ###  更新 周线记录日期
monthly_20131201 = pro.monthly(trade_date='20131231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20131201_tscode_list = list() 
for ts_code_sh in monthly_20131201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20131201_tscode_list.append(stock_name)
monthly_20131201_addname_dataframe=pd.DataFrame()
monthly_20131201_addname_dataframe['cname'] = monthly_20131201_tscode_list
for table_name in monthly_20131201.columns.values.tolist():
    monthly_20131201_addname_dataframe[table_name] = monthly_20131201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20131201 20131231 返回数据 row 行数 = "+str(monthly_20131201.shape[0]))
monthly_20131201_addname_dataframe.to_excel(monthly_2013_excel_writer,'12',index=False)
monthly_2013_excel_writer.save()
createexcel('monthly_2014.xlsx')
monthly_2014_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2014.xlsx')
monthly_2014_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2014.xlsx', engine='openpyxl')
monthly_2014_excel_writer.book = monthly_2014_book
monthly_2014_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2014_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20140101")       ###  更新 周线记录日期
monthly_20140101 = pro.monthly(trade_date='20140130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20140101_tscode_list = list() 
for ts_code_sh in monthly_20140101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20140101_tscode_list.append(stock_name)
monthly_20140101_addname_dataframe=pd.DataFrame()
monthly_20140101_addname_dataframe['cname'] = monthly_20140101_tscode_list
for table_name in monthly_20140101.columns.values.tolist():
    monthly_20140101_addname_dataframe[table_name] = monthly_20140101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20140101 20140130 返回数据 row 行数 = "+str(monthly_20140101.shape[0]))
monthly_20140101_addname_dataframe.to_excel(monthly_2014_excel_writer,'1',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140201")       ###  更新 周线记录日期
monthly_20140201 = pro.monthly(trade_date='20140228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20140201_tscode_list = list() 
for ts_code_sh in monthly_20140201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20140201_tscode_list.append(stock_name)
monthly_20140201_addname_dataframe=pd.DataFrame()
monthly_20140201_addname_dataframe['cname'] = monthly_20140201_tscode_list
for table_name in monthly_20140201.columns.values.tolist():
    monthly_20140201_addname_dataframe[table_name] = monthly_20140201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20140201 20140228 返回数据 row 行数 = "+str(monthly_20140201.shape[0]))
monthly_20140201_addname_dataframe.to_excel(monthly_2014_excel_writer,'2',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140301")       ###  更新 周线记录日期
monthly_20140301 = pro.monthly(trade_date='20140331', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20140301_tscode_list = list() 
for ts_code_sh in monthly_20140301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20140301_tscode_list.append(stock_name)
monthly_20140301_addname_dataframe=pd.DataFrame()
monthly_20140301_addname_dataframe['cname'] = monthly_20140301_tscode_list
for table_name in monthly_20140301.columns.values.tolist():
    monthly_20140301_addname_dataframe[table_name] = monthly_20140301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20140301 20140331 返回数据 row 行数 = "+str(monthly_20140301.shape[0]))
monthly_20140301_addname_dataframe.to_excel(monthly_2014_excel_writer,'3',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140401")       ###  更新 周线记录日期
monthly_20140401 = pro.monthly(trade_date='20140430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20140401_tscode_list = list() 
for ts_code_sh in monthly_20140401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20140401_tscode_list.append(stock_name)
monthly_20140401_addname_dataframe=pd.DataFrame()
monthly_20140401_addname_dataframe['cname'] = monthly_20140401_tscode_list
for table_name in monthly_20140401.columns.values.tolist():
    monthly_20140401_addname_dataframe[table_name] = monthly_20140401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20140401 20140430 返回数据 row 行数 = "+str(monthly_20140401.shape[0]))
monthly_20140401_addname_dataframe.to_excel(monthly_2014_excel_writer,'4',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140501")       ###  更新 周线记录日期
monthly_20140501 = pro.monthly(trade_date='20140530', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20140501_tscode_list = list() 
for ts_code_sh in monthly_20140501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20140501_tscode_list.append(stock_name)
monthly_20140501_addname_dataframe=pd.DataFrame()
monthly_20140501_addname_dataframe['cname'] = monthly_20140501_tscode_list
for table_name in monthly_20140501.columns.values.tolist():
    monthly_20140501_addname_dataframe[table_name] = monthly_20140501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20140501 20140530 返回数据 row 行数 = "+str(monthly_20140501.shape[0]))
monthly_20140501_addname_dataframe.to_excel(monthly_2014_excel_writer,'5',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140601")       ###  更新 周线记录日期
monthly_20140601 = pro.monthly(trade_date='20140630', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20140601_tscode_list = list() 
for ts_code_sh in monthly_20140601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20140601_tscode_list.append(stock_name)
monthly_20140601_addname_dataframe=pd.DataFrame()
monthly_20140601_addname_dataframe['cname'] = monthly_20140601_tscode_list
for table_name in monthly_20140601.columns.values.tolist():
    monthly_20140601_addname_dataframe[table_name] = monthly_20140601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20140601 20140630 返回数据 row 行数 = "+str(monthly_20140601.shape[0]))
monthly_20140601_addname_dataframe.to_excel(monthly_2014_excel_writer,'6',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140701")       ###  更新 周线记录日期
monthly_20140701 = pro.monthly(trade_date='20140731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20140701_tscode_list = list() 
for ts_code_sh in monthly_20140701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20140701_tscode_list.append(stock_name)
monthly_20140701_addname_dataframe=pd.DataFrame()
monthly_20140701_addname_dataframe['cname'] = monthly_20140701_tscode_list
for table_name in monthly_20140701.columns.values.tolist():
    monthly_20140701_addname_dataframe[table_name] = monthly_20140701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20140701 20140731 返回数据 row 行数 = "+str(monthly_20140701.shape[0]))
monthly_20140701_addname_dataframe.to_excel(monthly_2014_excel_writer,'7',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140801")       ###  更新 周线记录日期
monthly_20140801 = pro.monthly(trade_date='20140829', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20140801_tscode_list = list() 
for ts_code_sh in monthly_20140801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20140801_tscode_list.append(stock_name)
monthly_20140801_addname_dataframe=pd.DataFrame()
monthly_20140801_addname_dataframe['cname'] = monthly_20140801_tscode_list
for table_name in monthly_20140801.columns.values.tolist():
    monthly_20140801_addname_dataframe[table_name] = monthly_20140801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20140801 20140829 返回数据 row 行数 = "+str(monthly_20140801.shape[0]))
monthly_20140801_addname_dataframe.to_excel(monthly_2014_excel_writer,'8',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140901")       ###  更新 周线记录日期
monthly_20140901 = pro.monthly(trade_date='20140930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20140901_tscode_list = list() 
for ts_code_sh in monthly_20140901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20140901_tscode_list.append(stock_name)
monthly_20140901_addname_dataframe=pd.DataFrame()
monthly_20140901_addname_dataframe['cname'] = monthly_20140901_tscode_list
for table_name in monthly_20140901.columns.values.tolist():
    monthly_20140901_addname_dataframe[table_name] = monthly_20140901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20140901 20140930 返回数据 row 行数 = "+str(monthly_20140901.shape[0]))
monthly_20140901_addname_dataframe.to_excel(monthly_2014_excel_writer,'9',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141001")       ###  更新 周线记录日期
monthly_20141001 = pro.monthly(trade_date='20141031', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20141001_tscode_list = list() 
for ts_code_sh in monthly_20141001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20141001_tscode_list.append(stock_name)
monthly_20141001_addname_dataframe=pd.DataFrame()
monthly_20141001_addname_dataframe['cname'] = monthly_20141001_tscode_list
for table_name in monthly_20141001.columns.values.tolist():
    monthly_20141001_addname_dataframe[table_name] = monthly_20141001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20141001 20141031 返回数据 row 行数 = "+str(monthly_20141001.shape[0]))
monthly_20141001_addname_dataframe.to_excel(monthly_2014_excel_writer,'10',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141101")       ###  更新 周线记录日期
monthly_20141101 = pro.monthly(trade_date='20141128', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20141101_tscode_list = list() 
for ts_code_sh in monthly_20141101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20141101_tscode_list.append(stock_name)
monthly_20141101_addname_dataframe=pd.DataFrame()
monthly_20141101_addname_dataframe['cname'] = monthly_20141101_tscode_list
for table_name in monthly_20141101.columns.values.tolist():
    monthly_20141101_addname_dataframe[table_name] = monthly_20141101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20141101 20141128 返回数据 row 行数 = "+str(monthly_20141101.shape[0]))
monthly_20141101_addname_dataframe.to_excel(monthly_2014_excel_writer,'11',index=False)
monthly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141201")       ###  更新 周线记录日期
monthly_20141201 = pro.monthly(trade_date='20141231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20141201_tscode_list = list() 
for ts_code_sh in monthly_20141201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20141201_tscode_list.append(stock_name)
monthly_20141201_addname_dataframe=pd.DataFrame()
monthly_20141201_addname_dataframe['cname'] = monthly_20141201_tscode_list
for table_name in monthly_20141201.columns.values.tolist():
    monthly_20141201_addname_dataframe[table_name] = monthly_20141201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20141201 20141231 返回数据 row 行数 = "+str(monthly_20141201.shape[0]))
monthly_20141201_addname_dataframe.to_excel(monthly_2014_excel_writer,'12',index=False)
monthly_2014_excel_writer.save()
createexcel('monthly_2015.xlsx')
monthly_2015_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2015.xlsx')
monthly_2015_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2015.xlsx', engine='openpyxl')
monthly_2015_excel_writer.book = monthly_2015_book
monthly_2015_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2015_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20150101")       ###  更新 周线记录日期
monthly_20150101 = pro.monthly(trade_date='20150130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20150101_tscode_list = list() 
for ts_code_sh in monthly_20150101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20150101_tscode_list.append(stock_name)
monthly_20150101_addname_dataframe=pd.DataFrame()
monthly_20150101_addname_dataframe['cname'] = monthly_20150101_tscode_list
for table_name in monthly_20150101.columns.values.tolist():
    monthly_20150101_addname_dataframe[table_name] = monthly_20150101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20150101 20150130 返回数据 row 行数 = "+str(monthly_20150101.shape[0]))
monthly_20150101_addname_dataframe.to_excel(monthly_2015_excel_writer,'1',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150201")       ###  更新 周线记录日期
monthly_20150201 = pro.monthly(trade_date='20150227', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20150201_tscode_list = list() 
for ts_code_sh in monthly_20150201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20150201_tscode_list.append(stock_name)
monthly_20150201_addname_dataframe=pd.DataFrame()
monthly_20150201_addname_dataframe['cname'] = monthly_20150201_tscode_list
for table_name in monthly_20150201.columns.values.tolist():
    monthly_20150201_addname_dataframe[table_name] = monthly_20150201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20150201 20150227 返回数据 row 行数 = "+str(monthly_20150201.shape[0]))
monthly_20150201_addname_dataframe.to_excel(monthly_2015_excel_writer,'2',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150301")       ###  更新 周线记录日期
monthly_20150301 = pro.monthly(trade_date='20150331', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20150301_tscode_list = list() 
for ts_code_sh in monthly_20150301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20150301_tscode_list.append(stock_name)
monthly_20150301_addname_dataframe=pd.DataFrame()
monthly_20150301_addname_dataframe['cname'] = monthly_20150301_tscode_list
for table_name in monthly_20150301.columns.values.tolist():
    monthly_20150301_addname_dataframe[table_name] = monthly_20150301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20150301 20150331 返回数据 row 行数 = "+str(monthly_20150301.shape[0]))
monthly_20150301_addname_dataframe.to_excel(monthly_2015_excel_writer,'3',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150401")       ###  更新 周线记录日期
monthly_20150401 = pro.monthly(trade_date='20150430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20150401_tscode_list = list() 
for ts_code_sh in monthly_20150401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20150401_tscode_list.append(stock_name)
monthly_20150401_addname_dataframe=pd.DataFrame()
monthly_20150401_addname_dataframe['cname'] = monthly_20150401_tscode_list
for table_name in monthly_20150401.columns.values.tolist():
    monthly_20150401_addname_dataframe[table_name] = monthly_20150401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20150401 20150430 返回数据 row 行数 = "+str(monthly_20150401.shape[0]))
monthly_20150401_addname_dataframe.to_excel(monthly_2015_excel_writer,'4',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150501")       ###  更新 周线记录日期
monthly_20150501 = pro.monthly(trade_date='20150529', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20150501_tscode_list = list() 
for ts_code_sh in monthly_20150501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20150501_tscode_list.append(stock_name)
monthly_20150501_addname_dataframe=pd.DataFrame()
monthly_20150501_addname_dataframe['cname'] = monthly_20150501_tscode_list
for table_name in monthly_20150501.columns.values.tolist():
    monthly_20150501_addname_dataframe[table_name] = monthly_20150501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20150501 20150529 返回数据 row 行数 = "+str(monthly_20150501.shape[0]))
monthly_20150501_addname_dataframe.to_excel(monthly_2015_excel_writer,'5',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150601")       ###  更新 周线记录日期
monthly_20150601 = pro.monthly(trade_date='20150630', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20150601_tscode_list = list() 
for ts_code_sh in monthly_20150601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20150601_tscode_list.append(stock_name)
monthly_20150601_addname_dataframe=pd.DataFrame()
monthly_20150601_addname_dataframe['cname'] = monthly_20150601_tscode_list
for table_name in monthly_20150601.columns.values.tolist():
    monthly_20150601_addname_dataframe[table_name] = monthly_20150601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20150601 20150630 返回数据 row 行数 = "+str(monthly_20150601.shape[0]))
monthly_20150601_addname_dataframe.to_excel(monthly_2015_excel_writer,'6',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150701")       ###  更新 周线记录日期
monthly_20150701 = pro.monthly(trade_date='20150731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20150701_tscode_list = list() 
for ts_code_sh in monthly_20150701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20150701_tscode_list.append(stock_name)
monthly_20150701_addname_dataframe=pd.DataFrame()
monthly_20150701_addname_dataframe['cname'] = monthly_20150701_tscode_list
for table_name in monthly_20150701.columns.values.tolist():
    monthly_20150701_addname_dataframe[table_name] = monthly_20150701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20150701 20150731 返回数据 row 行数 = "+str(monthly_20150701.shape[0]))
monthly_20150701_addname_dataframe.to_excel(monthly_2015_excel_writer,'7',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150801")       ###  更新 周线记录日期
monthly_20150801 = pro.monthly(trade_date='20150831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20150801_tscode_list = list() 
for ts_code_sh in monthly_20150801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20150801_tscode_list.append(stock_name)
monthly_20150801_addname_dataframe=pd.DataFrame()
monthly_20150801_addname_dataframe['cname'] = monthly_20150801_tscode_list
for table_name in monthly_20150801.columns.values.tolist():
    monthly_20150801_addname_dataframe[table_name] = monthly_20150801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20150801 20150831 返回数据 row 行数 = "+str(monthly_20150801.shape[0]))
monthly_20150801_addname_dataframe.to_excel(monthly_2015_excel_writer,'8',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150901")       ###  更新 周线记录日期
monthly_20150901 = pro.monthly(trade_date='20150930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20150901_tscode_list = list() 
for ts_code_sh in monthly_20150901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20150901_tscode_list.append(stock_name)
monthly_20150901_addname_dataframe=pd.DataFrame()
monthly_20150901_addname_dataframe['cname'] = monthly_20150901_tscode_list
for table_name in monthly_20150901.columns.values.tolist():
    monthly_20150901_addname_dataframe[table_name] = monthly_20150901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20150901 20150930 返回数据 row 行数 = "+str(monthly_20150901.shape[0]))
monthly_20150901_addname_dataframe.to_excel(monthly_2015_excel_writer,'9',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151001")       ###  更新 周线记录日期
monthly_20151001 = pro.monthly(trade_date='20151030', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20151001_tscode_list = list() 
for ts_code_sh in monthly_20151001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20151001_tscode_list.append(stock_name)
monthly_20151001_addname_dataframe=pd.DataFrame()
monthly_20151001_addname_dataframe['cname'] = monthly_20151001_tscode_list
for table_name in monthly_20151001.columns.values.tolist():
    monthly_20151001_addname_dataframe[table_name] = monthly_20151001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20151001 20151030 返回数据 row 行数 = "+str(monthly_20151001.shape[0]))
monthly_20151001_addname_dataframe.to_excel(monthly_2015_excel_writer,'10',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151101")       ###  更新 周线记录日期
monthly_20151101 = pro.monthly(trade_date='20151130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20151101_tscode_list = list() 
for ts_code_sh in monthly_20151101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20151101_tscode_list.append(stock_name)
monthly_20151101_addname_dataframe=pd.DataFrame()
monthly_20151101_addname_dataframe['cname'] = monthly_20151101_tscode_list
for table_name in monthly_20151101.columns.values.tolist():
    monthly_20151101_addname_dataframe[table_name] = monthly_20151101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20151101 20151130 返回数据 row 行数 = "+str(monthly_20151101.shape[0]))
monthly_20151101_addname_dataframe.to_excel(monthly_2015_excel_writer,'11',index=False)
monthly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151201")       ###  更新 周线记录日期
monthly_20151201 = pro.monthly(trade_date='20151231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20151201_tscode_list = list() 
for ts_code_sh in monthly_20151201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20151201_tscode_list.append(stock_name)
monthly_20151201_addname_dataframe=pd.DataFrame()
monthly_20151201_addname_dataframe['cname'] = monthly_20151201_tscode_list
for table_name in monthly_20151201.columns.values.tolist():
    monthly_20151201_addname_dataframe[table_name] = monthly_20151201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20151201 20151231 返回数据 row 行数 = "+str(monthly_20151201.shape[0]))
monthly_20151201_addname_dataframe.to_excel(monthly_2015_excel_writer,'12',index=False)
monthly_2015_excel_writer.save()
createexcel('monthly_2016.xlsx')
monthly_2016_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2016.xlsx')
monthly_2016_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2016.xlsx', engine='openpyxl')
monthly_2016_excel_writer.book = monthly_2016_book
monthly_2016_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2016_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20160101")       ###  更新 周线记录日期
monthly_20160101 = pro.monthly(trade_date='20160129', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20160101_tscode_list = list() 
for ts_code_sh in monthly_20160101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20160101_tscode_list.append(stock_name)
monthly_20160101_addname_dataframe=pd.DataFrame()
monthly_20160101_addname_dataframe['cname'] = monthly_20160101_tscode_list
for table_name in monthly_20160101.columns.values.tolist():
    monthly_20160101_addname_dataframe[table_name] = monthly_20160101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20160101 20160129 返回数据 row 行数 = "+str(monthly_20160101.shape[0]))
monthly_20160101_addname_dataframe.to_excel(monthly_2016_excel_writer,'1',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160201")       ###  更新 周线记录日期
monthly_20160201 = pro.monthly(trade_date='20160229', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20160201_tscode_list = list() 
for ts_code_sh in monthly_20160201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20160201_tscode_list.append(stock_name)
monthly_20160201_addname_dataframe=pd.DataFrame()
monthly_20160201_addname_dataframe['cname'] = monthly_20160201_tscode_list
for table_name in monthly_20160201.columns.values.tolist():
    monthly_20160201_addname_dataframe[table_name] = monthly_20160201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20160201 20160229 返回数据 row 行数 = "+str(monthly_20160201.shape[0]))
monthly_20160201_addname_dataframe.to_excel(monthly_2016_excel_writer,'2',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160301")       ###  更新 周线记录日期
monthly_20160301 = pro.monthly(trade_date='20160331', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20160301_tscode_list = list() 
for ts_code_sh in monthly_20160301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20160301_tscode_list.append(stock_name)
monthly_20160301_addname_dataframe=pd.DataFrame()
monthly_20160301_addname_dataframe['cname'] = monthly_20160301_tscode_list
for table_name in monthly_20160301.columns.values.tolist():
    monthly_20160301_addname_dataframe[table_name] = monthly_20160301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20160301 20160331 返回数据 row 行数 = "+str(monthly_20160301.shape[0]))
monthly_20160301_addname_dataframe.to_excel(monthly_2016_excel_writer,'3',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160401")       ###  更新 周线记录日期
monthly_20160401 = pro.monthly(trade_date='20160429', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20160401_tscode_list = list() 
for ts_code_sh in monthly_20160401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20160401_tscode_list.append(stock_name)
monthly_20160401_addname_dataframe=pd.DataFrame()
monthly_20160401_addname_dataframe['cname'] = monthly_20160401_tscode_list
for table_name in monthly_20160401.columns.values.tolist():
    monthly_20160401_addname_dataframe[table_name] = monthly_20160401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20160401 20160429 返回数据 row 行数 = "+str(monthly_20160401.shape[0]))
monthly_20160401_addname_dataframe.to_excel(monthly_2016_excel_writer,'4',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160501")       ###  更新 周线记录日期
monthly_20160501 = pro.monthly(trade_date='20160531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20160501_tscode_list = list() 
for ts_code_sh in monthly_20160501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20160501_tscode_list.append(stock_name)
monthly_20160501_addname_dataframe=pd.DataFrame()
monthly_20160501_addname_dataframe['cname'] = monthly_20160501_tscode_list
for table_name in monthly_20160501.columns.values.tolist():
    monthly_20160501_addname_dataframe[table_name] = monthly_20160501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20160501 20160531 返回数据 row 行数 = "+str(monthly_20160501.shape[0]))
monthly_20160501_addname_dataframe.to_excel(monthly_2016_excel_writer,'5',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160601")       ###  更新 周线记录日期
monthly_20160601 = pro.monthly(trade_date='20160630', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20160601_tscode_list = list() 
for ts_code_sh in monthly_20160601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20160601_tscode_list.append(stock_name)
monthly_20160601_addname_dataframe=pd.DataFrame()
monthly_20160601_addname_dataframe['cname'] = monthly_20160601_tscode_list
for table_name in monthly_20160601.columns.values.tolist():
    monthly_20160601_addname_dataframe[table_name] = monthly_20160601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20160601 20160630 返回数据 row 行数 = "+str(monthly_20160601.shape[0]))
monthly_20160601_addname_dataframe.to_excel(monthly_2016_excel_writer,'6',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160701")       ###  更新 周线记录日期
monthly_20160701 = pro.monthly(trade_date='20160729', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20160701_tscode_list = list() 
for ts_code_sh in monthly_20160701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20160701_tscode_list.append(stock_name)
monthly_20160701_addname_dataframe=pd.DataFrame()
monthly_20160701_addname_dataframe['cname'] = monthly_20160701_tscode_list
for table_name in monthly_20160701.columns.values.tolist():
    monthly_20160701_addname_dataframe[table_name] = monthly_20160701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20160701 20160729 返回数据 row 行数 = "+str(monthly_20160701.shape[0]))
monthly_20160701_addname_dataframe.to_excel(monthly_2016_excel_writer,'7',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160801")       ###  更新 周线记录日期
monthly_20160801 = pro.monthly(trade_date='20160831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20160801_tscode_list = list() 
for ts_code_sh in monthly_20160801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20160801_tscode_list.append(stock_name)
monthly_20160801_addname_dataframe=pd.DataFrame()
monthly_20160801_addname_dataframe['cname'] = monthly_20160801_tscode_list
for table_name in monthly_20160801.columns.values.tolist():
    monthly_20160801_addname_dataframe[table_name] = monthly_20160801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20160801 20160831 返回数据 row 行数 = "+str(monthly_20160801.shape[0]))
monthly_20160801_addname_dataframe.to_excel(monthly_2016_excel_writer,'8',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160901")       ###  更新 周线记录日期
monthly_20160901 = pro.monthly(trade_date='20160930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20160901_tscode_list = list() 
for ts_code_sh in monthly_20160901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20160901_tscode_list.append(stock_name)
monthly_20160901_addname_dataframe=pd.DataFrame()
monthly_20160901_addname_dataframe['cname'] = monthly_20160901_tscode_list
for table_name in monthly_20160901.columns.values.tolist():
    monthly_20160901_addname_dataframe[table_name] = monthly_20160901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20160901 20160930 返回数据 row 行数 = "+str(monthly_20160901.shape[0]))
monthly_20160901_addname_dataframe.to_excel(monthly_2016_excel_writer,'9',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161001")       ###  更新 周线记录日期
monthly_20161001 = pro.monthly(trade_date='20161031', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20161001_tscode_list = list() 
for ts_code_sh in monthly_20161001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20161001_tscode_list.append(stock_name)
monthly_20161001_addname_dataframe=pd.DataFrame()
monthly_20161001_addname_dataframe['cname'] = monthly_20161001_tscode_list
for table_name in monthly_20161001.columns.values.tolist():
    monthly_20161001_addname_dataframe[table_name] = monthly_20161001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20161001 20161031 返回数据 row 行数 = "+str(monthly_20161001.shape[0]))
monthly_20161001_addname_dataframe.to_excel(monthly_2016_excel_writer,'10',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161101")       ###  更新 周线记录日期
monthly_20161101 = pro.monthly(trade_date='20161130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20161101_tscode_list = list() 
for ts_code_sh in monthly_20161101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20161101_tscode_list.append(stock_name)
monthly_20161101_addname_dataframe=pd.DataFrame()
monthly_20161101_addname_dataframe['cname'] = monthly_20161101_tscode_list
for table_name in monthly_20161101.columns.values.tolist():
    monthly_20161101_addname_dataframe[table_name] = monthly_20161101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20161101 20161130 返回数据 row 行数 = "+str(monthly_20161101.shape[0]))
monthly_20161101_addname_dataframe.to_excel(monthly_2016_excel_writer,'11',index=False)
monthly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161201")       ###  更新 周线记录日期
monthly_20161201 = pro.monthly(trade_date='20161230', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20161201_tscode_list = list() 
for ts_code_sh in monthly_20161201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20161201_tscode_list.append(stock_name)
monthly_20161201_addname_dataframe=pd.DataFrame()
monthly_20161201_addname_dataframe['cname'] = monthly_20161201_tscode_list
for table_name in monthly_20161201.columns.values.tolist():
    monthly_20161201_addname_dataframe[table_name] = monthly_20161201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20161201 20161230 返回数据 row 行数 = "+str(monthly_20161201.shape[0]))
monthly_20161201_addname_dataframe.to_excel(monthly_2016_excel_writer,'12',index=False)
monthly_2016_excel_writer.save()
createexcel('monthly_2017.xlsx')
monthly_2017_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2017.xlsx')
monthly_2017_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2017.xlsx', engine='openpyxl')
monthly_2017_excel_writer.book = monthly_2017_book
monthly_2017_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2017_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20170101")       ###  更新 周线记录日期
monthly_20170101 = pro.monthly(trade_date='20170126', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20170101_tscode_list = list() 
for ts_code_sh in monthly_20170101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20170101_tscode_list.append(stock_name)
monthly_20170101_addname_dataframe=pd.DataFrame()
monthly_20170101_addname_dataframe['cname'] = monthly_20170101_tscode_list
for table_name in monthly_20170101.columns.values.tolist():
    monthly_20170101_addname_dataframe[table_name] = monthly_20170101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20170101 20170126 返回数据 row 行数 = "+str(monthly_20170101.shape[0]))
monthly_20170101_addname_dataframe.to_excel(monthly_2017_excel_writer,'1',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170201")       ###  更新 周线记录日期
monthly_20170201 = pro.monthly(trade_date='20170228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20170201_tscode_list = list() 
for ts_code_sh in monthly_20170201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20170201_tscode_list.append(stock_name)
monthly_20170201_addname_dataframe=pd.DataFrame()
monthly_20170201_addname_dataframe['cname'] = monthly_20170201_tscode_list
for table_name in monthly_20170201.columns.values.tolist():
    monthly_20170201_addname_dataframe[table_name] = monthly_20170201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20170201 20170228 返回数据 row 行数 = "+str(monthly_20170201.shape[0]))
monthly_20170201_addname_dataframe.to_excel(monthly_2017_excel_writer,'2',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170301")       ###  更新 周线记录日期
monthly_20170301 = pro.monthly(trade_date='20170331', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20170301_tscode_list = list() 
for ts_code_sh in monthly_20170301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20170301_tscode_list.append(stock_name)
monthly_20170301_addname_dataframe=pd.DataFrame()
monthly_20170301_addname_dataframe['cname'] = monthly_20170301_tscode_list
for table_name in monthly_20170301.columns.values.tolist():
    monthly_20170301_addname_dataframe[table_name] = monthly_20170301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20170301 20170331 返回数据 row 行数 = "+str(monthly_20170301.shape[0]))
monthly_20170301_addname_dataframe.to_excel(monthly_2017_excel_writer,'3',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170401")       ###  更新 周线记录日期
monthly_20170401 = pro.monthly(trade_date='20170428', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20170401_tscode_list = list() 
for ts_code_sh in monthly_20170401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20170401_tscode_list.append(stock_name)
monthly_20170401_addname_dataframe=pd.DataFrame()
monthly_20170401_addname_dataframe['cname'] = monthly_20170401_tscode_list
for table_name in monthly_20170401.columns.values.tolist():
    monthly_20170401_addname_dataframe[table_name] = monthly_20170401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20170401 20170428 返回数据 row 行数 = "+str(monthly_20170401.shape[0]))
monthly_20170401_addname_dataframe.to_excel(monthly_2017_excel_writer,'4',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170501")       ###  更新 周线记录日期
monthly_20170501 = pro.monthly(trade_date='20170531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20170501_tscode_list = list() 
for ts_code_sh in monthly_20170501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20170501_tscode_list.append(stock_name)
monthly_20170501_addname_dataframe=pd.DataFrame()
monthly_20170501_addname_dataframe['cname'] = monthly_20170501_tscode_list
for table_name in monthly_20170501.columns.values.tolist():
    monthly_20170501_addname_dataframe[table_name] = monthly_20170501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20170501 20170531 返回数据 row 行数 = "+str(monthly_20170501.shape[0]))
monthly_20170501_addname_dataframe.to_excel(monthly_2017_excel_writer,'5',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170601")       ###  更新 周线记录日期
monthly_20170601 = pro.monthly(trade_date='20170630', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20170601_tscode_list = list() 
for ts_code_sh in monthly_20170601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20170601_tscode_list.append(stock_name)
monthly_20170601_addname_dataframe=pd.DataFrame()
monthly_20170601_addname_dataframe['cname'] = monthly_20170601_tscode_list
for table_name in monthly_20170601.columns.values.tolist():
    monthly_20170601_addname_dataframe[table_name] = monthly_20170601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20170601 20170630 返回数据 row 行数 = "+str(monthly_20170601.shape[0]))
monthly_20170601_addname_dataframe.to_excel(monthly_2017_excel_writer,'6',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170701")       ###  更新 周线记录日期
monthly_20170701 = pro.monthly(trade_date='20170731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20170701_tscode_list = list() 
for ts_code_sh in monthly_20170701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20170701_tscode_list.append(stock_name)
monthly_20170701_addname_dataframe=pd.DataFrame()
monthly_20170701_addname_dataframe['cname'] = monthly_20170701_tscode_list
for table_name in monthly_20170701.columns.values.tolist():
    monthly_20170701_addname_dataframe[table_name] = monthly_20170701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20170701 20170731 返回数据 row 行数 = "+str(monthly_20170701.shape[0]))
monthly_20170701_addname_dataframe.to_excel(monthly_2017_excel_writer,'7',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170801")       ###  更新 周线记录日期
monthly_20170801 = pro.monthly(trade_date='20170831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20170801_tscode_list = list() 
for ts_code_sh in monthly_20170801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20170801_tscode_list.append(stock_name)
monthly_20170801_addname_dataframe=pd.DataFrame()
monthly_20170801_addname_dataframe['cname'] = monthly_20170801_tscode_list
for table_name in monthly_20170801.columns.values.tolist():
    monthly_20170801_addname_dataframe[table_name] = monthly_20170801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20170801 20170831 返回数据 row 行数 = "+str(monthly_20170801.shape[0]))
monthly_20170801_addname_dataframe.to_excel(monthly_2017_excel_writer,'8',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170901")       ###  更新 周线记录日期
monthly_20170901 = pro.monthly(trade_date='20170929', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20170901_tscode_list = list() 
for ts_code_sh in monthly_20170901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20170901_tscode_list.append(stock_name)
monthly_20170901_addname_dataframe=pd.DataFrame()
monthly_20170901_addname_dataframe['cname'] = monthly_20170901_tscode_list
for table_name in monthly_20170901.columns.values.tolist():
    monthly_20170901_addname_dataframe[table_name] = monthly_20170901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20170901 20170929 返回数据 row 行数 = "+str(monthly_20170901.shape[0]))
monthly_20170901_addname_dataframe.to_excel(monthly_2017_excel_writer,'9',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171001")       ###  更新 周线记录日期
monthly_20171001 = pro.monthly(trade_date='20171031', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20171001_tscode_list = list() 
for ts_code_sh in monthly_20171001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20171001_tscode_list.append(stock_name)
monthly_20171001_addname_dataframe=pd.DataFrame()
monthly_20171001_addname_dataframe['cname'] = monthly_20171001_tscode_list
for table_name in monthly_20171001.columns.values.tolist():
    monthly_20171001_addname_dataframe[table_name] = monthly_20171001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20171001 20171031 返回数据 row 行数 = "+str(monthly_20171001.shape[0]))
monthly_20171001_addname_dataframe.to_excel(monthly_2017_excel_writer,'10',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171101")       ###  更新 周线记录日期
monthly_20171101 = pro.monthly(trade_date='20171130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20171101_tscode_list = list() 
for ts_code_sh in monthly_20171101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20171101_tscode_list.append(stock_name)
monthly_20171101_addname_dataframe=pd.DataFrame()
monthly_20171101_addname_dataframe['cname'] = monthly_20171101_tscode_list
for table_name in monthly_20171101.columns.values.tolist():
    monthly_20171101_addname_dataframe[table_name] = monthly_20171101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20171101 20171130 返回数据 row 行数 = "+str(monthly_20171101.shape[0]))
monthly_20171101_addname_dataframe.to_excel(monthly_2017_excel_writer,'11',index=False)
monthly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171201")       ###  更新 周线记录日期
monthly_20171201 = pro.monthly(trade_date='20171229', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20171201_tscode_list = list() 
for ts_code_sh in monthly_20171201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20171201_tscode_list.append(stock_name)
monthly_20171201_addname_dataframe=pd.DataFrame()
monthly_20171201_addname_dataframe['cname'] = monthly_20171201_tscode_list
for table_name in monthly_20171201.columns.values.tolist():
    monthly_20171201_addname_dataframe[table_name] = monthly_20171201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20171201 20171229 返回数据 row 行数 = "+str(monthly_20171201.shape[0]))
monthly_20171201_addname_dataframe.to_excel(monthly_2017_excel_writer,'12',index=False)
monthly_2017_excel_writer.save()
createexcel('monthly_2018.xlsx')
monthly_2018_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2018.xlsx')
monthly_2018_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2018.xlsx', engine='openpyxl')
monthly_2018_excel_writer.book = monthly_2018_book
monthly_2018_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2018_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20180101")       ###  更新 周线记录日期
monthly_20180101 = pro.monthly(trade_date='20180131', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20180101_tscode_list = list() 
for ts_code_sh in monthly_20180101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20180101_tscode_list.append(stock_name)
monthly_20180101_addname_dataframe=pd.DataFrame()
monthly_20180101_addname_dataframe['cname'] = monthly_20180101_tscode_list
for table_name in monthly_20180101.columns.values.tolist():
    monthly_20180101_addname_dataframe[table_name] = monthly_20180101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20180101 20180131 返回数据 row 行数 = "+str(monthly_20180101.shape[0]))
monthly_20180101_addname_dataframe.to_excel(monthly_2018_excel_writer,'1',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180201")       ###  更新 周线记录日期
monthly_20180201 = pro.monthly(trade_date='20180228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20180201_tscode_list = list() 
for ts_code_sh in monthly_20180201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20180201_tscode_list.append(stock_name)
monthly_20180201_addname_dataframe=pd.DataFrame()
monthly_20180201_addname_dataframe['cname'] = monthly_20180201_tscode_list
for table_name in monthly_20180201.columns.values.tolist():
    monthly_20180201_addname_dataframe[table_name] = monthly_20180201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20180201 20180228 返回数据 row 行数 = "+str(monthly_20180201.shape[0]))
monthly_20180201_addname_dataframe.to_excel(monthly_2018_excel_writer,'2',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180301")       ###  更新 周线记录日期
monthly_20180301 = pro.monthly(trade_date='20180330', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20180301_tscode_list = list() 
for ts_code_sh in monthly_20180301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20180301_tscode_list.append(stock_name)
monthly_20180301_addname_dataframe=pd.DataFrame()
monthly_20180301_addname_dataframe['cname'] = monthly_20180301_tscode_list
for table_name in monthly_20180301.columns.values.tolist():
    monthly_20180301_addname_dataframe[table_name] = monthly_20180301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20180301 20180330 返回数据 row 行数 = "+str(monthly_20180301.shape[0]))
monthly_20180301_addname_dataframe.to_excel(monthly_2018_excel_writer,'3',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180401")       ###  更新 周线记录日期
monthly_20180401 = pro.monthly(trade_date='20180427', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20180401_tscode_list = list() 
for ts_code_sh in monthly_20180401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20180401_tscode_list.append(stock_name)
monthly_20180401_addname_dataframe=pd.DataFrame()
monthly_20180401_addname_dataframe['cname'] = monthly_20180401_tscode_list
for table_name in monthly_20180401.columns.values.tolist():
    monthly_20180401_addname_dataframe[table_name] = monthly_20180401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20180401 20180427 返回数据 row 行数 = "+str(monthly_20180401.shape[0]))
monthly_20180401_addname_dataframe.to_excel(monthly_2018_excel_writer,'4',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180501")       ###  更新 周线记录日期
monthly_20180501 = pro.monthly(trade_date='20180531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20180501_tscode_list = list() 
for ts_code_sh in monthly_20180501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20180501_tscode_list.append(stock_name)
monthly_20180501_addname_dataframe=pd.DataFrame()
monthly_20180501_addname_dataframe['cname'] = monthly_20180501_tscode_list
for table_name in monthly_20180501.columns.values.tolist():
    monthly_20180501_addname_dataframe[table_name] = monthly_20180501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20180501 20180531 返回数据 row 行数 = "+str(monthly_20180501.shape[0]))
monthly_20180501_addname_dataframe.to_excel(monthly_2018_excel_writer,'5',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180601")       ###  更新 周线记录日期
monthly_20180601 = pro.monthly(trade_date='20180629', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20180601_tscode_list = list() 
for ts_code_sh in monthly_20180601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20180601_tscode_list.append(stock_name)
monthly_20180601_addname_dataframe=pd.DataFrame()
monthly_20180601_addname_dataframe['cname'] = monthly_20180601_tscode_list
for table_name in monthly_20180601.columns.values.tolist():
    monthly_20180601_addname_dataframe[table_name] = monthly_20180601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20180601 20180629 返回数据 row 行数 = "+str(monthly_20180601.shape[0]))
monthly_20180601_addname_dataframe.to_excel(monthly_2018_excel_writer,'6',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180701")       ###  更新 周线记录日期
monthly_20180701 = pro.monthly(trade_date='20180731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20180701_tscode_list = list() 
for ts_code_sh in monthly_20180701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20180701_tscode_list.append(stock_name)
monthly_20180701_addname_dataframe=pd.DataFrame()
monthly_20180701_addname_dataframe['cname'] = monthly_20180701_tscode_list
for table_name in monthly_20180701.columns.values.tolist():
    monthly_20180701_addname_dataframe[table_name] = monthly_20180701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20180701 20180731 返回数据 row 行数 = "+str(monthly_20180701.shape[0]))
monthly_20180701_addname_dataframe.to_excel(monthly_2018_excel_writer,'7',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180801")       ###  更新 周线记录日期
monthly_20180801 = pro.monthly(trade_date='20180831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20180801_tscode_list = list() 
for ts_code_sh in monthly_20180801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20180801_tscode_list.append(stock_name)
monthly_20180801_addname_dataframe=pd.DataFrame()
monthly_20180801_addname_dataframe['cname'] = monthly_20180801_tscode_list
for table_name in monthly_20180801.columns.values.tolist():
    monthly_20180801_addname_dataframe[table_name] = monthly_20180801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20180801 20180831 返回数据 row 行数 = "+str(monthly_20180801.shape[0]))
monthly_20180801_addname_dataframe.to_excel(monthly_2018_excel_writer,'8',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180901")       ###  更新 周线记录日期
monthly_20180901 = pro.monthly(trade_date='20180928', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20180901_tscode_list = list() 
for ts_code_sh in monthly_20180901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20180901_tscode_list.append(stock_name)
monthly_20180901_addname_dataframe=pd.DataFrame()
monthly_20180901_addname_dataframe['cname'] = monthly_20180901_tscode_list
for table_name in monthly_20180901.columns.values.tolist():
    monthly_20180901_addname_dataframe[table_name] = monthly_20180901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20180901 20180928 返回数据 row 行数 = "+str(monthly_20180901.shape[0]))
monthly_20180901_addname_dataframe.to_excel(monthly_2018_excel_writer,'9',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181001")       ###  更新 周线记录日期
monthly_20181001 = pro.monthly(trade_date='20181031', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20181001_tscode_list = list() 
for ts_code_sh in monthly_20181001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20181001_tscode_list.append(stock_name)
monthly_20181001_addname_dataframe=pd.DataFrame()
monthly_20181001_addname_dataframe['cname'] = monthly_20181001_tscode_list
for table_name in monthly_20181001.columns.values.tolist():
    monthly_20181001_addname_dataframe[table_name] = monthly_20181001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20181001 20181031 返回数据 row 行数 = "+str(monthly_20181001.shape[0]))
monthly_20181001_addname_dataframe.to_excel(monthly_2018_excel_writer,'10',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181101")       ###  更新 周线记录日期
monthly_20181101 = pro.monthly(trade_date='20181130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20181101_tscode_list = list() 
for ts_code_sh in monthly_20181101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20181101_tscode_list.append(stock_name)
monthly_20181101_addname_dataframe=pd.DataFrame()
monthly_20181101_addname_dataframe['cname'] = monthly_20181101_tscode_list
for table_name in monthly_20181101.columns.values.tolist():
    monthly_20181101_addname_dataframe[table_name] = monthly_20181101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20181101 20181130 返回数据 row 行数 = "+str(monthly_20181101.shape[0]))
monthly_20181101_addname_dataframe.to_excel(monthly_2018_excel_writer,'11',index=False)
monthly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181201")       ###  更新 周线记录日期
monthly_20181201 = pro.monthly(trade_date='20181228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20181201_tscode_list = list() 
for ts_code_sh in monthly_20181201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20181201_tscode_list.append(stock_name)
monthly_20181201_addname_dataframe=pd.DataFrame()
monthly_20181201_addname_dataframe['cname'] = monthly_20181201_tscode_list
for table_name in monthly_20181201.columns.values.tolist():
    monthly_20181201_addname_dataframe[table_name] = monthly_20181201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20181201 20181228 返回数据 row 行数 = "+str(monthly_20181201.shape[0]))
monthly_20181201_addname_dataframe.to_excel(monthly_2018_excel_writer,'12',index=False)
monthly_2018_excel_writer.save()
createexcel('monthly_2019.xlsx')
monthly_2019_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2019.xlsx')
monthly_2019_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2019.xlsx', engine='openpyxl')
monthly_2019_excel_writer.book = monthly_2019_book
monthly_2019_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2019_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20190101")       ###  更新 周线记录日期
monthly_20190101 = pro.monthly(trade_date='20190131', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20190101_tscode_list = list() 
for ts_code_sh in monthly_20190101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20190101_tscode_list.append(stock_name)
monthly_20190101_addname_dataframe=pd.DataFrame()
monthly_20190101_addname_dataframe['cname'] = monthly_20190101_tscode_list
for table_name in monthly_20190101.columns.values.tolist():
    monthly_20190101_addname_dataframe[table_name] = monthly_20190101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20190101 20190131 返回数据 row 行数 = "+str(monthly_20190101.shape[0]))
monthly_20190101_addname_dataframe.to_excel(monthly_2019_excel_writer,'1',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190201")       ###  更新 周线记录日期
monthly_20190201 = pro.monthly(trade_date='20190228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20190201_tscode_list = list() 
for ts_code_sh in monthly_20190201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20190201_tscode_list.append(stock_name)
monthly_20190201_addname_dataframe=pd.DataFrame()
monthly_20190201_addname_dataframe['cname'] = monthly_20190201_tscode_list
for table_name in monthly_20190201.columns.values.tolist():
    monthly_20190201_addname_dataframe[table_name] = monthly_20190201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20190201 20190228 返回数据 row 行数 = "+str(monthly_20190201.shape[0]))
monthly_20190201_addname_dataframe.to_excel(monthly_2019_excel_writer,'2',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190301")       ###  更新 周线记录日期
monthly_20190301 = pro.monthly(trade_date='20190329', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20190301_tscode_list = list() 
for ts_code_sh in monthly_20190301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20190301_tscode_list.append(stock_name)
monthly_20190301_addname_dataframe=pd.DataFrame()
monthly_20190301_addname_dataframe['cname'] = monthly_20190301_tscode_list
for table_name in monthly_20190301.columns.values.tolist():
    monthly_20190301_addname_dataframe[table_name] = monthly_20190301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20190301 20190329 返回数据 row 行数 = "+str(monthly_20190301.shape[0]))
monthly_20190301_addname_dataframe.to_excel(monthly_2019_excel_writer,'3',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190401")       ###  更新 周线记录日期
monthly_20190401 = pro.monthly(trade_date='20190430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20190401_tscode_list = list() 
for ts_code_sh in monthly_20190401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20190401_tscode_list.append(stock_name)
monthly_20190401_addname_dataframe=pd.DataFrame()
monthly_20190401_addname_dataframe['cname'] = monthly_20190401_tscode_list
for table_name in monthly_20190401.columns.values.tolist():
    monthly_20190401_addname_dataframe[table_name] = monthly_20190401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20190401 20190430 返回数据 row 行数 = "+str(monthly_20190401.shape[0]))
monthly_20190401_addname_dataframe.to_excel(monthly_2019_excel_writer,'4',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190501")       ###  更新 周线记录日期
monthly_20190501 = pro.monthly(trade_date='20190531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20190501_tscode_list = list() 
for ts_code_sh in monthly_20190501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20190501_tscode_list.append(stock_name)
monthly_20190501_addname_dataframe=pd.DataFrame()
monthly_20190501_addname_dataframe['cname'] = monthly_20190501_tscode_list
for table_name in monthly_20190501.columns.values.tolist():
    monthly_20190501_addname_dataframe[table_name] = monthly_20190501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20190501 20190531 返回数据 row 行数 = "+str(monthly_20190501.shape[0]))
monthly_20190501_addname_dataframe.to_excel(monthly_2019_excel_writer,'5',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190601")       ###  更新 周线记录日期
monthly_20190601 = pro.monthly(trade_date='20190628', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20190601_tscode_list = list() 
for ts_code_sh in monthly_20190601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20190601_tscode_list.append(stock_name)
monthly_20190601_addname_dataframe=pd.DataFrame()
monthly_20190601_addname_dataframe['cname'] = monthly_20190601_tscode_list
for table_name in monthly_20190601.columns.values.tolist():
    monthly_20190601_addname_dataframe[table_name] = monthly_20190601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20190601 20190628 返回数据 row 行数 = "+str(monthly_20190601.shape[0]))
monthly_20190601_addname_dataframe.to_excel(monthly_2019_excel_writer,'6',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190701")       ###  更新 周线记录日期
monthly_20190701 = pro.monthly(trade_date='20190731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20190701_tscode_list = list() 
for ts_code_sh in monthly_20190701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20190701_tscode_list.append(stock_name)
monthly_20190701_addname_dataframe=pd.DataFrame()
monthly_20190701_addname_dataframe['cname'] = monthly_20190701_tscode_list
for table_name in monthly_20190701.columns.values.tolist():
    monthly_20190701_addname_dataframe[table_name] = monthly_20190701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20190701 20190731 返回数据 row 行数 = "+str(monthly_20190701.shape[0]))
monthly_20190701_addname_dataframe.to_excel(monthly_2019_excel_writer,'7',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190801")       ###  更新 周线记录日期
monthly_20190801 = pro.monthly(trade_date='20190830', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20190801_tscode_list = list() 
for ts_code_sh in monthly_20190801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20190801_tscode_list.append(stock_name)
monthly_20190801_addname_dataframe=pd.DataFrame()
monthly_20190801_addname_dataframe['cname'] = monthly_20190801_tscode_list
for table_name in monthly_20190801.columns.values.tolist():
    monthly_20190801_addname_dataframe[table_name] = monthly_20190801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20190801 20190830 返回数据 row 行数 = "+str(monthly_20190801.shape[0]))
monthly_20190801_addname_dataframe.to_excel(monthly_2019_excel_writer,'8',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190901")       ###  更新 周线记录日期
monthly_20190901 = pro.monthly(trade_date='20190930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20190901_tscode_list = list() 
for ts_code_sh in monthly_20190901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20190901_tscode_list.append(stock_name)
monthly_20190901_addname_dataframe=pd.DataFrame()
monthly_20190901_addname_dataframe['cname'] = monthly_20190901_tscode_list
for table_name in monthly_20190901.columns.values.tolist():
    monthly_20190901_addname_dataframe[table_name] = monthly_20190901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20190901 20190930 返回数据 row 行数 = "+str(monthly_20190901.shape[0]))
monthly_20190901_addname_dataframe.to_excel(monthly_2019_excel_writer,'9',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191001")       ###  更新 周线记录日期
monthly_20191001 = pro.monthly(trade_date='20191031', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20191001_tscode_list = list() 
for ts_code_sh in monthly_20191001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20191001_tscode_list.append(stock_name)
monthly_20191001_addname_dataframe=pd.DataFrame()
monthly_20191001_addname_dataframe['cname'] = monthly_20191001_tscode_list
for table_name in monthly_20191001.columns.values.tolist():
    monthly_20191001_addname_dataframe[table_name] = monthly_20191001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20191001 20191031 返回数据 row 行数 = "+str(monthly_20191001.shape[0]))
monthly_20191001_addname_dataframe.to_excel(monthly_2019_excel_writer,'10',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191101")       ###  更新 周线记录日期
monthly_20191101 = pro.monthly(trade_date='20191129', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20191101_tscode_list = list() 
for ts_code_sh in monthly_20191101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20191101_tscode_list.append(stock_name)
monthly_20191101_addname_dataframe=pd.DataFrame()
monthly_20191101_addname_dataframe['cname'] = monthly_20191101_tscode_list
for table_name in monthly_20191101.columns.values.tolist():
    monthly_20191101_addname_dataframe[table_name] = monthly_20191101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20191101 20191129 返回数据 row 行数 = "+str(monthly_20191101.shape[0]))
monthly_20191101_addname_dataframe.to_excel(monthly_2019_excel_writer,'11',index=False)
monthly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191201")       ###  更新 周线记录日期
monthly_20191201 = pro.monthly(trade_date='20191231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20191201_tscode_list = list() 
for ts_code_sh in monthly_20191201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20191201_tscode_list.append(stock_name)
monthly_20191201_addname_dataframe=pd.DataFrame()
monthly_20191201_addname_dataframe['cname'] = monthly_20191201_tscode_list
for table_name in monthly_20191201.columns.values.tolist():
    monthly_20191201_addname_dataframe[table_name] = monthly_20191201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20191201 20191231 返回数据 row 行数 = "+str(monthly_20191201.shape[0]))
monthly_20191201_addname_dataframe.to_excel(monthly_2019_excel_writer,'12',index=False)
monthly_2019_excel_writer.save()
createexcel('monthly_2020.xlsx')
monthly_2020_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2020.xlsx')
monthly_2020_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\monthly_2020.xlsx', engine='openpyxl')
monthly_2020_excel_writer.book = monthly_2020_book
monthly_2020_excel_writer.sheets = dict((ws.title, ws) for ws in monthly_2020_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20200101")       ###  更新 周线记录日期
monthly_20200101 = pro.monthly(trade_date='20200123', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20200101_tscode_list = list() 
for ts_code_sh in monthly_20200101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20200101_tscode_list.append(stock_name)
monthly_20200101_addname_dataframe=pd.DataFrame()
monthly_20200101_addname_dataframe['cname'] = monthly_20200101_tscode_list
for table_name in monthly_20200101.columns.values.tolist():
    monthly_20200101_addname_dataframe[table_name] = monthly_20200101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20200101 20200123 返回数据 row 行数 = "+str(monthly_20200101.shape[0]))
monthly_20200101_addname_dataframe.to_excel(monthly_2020_excel_writer,'1',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200201")       ###  更新 周线记录日期
monthly_20200201 = pro.monthly(trade_date='20200228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20200201_tscode_list = list() 
for ts_code_sh in monthly_20200201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20200201_tscode_list.append(stock_name)
monthly_20200201_addname_dataframe=pd.DataFrame()
monthly_20200201_addname_dataframe['cname'] = monthly_20200201_tscode_list
for table_name in monthly_20200201.columns.values.tolist():
    monthly_20200201_addname_dataframe[table_name] = monthly_20200201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20200201 20200228 返回数据 row 行数 = "+str(monthly_20200201.shape[0]))
monthly_20200201_addname_dataframe.to_excel(monthly_2020_excel_writer,'2',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200301")       ###  更新 周线记录日期
monthly_20200301 = pro.monthly(trade_date='20200331', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20200301_tscode_list = list() 
for ts_code_sh in monthly_20200301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20200301_tscode_list.append(stock_name)
monthly_20200301_addname_dataframe=pd.DataFrame()
monthly_20200301_addname_dataframe['cname'] = monthly_20200301_tscode_list
for table_name in monthly_20200301.columns.values.tolist():
    monthly_20200301_addname_dataframe[table_name] = monthly_20200301[table_name]
print("月线行情-每个交易月的最后一日  monthly_20200301 20200331 返回数据 row 行数 = "+str(monthly_20200301.shape[0]))
monthly_20200301_addname_dataframe.to_excel(monthly_2020_excel_writer,'3',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200401")       ###  更新 周线记录日期
monthly_20200401 = pro.monthly(trade_date='20200430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20200401_tscode_list = list() 
for ts_code_sh in monthly_20200401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20200401_tscode_list.append(stock_name)
monthly_20200401_addname_dataframe=pd.DataFrame()
monthly_20200401_addname_dataframe['cname'] = monthly_20200401_tscode_list
for table_name in monthly_20200401.columns.values.tolist():
    monthly_20200401_addname_dataframe[table_name] = monthly_20200401[table_name]
print("月线行情-每个交易月的最后一日  monthly_20200401 20200430 返回数据 row 行数 = "+str(monthly_20200401.shape[0]))
monthly_20200401_addname_dataframe.to_excel(monthly_2020_excel_writer,'4',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200501")       ###  更新 周线记录日期
monthly_20200501 = pro.monthly(trade_date='20200529', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20200501_tscode_list = list() 
for ts_code_sh in monthly_20200501['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20200501_tscode_list.append(stock_name)
monthly_20200501_addname_dataframe=pd.DataFrame()
monthly_20200501_addname_dataframe['cname'] = monthly_20200501_tscode_list
for table_name in monthly_20200501.columns.values.tolist():
    monthly_20200501_addname_dataframe[table_name] = monthly_20200501[table_name]
print("月线行情-每个交易月的最后一日  monthly_20200501 20200529 返回数据 row 行数 = "+str(monthly_20200501.shape[0]))
monthly_20200501_addname_dataframe.to_excel(monthly_2020_excel_writer,'5',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200601")       ###  更新 周线记录日期
monthly_20200601 = pro.monthly(trade_date='20200630', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20200601_tscode_list = list() 
for ts_code_sh in monthly_20200601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20200601_tscode_list.append(stock_name)
monthly_20200601_addname_dataframe=pd.DataFrame()
monthly_20200601_addname_dataframe['cname'] = monthly_20200601_tscode_list
for table_name in monthly_20200601.columns.values.tolist():
    monthly_20200601_addname_dataframe[table_name] = monthly_20200601[table_name]
print("月线行情-每个交易月的最后一日  monthly_20200601 20200630 返回数据 row 行数 = "+str(monthly_20200601.shape[0]))
monthly_20200601_addname_dataframe.to_excel(monthly_2020_excel_writer,'6',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200701")       ###  更新 周线记录日期
monthly_20200701 = pro.monthly(trade_date='20200731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20200701_tscode_list = list() 
for ts_code_sh in monthly_20200701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20200701_tscode_list.append(stock_name)
monthly_20200701_addname_dataframe=pd.DataFrame()
monthly_20200701_addname_dataframe['cname'] = monthly_20200701_tscode_list
for table_name in monthly_20200701.columns.values.tolist():
    monthly_20200701_addname_dataframe[table_name] = monthly_20200701[table_name]
print("月线行情-每个交易月的最后一日  monthly_20200701 20200731 返回数据 row 行数 = "+str(monthly_20200701.shape[0]))
monthly_20200701_addname_dataframe.to_excel(monthly_2020_excel_writer,'7',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200801")       ###  更新 周线记录日期
monthly_20200801 = pro.monthly(trade_date='20200831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20200801_tscode_list = list() 
for ts_code_sh in monthly_20200801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20200801_tscode_list.append(stock_name)
monthly_20200801_addname_dataframe=pd.DataFrame()
monthly_20200801_addname_dataframe['cname'] = monthly_20200801_tscode_list
for table_name in monthly_20200801.columns.values.tolist():
    monthly_20200801_addname_dataframe[table_name] = monthly_20200801[table_name]
print("月线行情-每个交易月的最后一日  monthly_20200801 20200831 返回数据 row 行数 = "+str(monthly_20200801.shape[0]))
monthly_20200801_addname_dataframe.to_excel(monthly_2020_excel_writer,'8',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200901")       ###  更新 周线记录日期
monthly_20200901 = pro.monthly(trade_date='20200930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20200901_tscode_list = list() 
for ts_code_sh in monthly_20200901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20200901_tscode_list.append(stock_name)
monthly_20200901_addname_dataframe=pd.DataFrame()
monthly_20200901_addname_dataframe['cname'] = monthly_20200901_tscode_list
for table_name in monthly_20200901.columns.values.tolist():
    monthly_20200901_addname_dataframe[table_name] = monthly_20200901[table_name]
print("月线行情-每个交易月的最后一日  monthly_20200901 20200930 返回数据 row 行数 = "+str(monthly_20200901.shape[0]))
monthly_20200901_addname_dataframe.to_excel(monthly_2020_excel_writer,'9',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201001")       ###  更新 周线记录日期
monthly_20201001 = pro.monthly(trade_date='20201030', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20201001_tscode_list = list() 
for ts_code_sh in monthly_20201001['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20201001_tscode_list.append(stock_name)
monthly_20201001_addname_dataframe=pd.DataFrame()
monthly_20201001_addname_dataframe['cname'] = monthly_20201001_tscode_list
for table_name in monthly_20201001.columns.values.tolist():
    monthly_20201001_addname_dataframe[table_name] = monthly_20201001[table_name]
print("月线行情-每个交易月的最后一日  monthly_20201001 20201030 返回数据 row 行数 = "+str(monthly_20201001.shape[0]))
monthly_20201001_addname_dataframe.to_excel(monthly_2020_excel_writer,'10',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201101")       ###  更新 周线记录日期
monthly_20201101 = pro.monthly(trade_date='20201130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20201101_tscode_list = list() 
for ts_code_sh in monthly_20201101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20201101_tscode_list.append(stock_name)
monthly_20201101_addname_dataframe=pd.DataFrame()
monthly_20201101_addname_dataframe['cname'] = monthly_20201101_tscode_list
for table_name in monthly_20201101.columns.values.tolist():
    monthly_20201101_addname_dataframe[table_name] = monthly_20201101[table_name]
print("月线行情-每个交易月的最后一日  monthly_20201101 20201130 返回数据 row 行数 = "+str(monthly_20201101.shape[0]))
monthly_20201101_addname_dataframe.to_excel(monthly_2020_excel_writer,'11',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201201")       ###  更新 周线记录日期
monthly_20201201 = pro.monthly(trade_date='20201231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
monthly_20201201_tscode_list = list() 
for ts_code_sh in monthly_20201201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    monthly_20201201_tscode_list.append(stock_name)
monthly_20201201_addname_dataframe=pd.DataFrame()
monthly_20201201_addname_dataframe['cname'] = monthly_20201201_tscode_list
for table_name in monthly_20201201.columns.values.tolist():
    monthly_20201201_addname_dataframe[table_name] = monthly_20201201[table_name]
print("月线行情-每个交易月的最后一日  monthly_20201201 20201231 返回数据 row 行数 = "+str(monthly_20201201.shape[0]))
monthly_20201201_addname_dataframe.to_excel(monthly_2020_excel_writer,'12',index=False)
monthly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "19901219")       ###  更新 记录日期
