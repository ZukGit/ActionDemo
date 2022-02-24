#  encoding: utf-8

import tushare as ts
import os
import pandas as pd  # https://www.jianshu.com/p/5c0aa1fa19af
import openpyxl 
from openpyxl import load_workbook,Workbook
import time,datetime  # https://www.jb51.net/article/66019.htm
import re
import os
import tempfile

############################## 全局数据初始化块 数据初始化函数  End ##############################

##############  properities 函数 Begin ##############

class Properties:

    def __init__(self, file_name):
        self.file_name = file_name
        self.properties = {}
        try:
            fopen = open(self.file_name, 'r')
            for line in fopen:
                line = line.strip()
                if line.find('=') > 0 and not line.startswith('#'):
                    strs = line.split('=')
                    self.properties[strs[0].strip()] = strs[1].strip()
        except Exception :
            raise
        else:
            fopen.close()

    def has_key(self, key):
        return key in self.properties

    def get(self, key, default_value=''):
        if key in self.properties:
            return self.properties[key]
        return default_value

    def put(self, key, value):
        self.properties[key] = value
        replace_property(self.file_name, key + '=.*', key + '=' + value, True)


def replace_property(file_name, from_regex, to_str, append_on_not_exists=True):
    tmpfile = tempfile.TemporaryFile()

    if os.path.exists(file_name):
        r_open = open(file_name, 'r')
        pattern = re.compile(r'' + from_regex)
        found = None
        for line in r_open:
            if pattern.search(line) and not line.strip().startswith('#'):
                found = True
                line = re.sub(from_regex, to_str, line)
            tmpfile.write(line.encode())
        if not found and append_on_not_exists:
            tmpfile.write(('\n' + to_str).encode())
        r_open.close()
        tmpfile.seek(0)

        content = tmpfile.read()

        if os.path.exists(file_name):
            os.remove(file_name)

        w_open = open(file_name, 'wb')
        w_open.write(content)
        w_open.close()

        tmpfile.close()
    else:
        print ("file %s not found" % file_name)

##############  properities 函数 End ##############


##############  XLSX excel 函数 Begin ##############

def createexcel(filename):    ### 创建本地文件名称为 filename的文件
    wb = Workbook()
    curPath = os.path.abspath('.')
    cur_desktop_path=os.path.join(os.path.expanduser("~"), 'Desktop')
    curDir = cur_desktop_path+os.sep+"zbin"+os.sep+"J0_Data"+os.sep  ## 创建~/Desktop/zbin/J0_Data 这个 这个工作目录
    #curDir = curPath+os.sep
    if not os.path.exists(curDir):
        os.makedirs(curDir)
    #curPath = os.path.dirname(os.path.abspath('.'))
    #t = time.strftime('%Y%m%d%H%M%S', time.localtime(time.time()))
    #suffix = '.xlsx' # 文件类型
    #newfile = t + suffix
    xlsxPath = curDir + filename
    print("curPath curDir=" + curPath)
    print("createexcel zbin_JO_Dir=" + curDir)
    print("createexcel path=" + xlsxPath)
    if not os.path.exists(xlsxPath):
        wb.save(xlsxPath)
        print(" wb.save(xlsxPath)  xlsxPath =" + xlsxPath)
        time.sleep(1)
    return xlsxPath

def getColumnIndex(table, columnName):   ## 返回 table 中 名称为  columnName 的 那列 的索引
    columnIndex = None
    for i in range(table.ncols):
        if(table.cell_value(0, i) == columnName):
            columnIndex = i
            break
    return columnIndex
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007

##############  XLSX excel 函数 End ##############



############################## Prop初始化 Begin ##############################

#
#Thu Aug 13 22:33:45 CST 2020
#rixianhangqing-time_record_date=20200707
#rixianhangqing-time_start_date=20010101

desktop_path = os.path.join(os.path.expanduser("~"), 'Desktop')
zbin_path = str(desktop_path)+os.sep+"zbin"
j0_properties_path= str(zbin_path)+os.sep+"J0.properties"
print("desktop_path = "+ str(desktop_path))
print("zbin_path = "+ str(zbin_path))
print("j0_properties_path = "+ str(j0_properties_path))
J0_PROPS =  Properties(j0_properties_path)


############################## Prop初始化 End ##############################


############## tscode_股票列表的初始化  Begin ##############
#封装函数    https://blog.csdn.net/weixin_41267342/article/details/86634007
def init_tscode_data(book_name, sheet_name,ts_code_set,tscode_name_dict):
    # 读取excel
    wb = openpyxl.load_workbook(book_name)
    # 读取sheet表
    ws = wb[sheet_name]
    for i in range(ws.max_row):
         # 获取下拉框中的所有选择值
         if (i == 0 or i == 1):
             continue
         #print("i="+str(i)+" 总的列数:" + str(ws.max_row)+"  value:"+str(ws.cell(i,2).value))
         tscode_item = str(ws.cell(i,1).value)  ##  20201010--> 2:ts_code    4_name   ##  20201116--> 1:ts_code    3_name 
         tscode_name_item = str(ws.cell(i,3).value)
         #print("index="+str(i)+" tscode_item = "+ str(tscode_item) + "   tscode_name_item="+str(tscode_name_item))
         ts_code_set.add(str(ws.cell(i,1).value))
         tscode_name_dict[tscode_item]=tscode_name_item



tscode_set = set()    #### 股票代码的集合   000001.SZ   .... 999999.SH
tscode_name_dict = dict()  #### code-name 的 map的集合
init_tscode_data(zbin_path+os.sep+"J0_Python"+os.sep+"J0_股票列表.xlsx","股票列表",tscode_set,tscode_name_dict)
############## tscode_股票列表的初始化  End  ##############


############################## 运行属性 Begin ##############################
pd.set_option('display.max_rows', None)   ##  解决纵向出现...
#pd.set_option('display.width', 1000) 
pd.set_option('expand_frame_repr', False)  ##  解决横向出现...
Cur_Abs_Path=os.path.abspath('.')   # 表示当前所处的文件夹的绝对路径
print("当前绝对路径:"+Cur_Abs_Path)
Cur_Ref_Path=os.path.abspath('..')  # 表示当前所处的文件夹上一级文件夹的绝对路径
print("当前父目录绝对路径:"+Cur_Ref_Path)
############################## 运行属性 End ##############################

############################## 时间Date Begin ##############################

now_yyyymmdd=str(time.strftime('%Y%m%d', time.localtime(time.time())))
print("now_yyyymmdd = "+ str(now_yyyymmdd))

############################## 时间Date End ##############################





pro = ts.pro_api('43acb9a5ddc2cf73c6c4ea54796748f965457ed57daaa736bb778ea2')
# print(J0_PROPS.get(tree_node_name+"record_date"))           #根据key读取value
# J0_PROPS.put(tree_node_name+"record_date", now_yyyymmdd)       ###  覆盖原有的 key 和 value
#  zukgit
# zhouxianhangqing-time_zukgit_website  =   https://tushare.pro/document/2?doc_id=144
tree_node_name="zhouxianhangqing-time"+"_"
createexcel('weekly_2009.xlsx')
weekly_2009_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2009.xlsx')
weekly_2009_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2009.xlsx', engine='openpyxl')
weekly_2009_excel_writer.book = weekly_2009_book
weekly_2009_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2009_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20091231")       ###  更新 周线记录日期
weekly_20091231 = pro.weekly(trade_date='20091231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20091231_tscode_list = list() 
for ts_code_sh in weekly_20091231['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20091231_tscode_list.append(stock_name)
weekly_20091231_addname_dataframe=pd.DataFrame()
weekly_20091231_addname_dataframe['cname'] = weekly_20091231_tscode_list
for table_name in weekly_20091231.columns.values.tolist():
    weekly_20091231_addname_dataframe[table_name] = weekly_20091231[table_name]
print("周线行情-时间为序  weekly_20091231 53_20091231 返回数据 row 行数 = "+str(weekly_20091231.shape[0]))
weekly_20091231_addname_dataframe.to_excel(weekly_2009_excel_writer,'53_20091231',index=False)
weekly_2009_excel_writer.save()
createexcel('weekly_2010.xlsx')
weekly_2010_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2010.xlsx')
weekly_2010_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2010.xlsx', engine='openpyxl')
weekly_2010_excel_writer.book = weekly_2010_book
weekly_2010_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2010_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20100108")       ###  更新 周线记录日期
weekly_20100108 = pro.weekly(trade_date='20100108', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100108_tscode_list = list() 
for ts_code_sh in weekly_20100108['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100108_tscode_list.append(stock_name)
weekly_20100108_addname_dataframe=pd.DataFrame()
weekly_20100108_addname_dataframe['cname'] = weekly_20100108_tscode_list
for table_name in weekly_20100108.columns.values.tolist():
    weekly_20100108_addname_dataframe[table_name] = weekly_20100108[table_name]
print("周线行情-时间为序  weekly_20100108 2_20100108 返回数据 row 行数 = "+str(weekly_20100108.shape[0]))
weekly_20100108_addname_dataframe.to_excel(weekly_2010_excel_writer,'2_20100108',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100115")       ###  更新 周线记录日期
weekly_20100115 = pro.weekly(trade_date='20100115', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100115_tscode_list = list() 
for ts_code_sh in weekly_20100115['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100115_tscode_list.append(stock_name)
weekly_20100115_addname_dataframe=pd.DataFrame()
weekly_20100115_addname_dataframe['cname'] = weekly_20100115_tscode_list
for table_name in weekly_20100115.columns.values.tolist():
    weekly_20100115_addname_dataframe[table_name] = weekly_20100115[table_name]
print("周线行情-时间为序  weekly_20100115 3_20100115 返回数据 row 行数 = "+str(weekly_20100115.shape[0]))
weekly_20100115_addname_dataframe.to_excel(weekly_2010_excel_writer,'3_20100115',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100122")       ###  更新 周线记录日期
weekly_20100122 = pro.weekly(trade_date='20100122', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100122_tscode_list = list() 
for ts_code_sh in weekly_20100122['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100122_tscode_list.append(stock_name)
weekly_20100122_addname_dataframe=pd.DataFrame()
weekly_20100122_addname_dataframe['cname'] = weekly_20100122_tscode_list
for table_name in weekly_20100122.columns.values.tolist():
    weekly_20100122_addname_dataframe[table_name] = weekly_20100122[table_name]
print("周线行情-时间为序  weekly_20100122 4_20100122 返回数据 row 行数 = "+str(weekly_20100122.shape[0]))
weekly_20100122_addname_dataframe.to_excel(weekly_2010_excel_writer,'4_20100122',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100129")       ###  更新 周线记录日期
weekly_20100129 = pro.weekly(trade_date='20100129', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100129_tscode_list = list() 
for ts_code_sh in weekly_20100129['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100129_tscode_list.append(stock_name)
weekly_20100129_addname_dataframe=pd.DataFrame()
weekly_20100129_addname_dataframe['cname'] = weekly_20100129_tscode_list
for table_name in weekly_20100129.columns.values.tolist():
    weekly_20100129_addname_dataframe[table_name] = weekly_20100129[table_name]
print("周线行情-时间为序  weekly_20100129 5_20100129 返回数据 row 行数 = "+str(weekly_20100129.shape[0]))
weekly_20100129_addname_dataframe.to_excel(weekly_2010_excel_writer,'5_20100129',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100205")       ###  更新 周线记录日期
weekly_20100205 = pro.weekly(trade_date='20100205', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100205_tscode_list = list() 
for ts_code_sh in weekly_20100205['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100205_tscode_list.append(stock_name)
weekly_20100205_addname_dataframe=pd.DataFrame()
weekly_20100205_addname_dataframe['cname'] = weekly_20100205_tscode_list
for table_name in weekly_20100205.columns.values.tolist():
    weekly_20100205_addname_dataframe[table_name] = weekly_20100205[table_name]
print("周线行情-时间为序  weekly_20100205 6_20100205 返回数据 row 行数 = "+str(weekly_20100205.shape[0]))
weekly_20100205_addname_dataframe.to_excel(weekly_2010_excel_writer,'6_20100205',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100212")       ###  更新 周线记录日期
weekly_20100212 = pro.weekly(trade_date='20100212', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100212_tscode_list = list() 
for ts_code_sh in weekly_20100212['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100212_tscode_list.append(stock_name)
weekly_20100212_addname_dataframe=pd.DataFrame()
weekly_20100212_addname_dataframe['cname'] = weekly_20100212_tscode_list
for table_name in weekly_20100212.columns.values.tolist():
    weekly_20100212_addname_dataframe[table_name] = weekly_20100212[table_name]
print("周线行情-时间为序  weekly_20100212 7_20100212 返回数据 row 行数 = "+str(weekly_20100212.shape[0]))
weekly_20100212_addname_dataframe.to_excel(weekly_2010_excel_writer,'7_20100212',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100226")       ###  更新 周线记录日期
weekly_20100226 = pro.weekly(trade_date='20100226', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100226_tscode_list = list() 
for ts_code_sh in weekly_20100226['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100226_tscode_list.append(stock_name)
weekly_20100226_addname_dataframe=pd.DataFrame()
weekly_20100226_addname_dataframe['cname'] = weekly_20100226_tscode_list
for table_name in weekly_20100226.columns.values.tolist():
    weekly_20100226_addname_dataframe[table_name] = weekly_20100226[table_name]
print("周线行情-时间为序  weekly_20100226 9_20100226 返回数据 row 行数 = "+str(weekly_20100226.shape[0]))
weekly_20100226_addname_dataframe.to_excel(weekly_2010_excel_writer,'9_20100226',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100305")       ###  更新 周线记录日期
weekly_20100305 = pro.weekly(trade_date='20100305', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100305_tscode_list = list() 
for ts_code_sh in weekly_20100305['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100305_tscode_list.append(stock_name)
weekly_20100305_addname_dataframe=pd.DataFrame()
weekly_20100305_addname_dataframe['cname'] = weekly_20100305_tscode_list
for table_name in weekly_20100305.columns.values.tolist():
    weekly_20100305_addname_dataframe[table_name] = weekly_20100305[table_name]
print("周线行情-时间为序  weekly_20100305 10_20100305 返回数据 row 行数 = "+str(weekly_20100305.shape[0]))
weekly_20100305_addname_dataframe.to_excel(weekly_2010_excel_writer,'10_20100305',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100312")       ###  更新 周线记录日期
weekly_20100312 = pro.weekly(trade_date='20100312', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100312_tscode_list = list() 
for ts_code_sh in weekly_20100312['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100312_tscode_list.append(stock_name)
weekly_20100312_addname_dataframe=pd.DataFrame()
weekly_20100312_addname_dataframe['cname'] = weekly_20100312_tscode_list
for table_name in weekly_20100312.columns.values.tolist():
    weekly_20100312_addname_dataframe[table_name] = weekly_20100312[table_name]
print("周线行情-时间为序  weekly_20100312 11_20100312 返回数据 row 行数 = "+str(weekly_20100312.shape[0]))
weekly_20100312_addname_dataframe.to_excel(weekly_2010_excel_writer,'11_20100312',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100319")       ###  更新 周线记录日期
weekly_20100319 = pro.weekly(trade_date='20100319', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100319_tscode_list = list() 
for ts_code_sh in weekly_20100319['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100319_tscode_list.append(stock_name)
weekly_20100319_addname_dataframe=pd.DataFrame()
weekly_20100319_addname_dataframe['cname'] = weekly_20100319_tscode_list
for table_name in weekly_20100319.columns.values.tolist():
    weekly_20100319_addname_dataframe[table_name] = weekly_20100319[table_name]
print("周线行情-时间为序  weekly_20100319 12_20100319 返回数据 row 行数 = "+str(weekly_20100319.shape[0]))
weekly_20100319_addname_dataframe.to_excel(weekly_2010_excel_writer,'12_20100319',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100326")       ###  更新 周线记录日期
weekly_20100326 = pro.weekly(trade_date='20100326', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100326_tscode_list = list() 
for ts_code_sh in weekly_20100326['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100326_tscode_list.append(stock_name)
weekly_20100326_addname_dataframe=pd.DataFrame()
weekly_20100326_addname_dataframe['cname'] = weekly_20100326_tscode_list
for table_name in weekly_20100326.columns.values.tolist():
    weekly_20100326_addname_dataframe[table_name] = weekly_20100326[table_name]
print("周线行情-时间为序  weekly_20100326 13_20100326 返回数据 row 行数 = "+str(weekly_20100326.shape[0]))
weekly_20100326_addname_dataframe.to_excel(weekly_2010_excel_writer,'13_20100326',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100402")       ###  更新 周线记录日期
weekly_20100402 = pro.weekly(trade_date='20100402', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100402_tscode_list = list() 
for ts_code_sh in weekly_20100402['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100402_tscode_list.append(stock_name)
weekly_20100402_addname_dataframe=pd.DataFrame()
weekly_20100402_addname_dataframe['cname'] = weekly_20100402_tscode_list
for table_name in weekly_20100402.columns.values.tolist():
    weekly_20100402_addname_dataframe[table_name] = weekly_20100402[table_name]
print("周线行情-时间为序  weekly_20100402 14_20100402 返回数据 row 行数 = "+str(weekly_20100402.shape[0]))
weekly_20100402_addname_dataframe.to_excel(weekly_2010_excel_writer,'14_20100402',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100409")       ###  更新 周线记录日期
weekly_20100409 = pro.weekly(trade_date='20100409', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100409_tscode_list = list() 
for ts_code_sh in weekly_20100409['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100409_tscode_list.append(stock_name)
weekly_20100409_addname_dataframe=pd.DataFrame()
weekly_20100409_addname_dataframe['cname'] = weekly_20100409_tscode_list
for table_name in weekly_20100409.columns.values.tolist():
    weekly_20100409_addname_dataframe[table_name] = weekly_20100409[table_name]
print("周线行情-时间为序  weekly_20100409 15_20100409 返回数据 row 行数 = "+str(weekly_20100409.shape[0]))
weekly_20100409_addname_dataframe.to_excel(weekly_2010_excel_writer,'15_20100409',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100416")       ###  更新 周线记录日期
weekly_20100416 = pro.weekly(trade_date='20100416', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100416_tscode_list = list() 
for ts_code_sh in weekly_20100416['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100416_tscode_list.append(stock_name)
weekly_20100416_addname_dataframe=pd.DataFrame()
weekly_20100416_addname_dataframe['cname'] = weekly_20100416_tscode_list
for table_name in weekly_20100416.columns.values.tolist():
    weekly_20100416_addname_dataframe[table_name] = weekly_20100416[table_name]
print("周线行情-时间为序  weekly_20100416 16_20100416 返回数据 row 行数 = "+str(weekly_20100416.shape[0]))
weekly_20100416_addname_dataframe.to_excel(weekly_2010_excel_writer,'16_20100416',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100423")       ###  更新 周线记录日期
weekly_20100423 = pro.weekly(trade_date='20100423', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100423_tscode_list = list() 
for ts_code_sh in weekly_20100423['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100423_tscode_list.append(stock_name)
weekly_20100423_addname_dataframe=pd.DataFrame()
weekly_20100423_addname_dataframe['cname'] = weekly_20100423_tscode_list
for table_name in weekly_20100423.columns.values.tolist():
    weekly_20100423_addname_dataframe[table_name] = weekly_20100423[table_name]
print("周线行情-时间为序  weekly_20100423 17_20100423 返回数据 row 行数 = "+str(weekly_20100423.shape[0]))
weekly_20100423_addname_dataframe.to_excel(weekly_2010_excel_writer,'17_20100423',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100430")       ###  更新 周线记录日期
weekly_20100430 = pro.weekly(trade_date='20100430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100430_tscode_list = list() 
for ts_code_sh in weekly_20100430['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100430_tscode_list.append(stock_name)
weekly_20100430_addname_dataframe=pd.DataFrame()
weekly_20100430_addname_dataframe['cname'] = weekly_20100430_tscode_list
for table_name in weekly_20100430.columns.values.tolist():
    weekly_20100430_addname_dataframe[table_name] = weekly_20100430[table_name]
print("周线行情-时间为序  weekly_20100430 18_20100430 返回数据 row 行数 = "+str(weekly_20100430.shape[0]))
weekly_20100430_addname_dataframe.to_excel(weekly_2010_excel_writer,'18_20100430',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100507")       ###  更新 周线记录日期
weekly_20100507 = pro.weekly(trade_date='20100507', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100507_tscode_list = list() 
for ts_code_sh in weekly_20100507['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100507_tscode_list.append(stock_name)
weekly_20100507_addname_dataframe=pd.DataFrame()
weekly_20100507_addname_dataframe['cname'] = weekly_20100507_tscode_list
for table_name in weekly_20100507.columns.values.tolist():
    weekly_20100507_addname_dataframe[table_name] = weekly_20100507[table_name]
print("周线行情-时间为序  weekly_20100507 19_20100507 返回数据 row 行数 = "+str(weekly_20100507.shape[0]))
weekly_20100507_addname_dataframe.to_excel(weekly_2010_excel_writer,'19_20100507',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100514")       ###  更新 周线记录日期
weekly_20100514 = pro.weekly(trade_date='20100514', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100514_tscode_list = list() 
for ts_code_sh in weekly_20100514['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100514_tscode_list.append(stock_name)
weekly_20100514_addname_dataframe=pd.DataFrame()
weekly_20100514_addname_dataframe['cname'] = weekly_20100514_tscode_list
for table_name in weekly_20100514.columns.values.tolist():
    weekly_20100514_addname_dataframe[table_name] = weekly_20100514[table_name]
print("周线行情-时间为序  weekly_20100514 20_20100514 返回数据 row 行数 = "+str(weekly_20100514.shape[0]))
weekly_20100514_addname_dataframe.to_excel(weekly_2010_excel_writer,'20_20100514',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100521")       ###  更新 周线记录日期
weekly_20100521 = pro.weekly(trade_date='20100521', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100521_tscode_list = list() 
for ts_code_sh in weekly_20100521['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100521_tscode_list.append(stock_name)
weekly_20100521_addname_dataframe=pd.DataFrame()
weekly_20100521_addname_dataframe['cname'] = weekly_20100521_tscode_list
for table_name in weekly_20100521.columns.values.tolist():
    weekly_20100521_addname_dataframe[table_name] = weekly_20100521[table_name]
print("周线行情-时间为序  weekly_20100521 21_20100521 返回数据 row 行数 = "+str(weekly_20100521.shape[0]))
weekly_20100521_addname_dataframe.to_excel(weekly_2010_excel_writer,'21_20100521',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100528")       ###  更新 周线记录日期
weekly_20100528 = pro.weekly(trade_date='20100528', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100528_tscode_list = list() 
for ts_code_sh in weekly_20100528['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100528_tscode_list.append(stock_name)
weekly_20100528_addname_dataframe=pd.DataFrame()
weekly_20100528_addname_dataframe['cname'] = weekly_20100528_tscode_list
for table_name in weekly_20100528.columns.values.tolist():
    weekly_20100528_addname_dataframe[table_name] = weekly_20100528[table_name]
print("周线行情-时间为序  weekly_20100528 22_20100528 返回数据 row 行数 = "+str(weekly_20100528.shape[0]))
weekly_20100528_addname_dataframe.to_excel(weekly_2010_excel_writer,'22_20100528',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100604")       ###  更新 周线记录日期
weekly_20100604 = pro.weekly(trade_date='20100604', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100604_tscode_list = list() 
for ts_code_sh in weekly_20100604['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100604_tscode_list.append(stock_name)
weekly_20100604_addname_dataframe=pd.DataFrame()
weekly_20100604_addname_dataframe['cname'] = weekly_20100604_tscode_list
for table_name in weekly_20100604.columns.values.tolist():
    weekly_20100604_addname_dataframe[table_name] = weekly_20100604[table_name]
print("周线行情-时间为序  weekly_20100604 23_20100604 返回数据 row 行数 = "+str(weekly_20100604.shape[0]))
weekly_20100604_addname_dataframe.to_excel(weekly_2010_excel_writer,'23_20100604',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100611")       ###  更新 周线记录日期
weekly_20100611 = pro.weekly(trade_date='20100611', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100611_tscode_list = list() 
for ts_code_sh in weekly_20100611['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100611_tscode_list.append(stock_name)
weekly_20100611_addname_dataframe=pd.DataFrame()
weekly_20100611_addname_dataframe['cname'] = weekly_20100611_tscode_list
for table_name in weekly_20100611.columns.values.tolist():
    weekly_20100611_addname_dataframe[table_name] = weekly_20100611[table_name]
print("周线行情-时间为序  weekly_20100611 24_20100611 返回数据 row 行数 = "+str(weekly_20100611.shape[0]))
weekly_20100611_addname_dataframe.to_excel(weekly_2010_excel_writer,'24_20100611',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100618")       ###  更新 周线记录日期
weekly_20100618 = pro.weekly(trade_date='20100618', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100618_tscode_list = list() 
for ts_code_sh in weekly_20100618['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100618_tscode_list.append(stock_name)
weekly_20100618_addname_dataframe=pd.DataFrame()
weekly_20100618_addname_dataframe['cname'] = weekly_20100618_tscode_list
for table_name in weekly_20100618.columns.values.tolist():
    weekly_20100618_addname_dataframe[table_name] = weekly_20100618[table_name]
print("周线行情-时间为序  weekly_20100618 25_20100618 返回数据 row 行数 = "+str(weekly_20100618.shape[0]))
weekly_20100618_addname_dataframe.to_excel(weekly_2010_excel_writer,'25_20100618',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100625")       ###  更新 周线记录日期
weekly_20100625 = pro.weekly(trade_date='20100625', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100625_tscode_list = list() 
for ts_code_sh in weekly_20100625['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100625_tscode_list.append(stock_name)
weekly_20100625_addname_dataframe=pd.DataFrame()
weekly_20100625_addname_dataframe['cname'] = weekly_20100625_tscode_list
for table_name in weekly_20100625.columns.values.tolist():
    weekly_20100625_addname_dataframe[table_name] = weekly_20100625[table_name]
print("周线行情-时间为序  weekly_20100625 26_20100625 返回数据 row 行数 = "+str(weekly_20100625.shape[0]))
weekly_20100625_addname_dataframe.to_excel(weekly_2010_excel_writer,'26_20100625',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100702")       ###  更新 周线记录日期
weekly_20100702 = pro.weekly(trade_date='20100702', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100702_tscode_list = list() 
for ts_code_sh in weekly_20100702['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100702_tscode_list.append(stock_name)
weekly_20100702_addname_dataframe=pd.DataFrame()
weekly_20100702_addname_dataframe['cname'] = weekly_20100702_tscode_list
for table_name in weekly_20100702.columns.values.tolist():
    weekly_20100702_addname_dataframe[table_name] = weekly_20100702[table_name]
print("周线行情-时间为序  weekly_20100702 27_20100702 返回数据 row 行数 = "+str(weekly_20100702.shape[0]))
weekly_20100702_addname_dataframe.to_excel(weekly_2010_excel_writer,'27_20100702',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100709")       ###  更新 周线记录日期
weekly_20100709 = pro.weekly(trade_date='20100709', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100709_tscode_list = list() 
for ts_code_sh in weekly_20100709['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100709_tscode_list.append(stock_name)
weekly_20100709_addname_dataframe=pd.DataFrame()
weekly_20100709_addname_dataframe['cname'] = weekly_20100709_tscode_list
for table_name in weekly_20100709.columns.values.tolist():
    weekly_20100709_addname_dataframe[table_name] = weekly_20100709[table_name]
print("周线行情-时间为序  weekly_20100709 28_20100709 返回数据 row 行数 = "+str(weekly_20100709.shape[0]))
weekly_20100709_addname_dataframe.to_excel(weekly_2010_excel_writer,'28_20100709',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100716")       ###  更新 周线记录日期
weekly_20100716 = pro.weekly(trade_date='20100716', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100716_tscode_list = list() 
for ts_code_sh in weekly_20100716['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100716_tscode_list.append(stock_name)
weekly_20100716_addname_dataframe=pd.DataFrame()
weekly_20100716_addname_dataframe['cname'] = weekly_20100716_tscode_list
for table_name in weekly_20100716.columns.values.tolist():
    weekly_20100716_addname_dataframe[table_name] = weekly_20100716[table_name]
print("周线行情-时间为序  weekly_20100716 29_20100716 返回数据 row 行数 = "+str(weekly_20100716.shape[0]))
weekly_20100716_addname_dataframe.to_excel(weekly_2010_excel_writer,'29_20100716',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100723")       ###  更新 周线记录日期
weekly_20100723 = pro.weekly(trade_date='20100723', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100723_tscode_list = list() 
for ts_code_sh in weekly_20100723['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100723_tscode_list.append(stock_name)
weekly_20100723_addname_dataframe=pd.DataFrame()
weekly_20100723_addname_dataframe['cname'] = weekly_20100723_tscode_list
for table_name in weekly_20100723.columns.values.tolist():
    weekly_20100723_addname_dataframe[table_name] = weekly_20100723[table_name]
print("周线行情-时间为序  weekly_20100723 30_20100723 返回数据 row 行数 = "+str(weekly_20100723.shape[0]))
weekly_20100723_addname_dataframe.to_excel(weekly_2010_excel_writer,'30_20100723',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100730")       ###  更新 周线记录日期
weekly_20100730 = pro.weekly(trade_date='20100730', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100730_tscode_list = list() 
for ts_code_sh in weekly_20100730['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100730_tscode_list.append(stock_name)
weekly_20100730_addname_dataframe=pd.DataFrame()
weekly_20100730_addname_dataframe['cname'] = weekly_20100730_tscode_list
for table_name in weekly_20100730.columns.values.tolist():
    weekly_20100730_addname_dataframe[table_name] = weekly_20100730[table_name]
print("周线行情-时间为序  weekly_20100730 31_20100730 返回数据 row 行数 = "+str(weekly_20100730.shape[0]))
weekly_20100730_addname_dataframe.to_excel(weekly_2010_excel_writer,'31_20100730',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100806")       ###  更新 周线记录日期
weekly_20100806 = pro.weekly(trade_date='20100806', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100806_tscode_list = list() 
for ts_code_sh in weekly_20100806['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100806_tscode_list.append(stock_name)
weekly_20100806_addname_dataframe=pd.DataFrame()
weekly_20100806_addname_dataframe['cname'] = weekly_20100806_tscode_list
for table_name in weekly_20100806.columns.values.tolist():
    weekly_20100806_addname_dataframe[table_name] = weekly_20100806[table_name]
print("周线行情-时间为序  weekly_20100806 32_20100806 返回数据 row 行数 = "+str(weekly_20100806.shape[0]))
weekly_20100806_addname_dataframe.to_excel(weekly_2010_excel_writer,'32_20100806',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100813")       ###  更新 周线记录日期
weekly_20100813 = pro.weekly(trade_date='20100813', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100813_tscode_list = list() 
for ts_code_sh in weekly_20100813['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100813_tscode_list.append(stock_name)
weekly_20100813_addname_dataframe=pd.DataFrame()
weekly_20100813_addname_dataframe['cname'] = weekly_20100813_tscode_list
for table_name in weekly_20100813.columns.values.tolist():
    weekly_20100813_addname_dataframe[table_name] = weekly_20100813[table_name]
print("周线行情-时间为序  weekly_20100813 33_20100813 返回数据 row 行数 = "+str(weekly_20100813.shape[0]))
weekly_20100813_addname_dataframe.to_excel(weekly_2010_excel_writer,'33_20100813',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100820")       ###  更新 周线记录日期
weekly_20100820 = pro.weekly(trade_date='20100820', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100820_tscode_list = list() 
for ts_code_sh in weekly_20100820['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100820_tscode_list.append(stock_name)
weekly_20100820_addname_dataframe=pd.DataFrame()
weekly_20100820_addname_dataframe['cname'] = weekly_20100820_tscode_list
for table_name in weekly_20100820.columns.values.tolist():
    weekly_20100820_addname_dataframe[table_name] = weekly_20100820[table_name]
print("周线行情-时间为序  weekly_20100820 34_20100820 返回数据 row 行数 = "+str(weekly_20100820.shape[0]))
weekly_20100820_addname_dataframe.to_excel(weekly_2010_excel_writer,'34_20100820',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100827")       ###  更新 周线记录日期
weekly_20100827 = pro.weekly(trade_date='20100827', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100827_tscode_list = list() 
for ts_code_sh in weekly_20100827['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100827_tscode_list.append(stock_name)
weekly_20100827_addname_dataframe=pd.DataFrame()
weekly_20100827_addname_dataframe['cname'] = weekly_20100827_tscode_list
for table_name in weekly_20100827.columns.values.tolist():
    weekly_20100827_addname_dataframe[table_name] = weekly_20100827[table_name]
print("周线行情-时间为序  weekly_20100827 35_20100827 返回数据 row 行数 = "+str(weekly_20100827.shape[0]))
weekly_20100827_addname_dataframe.to_excel(weekly_2010_excel_writer,'35_20100827',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100903")       ###  更新 周线记录日期
weekly_20100903 = pro.weekly(trade_date='20100903', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100903_tscode_list = list() 
for ts_code_sh in weekly_20100903['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100903_tscode_list.append(stock_name)
weekly_20100903_addname_dataframe=pd.DataFrame()
weekly_20100903_addname_dataframe['cname'] = weekly_20100903_tscode_list
for table_name in weekly_20100903.columns.values.tolist():
    weekly_20100903_addname_dataframe[table_name] = weekly_20100903[table_name]
print("周线行情-时间为序  weekly_20100903 36_20100903 返回数据 row 行数 = "+str(weekly_20100903.shape[0]))
weekly_20100903_addname_dataframe.to_excel(weekly_2010_excel_writer,'36_20100903',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100910")       ###  更新 周线记录日期
weekly_20100910 = pro.weekly(trade_date='20100910', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100910_tscode_list = list() 
for ts_code_sh in weekly_20100910['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100910_tscode_list.append(stock_name)
weekly_20100910_addname_dataframe=pd.DataFrame()
weekly_20100910_addname_dataframe['cname'] = weekly_20100910_tscode_list
for table_name in weekly_20100910.columns.values.tolist():
    weekly_20100910_addname_dataframe[table_name] = weekly_20100910[table_name]
print("周线行情-时间为序  weekly_20100910 37_20100910 返回数据 row 行数 = "+str(weekly_20100910.shape[0]))
weekly_20100910_addname_dataframe.to_excel(weekly_2010_excel_writer,'37_20100910',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100917")       ###  更新 周线记录日期
weekly_20100917 = pro.weekly(trade_date='20100917', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100917_tscode_list = list() 
for ts_code_sh in weekly_20100917['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100917_tscode_list.append(stock_name)
weekly_20100917_addname_dataframe=pd.DataFrame()
weekly_20100917_addname_dataframe['cname'] = weekly_20100917_tscode_list
for table_name in weekly_20100917.columns.values.tolist():
    weekly_20100917_addname_dataframe[table_name] = weekly_20100917[table_name]
print("周线行情-时间为序  weekly_20100917 38_20100917 返回数据 row 行数 = "+str(weekly_20100917.shape[0]))
weekly_20100917_addname_dataframe.to_excel(weekly_2010_excel_writer,'38_20100917',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100921")       ###  更新 周线记录日期
weekly_20100921 = pro.weekly(trade_date='20100921', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100921_tscode_list = list() 
for ts_code_sh in weekly_20100921['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100921_tscode_list.append(stock_name)
weekly_20100921_addname_dataframe=pd.DataFrame()
weekly_20100921_addname_dataframe['cname'] = weekly_20100921_tscode_list
for table_name in weekly_20100921.columns.values.tolist():
    weekly_20100921_addname_dataframe[table_name] = weekly_20100921[table_name]
print("周线行情-时间为序  weekly_20100921 38_20100921 返回数据 row 行数 = "+str(weekly_20100921.shape[0]))
weekly_20100921_addname_dataframe.to_excel(weekly_2010_excel_writer,'38_20100921',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20100930")       ###  更新 周线记录日期
weekly_20100930 = pro.weekly(trade_date='20100930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20100930_tscode_list = list() 
for ts_code_sh in weekly_20100930['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20100930_tscode_list.append(stock_name)
weekly_20100930_addname_dataframe=pd.DataFrame()
weekly_20100930_addname_dataframe['cname'] = weekly_20100930_tscode_list
for table_name in weekly_20100930.columns.values.tolist():
    weekly_20100930_addname_dataframe[table_name] = weekly_20100930[table_name]
print("周线行情-时间为序  weekly_20100930 39_20100930 返回数据 row 行数 = "+str(weekly_20100930.shape[0]))
weekly_20100930_addname_dataframe.to_excel(weekly_2010_excel_writer,'39_20100930',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101008")       ###  更新 周线记录日期
weekly_20101008 = pro.weekly(trade_date='20101008', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101008_tscode_list = list() 
for ts_code_sh in weekly_20101008['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101008_tscode_list.append(stock_name)
weekly_20101008_addname_dataframe=pd.DataFrame()
weekly_20101008_addname_dataframe['cname'] = weekly_20101008_tscode_list
for table_name in weekly_20101008.columns.values.tolist():
    weekly_20101008_addname_dataframe[table_name] = weekly_20101008[table_name]
print("周线行情-时间为序  weekly_20101008 41_20101008 返回数据 row 行数 = "+str(weekly_20101008.shape[0]))
weekly_20101008_addname_dataframe.to_excel(weekly_2010_excel_writer,'41_20101008',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101015")       ###  更新 周线记录日期
weekly_20101015 = pro.weekly(trade_date='20101015', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101015_tscode_list = list() 
for ts_code_sh in weekly_20101015['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101015_tscode_list.append(stock_name)
weekly_20101015_addname_dataframe=pd.DataFrame()
weekly_20101015_addname_dataframe['cname'] = weekly_20101015_tscode_list
for table_name in weekly_20101015.columns.values.tolist():
    weekly_20101015_addname_dataframe[table_name] = weekly_20101015[table_name]
print("周线行情-时间为序  weekly_20101015 42_20101015 返回数据 row 行数 = "+str(weekly_20101015.shape[0]))
weekly_20101015_addname_dataframe.to_excel(weekly_2010_excel_writer,'42_20101015',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101022")       ###  更新 周线记录日期
weekly_20101022 = pro.weekly(trade_date='20101022', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101022_tscode_list = list() 
for ts_code_sh in weekly_20101022['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101022_tscode_list.append(stock_name)
weekly_20101022_addname_dataframe=pd.DataFrame()
weekly_20101022_addname_dataframe['cname'] = weekly_20101022_tscode_list
for table_name in weekly_20101022.columns.values.tolist():
    weekly_20101022_addname_dataframe[table_name] = weekly_20101022[table_name]
print("周线行情-时间为序  weekly_20101022 43_20101022 返回数据 row 行数 = "+str(weekly_20101022.shape[0]))
weekly_20101022_addname_dataframe.to_excel(weekly_2010_excel_writer,'43_20101022',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101029")       ###  更新 周线记录日期
weekly_20101029 = pro.weekly(trade_date='20101029', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101029_tscode_list = list() 
for ts_code_sh in weekly_20101029['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101029_tscode_list.append(stock_name)
weekly_20101029_addname_dataframe=pd.DataFrame()
weekly_20101029_addname_dataframe['cname'] = weekly_20101029_tscode_list
for table_name in weekly_20101029.columns.values.tolist():
    weekly_20101029_addname_dataframe[table_name] = weekly_20101029[table_name]
print("周线行情-时间为序  weekly_20101029 44_20101029 返回数据 row 行数 = "+str(weekly_20101029.shape[0]))
weekly_20101029_addname_dataframe.to_excel(weekly_2010_excel_writer,'44_20101029',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101105")       ###  更新 周线记录日期
weekly_20101105 = pro.weekly(trade_date='20101105', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101105_tscode_list = list() 
for ts_code_sh in weekly_20101105['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101105_tscode_list.append(stock_name)
weekly_20101105_addname_dataframe=pd.DataFrame()
weekly_20101105_addname_dataframe['cname'] = weekly_20101105_tscode_list
for table_name in weekly_20101105.columns.values.tolist():
    weekly_20101105_addname_dataframe[table_name] = weekly_20101105[table_name]
print("周线行情-时间为序  weekly_20101105 45_20101105 返回数据 row 行数 = "+str(weekly_20101105.shape[0]))
weekly_20101105_addname_dataframe.to_excel(weekly_2010_excel_writer,'45_20101105',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101112")       ###  更新 周线记录日期
weekly_20101112 = pro.weekly(trade_date='20101112', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101112_tscode_list = list() 
for ts_code_sh in weekly_20101112['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101112_tscode_list.append(stock_name)
weekly_20101112_addname_dataframe=pd.DataFrame()
weekly_20101112_addname_dataframe['cname'] = weekly_20101112_tscode_list
for table_name in weekly_20101112.columns.values.tolist():
    weekly_20101112_addname_dataframe[table_name] = weekly_20101112[table_name]
print("周线行情-时间为序  weekly_20101112 46_20101112 返回数据 row 行数 = "+str(weekly_20101112.shape[0]))
weekly_20101112_addname_dataframe.to_excel(weekly_2010_excel_writer,'46_20101112',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101119")       ###  更新 周线记录日期
weekly_20101119 = pro.weekly(trade_date='20101119', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101119_tscode_list = list() 
for ts_code_sh in weekly_20101119['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101119_tscode_list.append(stock_name)
weekly_20101119_addname_dataframe=pd.DataFrame()
weekly_20101119_addname_dataframe['cname'] = weekly_20101119_tscode_list
for table_name in weekly_20101119.columns.values.tolist():
    weekly_20101119_addname_dataframe[table_name] = weekly_20101119[table_name]
print("周线行情-时间为序  weekly_20101119 47_20101119 返回数据 row 行数 = "+str(weekly_20101119.shape[0]))
weekly_20101119_addname_dataframe.to_excel(weekly_2010_excel_writer,'47_20101119',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101126")       ###  更新 周线记录日期
weekly_20101126 = pro.weekly(trade_date='20101126', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101126_tscode_list = list() 
for ts_code_sh in weekly_20101126['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101126_tscode_list.append(stock_name)
weekly_20101126_addname_dataframe=pd.DataFrame()
weekly_20101126_addname_dataframe['cname'] = weekly_20101126_tscode_list
for table_name in weekly_20101126.columns.values.tolist():
    weekly_20101126_addname_dataframe[table_name] = weekly_20101126[table_name]
print("周线行情-时间为序  weekly_20101126 48_20101126 返回数据 row 行数 = "+str(weekly_20101126.shape[0]))
weekly_20101126_addname_dataframe.to_excel(weekly_2010_excel_writer,'48_20101126',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101203")       ###  更新 周线记录日期
weekly_20101203 = pro.weekly(trade_date='20101203', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101203_tscode_list = list() 
for ts_code_sh in weekly_20101203['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101203_tscode_list.append(stock_name)
weekly_20101203_addname_dataframe=pd.DataFrame()
weekly_20101203_addname_dataframe['cname'] = weekly_20101203_tscode_list
for table_name in weekly_20101203.columns.values.tolist():
    weekly_20101203_addname_dataframe[table_name] = weekly_20101203[table_name]
print("周线行情-时间为序  weekly_20101203 49_20101203 返回数据 row 行数 = "+str(weekly_20101203.shape[0]))
weekly_20101203_addname_dataframe.to_excel(weekly_2010_excel_writer,'49_20101203',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101210")       ###  更新 周线记录日期
weekly_20101210 = pro.weekly(trade_date='20101210', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101210_tscode_list = list() 
for ts_code_sh in weekly_20101210['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101210_tscode_list.append(stock_name)
weekly_20101210_addname_dataframe=pd.DataFrame()
weekly_20101210_addname_dataframe['cname'] = weekly_20101210_tscode_list
for table_name in weekly_20101210.columns.values.tolist():
    weekly_20101210_addname_dataframe[table_name] = weekly_20101210[table_name]
print("周线行情-时间为序  weekly_20101210 50_20101210 返回数据 row 行数 = "+str(weekly_20101210.shape[0]))
weekly_20101210_addname_dataframe.to_excel(weekly_2010_excel_writer,'50_20101210',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101217")       ###  更新 周线记录日期
weekly_20101217 = pro.weekly(trade_date='20101217', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101217_tscode_list = list() 
for ts_code_sh in weekly_20101217['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101217_tscode_list.append(stock_name)
weekly_20101217_addname_dataframe=pd.DataFrame()
weekly_20101217_addname_dataframe['cname'] = weekly_20101217_tscode_list
for table_name in weekly_20101217.columns.values.tolist():
    weekly_20101217_addname_dataframe[table_name] = weekly_20101217[table_name]
print("周线行情-时间为序  weekly_20101217 51_20101217 返回数据 row 行数 = "+str(weekly_20101217.shape[0]))
weekly_20101217_addname_dataframe.to_excel(weekly_2010_excel_writer,'51_20101217',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101224")       ###  更新 周线记录日期
weekly_20101224 = pro.weekly(trade_date='20101224', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101224_tscode_list = list() 
for ts_code_sh in weekly_20101224['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101224_tscode_list.append(stock_name)
weekly_20101224_addname_dataframe=pd.DataFrame()
weekly_20101224_addname_dataframe['cname'] = weekly_20101224_tscode_list
for table_name in weekly_20101224.columns.values.tolist():
    weekly_20101224_addname_dataframe[table_name] = weekly_20101224[table_name]
print("周线行情-时间为序  weekly_20101224 52_20101224 返回数据 row 行数 = "+str(weekly_20101224.shape[0]))
weekly_20101224_addname_dataframe.to_excel(weekly_2010_excel_writer,'52_20101224',index=False)
weekly_2010_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20101231")       ###  更新 周线记录日期
weekly_20101231 = pro.weekly(trade_date='20101231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20101231_tscode_list = list() 
for ts_code_sh in weekly_20101231['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20101231_tscode_list.append(stock_name)
weekly_20101231_addname_dataframe=pd.DataFrame()
weekly_20101231_addname_dataframe['cname'] = weekly_20101231_tscode_list
for table_name in weekly_20101231.columns.values.tolist():
    weekly_20101231_addname_dataframe[table_name] = weekly_20101231[table_name]
print("周线行情-时间为序  weekly_20101231 53_20101231 返回数据 row 行数 = "+str(weekly_20101231.shape[0]))
weekly_20101231_addname_dataframe.to_excel(weekly_2010_excel_writer,'53_20101231',index=False)
weekly_2010_excel_writer.save()
createexcel('weekly_2011.xlsx')
weekly_2011_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2011.xlsx')
weekly_2011_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2011.xlsx', engine='openpyxl')
weekly_2011_excel_writer.book = weekly_2011_book
weekly_2011_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2011_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20110107")       ###  更新 周线记录日期
weekly_20110107 = pro.weekly(trade_date='20110107', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110107_tscode_list = list() 
for ts_code_sh in weekly_20110107['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110107_tscode_list.append(stock_name)
weekly_20110107_addname_dataframe=pd.DataFrame()
weekly_20110107_addname_dataframe['cname'] = weekly_20110107_tscode_list
for table_name in weekly_20110107.columns.values.tolist():
    weekly_20110107_addname_dataframe[table_name] = weekly_20110107[table_name]
print("周线行情-时间为序  weekly_20110107 1_20110107 返回数据 row 行数 = "+str(weekly_20110107.shape[0]))
weekly_20110107_addname_dataframe.to_excel(weekly_2011_excel_writer,'1_20110107',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110114")       ###  更新 周线记录日期
weekly_20110114 = pro.weekly(trade_date='20110114', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110114_tscode_list = list() 
for ts_code_sh in weekly_20110114['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110114_tscode_list.append(stock_name)
weekly_20110114_addname_dataframe=pd.DataFrame()
weekly_20110114_addname_dataframe['cname'] = weekly_20110114_tscode_list
for table_name in weekly_20110114.columns.values.tolist():
    weekly_20110114_addname_dataframe[table_name] = weekly_20110114[table_name]
print("周线行情-时间为序  weekly_20110114 2_20110114 返回数据 row 行数 = "+str(weekly_20110114.shape[0]))
weekly_20110114_addname_dataframe.to_excel(weekly_2011_excel_writer,'2_20110114',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110121")       ###  更新 周线记录日期
weekly_20110121 = pro.weekly(trade_date='20110121', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110121_tscode_list = list() 
for ts_code_sh in weekly_20110121['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110121_tscode_list.append(stock_name)
weekly_20110121_addname_dataframe=pd.DataFrame()
weekly_20110121_addname_dataframe['cname'] = weekly_20110121_tscode_list
for table_name in weekly_20110121.columns.values.tolist():
    weekly_20110121_addname_dataframe[table_name] = weekly_20110121[table_name]
print("周线行情-时间为序  weekly_20110121 3_20110121 返回数据 row 行数 = "+str(weekly_20110121.shape[0]))
weekly_20110121_addname_dataframe.to_excel(weekly_2011_excel_writer,'3_20110121',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110128")       ###  更新 周线记录日期
weekly_20110128 = pro.weekly(trade_date='20110128', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110128_tscode_list = list() 
for ts_code_sh in weekly_20110128['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110128_tscode_list.append(stock_name)
weekly_20110128_addname_dataframe=pd.DataFrame()
weekly_20110128_addname_dataframe['cname'] = weekly_20110128_tscode_list
for table_name in weekly_20110128.columns.values.tolist():
    weekly_20110128_addname_dataframe[table_name] = weekly_20110128[table_name]
print("周线行情-时间为序  weekly_20110128 4_20110128 返回数据 row 行数 = "+str(weekly_20110128.shape[0]))
weekly_20110128_addname_dataframe.to_excel(weekly_2011_excel_writer,'4_20110128',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110201")       ###  更新 周线记录日期
weekly_20110201 = pro.weekly(trade_date='20110201', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110201_tscode_list = list() 
for ts_code_sh in weekly_20110201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110201_tscode_list.append(stock_name)
weekly_20110201_addname_dataframe=pd.DataFrame()
weekly_20110201_addname_dataframe['cname'] = weekly_20110201_tscode_list
for table_name in weekly_20110201.columns.values.tolist():
    weekly_20110201_addname_dataframe[table_name] = weekly_20110201[table_name]
print("周线行情-时间为序  weekly_20110201 5_20110201 返回数据 row 行数 = "+str(weekly_20110201.shape[0]))
weekly_20110201_addname_dataframe.to_excel(weekly_2011_excel_writer,'5_20110201',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110211")       ###  更新 周线记录日期
weekly_20110211 = pro.weekly(trade_date='20110211', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110211_tscode_list = list() 
for ts_code_sh in weekly_20110211['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110211_tscode_list.append(stock_name)
weekly_20110211_addname_dataframe=pd.DataFrame()
weekly_20110211_addname_dataframe['cname'] = weekly_20110211_tscode_list
for table_name in weekly_20110211.columns.values.tolist():
    weekly_20110211_addname_dataframe[table_name] = weekly_20110211[table_name]
print("周线行情-时间为序  weekly_20110211 6_20110211 返回数据 row 行数 = "+str(weekly_20110211.shape[0]))
weekly_20110211_addname_dataframe.to_excel(weekly_2011_excel_writer,'6_20110211',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110218")       ###  更新 周线记录日期
weekly_20110218 = pro.weekly(trade_date='20110218', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110218_tscode_list = list() 
for ts_code_sh in weekly_20110218['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110218_tscode_list.append(stock_name)
weekly_20110218_addname_dataframe=pd.DataFrame()
weekly_20110218_addname_dataframe['cname'] = weekly_20110218_tscode_list
for table_name in weekly_20110218.columns.values.tolist():
    weekly_20110218_addname_dataframe[table_name] = weekly_20110218[table_name]
print("周线行情-时间为序  weekly_20110218 7_20110218 返回数据 row 行数 = "+str(weekly_20110218.shape[0]))
weekly_20110218_addname_dataframe.to_excel(weekly_2011_excel_writer,'7_20110218',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110225")       ###  更新 周线记录日期
weekly_20110225 = pro.weekly(trade_date='20110225', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110225_tscode_list = list() 
for ts_code_sh in weekly_20110225['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110225_tscode_list.append(stock_name)
weekly_20110225_addname_dataframe=pd.DataFrame()
weekly_20110225_addname_dataframe['cname'] = weekly_20110225_tscode_list
for table_name in weekly_20110225.columns.values.tolist():
    weekly_20110225_addname_dataframe[table_name] = weekly_20110225[table_name]
print("周线行情-时间为序  weekly_20110225 8_20110225 返回数据 row 行数 = "+str(weekly_20110225.shape[0]))
weekly_20110225_addname_dataframe.to_excel(weekly_2011_excel_writer,'8_20110225',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110304")       ###  更新 周线记录日期
weekly_20110304 = pro.weekly(trade_date='20110304', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110304_tscode_list = list() 
for ts_code_sh in weekly_20110304['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110304_tscode_list.append(stock_name)
weekly_20110304_addname_dataframe=pd.DataFrame()
weekly_20110304_addname_dataframe['cname'] = weekly_20110304_tscode_list
for table_name in weekly_20110304.columns.values.tolist():
    weekly_20110304_addname_dataframe[table_name] = weekly_20110304[table_name]
print("周线行情-时间为序  weekly_20110304 9_20110304 返回数据 row 行数 = "+str(weekly_20110304.shape[0]))
weekly_20110304_addname_dataframe.to_excel(weekly_2011_excel_writer,'9_20110304',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110311")       ###  更新 周线记录日期
weekly_20110311 = pro.weekly(trade_date='20110311', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110311_tscode_list = list() 
for ts_code_sh in weekly_20110311['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110311_tscode_list.append(stock_name)
weekly_20110311_addname_dataframe=pd.DataFrame()
weekly_20110311_addname_dataframe['cname'] = weekly_20110311_tscode_list
for table_name in weekly_20110311.columns.values.tolist():
    weekly_20110311_addname_dataframe[table_name] = weekly_20110311[table_name]
print("周线行情-时间为序  weekly_20110311 10_20110311 返回数据 row 行数 = "+str(weekly_20110311.shape[0]))
weekly_20110311_addname_dataframe.to_excel(weekly_2011_excel_writer,'10_20110311',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110318")       ###  更新 周线记录日期
weekly_20110318 = pro.weekly(trade_date='20110318', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110318_tscode_list = list() 
for ts_code_sh in weekly_20110318['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110318_tscode_list.append(stock_name)
weekly_20110318_addname_dataframe=pd.DataFrame()
weekly_20110318_addname_dataframe['cname'] = weekly_20110318_tscode_list
for table_name in weekly_20110318.columns.values.tolist():
    weekly_20110318_addname_dataframe[table_name] = weekly_20110318[table_name]
print("周线行情-时间为序  weekly_20110318 11_20110318 返回数据 row 行数 = "+str(weekly_20110318.shape[0]))
weekly_20110318_addname_dataframe.to_excel(weekly_2011_excel_writer,'11_20110318',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110325")       ###  更新 周线记录日期
weekly_20110325 = pro.weekly(trade_date='20110325', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110325_tscode_list = list() 
for ts_code_sh in weekly_20110325['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110325_tscode_list.append(stock_name)
weekly_20110325_addname_dataframe=pd.DataFrame()
weekly_20110325_addname_dataframe['cname'] = weekly_20110325_tscode_list
for table_name in weekly_20110325.columns.values.tolist():
    weekly_20110325_addname_dataframe[table_name] = weekly_20110325[table_name]
print("周线行情-时间为序  weekly_20110325 12_20110325 返回数据 row 行数 = "+str(weekly_20110325.shape[0]))
weekly_20110325_addname_dataframe.to_excel(weekly_2011_excel_writer,'12_20110325',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110401")       ###  更新 周线记录日期
weekly_20110401 = pro.weekly(trade_date='20110401', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110401_tscode_list = list() 
for ts_code_sh in weekly_20110401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110401_tscode_list.append(stock_name)
weekly_20110401_addname_dataframe=pd.DataFrame()
weekly_20110401_addname_dataframe['cname'] = weekly_20110401_tscode_list
for table_name in weekly_20110401.columns.values.tolist():
    weekly_20110401_addname_dataframe[table_name] = weekly_20110401[table_name]
print("周线行情-时间为序  weekly_20110401 13_20110401 返回数据 row 行数 = "+str(weekly_20110401.shape[0]))
weekly_20110401_addname_dataframe.to_excel(weekly_2011_excel_writer,'13_20110401',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110408")       ###  更新 周线记录日期
weekly_20110408 = pro.weekly(trade_date='20110408', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110408_tscode_list = list() 
for ts_code_sh in weekly_20110408['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110408_tscode_list.append(stock_name)
weekly_20110408_addname_dataframe=pd.DataFrame()
weekly_20110408_addname_dataframe['cname'] = weekly_20110408_tscode_list
for table_name in weekly_20110408.columns.values.tolist():
    weekly_20110408_addname_dataframe[table_name] = weekly_20110408[table_name]
print("周线行情-时间为序  weekly_20110408 14_20110408 返回数据 row 行数 = "+str(weekly_20110408.shape[0]))
weekly_20110408_addname_dataframe.to_excel(weekly_2011_excel_writer,'14_20110408',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110415")       ###  更新 周线记录日期
weekly_20110415 = pro.weekly(trade_date='20110415', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110415_tscode_list = list() 
for ts_code_sh in weekly_20110415['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110415_tscode_list.append(stock_name)
weekly_20110415_addname_dataframe=pd.DataFrame()
weekly_20110415_addname_dataframe['cname'] = weekly_20110415_tscode_list
for table_name in weekly_20110415.columns.values.tolist():
    weekly_20110415_addname_dataframe[table_name] = weekly_20110415[table_name]
print("周线行情-时间为序  weekly_20110415 15_20110415 返回数据 row 行数 = "+str(weekly_20110415.shape[0]))
weekly_20110415_addname_dataframe.to_excel(weekly_2011_excel_writer,'15_20110415',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110422")       ###  更新 周线记录日期
weekly_20110422 = pro.weekly(trade_date='20110422', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110422_tscode_list = list() 
for ts_code_sh in weekly_20110422['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110422_tscode_list.append(stock_name)
weekly_20110422_addname_dataframe=pd.DataFrame()
weekly_20110422_addname_dataframe['cname'] = weekly_20110422_tscode_list
for table_name in weekly_20110422.columns.values.tolist():
    weekly_20110422_addname_dataframe[table_name] = weekly_20110422[table_name]
print("周线行情-时间为序  weekly_20110422 16_20110422 返回数据 row 行数 = "+str(weekly_20110422.shape[0]))
weekly_20110422_addname_dataframe.to_excel(weekly_2011_excel_writer,'16_20110422',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110429")       ###  更新 周线记录日期
weekly_20110429 = pro.weekly(trade_date='20110429', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110429_tscode_list = list() 
for ts_code_sh in weekly_20110429['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110429_tscode_list.append(stock_name)
weekly_20110429_addname_dataframe=pd.DataFrame()
weekly_20110429_addname_dataframe['cname'] = weekly_20110429_tscode_list
for table_name in weekly_20110429.columns.values.tolist():
    weekly_20110429_addname_dataframe[table_name] = weekly_20110429[table_name]
print("周线行情-时间为序  weekly_20110429 17_20110429 返回数据 row 行数 = "+str(weekly_20110429.shape[0]))
weekly_20110429_addname_dataframe.to_excel(weekly_2011_excel_writer,'17_20110429',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110506")       ###  更新 周线记录日期
weekly_20110506 = pro.weekly(trade_date='20110506', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110506_tscode_list = list() 
for ts_code_sh in weekly_20110506['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110506_tscode_list.append(stock_name)
weekly_20110506_addname_dataframe=pd.DataFrame()
weekly_20110506_addname_dataframe['cname'] = weekly_20110506_tscode_list
for table_name in weekly_20110506.columns.values.tolist():
    weekly_20110506_addname_dataframe[table_name] = weekly_20110506[table_name]
print("周线行情-时间为序  weekly_20110506 18_20110506 返回数据 row 行数 = "+str(weekly_20110506.shape[0]))
weekly_20110506_addname_dataframe.to_excel(weekly_2011_excel_writer,'18_20110506',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110513")       ###  更新 周线记录日期
weekly_20110513 = pro.weekly(trade_date='20110513', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110513_tscode_list = list() 
for ts_code_sh in weekly_20110513['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110513_tscode_list.append(stock_name)
weekly_20110513_addname_dataframe=pd.DataFrame()
weekly_20110513_addname_dataframe['cname'] = weekly_20110513_tscode_list
for table_name in weekly_20110513.columns.values.tolist():
    weekly_20110513_addname_dataframe[table_name] = weekly_20110513[table_name]
print("周线行情-时间为序  weekly_20110513 19_20110513 返回数据 row 行数 = "+str(weekly_20110513.shape[0]))
weekly_20110513_addname_dataframe.to_excel(weekly_2011_excel_writer,'19_20110513',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110520")       ###  更新 周线记录日期
weekly_20110520 = pro.weekly(trade_date='20110520', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110520_tscode_list = list() 
for ts_code_sh in weekly_20110520['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110520_tscode_list.append(stock_name)
weekly_20110520_addname_dataframe=pd.DataFrame()
weekly_20110520_addname_dataframe['cname'] = weekly_20110520_tscode_list
for table_name in weekly_20110520.columns.values.tolist():
    weekly_20110520_addname_dataframe[table_name] = weekly_20110520[table_name]
print("周线行情-时间为序  weekly_20110520 20_20110520 返回数据 row 行数 = "+str(weekly_20110520.shape[0]))
weekly_20110520_addname_dataframe.to_excel(weekly_2011_excel_writer,'20_20110520',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110527")       ###  更新 周线记录日期
weekly_20110527 = pro.weekly(trade_date='20110527', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110527_tscode_list = list() 
for ts_code_sh in weekly_20110527['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110527_tscode_list.append(stock_name)
weekly_20110527_addname_dataframe=pd.DataFrame()
weekly_20110527_addname_dataframe['cname'] = weekly_20110527_tscode_list
for table_name in weekly_20110527.columns.values.tolist():
    weekly_20110527_addname_dataframe[table_name] = weekly_20110527[table_name]
print("周线行情-时间为序  weekly_20110527 21_20110527 返回数据 row 行数 = "+str(weekly_20110527.shape[0]))
weekly_20110527_addname_dataframe.to_excel(weekly_2011_excel_writer,'21_20110527',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110603")       ###  更新 周线记录日期
weekly_20110603 = pro.weekly(trade_date='20110603', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110603_tscode_list = list() 
for ts_code_sh in weekly_20110603['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110603_tscode_list.append(stock_name)
weekly_20110603_addname_dataframe=pd.DataFrame()
weekly_20110603_addname_dataframe['cname'] = weekly_20110603_tscode_list
for table_name in weekly_20110603.columns.values.tolist():
    weekly_20110603_addname_dataframe[table_name] = weekly_20110603[table_name]
print("周线行情-时间为序  weekly_20110603 22_20110603 返回数据 row 行数 = "+str(weekly_20110603.shape[0]))
weekly_20110603_addname_dataframe.to_excel(weekly_2011_excel_writer,'22_20110603',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110610")       ###  更新 周线记录日期
weekly_20110610 = pro.weekly(trade_date='20110610', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110610_tscode_list = list() 
for ts_code_sh in weekly_20110610['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110610_tscode_list.append(stock_name)
weekly_20110610_addname_dataframe=pd.DataFrame()
weekly_20110610_addname_dataframe['cname'] = weekly_20110610_tscode_list
for table_name in weekly_20110610.columns.values.tolist():
    weekly_20110610_addname_dataframe[table_name] = weekly_20110610[table_name]
print("周线行情-时间为序  weekly_20110610 23_20110610 返回数据 row 行数 = "+str(weekly_20110610.shape[0]))
weekly_20110610_addname_dataframe.to_excel(weekly_2011_excel_writer,'23_20110610',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110617")       ###  更新 周线记录日期
weekly_20110617 = pro.weekly(trade_date='20110617', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110617_tscode_list = list() 
for ts_code_sh in weekly_20110617['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110617_tscode_list.append(stock_name)
weekly_20110617_addname_dataframe=pd.DataFrame()
weekly_20110617_addname_dataframe['cname'] = weekly_20110617_tscode_list
for table_name in weekly_20110617.columns.values.tolist():
    weekly_20110617_addname_dataframe[table_name] = weekly_20110617[table_name]
print("周线行情-时间为序  weekly_20110617 24_20110617 返回数据 row 行数 = "+str(weekly_20110617.shape[0]))
weekly_20110617_addname_dataframe.to_excel(weekly_2011_excel_writer,'24_20110617',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110624")       ###  更新 周线记录日期
weekly_20110624 = pro.weekly(trade_date='20110624', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110624_tscode_list = list() 
for ts_code_sh in weekly_20110624['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110624_tscode_list.append(stock_name)
weekly_20110624_addname_dataframe=pd.DataFrame()
weekly_20110624_addname_dataframe['cname'] = weekly_20110624_tscode_list
for table_name in weekly_20110624.columns.values.tolist():
    weekly_20110624_addname_dataframe[table_name] = weekly_20110624[table_name]
print("周线行情-时间为序  weekly_20110624 25_20110624 返回数据 row 行数 = "+str(weekly_20110624.shape[0]))
weekly_20110624_addname_dataframe.to_excel(weekly_2011_excel_writer,'25_20110624',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110701")       ###  更新 周线记录日期
weekly_20110701 = pro.weekly(trade_date='20110701', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110701_tscode_list = list() 
for ts_code_sh in weekly_20110701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110701_tscode_list.append(stock_name)
weekly_20110701_addname_dataframe=pd.DataFrame()
weekly_20110701_addname_dataframe['cname'] = weekly_20110701_tscode_list
for table_name in weekly_20110701.columns.values.tolist():
    weekly_20110701_addname_dataframe[table_name] = weekly_20110701[table_name]
print("周线行情-时间为序  weekly_20110701 26_20110701 返回数据 row 行数 = "+str(weekly_20110701.shape[0]))
weekly_20110701_addname_dataframe.to_excel(weekly_2011_excel_writer,'26_20110701',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110708")       ###  更新 周线记录日期
weekly_20110708 = pro.weekly(trade_date='20110708', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110708_tscode_list = list() 
for ts_code_sh in weekly_20110708['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110708_tscode_list.append(stock_name)
weekly_20110708_addname_dataframe=pd.DataFrame()
weekly_20110708_addname_dataframe['cname'] = weekly_20110708_tscode_list
for table_name in weekly_20110708.columns.values.tolist():
    weekly_20110708_addname_dataframe[table_name] = weekly_20110708[table_name]
print("周线行情-时间为序  weekly_20110708 27_20110708 返回数据 row 行数 = "+str(weekly_20110708.shape[0]))
weekly_20110708_addname_dataframe.to_excel(weekly_2011_excel_writer,'27_20110708',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110715")       ###  更新 周线记录日期
weekly_20110715 = pro.weekly(trade_date='20110715', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110715_tscode_list = list() 
for ts_code_sh in weekly_20110715['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110715_tscode_list.append(stock_name)
weekly_20110715_addname_dataframe=pd.DataFrame()
weekly_20110715_addname_dataframe['cname'] = weekly_20110715_tscode_list
for table_name in weekly_20110715.columns.values.tolist():
    weekly_20110715_addname_dataframe[table_name] = weekly_20110715[table_name]
print("周线行情-时间为序  weekly_20110715 28_20110715 返回数据 row 行数 = "+str(weekly_20110715.shape[0]))
weekly_20110715_addname_dataframe.to_excel(weekly_2011_excel_writer,'28_20110715',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110722")       ###  更新 周线记录日期
weekly_20110722 = pro.weekly(trade_date='20110722', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110722_tscode_list = list() 
for ts_code_sh in weekly_20110722['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110722_tscode_list.append(stock_name)
weekly_20110722_addname_dataframe=pd.DataFrame()
weekly_20110722_addname_dataframe['cname'] = weekly_20110722_tscode_list
for table_name in weekly_20110722.columns.values.tolist():
    weekly_20110722_addname_dataframe[table_name] = weekly_20110722[table_name]
print("周线行情-时间为序  weekly_20110722 29_20110722 返回数据 row 行数 = "+str(weekly_20110722.shape[0]))
weekly_20110722_addname_dataframe.to_excel(weekly_2011_excel_writer,'29_20110722',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110729")       ###  更新 周线记录日期
weekly_20110729 = pro.weekly(trade_date='20110729', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110729_tscode_list = list() 
for ts_code_sh in weekly_20110729['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110729_tscode_list.append(stock_name)
weekly_20110729_addname_dataframe=pd.DataFrame()
weekly_20110729_addname_dataframe['cname'] = weekly_20110729_tscode_list
for table_name in weekly_20110729.columns.values.tolist():
    weekly_20110729_addname_dataframe[table_name] = weekly_20110729[table_name]
print("周线行情-时间为序  weekly_20110729 30_20110729 返回数据 row 行数 = "+str(weekly_20110729.shape[0]))
weekly_20110729_addname_dataframe.to_excel(weekly_2011_excel_writer,'30_20110729',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110805")       ###  更新 周线记录日期
weekly_20110805 = pro.weekly(trade_date='20110805', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110805_tscode_list = list() 
for ts_code_sh in weekly_20110805['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110805_tscode_list.append(stock_name)
weekly_20110805_addname_dataframe=pd.DataFrame()
weekly_20110805_addname_dataframe['cname'] = weekly_20110805_tscode_list
for table_name in weekly_20110805.columns.values.tolist():
    weekly_20110805_addname_dataframe[table_name] = weekly_20110805[table_name]
print("周线行情-时间为序  weekly_20110805 31_20110805 返回数据 row 行数 = "+str(weekly_20110805.shape[0]))
weekly_20110805_addname_dataframe.to_excel(weekly_2011_excel_writer,'31_20110805',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110812")       ###  更新 周线记录日期
weekly_20110812 = pro.weekly(trade_date='20110812', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110812_tscode_list = list() 
for ts_code_sh in weekly_20110812['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110812_tscode_list.append(stock_name)
weekly_20110812_addname_dataframe=pd.DataFrame()
weekly_20110812_addname_dataframe['cname'] = weekly_20110812_tscode_list
for table_name in weekly_20110812.columns.values.tolist():
    weekly_20110812_addname_dataframe[table_name] = weekly_20110812[table_name]
print("周线行情-时间为序  weekly_20110812 32_20110812 返回数据 row 行数 = "+str(weekly_20110812.shape[0]))
weekly_20110812_addname_dataframe.to_excel(weekly_2011_excel_writer,'32_20110812',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110819")       ###  更新 周线记录日期
weekly_20110819 = pro.weekly(trade_date='20110819', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110819_tscode_list = list() 
for ts_code_sh in weekly_20110819['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110819_tscode_list.append(stock_name)
weekly_20110819_addname_dataframe=pd.DataFrame()
weekly_20110819_addname_dataframe['cname'] = weekly_20110819_tscode_list
for table_name in weekly_20110819.columns.values.tolist():
    weekly_20110819_addname_dataframe[table_name] = weekly_20110819[table_name]
print("周线行情-时间为序  weekly_20110819 33_20110819 返回数据 row 行数 = "+str(weekly_20110819.shape[0]))
weekly_20110819_addname_dataframe.to_excel(weekly_2011_excel_writer,'33_20110819',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110826")       ###  更新 周线记录日期
weekly_20110826 = pro.weekly(trade_date='20110826', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110826_tscode_list = list() 
for ts_code_sh in weekly_20110826['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110826_tscode_list.append(stock_name)
weekly_20110826_addname_dataframe=pd.DataFrame()
weekly_20110826_addname_dataframe['cname'] = weekly_20110826_tscode_list
for table_name in weekly_20110826.columns.values.tolist():
    weekly_20110826_addname_dataframe[table_name] = weekly_20110826[table_name]
print("周线行情-时间为序  weekly_20110826 34_20110826 返回数据 row 行数 = "+str(weekly_20110826.shape[0]))
weekly_20110826_addname_dataframe.to_excel(weekly_2011_excel_writer,'34_20110826',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110902")       ###  更新 周线记录日期
weekly_20110902 = pro.weekly(trade_date='20110902', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110902_tscode_list = list() 
for ts_code_sh in weekly_20110902['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110902_tscode_list.append(stock_name)
weekly_20110902_addname_dataframe=pd.DataFrame()
weekly_20110902_addname_dataframe['cname'] = weekly_20110902_tscode_list
for table_name in weekly_20110902.columns.values.tolist():
    weekly_20110902_addname_dataframe[table_name] = weekly_20110902[table_name]
print("周线行情-时间为序  weekly_20110902 35_20110902 返回数据 row 行数 = "+str(weekly_20110902.shape[0]))
weekly_20110902_addname_dataframe.to_excel(weekly_2011_excel_writer,'35_20110902',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110909")       ###  更新 周线记录日期
weekly_20110909 = pro.weekly(trade_date='20110909', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110909_tscode_list = list() 
for ts_code_sh in weekly_20110909['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110909_tscode_list.append(stock_name)
weekly_20110909_addname_dataframe=pd.DataFrame()
weekly_20110909_addname_dataframe['cname'] = weekly_20110909_tscode_list
for table_name in weekly_20110909.columns.values.tolist():
    weekly_20110909_addname_dataframe[table_name] = weekly_20110909[table_name]
print("周线行情-时间为序  weekly_20110909 36_20110909 返回数据 row 行数 = "+str(weekly_20110909.shape[0]))
weekly_20110909_addname_dataframe.to_excel(weekly_2011_excel_writer,'36_20110909',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110916")       ###  更新 周线记录日期
weekly_20110916 = pro.weekly(trade_date='20110916', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110916_tscode_list = list() 
for ts_code_sh in weekly_20110916['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110916_tscode_list.append(stock_name)
weekly_20110916_addname_dataframe=pd.DataFrame()
weekly_20110916_addname_dataframe['cname'] = weekly_20110916_tscode_list
for table_name in weekly_20110916.columns.values.tolist():
    weekly_20110916_addname_dataframe[table_name] = weekly_20110916[table_name]
print("周线行情-时间为序  weekly_20110916 37_20110916 返回数据 row 行数 = "+str(weekly_20110916.shape[0]))
weekly_20110916_addname_dataframe.to_excel(weekly_2011_excel_writer,'37_20110916',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110923")       ###  更新 周线记录日期
weekly_20110923 = pro.weekly(trade_date='20110923', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110923_tscode_list = list() 
for ts_code_sh in weekly_20110923['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110923_tscode_list.append(stock_name)
weekly_20110923_addname_dataframe=pd.DataFrame()
weekly_20110923_addname_dataframe['cname'] = weekly_20110923_tscode_list
for table_name in weekly_20110923.columns.values.tolist():
    weekly_20110923_addname_dataframe[table_name] = weekly_20110923[table_name]
print("周线行情-时间为序  weekly_20110923 38_20110923 返回数据 row 行数 = "+str(weekly_20110923.shape[0]))
weekly_20110923_addname_dataframe.to_excel(weekly_2011_excel_writer,'38_20110923',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20110930")       ###  更新 周线记录日期
weekly_20110930 = pro.weekly(trade_date='20110930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20110930_tscode_list = list() 
for ts_code_sh in weekly_20110930['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20110930_tscode_list.append(stock_name)
weekly_20110930_addname_dataframe=pd.DataFrame()
weekly_20110930_addname_dataframe['cname'] = weekly_20110930_tscode_list
for table_name in weekly_20110930.columns.values.tolist():
    weekly_20110930_addname_dataframe[table_name] = weekly_20110930[table_name]
print("周线行情-时间为序  weekly_20110930 39_20110930 返回数据 row 行数 = "+str(weekly_20110930.shape[0]))
weekly_20110930_addname_dataframe.to_excel(weekly_2011_excel_writer,'39_20110930',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111014")       ###  更新 周线记录日期
weekly_20111014 = pro.weekly(trade_date='20111014', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111014_tscode_list = list() 
for ts_code_sh in weekly_20111014['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111014_tscode_list.append(stock_name)
weekly_20111014_addname_dataframe=pd.DataFrame()
weekly_20111014_addname_dataframe['cname'] = weekly_20111014_tscode_list
for table_name in weekly_20111014.columns.values.tolist():
    weekly_20111014_addname_dataframe[table_name] = weekly_20111014[table_name]
print("周线行情-时间为序  weekly_20111014 41_20111014 返回数据 row 行数 = "+str(weekly_20111014.shape[0]))
weekly_20111014_addname_dataframe.to_excel(weekly_2011_excel_writer,'41_20111014',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111021")       ###  更新 周线记录日期
weekly_20111021 = pro.weekly(trade_date='20111021', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111021_tscode_list = list() 
for ts_code_sh in weekly_20111021['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111021_tscode_list.append(stock_name)
weekly_20111021_addname_dataframe=pd.DataFrame()
weekly_20111021_addname_dataframe['cname'] = weekly_20111021_tscode_list
for table_name in weekly_20111021.columns.values.tolist():
    weekly_20111021_addname_dataframe[table_name] = weekly_20111021[table_name]
print("周线行情-时间为序  weekly_20111021 42_20111021 返回数据 row 行数 = "+str(weekly_20111021.shape[0]))
weekly_20111021_addname_dataframe.to_excel(weekly_2011_excel_writer,'42_20111021',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111028")       ###  更新 周线记录日期
weekly_20111028 = pro.weekly(trade_date='20111028', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111028_tscode_list = list() 
for ts_code_sh in weekly_20111028['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111028_tscode_list.append(stock_name)
weekly_20111028_addname_dataframe=pd.DataFrame()
weekly_20111028_addname_dataframe['cname'] = weekly_20111028_tscode_list
for table_name in weekly_20111028.columns.values.tolist():
    weekly_20111028_addname_dataframe[table_name] = weekly_20111028[table_name]
print("周线行情-时间为序  weekly_20111028 43_20111028 返回数据 row 行数 = "+str(weekly_20111028.shape[0]))
weekly_20111028_addname_dataframe.to_excel(weekly_2011_excel_writer,'43_20111028',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111104")       ###  更新 周线记录日期
weekly_20111104 = pro.weekly(trade_date='20111104', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111104_tscode_list = list() 
for ts_code_sh in weekly_20111104['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111104_tscode_list.append(stock_name)
weekly_20111104_addname_dataframe=pd.DataFrame()
weekly_20111104_addname_dataframe['cname'] = weekly_20111104_tscode_list
for table_name in weekly_20111104.columns.values.tolist():
    weekly_20111104_addname_dataframe[table_name] = weekly_20111104[table_name]
print("周线行情-时间为序  weekly_20111104 44_20111104 返回数据 row 行数 = "+str(weekly_20111104.shape[0]))
weekly_20111104_addname_dataframe.to_excel(weekly_2011_excel_writer,'44_20111104',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111111")       ###  更新 周线记录日期
weekly_20111111 = pro.weekly(trade_date='20111111', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111111_tscode_list = list() 
for ts_code_sh in weekly_20111111['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111111_tscode_list.append(stock_name)
weekly_20111111_addname_dataframe=pd.DataFrame()
weekly_20111111_addname_dataframe['cname'] = weekly_20111111_tscode_list
for table_name in weekly_20111111.columns.values.tolist():
    weekly_20111111_addname_dataframe[table_name] = weekly_20111111[table_name]
print("周线行情-时间为序  weekly_20111111 45_20111111 返回数据 row 行数 = "+str(weekly_20111111.shape[0]))
weekly_20111111_addname_dataframe.to_excel(weekly_2011_excel_writer,'45_20111111',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111118")       ###  更新 周线记录日期
weekly_20111118 = pro.weekly(trade_date='20111118', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111118_tscode_list = list() 
for ts_code_sh in weekly_20111118['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111118_tscode_list.append(stock_name)
weekly_20111118_addname_dataframe=pd.DataFrame()
weekly_20111118_addname_dataframe['cname'] = weekly_20111118_tscode_list
for table_name in weekly_20111118.columns.values.tolist():
    weekly_20111118_addname_dataframe[table_name] = weekly_20111118[table_name]
print("周线行情-时间为序  weekly_20111118 46_20111118 返回数据 row 行数 = "+str(weekly_20111118.shape[0]))
weekly_20111118_addname_dataframe.to_excel(weekly_2011_excel_writer,'46_20111118',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111125")       ###  更新 周线记录日期
weekly_20111125 = pro.weekly(trade_date='20111125', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111125_tscode_list = list() 
for ts_code_sh in weekly_20111125['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111125_tscode_list.append(stock_name)
weekly_20111125_addname_dataframe=pd.DataFrame()
weekly_20111125_addname_dataframe['cname'] = weekly_20111125_tscode_list
for table_name in weekly_20111125.columns.values.tolist():
    weekly_20111125_addname_dataframe[table_name] = weekly_20111125[table_name]
print("周线行情-时间为序  weekly_20111125 47_20111125 返回数据 row 行数 = "+str(weekly_20111125.shape[0]))
weekly_20111125_addname_dataframe.to_excel(weekly_2011_excel_writer,'47_20111125',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111202")       ###  更新 周线记录日期
weekly_20111202 = pro.weekly(trade_date='20111202', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111202_tscode_list = list() 
for ts_code_sh in weekly_20111202['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111202_tscode_list.append(stock_name)
weekly_20111202_addname_dataframe=pd.DataFrame()
weekly_20111202_addname_dataframe['cname'] = weekly_20111202_tscode_list
for table_name in weekly_20111202.columns.values.tolist():
    weekly_20111202_addname_dataframe[table_name] = weekly_20111202[table_name]
print("周线行情-时间为序  weekly_20111202 48_20111202 返回数据 row 行数 = "+str(weekly_20111202.shape[0]))
weekly_20111202_addname_dataframe.to_excel(weekly_2011_excel_writer,'48_20111202',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111209")       ###  更新 周线记录日期
weekly_20111209 = pro.weekly(trade_date='20111209', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111209_tscode_list = list() 
for ts_code_sh in weekly_20111209['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111209_tscode_list.append(stock_name)
weekly_20111209_addname_dataframe=pd.DataFrame()
weekly_20111209_addname_dataframe['cname'] = weekly_20111209_tscode_list
for table_name in weekly_20111209.columns.values.tolist():
    weekly_20111209_addname_dataframe[table_name] = weekly_20111209[table_name]
print("周线行情-时间为序  weekly_20111209 49_20111209 返回数据 row 行数 = "+str(weekly_20111209.shape[0]))
weekly_20111209_addname_dataframe.to_excel(weekly_2011_excel_writer,'49_20111209',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111216")       ###  更新 周线记录日期
weekly_20111216 = pro.weekly(trade_date='20111216', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111216_tscode_list = list() 
for ts_code_sh in weekly_20111216['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111216_tscode_list.append(stock_name)
weekly_20111216_addname_dataframe=pd.DataFrame()
weekly_20111216_addname_dataframe['cname'] = weekly_20111216_tscode_list
for table_name in weekly_20111216.columns.values.tolist():
    weekly_20111216_addname_dataframe[table_name] = weekly_20111216[table_name]
print("周线行情-时间为序  weekly_20111216 50_20111216 返回数据 row 行数 = "+str(weekly_20111216.shape[0]))
weekly_20111216_addname_dataframe.to_excel(weekly_2011_excel_writer,'50_20111216',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111223")       ###  更新 周线记录日期
weekly_20111223 = pro.weekly(trade_date='20111223', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111223_tscode_list = list() 
for ts_code_sh in weekly_20111223['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111223_tscode_list.append(stock_name)
weekly_20111223_addname_dataframe=pd.DataFrame()
weekly_20111223_addname_dataframe['cname'] = weekly_20111223_tscode_list
for table_name in weekly_20111223.columns.values.tolist():
    weekly_20111223_addname_dataframe[table_name] = weekly_20111223[table_name]
print("周线行情-时间为序  weekly_20111223 51_20111223 返回数据 row 行数 = "+str(weekly_20111223.shape[0]))
weekly_20111223_addname_dataframe.to_excel(weekly_2011_excel_writer,'51_20111223',index=False)
weekly_2011_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20111230")       ###  更新 周线记录日期
weekly_20111230 = pro.weekly(trade_date='20111230', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20111230_tscode_list = list() 
for ts_code_sh in weekly_20111230['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20111230_tscode_list.append(stock_name)
weekly_20111230_addname_dataframe=pd.DataFrame()
weekly_20111230_addname_dataframe['cname'] = weekly_20111230_tscode_list
for table_name in weekly_20111230.columns.values.tolist():
    weekly_20111230_addname_dataframe[table_name] = weekly_20111230[table_name]
print("周线行情-时间为序  weekly_20111230 52_20111230 返回数据 row 行数 = "+str(weekly_20111230.shape[0]))
weekly_20111230_addname_dataframe.to_excel(weekly_2011_excel_writer,'52_20111230',index=False)
weekly_2011_excel_writer.save()
createexcel('weekly_2012.xlsx')
weekly_2012_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2012.xlsx')
weekly_2012_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2012.xlsx', engine='openpyxl')
weekly_2012_excel_writer.book = weekly_2012_book
weekly_2012_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2012_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20120106")       ###  更新 周线记录日期
weekly_20120106 = pro.weekly(trade_date='20120106', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120106_tscode_list = list() 
for ts_code_sh in weekly_20120106['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120106_tscode_list.append(stock_name)
weekly_20120106_addname_dataframe=pd.DataFrame()
weekly_20120106_addname_dataframe['cname'] = weekly_20120106_tscode_list
for table_name in weekly_20120106.columns.values.tolist():
    weekly_20120106_addname_dataframe[table_name] = weekly_20120106[table_name]
print("周线行情-时间为序  weekly_20120106 1_20120106 返回数据 row 行数 = "+str(weekly_20120106.shape[0]))
weekly_20120106_addname_dataframe.to_excel(weekly_2012_excel_writer,'1_20120106',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120113")       ###  更新 周线记录日期
weekly_20120113 = pro.weekly(trade_date='20120113', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120113_tscode_list = list() 
for ts_code_sh in weekly_20120113['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120113_tscode_list.append(stock_name)
weekly_20120113_addname_dataframe=pd.DataFrame()
weekly_20120113_addname_dataframe['cname'] = weekly_20120113_tscode_list
for table_name in weekly_20120113.columns.values.tolist():
    weekly_20120113_addname_dataframe[table_name] = weekly_20120113[table_name]
print("周线行情-时间为序  weekly_20120113 2_20120113 返回数据 row 行数 = "+str(weekly_20120113.shape[0]))
weekly_20120113_addname_dataframe.to_excel(weekly_2012_excel_writer,'2_20120113',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120120")       ###  更新 周线记录日期
weekly_20120120 = pro.weekly(trade_date='20120120', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120120_tscode_list = list() 
for ts_code_sh in weekly_20120120['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120120_tscode_list.append(stock_name)
weekly_20120120_addname_dataframe=pd.DataFrame()
weekly_20120120_addname_dataframe['cname'] = weekly_20120120_tscode_list
for table_name in weekly_20120120.columns.values.tolist():
    weekly_20120120_addname_dataframe[table_name] = weekly_20120120[table_name]
print("周线行情-时间为序  weekly_20120120 3_20120120 返回数据 row 行数 = "+str(weekly_20120120.shape[0]))
weekly_20120120_addname_dataframe.to_excel(weekly_2012_excel_writer,'3_20120120',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120203")       ###  更新 周线记录日期
weekly_20120203 = pro.weekly(trade_date='20120203', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120203_tscode_list = list() 
for ts_code_sh in weekly_20120203['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120203_tscode_list.append(stock_name)
weekly_20120203_addname_dataframe=pd.DataFrame()
weekly_20120203_addname_dataframe['cname'] = weekly_20120203_tscode_list
for table_name in weekly_20120203.columns.values.tolist():
    weekly_20120203_addname_dataframe[table_name] = weekly_20120203[table_name]
print("周线行情-时间为序  weekly_20120203 5_20120203 返回数据 row 行数 = "+str(weekly_20120203.shape[0]))
weekly_20120203_addname_dataframe.to_excel(weekly_2012_excel_writer,'5_20120203',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120210")       ###  更新 周线记录日期
weekly_20120210 = pro.weekly(trade_date='20120210', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120210_tscode_list = list() 
for ts_code_sh in weekly_20120210['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120210_tscode_list.append(stock_name)
weekly_20120210_addname_dataframe=pd.DataFrame()
weekly_20120210_addname_dataframe['cname'] = weekly_20120210_tscode_list
for table_name in weekly_20120210.columns.values.tolist():
    weekly_20120210_addname_dataframe[table_name] = weekly_20120210[table_name]
print("周线行情-时间为序  weekly_20120210 6_20120210 返回数据 row 行数 = "+str(weekly_20120210.shape[0]))
weekly_20120210_addname_dataframe.to_excel(weekly_2012_excel_writer,'6_20120210',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120217")       ###  更新 周线记录日期
weekly_20120217 = pro.weekly(trade_date='20120217', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120217_tscode_list = list() 
for ts_code_sh in weekly_20120217['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120217_tscode_list.append(stock_name)
weekly_20120217_addname_dataframe=pd.DataFrame()
weekly_20120217_addname_dataframe['cname'] = weekly_20120217_tscode_list
for table_name in weekly_20120217.columns.values.tolist():
    weekly_20120217_addname_dataframe[table_name] = weekly_20120217[table_name]
print("周线行情-时间为序  weekly_20120217 7_20120217 返回数据 row 行数 = "+str(weekly_20120217.shape[0]))
weekly_20120217_addname_dataframe.to_excel(weekly_2012_excel_writer,'7_20120217',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120224")       ###  更新 周线记录日期
weekly_20120224 = pro.weekly(trade_date='20120224', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120224_tscode_list = list() 
for ts_code_sh in weekly_20120224['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120224_tscode_list.append(stock_name)
weekly_20120224_addname_dataframe=pd.DataFrame()
weekly_20120224_addname_dataframe['cname'] = weekly_20120224_tscode_list
for table_name in weekly_20120224.columns.values.tolist():
    weekly_20120224_addname_dataframe[table_name] = weekly_20120224[table_name]
print("周线行情-时间为序  weekly_20120224 8_20120224 返回数据 row 行数 = "+str(weekly_20120224.shape[0]))
weekly_20120224_addname_dataframe.to_excel(weekly_2012_excel_writer,'8_20120224',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120302")       ###  更新 周线记录日期
weekly_20120302 = pro.weekly(trade_date='20120302', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120302_tscode_list = list() 
for ts_code_sh in weekly_20120302['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120302_tscode_list.append(stock_name)
weekly_20120302_addname_dataframe=pd.DataFrame()
weekly_20120302_addname_dataframe['cname'] = weekly_20120302_tscode_list
for table_name in weekly_20120302.columns.values.tolist():
    weekly_20120302_addname_dataframe[table_name] = weekly_20120302[table_name]
print("周线行情-时间为序  weekly_20120302 9_20120302 返回数据 row 行数 = "+str(weekly_20120302.shape[0]))
weekly_20120302_addname_dataframe.to_excel(weekly_2012_excel_writer,'9_20120302',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120309")       ###  更新 周线记录日期
weekly_20120309 = pro.weekly(trade_date='20120309', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120309_tscode_list = list() 
for ts_code_sh in weekly_20120309['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120309_tscode_list.append(stock_name)
weekly_20120309_addname_dataframe=pd.DataFrame()
weekly_20120309_addname_dataframe['cname'] = weekly_20120309_tscode_list
for table_name in weekly_20120309.columns.values.tolist():
    weekly_20120309_addname_dataframe[table_name] = weekly_20120309[table_name]
print("周线行情-时间为序  weekly_20120309 10_20120309 返回数据 row 行数 = "+str(weekly_20120309.shape[0]))
weekly_20120309_addname_dataframe.to_excel(weekly_2012_excel_writer,'10_20120309',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120316")       ###  更新 周线记录日期
weekly_20120316 = pro.weekly(trade_date='20120316', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120316_tscode_list = list() 
for ts_code_sh in weekly_20120316['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120316_tscode_list.append(stock_name)
weekly_20120316_addname_dataframe=pd.DataFrame()
weekly_20120316_addname_dataframe['cname'] = weekly_20120316_tscode_list
for table_name in weekly_20120316.columns.values.tolist():
    weekly_20120316_addname_dataframe[table_name] = weekly_20120316[table_name]
print("周线行情-时间为序  weekly_20120316 11_20120316 返回数据 row 行数 = "+str(weekly_20120316.shape[0]))
weekly_20120316_addname_dataframe.to_excel(weekly_2012_excel_writer,'11_20120316',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120323")       ###  更新 周线记录日期
weekly_20120323 = pro.weekly(trade_date='20120323', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120323_tscode_list = list() 
for ts_code_sh in weekly_20120323['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120323_tscode_list.append(stock_name)
weekly_20120323_addname_dataframe=pd.DataFrame()
weekly_20120323_addname_dataframe['cname'] = weekly_20120323_tscode_list
for table_name in weekly_20120323.columns.values.tolist():
    weekly_20120323_addname_dataframe[table_name] = weekly_20120323[table_name]
print("周线行情-时间为序  weekly_20120323 12_20120323 返回数据 row 行数 = "+str(weekly_20120323.shape[0]))
weekly_20120323_addname_dataframe.to_excel(weekly_2012_excel_writer,'12_20120323',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120330")       ###  更新 周线记录日期
weekly_20120330 = pro.weekly(trade_date='20120330', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120330_tscode_list = list() 
for ts_code_sh in weekly_20120330['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120330_tscode_list.append(stock_name)
weekly_20120330_addname_dataframe=pd.DataFrame()
weekly_20120330_addname_dataframe['cname'] = weekly_20120330_tscode_list
for table_name in weekly_20120330.columns.values.tolist():
    weekly_20120330_addname_dataframe[table_name] = weekly_20120330[table_name]
print("周线行情-时间为序  weekly_20120330 13_20120330 返回数据 row 行数 = "+str(weekly_20120330.shape[0]))
weekly_20120330_addname_dataframe.to_excel(weekly_2012_excel_writer,'13_20120330',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120406")       ###  更新 周线记录日期
weekly_20120406 = pro.weekly(trade_date='20120406', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120406_tscode_list = list() 
for ts_code_sh in weekly_20120406['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120406_tscode_list.append(stock_name)
weekly_20120406_addname_dataframe=pd.DataFrame()
weekly_20120406_addname_dataframe['cname'] = weekly_20120406_tscode_list
for table_name in weekly_20120406.columns.values.tolist():
    weekly_20120406_addname_dataframe[table_name] = weekly_20120406[table_name]
print("周线行情-时间为序  weekly_20120406 14_20120406 返回数据 row 行数 = "+str(weekly_20120406.shape[0]))
weekly_20120406_addname_dataframe.to_excel(weekly_2012_excel_writer,'14_20120406',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120413")       ###  更新 周线记录日期
weekly_20120413 = pro.weekly(trade_date='20120413', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120413_tscode_list = list() 
for ts_code_sh in weekly_20120413['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120413_tscode_list.append(stock_name)
weekly_20120413_addname_dataframe=pd.DataFrame()
weekly_20120413_addname_dataframe['cname'] = weekly_20120413_tscode_list
for table_name in weekly_20120413.columns.values.tolist():
    weekly_20120413_addname_dataframe[table_name] = weekly_20120413[table_name]
print("周线行情-时间为序  weekly_20120413 15_20120413 返回数据 row 行数 = "+str(weekly_20120413.shape[0]))
weekly_20120413_addname_dataframe.to_excel(weekly_2012_excel_writer,'15_20120413',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120420")       ###  更新 周线记录日期
weekly_20120420 = pro.weekly(trade_date='20120420', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120420_tscode_list = list() 
for ts_code_sh in weekly_20120420['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120420_tscode_list.append(stock_name)
weekly_20120420_addname_dataframe=pd.DataFrame()
weekly_20120420_addname_dataframe['cname'] = weekly_20120420_tscode_list
for table_name in weekly_20120420.columns.values.tolist():
    weekly_20120420_addname_dataframe[table_name] = weekly_20120420[table_name]
print("周线行情-时间为序  weekly_20120420 16_20120420 返回数据 row 行数 = "+str(weekly_20120420.shape[0]))
weekly_20120420_addname_dataframe.to_excel(weekly_2012_excel_writer,'16_20120420',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120427")       ###  更新 周线记录日期
weekly_20120427 = pro.weekly(trade_date='20120427', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120427_tscode_list = list() 
for ts_code_sh in weekly_20120427['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120427_tscode_list.append(stock_name)
weekly_20120427_addname_dataframe=pd.DataFrame()
weekly_20120427_addname_dataframe['cname'] = weekly_20120427_tscode_list
for table_name in weekly_20120427.columns.values.tolist():
    weekly_20120427_addname_dataframe[table_name] = weekly_20120427[table_name]
print("周线行情-时间为序  weekly_20120427 17_20120427 返回数据 row 行数 = "+str(weekly_20120427.shape[0]))
weekly_20120427_addname_dataframe.to_excel(weekly_2012_excel_writer,'17_20120427',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120504")       ###  更新 周线记录日期
weekly_20120504 = pro.weekly(trade_date='20120504', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120504_tscode_list = list() 
for ts_code_sh in weekly_20120504['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120504_tscode_list.append(stock_name)
weekly_20120504_addname_dataframe=pd.DataFrame()
weekly_20120504_addname_dataframe['cname'] = weekly_20120504_tscode_list
for table_name in weekly_20120504.columns.values.tolist():
    weekly_20120504_addname_dataframe[table_name] = weekly_20120504[table_name]
print("周线行情-时间为序  weekly_20120504 18_20120504 返回数据 row 行数 = "+str(weekly_20120504.shape[0]))
weekly_20120504_addname_dataframe.to_excel(weekly_2012_excel_writer,'18_20120504',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120511")       ###  更新 周线记录日期
weekly_20120511 = pro.weekly(trade_date='20120511', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120511_tscode_list = list() 
for ts_code_sh in weekly_20120511['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120511_tscode_list.append(stock_name)
weekly_20120511_addname_dataframe=pd.DataFrame()
weekly_20120511_addname_dataframe['cname'] = weekly_20120511_tscode_list
for table_name in weekly_20120511.columns.values.tolist():
    weekly_20120511_addname_dataframe[table_name] = weekly_20120511[table_name]
print("周线行情-时间为序  weekly_20120511 19_20120511 返回数据 row 行数 = "+str(weekly_20120511.shape[0]))
weekly_20120511_addname_dataframe.to_excel(weekly_2012_excel_writer,'19_20120511',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120518")       ###  更新 周线记录日期
weekly_20120518 = pro.weekly(trade_date='20120518', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120518_tscode_list = list() 
for ts_code_sh in weekly_20120518['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120518_tscode_list.append(stock_name)
weekly_20120518_addname_dataframe=pd.DataFrame()
weekly_20120518_addname_dataframe['cname'] = weekly_20120518_tscode_list
for table_name in weekly_20120518.columns.values.tolist():
    weekly_20120518_addname_dataframe[table_name] = weekly_20120518[table_name]
print("周线行情-时间为序  weekly_20120518 20_20120518 返回数据 row 行数 = "+str(weekly_20120518.shape[0]))
weekly_20120518_addname_dataframe.to_excel(weekly_2012_excel_writer,'20_20120518',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120525")       ###  更新 周线记录日期
weekly_20120525 = pro.weekly(trade_date='20120525', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120525_tscode_list = list() 
for ts_code_sh in weekly_20120525['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120525_tscode_list.append(stock_name)
weekly_20120525_addname_dataframe=pd.DataFrame()
weekly_20120525_addname_dataframe['cname'] = weekly_20120525_tscode_list
for table_name in weekly_20120525.columns.values.tolist():
    weekly_20120525_addname_dataframe[table_name] = weekly_20120525[table_name]
print("周线行情-时间为序  weekly_20120525 21_20120525 返回数据 row 行数 = "+str(weekly_20120525.shape[0]))
weekly_20120525_addname_dataframe.to_excel(weekly_2012_excel_writer,'21_20120525',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120601")       ###  更新 周线记录日期
weekly_20120601 = pro.weekly(trade_date='20120601', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120601_tscode_list = list() 
for ts_code_sh in weekly_20120601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120601_tscode_list.append(stock_name)
weekly_20120601_addname_dataframe=pd.DataFrame()
weekly_20120601_addname_dataframe['cname'] = weekly_20120601_tscode_list
for table_name in weekly_20120601.columns.values.tolist():
    weekly_20120601_addname_dataframe[table_name] = weekly_20120601[table_name]
print("周线行情-时间为序  weekly_20120601 22_20120601 返回数据 row 行数 = "+str(weekly_20120601.shape[0]))
weekly_20120601_addname_dataframe.to_excel(weekly_2012_excel_writer,'22_20120601',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120608")       ###  更新 周线记录日期
weekly_20120608 = pro.weekly(trade_date='20120608', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120608_tscode_list = list() 
for ts_code_sh in weekly_20120608['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120608_tscode_list.append(stock_name)
weekly_20120608_addname_dataframe=pd.DataFrame()
weekly_20120608_addname_dataframe['cname'] = weekly_20120608_tscode_list
for table_name in weekly_20120608.columns.values.tolist():
    weekly_20120608_addname_dataframe[table_name] = weekly_20120608[table_name]
print("周线行情-时间为序  weekly_20120608 23_20120608 返回数据 row 行数 = "+str(weekly_20120608.shape[0]))
weekly_20120608_addname_dataframe.to_excel(weekly_2012_excel_writer,'23_20120608',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120615")       ###  更新 周线记录日期
weekly_20120615 = pro.weekly(trade_date='20120615', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120615_tscode_list = list() 
for ts_code_sh in weekly_20120615['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120615_tscode_list.append(stock_name)
weekly_20120615_addname_dataframe=pd.DataFrame()
weekly_20120615_addname_dataframe['cname'] = weekly_20120615_tscode_list
for table_name in weekly_20120615.columns.values.tolist():
    weekly_20120615_addname_dataframe[table_name] = weekly_20120615[table_name]
print("周线行情-时间为序  weekly_20120615 24_20120615 返回数据 row 行数 = "+str(weekly_20120615.shape[0]))
weekly_20120615_addname_dataframe.to_excel(weekly_2012_excel_writer,'24_20120615',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120621")       ###  更新 周线记录日期
weekly_20120621 = pro.weekly(trade_date='20120621', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120621_tscode_list = list() 
for ts_code_sh in weekly_20120621['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120621_tscode_list.append(stock_name)
weekly_20120621_addname_dataframe=pd.DataFrame()
weekly_20120621_addname_dataframe['cname'] = weekly_20120621_tscode_list
for table_name in weekly_20120621.columns.values.tolist():
    weekly_20120621_addname_dataframe[table_name] = weekly_20120621[table_name]
print("周线行情-时间为序  weekly_20120621 25_20120621 返回数据 row 行数 = "+str(weekly_20120621.shape[0]))
weekly_20120621_addname_dataframe.to_excel(weekly_2012_excel_writer,'25_20120621',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120629")       ###  更新 周线记录日期
weekly_20120629 = pro.weekly(trade_date='20120629', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120629_tscode_list = list() 
for ts_code_sh in weekly_20120629['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120629_tscode_list.append(stock_name)
weekly_20120629_addname_dataframe=pd.DataFrame()
weekly_20120629_addname_dataframe['cname'] = weekly_20120629_tscode_list
for table_name in weekly_20120629.columns.values.tolist():
    weekly_20120629_addname_dataframe[table_name] = weekly_20120629[table_name]
print("周线行情-时间为序  weekly_20120629 26_20120629 返回数据 row 行数 = "+str(weekly_20120629.shape[0]))
weekly_20120629_addname_dataframe.to_excel(weekly_2012_excel_writer,'26_20120629',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120706")       ###  更新 周线记录日期
weekly_20120706 = pro.weekly(trade_date='20120706', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120706_tscode_list = list() 
for ts_code_sh in weekly_20120706['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120706_tscode_list.append(stock_name)
weekly_20120706_addname_dataframe=pd.DataFrame()
weekly_20120706_addname_dataframe['cname'] = weekly_20120706_tscode_list
for table_name in weekly_20120706.columns.values.tolist():
    weekly_20120706_addname_dataframe[table_name] = weekly_20120706[table_name]
print("周线行情-时间为序  weekly_20120706 27_20120706 返回数据 row 行数 = "+str(weekly_20120706.shape[0]))
weekly_20120706_addname_dataframe.to_excel(weekly_2012_excel_writer,'27_20120706',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120713")       ###  更新 周线记录日期
weekly_20120713 = pro.weekly(trade_date='20120713', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120713_tscode_list = list() 
for ts_code_sh in weekly_20120713['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120713_tscode_list.append(stock_name)
weekly_20120713_addname_dataframe=pd.DataFrame()
weekly_20120713_addname_dataframe['cname'] = weekly_20120713_tscode_list
for table_name in weekly_20120713.columns.values.tolist():
    weekly_20120713_addname_dataframe[table_name] = weekly_20120713[table_name]
print("周线行情-时间为序  weekly_20120713 28_20120713 返回数据 row 行数 = "+str(weekly_20120713.shape[0]))
weekly_20120713_addname_dataframe.to_excel(weekly_2012_excel_writer,'28_20120713',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120720")       ###  更新 周线记录日期
weekly_20120720 = pro.weekly(trade_date='20120720', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120720_tscode_list = list() 
for ts_code_sh in weekly_20120720['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120720_tscode_list.append(stock_name)
weekly_20120720_addname_dataframe=pd.DataFrame()
weekly_20120720_addname_dataframe['cname'] = weekly_20120720_tscode_list
for table_name in weekly_20120720.columns.values.tolist():
    weekly_20120720_addname_dataframe[table_name] = weekly_20120720[table_name]
print("周线行情-时间为序  weekly_20120720 29_20120720 返回数据 row 行数 = "+str(weekly_20120720.shape[0]))
weekly_20120720_addname_dataframe.to_excel(weekly_2012_excel_writer,'29_20120720',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120727")       ###  更新 周线记录日期
weekly_20120727 = pro.weekly(trade_date='20120727', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120727_tscode_list = list() 
for ts_code_sh in weekly_20120727['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120727_tscode_list.append(stock_name)
weekly_20120727_addname_dataframe=pd.DataFrame()
weekly_20120727_addname_dataframe['cname'] = weekly_20120727_tscode_list
for table_name in weekly_20120727.columns.values.tolist():
    weekly_20120727_addname_dataframe[table_name] = weekly_20120727[table_name]
print("周线行情-时间为序  weekly_20120727 30_20120727 返回数据 row 行数 = "+str(weekly_20120727.shape[0]))
weekly_20120727_addname_dataframe.to_excel(weekly_2012_excel_writer,'30_20120727',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120803")       ###  更新 周线记录日期
weekly_20120803 = pro.weekly(trade_date='20120803', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120803_tscode_list = list() 
for ts_code_sh in weekly_20120803['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120803_tscode_list.append(stock_name)
weekly_20120803_addname_dataframe=pd.DataFrame()
weekly_20120803_addname_dataframe['cname'] = weekly_20120803_tscode_list
for table_name in weekly_20120803.columns.values.tolist():
    weekly_20120803_addname_dataframe[table_name] = weekly_20120803[table_name]
print("周线行情-时间为序  weekly_20120803 31_20120803 返回数据 row 行数 = "+str(weekly_20120803.shape[0]))
weekly_20120803_addname_dataframe.to_excel(weekly_2012_excel_writer,'31_20120803',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120810")       ###  更新 周线记录日期
weekly_20120810 = pro.weekly(trade_date='20120810', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120810_tscode_list = list() 
for ts_code_sh in weekly_20120810['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120810_tscode_list.append(stock_name)
weekly_20120810_addname_dataframe=pd.DataFrame()
weekly_20120810_addname_dataframe['cname'] = weekly_20120810_tscode_list
for table_name in weekly_20120810.columns.values.tolist():
    weekly_20120810_addname_dataframe[table_name] = weekly_20120810[table_name]
print("周线行情-时间为序  weekly_20120810 32_20120810 返回数据 row 行数 = "+str(weekly_20120810.shape[0]))
weekly_20120810_addname_dataframe.to_excel(weekly_2012_excel_writer,'32_20120810',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120817")       ###  更新 周线记录日期
weekly_20120817 = pro.weekly(trade_date='20120817', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120817_tscode_list = list() 
for ts_code_sh in weekly_20120817['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120817_tscode_list.append(stock_name)
weekly_20120817_addname_dataframe=pd.DataFrame()
weekly_20120817_addname_dataframe['cname'] = weekly_20120817_tscode_list
for table_name in weekly_20120817.columns.values.tolist():
    weekly_20120817_addname_dataframe[table_name] = weekly_20120817[table_name]
print("周线行情-时间为序  weekly_20120817 33_20120817 返回数据 row 行数 = "+str(weekly_20120817.shape[0]))
weekly_20120817_addname_dataframe.to_excel(weekly_2012_excel_writer,'33_20120817',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120824")       ###  更新 周线记录日期
weekly_20120824 = pro.weekly(trade_date='20120824', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120824_tscode_list = list() 
for ts_code_sh in weekly_20120824['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120824_tscode_list.append(stock_name)
weekly_20120824_addname_dataframe=pd.DataFrame()
weekly_20120824_addname_dataframe['cname'] = weekly_20120824_tscode_list
for table_name in weekly_20120824.columns.values.tolist():
    weekly_20120824_addname_dataframe[table_name] = weekly_20120824[table_name]
print("周线行情-时间为序  weekly_20120824 34_20120824 返回数据 row 行数 = "+str(weekly_20120824.shape[0]))
weekly_20120824_addname_dataframe.to_excel(weekly_2012_excel_writer,'34_20120824',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120831")       ###  更新 周线记录日期
weekly_20120831 = pro.weekly(trade_date='20120831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120831_tscode_list = list() 
for ts_code_sh in weekly_20120831['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120831_tscode_list.append(stock_name)
weekly_20120831_addname_dataframe=pd.DataFrame()
weekly_20120831_addname_dataframe['cname'] = weekly_20120831_tscode_list
for table_name in weekly_20120831.columns.values.tolist():
    weekly_20120831_addname_dataframe[table_name] = weekly_20120831[table_name]
print("周线行情-时间为序  weekly_20120831 35_20120831 返回数据 row 行数 = "+str(weekly_20120831.shape[0]))
weekly_20120831_addname_dataframe.to_excel(weekly_2012_excel_writer,'35_20120831',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120907")       ###  更新 周线记录日期
weekly_20120907 = pro.weekly(trade_date='20120907', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120907_tscode_list = list() 
for ts_code_sh in weekly_20120907['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120907_tscode_list.append(stock_name)
weekly_20120907_addname_dataframe=pd.DataFrame()
weekly_20120907_addname_dataframe['cname'] = weekly_20120907_tscode_list
for table_name in weekly_20120907.columns.values.tolist():
    weekly_20120907_addname_dataframe[table_name] = weekly_20120907[table_name]
print("周线行情-时间为序  weekly_20120907 36_20120907 返回数据 row 行数 = "+str(weekly_20120907.shape[0]))
weekly_20120907_addname_dataframe.to_excel(weekly_2012_excel_writer,'36_20120907',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120914")       ###  更新 周线记录日期
weekly_20120914 = pro.weekly(trade_date='20120914', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120914_tscode_list = list() 
for ts_code_sh in weekly_20120914['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120914_tscode_list.append(stock_name)
weekly_20120914_addname_dataframe=pd.DataFrame()
weekly_20120914_addname_dataframe['cname'] = weekly_20120914_tscode_list
for table_name in weekly_20120914.columns.values.tolist():
    weekly_20120914_addname_dataframe[table_name] = weekly_20120914[table_name]
print("周线行情-时间为序  weekly_20120914 37_20120914 返回数据 row 行数 = "+str(weekly_20120914.shape[0]))
weekly_20120914_addname_dataframe.to_excel(weekly_2012_excel_writer,'37_20120914',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120921")       ###  更新 周线记录日期
weekly_20120921 = pro.weekly(trade_date='20120921', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120921_tscode_list = list() 
for ts_code_sh in weekly_20120921['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120921_tscode_list.append(stock_name)
weekly_20120921_addname_dataframe=pd.DataFrame()
weekly_20120921_addname_dataframe['cname'] = weekly_20120921_tscode_list
for table_name in weekly_20120921.columns.values.tolist():
    weekly_20120921_addname_dataframe[table_name] = weekly_20120921[table_name]
print("周线行情-时间为序  weekly_20120921 38_20120921 返回数据 row 行数 = "+str(weekly_20120921.shape[0]))
weekly_20120921_addname_dataframe.to_excel(weekly_2012_excel_writer,'38_20120921',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20120928")       ###  更新 周线记录日期
weekly_20120928 = pro.weekly(trade_date='20120928', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20120928_tscode_list = list() 
for ts_code_sh in weekly_20120928['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20120928_tscode_list.append(stock_name)
weekly_20120928_addname_dataframe=pd.DataFrame()
weekly_20120928_addname_dataframe['cname'] = weekly_20120928_tscode_list
for table_name in weekly_20120928.columns.values.tolist():
    weekly_20120928_addname_dataframe[table_name] = weekly_20120928[table_name]
print("周线行情-时间为序  weekly_20120928 39_20120928 返回数据 row 行数 = "+str(weekly_20120928.shape[0]))
weekly_20120928_addname_dataframe.to_excel(weekly_2012_excel_writer,'39_20120928',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121012")       ###  更新 周线记录日期
weekly_20121012 = pro.weekly(trade_date='20121012', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121012_tscode_list = list() 
for ts_code_sh in weekly_20121012['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121012_tscode_list.append(stock_name)
weekly_20121012_addname_dataframe=pd.DataFrame()
weekly_20121012_addname_dataframe['cname'] = weekly_20121012_tscode_list
for table_name in weekly_20121012.columns.values.tolist():
    weekly_20121012_addname_dataframe[table_name] = weekly_20121012[table_name]
print("周线行情-时间为序  weekly_20121012 41_20121012 返回数据 row 行数 = "+str(weekly_20121012.shape[0]))
weekly_20121012_addname_dataframe.to_excel(weekly_2012_excel_writer,'41_20121012',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121019")       ###  更新 周线记录日期
weekly_20121019 = pro.weekly(trade_date='20121019', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121019_tscode_list = list() 
for ts_code_sh in weekly_20121019['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121019_tscode_list.append(stock_name)
weekly_20121019_addname_dataframe=pd.DataFrame()
weekly_20121019_addname_dataframe['cname'] = weekly_20121019_tscode_list
for table_name in weekly_20121019.columns.values.tolist():
    weekly_20121019_addname_dataframe[table_name] = weekly_20121019[table_name]
print("周线行情-时间为序  weekly_20121019 42_20121019 返回数据 row 行数 = "+str(weekly_20121019.shape[0]))
weekly_20121019_addname_dataframe.to_excel(weekly_2012_excel_writer,'42_20121019',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121026")       ###  更新 周线记录日期
weekly_20121026 = pro.weekly(trade_date='20121026', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121026_tscode_list = list() 
for ts_code_sh in weekly_20121026['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121026_tscode_list.append(stock_name)
weekly_20121026_addname_dataframe=pd.DataFrame()
weekly_20121026_addname_dataframe['cname'] = weekly_20121026_tscode_list
for table_name in weekly_20121026.columns.values.tolist():
    weekly_20121026_addname_dataframe[table_name] = weekly_20121026[table_name]
print("周线行情-时间为序  weekly_20121026 43_20121026 返回数据 row 行数 = "+str(weekly_20121026.shape[0]))
weekly_20121026_addname_dataframe.to_excel(weekly_2012_excel_writer,'43_20121026',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121102")       ###  更新 周线记录日期
weekly_20121102 = pro.weekly(trade_date='20121102', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121102_tscode_list = list() 
for ts_code_sh in weekly_20121102['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121102_tscode_list.append(stock_name)
weekly_20121102_addname_dataframe=pd.DataFrame()
weekly_20121102_addname_dataframe['cname'] = weekly_20121102_tscode_list
for table_name in weekly_20121102.columns.values.tolist():
    weekly_20121102_addname_dataframe[table_name] = weekly_20121102[table_name]
print("周线行情-时间为序  weekly_20121102 44_20121102 返回数据 row 行数 = "+str(weekly_20121102.shape[0]))
weekly_20121102_addname_dataframe.to_excel(weekly_2012_excel_writer,'44_20121102',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121109")       ###  更新 周线记录日期
weekly_20121109 = pro.weekly(trade_date='20121109', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121109_tscode_list = list() 
for ts_code_sh in weekly_20121109['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121109_tscode_list.append(stock_name)
weekly_20121109_addname_dataframe=pd.DataFrame()
weekly_20121109_addname_dataframe['cname'] = weekly_20121109_tscode_list
for table_name in weekly_20121109.columns.values.tolist():
    weekly_20121109_addname_dataframe[table_name] = weekly_20121109[table_name]
print("周线行情-时间为序  weekly_20121109 45_20121109 返回数据 row 行数 = "+str(weekly_20121109.shape[0]))
weekly_20121109_addname_dataframe.to_excel(weekly_2012_excel_writer,'45_20121109',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121116")       ###  更新 周线记录日期
weekly_20121116 = pro.weekly(trade_date='20121116', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121116_tscode_list = list() 
for ts_code_sh in weekly_20121116['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121116_tscode_list.append(stock_name)
weekly_20121116_addname_dataframe=pd.DataFrame()
weekly_20121116_addname_dataframe['cname'] = weekly_20121116_tscode_list
for table_name in weekly_20121116.columns.values.tolist():
    weekly_20121116_addname_dataframe[table_name] = weekly_20121116[table_name]
print("周线行情-时间为序  weekly_20121116 46_20121116 返回数据 row 行数 = "+str(weekly_20121116.shape[0]))
weekly_20121116_addname_dataframe.to_excel(weekly_2012_excel_writer,'46_20121116',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121123")       ###  更新 周线记录日期
weekly_20121123 = pro.weekly(trade_date='20121123', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121123_tscode_list = list() 
for ts_code_sh in weekly_20121123['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121123_tscode_list.append(stock_name)
weekly_20121123_addname_dataframe=pd.DataFrame()
weekly_20121123_addname_dataframe['cname'] = weekly_20121123_tscode_list
for table_name in weekly_20121123.columns.values.tolist():
    weekly_20121123_addname_dataframe[table_name] = weekly_20121123[table_name]
print("周线行情-时间为序  weekly_20121123 47_20121123 返回数据 row 行数 = "+str(weekly_20121123.shape[0]))
weekly_20121123_addname_dataframe.to_excel(weekly_2012_excel_writer,'47_20121123',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121130")       ###  更新 周线记录日期
weekly_20121130 = pro.weekly(trade_date='20121130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121130_tscode_list = list() 
for ts_code_sh in weekly_20121130['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121130_tscode_list.append(stock_name)
weekly_20121130_addname_dataframe=pd.DataFrame()
weekly_20121130_addname_dataframe['cname'] = weekly_20121130_tscode_list
for table_name in weekly_20121130.columns.values.tolist():
    weekly_20121130_addname_dataframe[table_name] = weekly_20121130[table_name]
print("周线行情-时间为序  weekly_20121130 48_20121130 返回数据 row 行数 = "+str(weekly_20121130.shape[0]))
weekly_20121130_addname_dataframe.to_excel(weekly_2012_excel_writer,'48_20121130',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121207")       ###  更新 周线记录日期
weekly_20121207 = pro.weekly(trade_date='20121207', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121207_tscode_list = list() 
for ts_code_sh in weekly_20121207['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121207_tscode_list.append(stock_name)
weekly_20121207_addname_dataframe=pd.DataFrame()
weekly_20121207_addname_dataframe['cname'] = weekly_20121207_tscode_list
for table_name in weekly_20121207.columns.values.tolist():
    weekly_20121207_addname_dataframe[table_name] = weekly_20121207[table_name]
print("周线行情-时间为序  weekly_20121207 49_20121207 返回数据 row 行数 = "+str(weekly_20121207.shape[0]))
weekly_20121207_addname_dataframe.to_excel(weekly_2012_excel_writer,'49_20121207',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121214")       ###  更新 周线记录日期
weekly_20121214 = pro.weekly(trade_date='20121214', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121214_tscode_list = list() 
for ts_code_sh in weekly_20121214['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121214_tscode_list.append(stock_name)
weekly_20121214_addname_dataframe=pd.DataFrame()
weekly_20121214_addname_dataframe['cname'] = weekly_20121214_tscode_list
for table_name in weekly_20121214.columns.values.tolist():
    weekly_20121214_addname_dataframe[table_name] = weekly_20121214[table_name]
print("周线行情-时间为序  weekly_20121214 50_20121214 返回数据 row 行数 = "+str(weekly_20121214.shape[0]))
weekly_20121214_addname_dataframe.to_excel(weekly_2012_excel_writer,'50_20121214',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121221")       ###  更新 周线记录日期
weekly_20121221 = pro.weekly(trade_date='20121221', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121221_tscode_list = list() 
for ts_code_sh in weekly_20121221['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121221_tscode_list.append(stock_name)
weekly_20121221_addname_dataframe=pd.DataFrame()
weekly_20121221_addname_dataframe['cname'] = weekly_20121221_tscode_list
for table_name in weekly_20121221.columns.values.tolist():
    weekly_20121221_addname_dataframe[table_name] = weekly_20121221[table_name]
print("周线行情-时间为序  weekly_20121221 51_20121221 返回数据 row 行数 = "+str(weekly_20121221.shape[0]))
weekly_20121221_addname_dataframe.to_excel(weekly_2012_excel_writer,'51_20121221',index=False)
weekly_2012_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20121228")       ###  更新 周线记录日期
weekly_20121228 = pro.weekly(trade_date='20121228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20121228_tscode_list = list() 
for ts_code_sh in weekly_20121228['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20121228_tscode_list.append(stock_name)
weekly_20121228_addname_dataframe=pd.DataFrame()
weekly_20121228_addname_dataframe['cname'] = weekly_20121228_tscode_list
for table_name in weekly_20121228.columns.values.tolist():
    weekly_20121228_addname_dataframe[table_name] = weekly_20121228[table_name]
print("周线行情-时间为序  weekly_20121228 52_20121228 返回数据 row 行数 = "+str(weekly_20121228.shape[0]))
weekly_20121228_addname_dataframe.to_excel(weekly_2012_excel_writer,'52_20121228',index=False)
weekly_2012_excel_writer.save()
createexcel('weekly_2013.xlsx')
weekly_2013_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2013.xlsx')
weekly_2013_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2013.xlsx', engine='openpyxl')
weekly_2013_excel_writer.book = weekly_2013_book
weekly_2013_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2013_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20130104")       ###  更新 周线记录日期
weekly_20130104 = pro.weekly(trade_date='20130104', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130104_tscode_list = list() 
for ts_code_sh in weekly_20130104['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130104_tscode_list.append(stock_name)
weekly_20130104_addname_dataframe=pd.DataFrame()
weekly_20130104_addname_dataframe['cname'] = weekly_20130104_tscode_list
for table_name in weekly_20130104.columns.values.tolist():
    weekly_20130104_addname_dataframe[table_name] = weekly_20130104[table_name]
print("周线行情-时间为序  weekly_20130104 1_20130104 返回数据 row 行数 = "+str(weekly_20130104.shape[0]))
weekly_20130104_addname_dataframe.to_excel(weekly_2013_excel_writer,'1_20130104',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130111")       ###  更新 周线记录日期
weekly_20130111 = pro.weekly(trade_date='20130111', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130111_tscode_list = list() 
for ts_code_sh in weekly_20130111['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130111_tscode_list.append(stock_name)
weekly_20130111_addname_dataframe=pd.DataFrame()
weekly_20130111_addname_dataframe['cname'] = weekly_20130111_tscode_list
for table_name in weekly_20130111.columns.values.tolist():
    weekly_20130111_addname_dataframe[table_name] = weekly_20130111[table_name]
print("周线行情-时间为序  weekly_20130111 2_20130111 返回数据 row 行数 = "+str(weekly_20130111.shape[0]))
weekly_20130111_addname_dataframe.to_excel(weekly_2013_excel_writer,'2_20130111',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130118")       ###  更新 周线记录日期
weekly_20130118 = pro.weekly(trade_date='20130118', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130118_tscode_list = list() 
for ts_code_sh in weekly_20130118['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130118_tscode_list.append(stock_name)
weekly_20130118_addname_dataframe=pd.DataFrame()
weekly_20130118_addname_dataframe['cname'] = weekly_20130118_tscode_list
for table_name in weekly_20130118.columns.values.tolist():
    weekly_20130118_addname_dataframe[table_name] = weekly_20130118[table_name]
print("周线行情-时间为序  weekly_20130118 3_20130118 返回数据 row 行数 = "+str(weekly_20130118.shape[0]))
weekly_20130118_addname_dataframe.to_excel(weekly_2013_excel_writer,'3_20130118',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130125")       ###  更新 周线记录日期
weekly_20130125 = pro.weekly(trade_date='20130125', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130125_tscode_list = list() 
for ts_code_sh in weekly_20130125['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130125_tscode_list.append(stock_name)
weekly_20130125_addname_dataframe=pd.DataFrame()
weekly_20130125_addname_dataframe['cname'] = weekly_20130125_tscode_list
for table_name in weekly_20130125.columns.values.tolist():
    weekly_20130125_addname_dataframe[table_name] = weekly_20130125[table_name]
print("周线行情-时间为序  weekly_20130125 4_20130125 返回数据 row 行数 = "+str(weekly_20130125.shape[0]))
weekly_20130125_addname_dataframe.to_excel(weekly_2013_excel_writer,'4_20130125',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130201")       ###  更新 周线记录日期
weekly_20130201 = pro.weekly(trade_date='20130201', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130201_tscode_list = list() 
for ts_code_sh in weekly_20130201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130201_tscode_list.append(stock_name)
weekly_20130201_addname_dataframe=pd.DataFrame()
weekly_20130201_addname_dataframe['cname'] = weekly_20130201_tscode_list
for table_name in weekly_20130201.columns.values.tolist():
    weekly_20130201_addname_dataframe[table_name] = weekly_20130201[table_name]
print("周线行情-时间为序  weekly_20130201 5_20130201 返回数据 row 行数 = "+str(weekly_20130201.shape[0]))
weekly_20130201_addname_dataframe.to_excel(weekly_2013_excel_writer,'5_20130201',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130208")       ###  更新 周线记录日期
weekly_20130208 = pro.weekly(trade_date='20130208', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130208_tscode_list = list() 
for ts_code_sh in weekly_20130208['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130208_tscode_list.append(stock_name)
weekly_20130208_addname_dataframe=pd.DataFrame()
weekly_20130208_addname_dataframe['cname'] = weekly_20130208_tscode_list
for table_name in weekly_20130208.columns.values.tolist():
    weekly_20130208_addname_dataframe[table_name] = weekly_20130208[table_name]
print("周线行情-时间为序  weekly_20130208 6_20130208 返回数据 row 行数 = "+str(weekly_20130208.shape[0]))
weekly_20130208_addname_dataframe.to_excel(weekly_2013_excel_writer,'6_20130208',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130222")       ###  更新 周线记录日期
weekly_20130222 = pro.weekly(trade_date='20130222', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130222_tscode_list = list() 
for ts_code_sh in weekly_20130222['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130222_tscode_list.append(stock_name)
weekly_20130222_addname_dataframe=pd.DataFrame()
weekly_20130222_addname_dataframe['cname'] = weekly_20130222_tscode_list
for table_name in weekly_20130222.columns.values.tolist():
    weekly_20130222_addname_dataframe[table_name] = weekly_20130222[table_name]
print("周线行情-时间为序  weekly_20130222 8_20130222 返回数据 row 行数 = "+str(weekly_20130222.shape[0]))
weekly_20130222_addname_dataframe.to_excel(weekly_2013_excel_writer,'8_20130222',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130301")       ###  更新 周线记录日期
weekly_20130301 = pro.weekly(trade_date='20130301', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130301_tscode_list = list() 
for ts_code_sh in weekly_20130301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130301_tscode_list.append(stock_name)
weekly_20130301_addname_dataframe=pd.DataFrame()
weekly_20130301_addname_dataframe['cname'] = weekly_20130301_tscode_list
for table_name in weekly_20130301.columns.values.tolist():
    weekly_20130301_addname_dataframe[table_name] = weekly_20130301[table_name]
print("周线行情-时间为序  weekly_20130301 9_20130301 返回数据 row 行数 = "+str(weekly_20130301.shape[0]))
weekly_20130301_addname_dataframe.to_excel(weekly_2013_excel_writer,'9_20130301',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130308")       ###  更新 周线记录日期
weekly_20130308 = pro.weekly(trade_date='20130308', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130308_tscode_list = list() 
for ts_code_sh in weekly_20130308['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130308_tscode_list.append(stock_name)
weekly_20130308_addname_dataframe=pd.DataFrame()
weekly_20130308_addname_dataframe['cname'] = weekly_20130308_tscode_list
for table_name in weekly_20130308.columns.values.tolist():
    weekly_20130308_addname_dataframe[table_name] = weekly_20130308[table_name]
print("周线行情-时间为序  weekly_20130308 10_20130308 返回数据 row 行数 = "+str(weekly_20130308.shape[0]))
weekly_20130308_addname_dataframe.to_excel(weekly_2013_excel_writer,'10_20130308',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130315")       ###  更新 周线记录日期
weekly_20130315 = pro.weekly(trade_date='20130315', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130315_tscode_list = list() 
for ts_code_sh in weekly_20130315['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130315_tscode_list.append(stock_name)
weekly_20130315_addname_dataframe=pd.DataFrame()
weekly_20130315_addname_dataframe['cname'] = weekly_20130315_tscode_list
for table_name in weekly_20130315.columns.values.tolist():
    weekly_20130315_addname_dataframe[table_name] = weekly_20130315[table_name]
print("周线行情-时间为序  weekly_20130315 11_20130315 返回数据 row 行数 = "+str(weekly_20130315.shape[0]))
weekly_20130315_addname_dataframe.to_excel(weekly_2013_excel_writer,'11_20130315',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130322")       ###  更新 周线记录日期
weekly_20130322 = pro.weekly(trade_date='20130322', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130322_tscode_list = list() 
for ts_code_sh in weekly_20130322['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130322_tscode_list.append(stock_name)
weekly_20130322_addname_dataframe=pd.DataFrame()
weekly_20130322_addname_dataframe['cname'] = weekly_20130322_tscode_list
for table_name in weekly_20130322.columns.values.tolist():
    weekly_20130322_addname_dataframe[table_name] = weekly_20130322[table_name]
print("周线行情-时间为序  weekly_20130322 12_20130322 返回数据 row 行数 = "+str(weekly_20130322.shape[0]))
weekly_20130322_addname_dataframe.to_excel(weekly_2013_excel_writer,'12_20130322',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130329")       ###  更新 周线记录日期
weekly_20130329 = pro.weekly(trade_date='20130329', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130329_tscode_list = list() 
for ts_code_sh in weekly_20130329['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130329_tscode_list.append(stock_name)
weekly_20130329_addname_dataframe=pd.DataFrame()
weekly_20130329_addname_dataframe['cname'] = weekly_20130329_tscode_list
for table_name in weekly_20130329.columns.values.tolist():
    weekly_20130329_addname_dataframe[table_name] = weekly_20130329[table_name]
print("周线行情-时间为序  weekly_20130329 13_20130329 返回数据 row 行数 = "+str(weekly_20130329.shape[0]))
weekly_20130329_addname_dataframe.to_excel(weekly_2013_excel_writer,'13_20130329',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130403")       ###  更新 周线记录日期
weekly_20130403 = pro.weekly(trade_date='20130403', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130403_tscode_list = list() 
for ts_code_sh in weekly_20130403['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130403_tscode_list.append(stock_name)
weekly_20130403_addname_dataframe=pd.DataFrame()
weekly_20130403_addname_dataframe['cname'] = weekly_20130403_tscode_list
for table_name in weekly_20130403.columns.values.tolist():
    weekly_20130403_addname_dataframe[table_name] = weekly_20130403[table_name]
print("周线行情-时间为序  weekly_20130403 14_20130403 返回数据 row 行数 = "+str(weekly_20130403.shape[0]))
weekly_20130403_addname_dataframe.to_excel(weekly_2013_excel_writer,'14_20130403',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130412")       ###  更新 周线记录日期
weekly_20130412 = pro.weekly(trade_date='20130412', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130412_tscode_list = list() 
for ts_code_sh in weekly_20130412['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130412_tscode_list.append(stock_name)
weekly_20130412_addname_dataframe=pd.DataFrame()
weekly_20130412_addname_dataframe['cname'] = weekly_20130412_tscode_list
for table_name in weekly_20130412.columns.values.tolist():
    weekly_20130412_addname_dataframe[table_name] = weekly_20130412[table_name]
print("周线行情-时间为序  weekly_20130412 15_20130412 返回数据 row 行数 = "+str(weekly_20130412.shape[0]))
weekly_20130412_addname_dataframe.to_excel(weekly_2013_excel_writer,'15_20130412',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130419")       ###  更新 周线记录日期
weekly_20130419 = pro.weekly(trade_date='20130419', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130419_tscode_list = list() 
for ts_code_sh in weekly_20130419['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130419_tscode_list.append(stock_name)
weekly_20130419_addname_dataframe=pd.DataFrame()
weekly_20130419_addname_dataframe['cname'] = weekly_20130419_tscode_list
for table_name in weekly_20130419.columns.values.tolist():
    weekly_20130419_addname_dataframe[table_name] = weekly_20130419[table_name]
print("周线行情-时间为序  weekly_20130419 16_20130419 返回数据 row 行数 = "+str(weekly_20130419.shape[0]))
weekly_20130419_addname_dataframe.to_excel(weekly_2013_excel_writer,'16_20130419',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130426")       ###  更新 周线记录日期
weekly_20130426 = pro.weekly(trade_date='20130426', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130426_tscode_list = list() 
for ts_code_sh in weekly_20130426['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130426_tscode_list.append(stock_name)
weekly_20130426_addname_dataframe=pd.DataFrame()
weekly_20130426_addname_dataframe['cname'] = weekly_20130426_tscode_list
for table_name in weekly_20130426.columns.values.tolist():
    weekly_20130426_addname_dataframe[table_name] = weekly_20130426[table_name]
print("周线行情-时间为序  weekly_20130426 17_20130426 返回数据 row 行数 = "+str(weekly_20130426.shape[0]))
weekly_20130426_addname_dataframe.to_excel(weekly_2013_excel_writer,'17_20130426',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130503")       ###  更新 周线记录日期
weekly_20130503 = pro.weekly(trade_date='20130503', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130503_tscode_list = list() 
for ts_code_sh in weekly_20130503['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130503_tscode_list.append(stock_name)
weekly_20130503_addname_dataframe=pd.DataFrame()
weekly_20130503_addname_dataframe['cname'] = weekly_20130503_tscode_list
for table_name in weekly_20130503.columns.values.tolist():
    weekly_20130503_addname_dataframe[table_name] = weekly_20130503[table_name]
print("周线行情-时间为序  weekly_20130503 18_20130503 返回数据 row 行数 = "+str(weekly_20130503.shape[0]))
weekly_20130503_addname_dataframe.to_excel(weekly_2013_excel_writer,'18_20130503',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130510")       ###  更新 周线记录日期
weekly_20130510 = pro.weekly(trade_date='20130510', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130510_tscode_list = list() 
for ts_code_sh in weekly_20130510['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130510_tscode_list.append(stock_name)
weekly_20130510_addname_dataframe=pd.DataFrame()
weekly_20130510_addname_dataframe['cname'] = weekly_20130510_tscode_list
for table_name in weekly_20130510.columns.values.tolist():
    weekly_20130510_addname_dataframe[table_name] = weekly_20130510[table_name]
print("周线行情-时间为序  weekly_20130510 19_20130510 返回数据 row 行数 = "+str(weekly_20130510.shape[0]))
weekly_20130510_addname_dataframe.to_excel(weekly_2013_excel_writer,'19_20130510',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130517")       ###  更新 周线记录日期
weekly_20130517 = pro.weekly(trade_date='20130517', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130517_tscode_list = list() 
for ts_code_sh in weekly_20130517['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130517_tscode_list.append(stock_name)
weekly_20130517_addname_dataframe=pd.DataFrame()
weekly_20130517_addname_dataframe['cname'] = weekly_20130517_tscode_list
for table_name in weekly_20130517.columns.values.tolist():
    weekly_20130517_addname_dataframe[table_name] = weekly_20130517[table_name]
print("周线行情-时间为序  weekly_20130517 20_20130517 返回数据 row 行数 = "+str(weekly_20130517.shape[0]))
weekly_20130517_addname_dataframe.to_excel(weekly_2013_excel_writer,'20_20130517',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130524")       ###  更新 周线记录日期
weekly_20130524 = pro.weekly(trade_date='20130524', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130524_tscode_list = list() 
for ts_code_sh in weekly_20130524['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130524_tscode_list.append(stock_name)
weekly_20130524_addname_dataframe=pd.DataFrame()
weekly_20130524_addname_dataframe['cname'] = weekly_20130524_tscode_list
for table_name in weekly_20130524.columns.values.tolist():
    weekly_20130524_addname_dataframe[table_name] = weekly_20130524[table_name]
print("周线行情-时间为序  weekly_20130524 21_20130524 返回数据 row 行数 = "+str(weekly_20130524.shape[0]))
weekly_20130524_addname_dataframe.to_excel(weekly_2013_excel_writer,'21_20130524',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130531")       ###  更新 周线记录日期
weekly_20130531 = pro.weekly(trade_date='20130531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130531_tscode_list = list() 
for ts_code_sh in weekly_20130531['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130531_tscode_list.append(stock_name)
weekly_20130531_addname_dataframe=pd.DataFrame()
weekly_20130531_addname_dataframe['cname'] = weekly_20130531_tscode_list
for table_name in weekly_20130531.columns.values.tolist():
    weekly_20130531_addname_dataframe[table_name] = weekly_20130531[table_name]
print("周线行情-时间为序  weekly_20130531 22_20130531 返回数据 row 行数 = "+str(weekly_20130531.shape[0]))
weekly_20130531_addname_dataframe.to_excel(weekly_2013_excel_writer,'22_20130531',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130607")       ###  更新 周线记录日期
weekly_20130607 = pro.weekly(trade_date='20130607', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130607_tscode_list = list() 
for ts_code_sh in weekly_20130607['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130607_tscode_list.append(stock_name)
weekly_20130607_addname_dataframe=pd.DataFrame()
weekly_20130607_addname_dataframe['cname'] = weekly_20130607_tscode_list
for table_name in weekly_20130607.columns.values.tolist():
    weekly_20130607_addname_dataframe[table_name] = weekly_20130607[table_name]
print("周线行情-时间为序  weekly_20130607 23_20130607 返回数据 row 行数 = "+str(weekly_20130607.shape[0]))
weekly_20130607_addname_dataframe.to_excel(weekly_2013_excel_writer,'23_20130607',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130614")       ###  更新 周线记录日期
weekly_20130614 = pro.weekly(trade_date='20130614', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130614_tscode_list = list() 
for ts_code_sh in weekly_20130614['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130614_tscode_list.append(stock_name)
weekly_20130614_addname_dataframe=pd.DataFrame()
weekly_20130614_addname_dataframe['cname'] = weekly_20130614_tscode_list
for table_name in weekly_20130614.columns.values.tolist():
    weekly_20130614_addname_dataframe[table_name] = weekly_20130614[table_name]
print("周线行情-时间为序  weekly_20130614 24_20130614 返回数据 row 行数 = "+str(weekly_20130614.shape[0]))
weekly_20130614_addname_dataframe.to_excel(weekly_2013_excel_writer,'24_20130614',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130621")       ###  更新 周线记录日期
weekly_20130621 = pro.weekly(trade_date='20130621', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130621_tscode_list = list() 
for ts_code_sh in weekly_20130621['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130621_tscode_list.append(stock_name)
weekly_20130621_addname_dataframe=pd.DataFrame()
weekly_20130621_addname_dataframe['cname'] = weekly_20130621_tscode_list
for table_name in weekly_20130621.columns.values.tolist():
    weekly_20130621_addname_dataframe[table_name] = weekly_20130621[table_name]
print("周线行情-时间为序  weekly_20130621 25_20130621 返回数据 row 行数 = "+str(weekly_20130621.shape[0]))
weekly_20130621_addname_dataframe.to_excel(weekly_2013_excel_writer,'25_20130621',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130628")       ###  更新 周线记录日期
weekly_20130628 = pro.weekly(trade_date='20130628', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130628_tscode_list = list() 
for ts_code_sh in weekly_20130628['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130628_tscode_list.append(stock_name)
weekly_20130628_addname_dataframe=pd.DataFrame()
weekly_20130628_addname_dataframe['cname'] = weekly_20130628_tscode_list
for table_name in weekly_20130628.columns.values.tolist():
    weekly_20130628_addname_dataframe[table_name] = weekly_20130628[table_name]
print("周线行情-时间为序  weekly_20130628 26_20130628 返回数据 row 行数 = "+str(weekly_20130628.shape[0]))
weekly_20130628_addname_dataframe.to_excel(weekly_2013_excel_writer,'26_20130628',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130705")       ###  更新 周线记录日期
weekly_20130705 = pro.weekly(trade_date='20130705', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130705_tscode_list = list() 
for ts_code_sh in weekly_20130705['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130705_tscode_list.append(stock_name)
weekly_20130705_addname_dataframe=pd.DataFrame()
weekly_20130705_addname_dataframe['cname'] = weekly_20130705_tscode_list
for table_name in weekly_20130705.columns.values.tolist():
    weekly_20130705_addname_dataframe[table_name] = weekly_20130705[table_name]
print("周线行情-时间为序  weekly_20130705 27_20130705 返回数据 row 行数 = "+str(weekly_20130705.shape[0]))
weekly_20130705_addname_dataframe.to_excel(weekly_2013_excel_writer,'27_20130705',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130712")       ###  更新 周线记录日期
weekly_20130712 = pro.weekly(trade_date='20130712', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130712_tscode_list = list() 
for ts_code_sh in weekly_20130712['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130712_tscode_list.append(stock_name)
weekly_20130712_addname_dataframe=pd.DataFrame()
weekly_20130712_addname_dataframe['cname'] = weekly_20130712_tscode_list
for table_name in weekly_20130712.columns.values.tolist():
    weekly_20130712_addname_dataframe[table_name] = weekly_20130712[table_name]
print("周线行情-时间为序  weekly_20130712 28_20130712 返回数据 row 行数 = "+str(weekly_20130712.shape[0]))
weekly_20130712_addname_dataframe.to_excel(weekly_2013_excel_writer,'28_20130712',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130719")       ###  更新 周线记录日期
weekly_20130719 = pro.weekly(trade_date='20130719', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130719_tscode_list = list() 
for ts_code_sh in weekly_20130719['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130719_tscode_list.append(stock_name)
weekly_20130719_addname_dataframe=pd.DataFrame()
weekly_20130719_addname_dataframe['cname'] = weekly_20130719_tscode_list
for table_name in weekly_20130719.columns.values.tolist():
    weekly_20130719_addname_dataframe[table_name] = weekly_20130719[table_name]
print("周线行情-时间为序  weekly_20130719 29_20130719 返回数据 row 行数 = "+str(weekly_20130719.shape[0]))
weekly_20130719_addname_dataframe.to_excel(weekly_2013_excel_writer,'29_20130719',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130726")       ###  更新 周线记录日期
weekly_20130726 = pro.weekly(trade_date='20130726', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130726_tscode_list = list() 
for ts_code_sh in weekly_20130726['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130726_tscode_list.append(stock_name)
weekly_20130726_addname_dataframe=pd.DataFrame()
weekly_20130726_addname_dataframe['cname'] = weekly_20130726_tscode_list
for table_name in weekly_20130726.columns.values.tolist():
    weekly_20130726_addname_dataframe[table_name] = weekly_20130726[table_name]
print("周线行情-时间为序  weekly_20130726 30_20130726 返回数据 row 行数 = "+str(weekly_20130726.shape[0]))
weekly_20130726_addname_dataframe.to_excel(weekly_2013_excel_writer,'30_20130726',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130802")       ###  更新 周线记录日期
weekly_20130802 = pro.weekly(trade_date='20130802', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130802_tscode_list = list() 
for ts_code_sh in weekly_20130802['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130802_tscode_list.append(stock_name)
weekly_20130802_addname_dataframe=pd.DataFrame()
weekly_20130802_addname_dataframe['cname'] = weekly_20130802_tscode_list
for table_name in weekly_20130802.columns.values.tolist():
    weekly_20130802_addname_dataframe[table_name] = weekly_20130802[table_name]
print("周线行情-时间为序  weekly_20130802 31_20130802 返回数据 row 行数 = "+str(weekly_20130802.shape[0]))
weekly_20130802_addname_dataframe.to_excel(weekly_2013_excel_writer,'31_20130802',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130809")       ###  更新 周线记录日期
weekly_20130809 = pro.weekly(trade_date='20130809', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130809_tscode_list = list() 
for ts_code_sh in weekly_20130809['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130809_tscode_list.append(stock_name)
weekly_20130809_addname_dataframe=pd.DataFrame()
weekly_20130809_addname_dataframe['cname'] = weekly_20130809_tscode_list
for table_name in weekly_20130809.columns.values.tolist():
    weekly_20130809_addname_dataframe[table_name] = weekly_20130809[table_name]
print("周线行情-时间为序  weekly_20130809 32_20130809 返回数据 row 行数 = "+str(weekly_20130809.shape[0]))
weekly_20130809_addname_dataframe.to_excel(weekly_2013_excel_writer,'32_20130809',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130816")       ###  更新 周线记录日期
weekly_20130816 = pro.weekly(trade_date='20130816', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130816_tscode_list = list() 
for ts_code_sh in weekly_20130816['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130816_tscode_list.append(stock_name)
weekly_20130816_addname_dataframe=pd.DataFrame()
weekly_20130816_addname_dataframe['cname'] = weekly_20130816_tscode_list
for table_name in weekly_20130816.columns.values.tolist():
    weekly_20130816_addname_dataframe[table_name] = weekly_20130816[table_name]
print("周线行情-时间为序  weekly_20130816 33_20130816 返回数据 row 行数 = "+str(weekly_20130816.shape[0]))
weekly_20130816_addname_dataframe.to_excel(weekly_2013_excel_writer,'33_20130816',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130823")       ###  更新 周线记录日期
weekly_20130823 = pro.weekly(trade_date='20130823', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130823_tscode_list = list() 
for ts_code_sh in weekly_20130823['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130823_tscode_list.append(stock_name)
weekly_20130823_addname_dataframe=pd.DataFrame()
weekly_20130823_addname_dataframe['cname'] = weekly_20130823_tscode_list
for table_name in weekly_20130823.columns.values.tolist():
    weekly_20130823_addname_dataframe[table_name] = weekly_20130823[table_name]
print("周线行情-时间为序  weekly_20130823 34_20130823 返回数据 row 行数 = "+str(weekly_20130823.shape[0]))
weekly_20130823_addname_dataframe.to_excel(weekly_2013_excel_writer,'34_20130823',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130830")       ###  更新 周线记录日期
weekly_20130830 = pro.weekly(trade_date='20130830', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130830_tscode_list = list() 
for ts_code_sh in weekly_20130830['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130830_tscode_list.append(stock_name)
weekly_20130830_addname_dataframe=pd.DataFrame()
weekly_20130830_addname_dataframe['cname'] = weekly_20130830_tscode_list
for table_name in weekly_20130830.columns.values.tolist():
    weekly_20130830_addname_dataframe[table_name] = weekly_20130830[table_name]
print("周线行情-时间为序  weekly_20130830 35_20130830 返回数据 row 行数 = "+str(weekly_20130830.shape[0]))
weekly_20130830_addname_dataframe.to_excel(weekly_2013_excel_writer,'35_20130830',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130906")       ###  更新 周线记录日期
weekly_20130906 = pro.weekly(trade_date='20130906', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130906_tscode_list = list() 
for ts_code_sh in weekly_20130906['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130906_tscode_list.append(stock_name)
weekly_20130906_addname_dataframe=pd.DataFrame()
weekly_20130906_addname_dataframe['cname'] = weekly_20130906_tscode_list
for table_name in weekly_20130906.columns.values.tolist():
    weekly_20130906_addname_dataframe[table_name] = weekly_20130906[table_name]
print("周线行情-时间为序  weekly_20130906 36_20130906 返回数据 row 行数 = "+str(weekly_20130906.shape[0]))
weekly_20130906_addname_dataframe.to_excel(weekly_2013_excel_writer,'36_20130906',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130913")       ###  更新 周线记录日期
weekly_20130913 = pro.weekly(trade_date='20130913', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130913_tscode_list = list() 
for ts_code_sh in weekly_20130913['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130913_tscode_list.append(stock_name)
weekly_20130913_addname_dataframe=pd.DataFrame()
weekly_20130913_addname_dataframe['cname'] = weekly_20130913_tscode_list
for table_name in weekly_20130913.columns.values.tolist():
    weekly_20130913_addname_dataframe[table_name] = weekly_20130913[table_name]
print("周线行情-时间为序  weekly_20130913 37_20130913 返回数据 row 行数 = "+str(weekly_20130913.shape[0]))
weekly_20130913_addname_dataframe.to_excel(weekly_2013_excel_writer,'37_20130913',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130918")       ###  更新 周线记录日期
weekly_20130918 = pro.weekly(trade_date='20130918', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130918_tscode_list = list() 
for ts_code_sh in weekly_20130918['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130918_tscode_list.append(stock_name)
weekly_20130918_addname_dataframe=pd.DataFrame()
weekly_20130918_addname_dataframe['cname'] = weekly_20130918_tscode_list
for table_name in weekly_20130918.columns.values.tolist():
    weekly_20130918_addname_dataframe[table_name] = weekly_20130918[table_name]
print("周线行情-时间为序  weekly_20130918 38_20130918 返回数据 row 行数 = "+str(weekly_20130918.shape[0]))
weekly_20130918_addname_dataframe.to_excel(weekly_2013_excel_writer,'38_20130918',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130927")       ###  更新 周线记录日期
weekly_20130927 = pro.weekly(trade_date='20130927', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130927_tscode_list = list() 
for ts_code_sh in weekly_20130927['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130927_tscode_list.append(stock_name)
weekly_20130927_addname_dataframe=pd.DataFrame()
weekly_20130927_addname_dataframe['cname'] = weekly_20130927_tscode_list
for table_name in weekly_20130927.columns.values.tolist():
    weekly_20130927_addname_dataframe[table_name] = weekly_20130927[table_name]
print("周线行情-时间为序  weekly_20130927 39_20130927 返回数据 row 行数 = "+str(weekly_20130927.shape[0]))
weekly_20130927_addname_dataframe.to_excel(weekly_2013_excel_writer,'39_20130927',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20130930")       ###  更新 周线记录日期
weekly_20130930 = pro.weekly(trade_date='20130930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20130930_tscode_list = list() 
for ts_code_sh in weekly_20130930['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20130930_tscode_list.append(stock_name)
weekly_20130930_addname_dataframe=pd.DataFrame()
weekly_20130930_addname_dataframe['cname'] = weekly_20130930_tscode_list
for table_name in weekly_20130930.columns.values.tolist():
    weekly_20130930_addname_dataframe[table_name] = weekly_20130930[table_name]
print("周线行情-时间为序  weekly_20130930 39_20130930 返回数据 row 行数 = "+str(weekly_20130930.shape[0]))
weekly_20130930_addname_dataframe.to_excel(weekly_2013_excel_writer,'39_20130930',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131011")       ###  更新 周线记录日期
weekly_20131011 = pro.weekly(trade_date='20131011', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131011_tscode_list = list() 
for ts_code_sh in weekly_20131011['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131011_tscode_list.append(stock_name)
weekly_20131011_addname_dataframe=pd.DataFrame()
weekly_20131011_addname_dataframe['cname'] = weekly_20131011_tscode_list
for table_name in weekly_20131011.columns.values.tolist():
    weekly_20131011_addname_dataframe[table_name] = weekly_20131011[table_name]
print("周线行情-时间为序  weekly_20131011 41_20131011 返回数据 row 行数 = "+str(weekly_20131011.shape[0]))
weekly_20131011_addname_dataframe.to_excel(weekly_2013_excel_writer,'41_20131011',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131018")       ###  更新 周线记录日期
weekly_20131018 = pro.weekly(trade_date='20131018', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131018_tscode_list = list() 
for ts_code_sh in weekly_20131018['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131018_tscode_list.append(stock_name)
weekly_20131018_addname_dataframe=pd.DataFrame()
weekly_20131018_addname_dataframe['cname'] = weekly_20131018_tscode_list
for table_name in weekly_20131018.columns.values.tolist():
    weekly_20131018_addname_dataframe[table_name] = weekly_20131018[table_name]
print("周线行情-时间为序  weekly_20131018 42_20131018 返回数据 row 行数 = "+str(weekly_20131018.shape[0]))
weekly_20131018_addname_dataframe.to_excel(weekly_2013_excel_writer,'42_20131018',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131025")       ###  更新 周线记录日期
weekly_20131025 = pro.weekly(trade_date='20131025', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131025_tscode_list = list() 
for ts_code_sh in weekly_20131025['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131025_tscode_list.append(stock_name)
weekly_20131025_addname_dataframe=pd.DataFrame()
weekly_20131025_addname_dataframe['cname'] = weekly_20131025_tscode_list
for table_name in weekly_20131025.columns.values.tolist():
    weekly_20131025_addname_dataframe[table_name] = weekly_20131025[table_name]
print("周线行情-时间为序  weekly_20131025 43_20131025 返回数据 row 行数 = "+str(weekly_20131025.shape[0]))
weekly_20131025_addname_dataframe.to_excel(weekly_2013_excel_writer,'43_20131025',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131101")       ###  更新 周线记录日期
weekly_20131101 = pro.weekly(trade_date='20131101', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131101_tscode_list = list() 
for ts_code_sh in weekly_20131101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131101_tscode_list.append(stock_name)
weekly_20131101_addname_dataframe=pd.DataFrame()
weekly_20131101_addname_dataframe['cname'] = weekly_20131101_tscode_list
for table_name in weekly_20131101.columns.values.tolist():
    weekly_20131101_addname_dataframe[table_name] = weekly_20131101[table_name]
print("周线行情-时间为序  weekly_20131101 44_20131101 返回数据 row 行数 = "+str(weekly_20131101.shape[0]))
weekly_20131101_addname_dataframe.to_excel(weekly_2013_excel_writer,'44_20131101',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131108")       ###  更新 周线记录日期
weekly_20131108 = pro.weekly(trade_date='20131108', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131108_tscode_list = list() 
for ts_code_sh in weekly_20131108['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131108_tscode_list.append(stock_name)
weekly_20131108_addname_dataframe=pd.DataFrame()
weekly_20131108_addname_dataframe['cname'] = weekly_20131108_tscode_list
for table_name in weekly_20131108.columns.values.tolist():
    weekly_20131108_addname_dataframe[table_name] = weekly_20131108[table_name]
print("周线行情-时间为序  weekly_20131108 45_20131108 返回数据 row 行数 = "+str(weekly_20131108.shape[0]))
weekly_20131108_addname_dataframe.to_excel(weekly_2013_excel_writer,'45_20131108',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131115")       ###  更新 周线记录日期
weekly_20131115 = pro.weekly(trade_date='20131115', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131115_tscode_list = list() 
for ts_code_sh in weekly_20131115['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131115_tscode_list.append(stock_name)
weekly_20131115_addname_dataframe=pd.DataFrame()
weekly_20131115_addname_dataframe['cname'] = weekly_20131115_tscode_list
for table_name in weekly_20131115.columns.values.tolist():
    weekly_20131115_addname_dataframe[table_name] = weekly_20131115[table_name]
print("周线行情-时间为序  weekly_20131115 46_20131115 返回数据 row 行数 = "+str(weekly_20131115.shape[0]))
weekly_20131115_addname_dataframe.to_excel(weekly_2013_excel_writer,'46_20131115',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131122")       ###  更新 周线记录日期
weekly_20131122 = pro.weekly(trade_date='20131122', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131122_tscode_list = list() 
for ts_code_sh in weekly_20131122['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131122_tscode_list.append(stock_name)
weekly_20131122_addname_dataframe=pd.DataFrame()
weekly_20131122_addname_dataframe['cname'] = weekly_20131122_tscode_list
for table_name in weekly_20131122.columns.values.tolist():
    weekly_20131122_addname_dataframe[table_name] = weekly_20131122[table_name]
print("周线行情-时间为序  weekly_20131122 47_20131122 返回数据 row 行数 = "+str(weekly_20131122.shape[0]))
weekly_20131122_addname_dataframe.to_excel(weekly_2013_excel_writer,'47_20131122',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131129")       ###  更新 周线记录日期
weekly_20131129 = pro.weekly(trade_date='20131129', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131129_tscode_list = list() 
for ts_code_sh in weekly_20131129['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131129_tscode_list.append(stock_name)
weekly_20131129_addname_dataframe=pd.DataFrame()
weekly_20131129_addname_dataframe['cname'] = weekly_20131129_tscode_list
for table_name in weekly_20131129.columns.values.tolist():
    weekly_20131129_addname_dataframe[table_name] = weekly_20131129[table_name]
print("周线行情-时间为序  weekly_20131129 48_20131129 返回数据 row 行数 = "+str(weekly_20131129.shape[0]))
weekly_20131129_addname_dataframe.to_excel(weekly_2013_excel_writer,'48_20131129',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131206")       ###  更新 周线记录日期
weekly_20131206 = pro.weekly(trade_date='20131206', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131206_tscode_list = list() 
for ts_code_sh in weekly_20131206['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131206_tscode_list.append(stock_name)
weekly_20131206_addname_dataframe=pd.DataFrame()
weekly_20131206_addname_dataframe['cname'] = weekly_20131206_tscode_list
for table_name in weekly_20131206.columns.values.tolist():
    weekly_20131206_addname_dataframe[table_name] = weekly_20131206[table_name]
print("周线行情-时间为序  weekly_20131206 49_20131206 返回数据 row 行数 = "+str(weekly_20131206.shape[0]))
weekly_20131206_addname_dataframe.to_excel(weekly_2013_excel_writer,'49_20131206',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131213")       ###  更新 周线记录日期
weekly_20131213 = pro.weekly(trade_date='20131213', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131213_tscode_list = list() 
for ts_code_sh in weekly_20131213['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131213_tscode_list.append(stock_name)
weekly_20131213_addname_dataframe=pd.DataFrame()
weekly_20131213_addname_dataframe['cname'] = weekly_20131213_tscode_list
for table_name in weekly_20131213.columns.values.tolist():
    weekly_20131213_addname_dataframe[table_name] = weekly_20131213[table_name]
print("周线行情-时间为序  weekly_20131213 50_20131213 返回数据 row 行数 = "+str(weekly_20131213.shape[0]))
weekly_20131213_addname_dataframe.to_excel(weekly_2013_excel_writer,'50_20131213',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131220")       ###  更新 周线记录日期
weekly_20131220 = pro.weekly(trade_date='20131220', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131220_tscode_list = list() 
for ts_code_sh in weekly_20131220['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131220_tscode_list.append(stock_name)
weekly_20131220_addname_dataframe=pd.DataFrame()
weekly_20131220_addname_dataframe['cname'] = weekly_20131220_tscode_list
for table_name in weekly_20131220.columns.values.tolist():
    weekly_20131220_addname_dataframe[table_name] = weekly_20131220[table_name]
print("周线行情-时间为序  weekly_20131220 51_20131220 返回数据 row 行数 = "+str(weekly_20131220.shape[0]))
weekly_20131220_addname_dataframe.to_excel(weekly_2013_excel_writer,'51_20131220',index=False)
weekly_2013_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20131227")       ###  更新 周线记录日期
weekly_20131227 = pro.weekly(trade_date='20131227', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20131227_tscode_list = list() 
for ts_code_sh in weekly_20131227['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20131227_tscode_list.append(stock_name)
weekly_20131227_addname_dataframe=pd.DataFrame()
weekly_20131227_addname_dataframe['cname'] = weekly_20131227_tscode_list
for table_name in weekly_20131227.columns.values.tolist():
    weekly_20131227_addname_dataframe[table_name] = weekly_20131227[table_name]
print("周线行情-时间为序  weekly_20131227 52_20131227 返回数据 row 行数 = "+str(weekly_20131227.shape[0]))
weekly_20131227_addname_dataframe.to_excel(weekly_2013_excel_writer,'52_20131227',index=False)
weekly_2013_excel_writer.save()
createexcel('weekly_2014.xlsx')
weekly_2014_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2014.xlsx')
weekly_2014_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2014.xlsx', engine='openpyxl')
weekly_2014_excel_writer.book = weekly_2014_book
weekly_2014_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2014_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20140103")       ###  更新 周线记录日期
weekly_20140103 = pro.weekly(trade_date='20140103', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140103_tscode_list = list() 
for ts_code_sh in weekly_20140103['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140103_tscode_list.append(stock_name)
weekly_20140103_addname_dataframe=pd.DataFrame()
weekly_20140103_addname_dataframe['cname'] = weekly_20140103_tscode_list
for table_name in weekly_20140103.columns.values.tolist():
    weekly_20140103_addname_dataframe[table_name] = weekly_20140103[table_name]
print("周线行情-时间为序  weekly_20140103 1_20140103 返回数据 row 行数 = "+str(weekly_20140103.shape[0]))
weekly_20140103_addname_dataframe.to_excel(weekly_2014_excel_writer,'1_20140103',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140110")       ###  更新 周线记录日期
weekly_20140110 = pro.weekly(trade_date='20140110', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140110_tscode_list = list() 
for ts_code_sh in weekly_20140110['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140110_tscode_list.append(stock_name)
weekly_20140110_addname_dataframe=pd.DataFrame()
weekly_20140110_addname_dataframe['cname'] = weekly_20140110_tscode_list
for table_name in weekly_20140110.columns.values.tolist():
    weekly_20140110_addname_dataframe[table_name] = weekly_20140110[table_name]
print("周线行情-时间为序  weekly_20140110 2_20140110 返回数据 row 行数 = "+str(weekly_20140110.shape[0]))
weekly_20140110_addname_dataframe.to_excel(weekly_2014_excel_writer,'2_20140110',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140117")       ###  更新 周线记录日期
weekly_20140117 = pro.weekly(trade_date='20140117', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140117_tscode_list = list() 
for ts_code_sh in weekly_20140117['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140117_tscode_list.append(stock_name)
weekly_20140117_addname_dataframe=pd.DataFrame()
weekly_20140117_addname_dataframe['cname'] = weekly_20140117_tscode_list
for table_name in weekly_20140117.columns.values.tolist():
    weekly_20140117_addname_dataframe[table_name] = weekly_20140117[table_name]
print("周线行情-时间为序  weekly_20140117 3_20140117 返回数据 row 行数 = "+str(weekly_20140117.shape[0]))
weekly_20140117_addname_dataframe.to_excel(weekly_2014_excel_writer,'3_20140117',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140124")       ###  更新 周线记录日期
weekly_20140124 = pro.weekly(trade_date='20140124', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140124_tscode_list = list() 
for ts_code_sh in weekly_20140124['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140124_tscode_list.append(stock_name)
weekly_20140124_addname_dataframe=pd.DataFrame()
weekly_20140124_addname_dataframe['cname'] = weekly_20140124_tscode_list
for table_name in weekly_20140124.columns.values.tolist():
    weekly_20140124_addname_dataframe[table_name] = weekly_20140124[table_name]
print("周线行情-时间为序  weekly_20140124 4_20140124 返回数据 row 行数 = "+str(weekly_20140124.shape[0]))
weekly_20140124_addname_dataframe.to_excel(weekly_2014_excel_writer,'4_20140124',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140130")       ###  更新 周线记录日期
weekly_20140130 = pro.weekly(trade_date='20140130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140130_tscode_list = list() 
for ts_code_sh in weekly_20140130['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140130_tscode_list.append(stock_name)
weekly_20140130_addname_dataframe=pd.DataFrame()
weekly_20140130_addname_dataframe['cname'] = weekly_20140130_tscode_list
for table_name in weekly_20140130.columns.values.tolist():
    weekly_20140130_addname_dataframe[table_name] = weekly_20140130[table_name]
print("周线行情-时间为序  weekly_20140130 5_20140130 返回数据 row 行数 = "+str(weekly_20140130.shape[0]))
weekly_20140130_addname_dataframe.to_excel(weekly_2014_excel_writer,'5_20140130',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140207")       ###  更新 周线记录日期
weekly_20140207 = pro.weekly(trade_date='20140207', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140207_tscode_list = list() 
for ts_code_sh in weekly_20140207['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140207_tscode_list.append(stock_name)
weekly_20140207_addname_dataframe=pd.DataFrame()
weekly_20140207_addname_dataframe['cname'] = weekly_20140207_tscode_list
for table_name in weekly_20140207.columns.values.tolist():
    weekly_20140207_addname_dataframe[table_name] = weekly_20140207[table_name]
print("周线行情-时间为序  weekly_20140207 6_20140207 返回数据 row 行数 = "+str(weekly_20140207.shape[0]))
weekly_20140207_addname_dataframe.to_excel(weekly_2014_excel_writer,'6_20140207',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140214")       ###  更新 周线记录日期
weekly_20140214 = pro.weekly(trade_date='20140214', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140214_tscode_list = list() 
for ts_code_sh in weekly_20140214['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140214_tscode_list.append(stock_name)
weekly_20140214_addname_dataframe=pd.DataFrame()
weekly_20140214_addname_dataframe['cname'] = weekly_20140214_tscode_list
for table_name in weekly_20140214.columns.values.tolist():
    weekly_20140214_addname_dataframe[table_name] = weekly_20140214[table_name]
print("周线行情-时间为序  weekly_20140214 7_20140214 返回数据 row 行数 = "+str(weekly_20140214.shape[0]))
weekly_20140214_addname_dataframe.to_excel(weekly_2014_excel_writer,'7_20140214',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140221")       ###  更新 周线记录日期
weekly_20140221 = pro.weekly(trade_date='20140221', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140221_tscode_list = list() 
for ts_code_sh in weekly_20140221['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140221_tscode_list.append(stock_name)
weekly_20140221_addname_dataframe=pd.DataFrame()
weekly_20140221_addname_dataframe['cname'] = weekly_20140221_tscode_list
for table_name in weekly_20140221.columns.values.tolist():
    weekly_20140221_addname_dataframe[table_name] = weekly_20140221[table_name]
print("周线行情-时间为序  weekly_20140221 8_20140221 返回数据 row 行数 = "+str(weekly_20140221.shape[0]))
weekly_20140221_addname_dataframe.to_excel(weekly_2014_excel_writer,'8_20140221',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140228")       ###  更新 周线记录日期
weekly_20140228 = pro.weekly(trade_date='20140228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140228_tscode_list = list() 
for ts_code_sh in weekly_20140228['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140228_tscode_list.append(stock_name)
weekly_20140228_addname_dataframe=pd.DataFrame()
weekly_20140228_addname_dataframe['cname'] = weekly_20140228_tscode_list
for table_name in weekly_20140228.columns.values.tolist():
    weekly_20140228_addname_dataframe[table_name] = weekly_20140228[table_name]
print("周线行情-时间为序  weekly_20140228 9_20140228 返回数据 row 行数 = "+str(weekly_20140228.shape[0]))
weekly_20140228_addname_dataframe.to_excel(weekly_2014_excel_writer,'9_20140228',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140307")       ###  更新 周线记录日期
weekly_20140307 = pro.weekly(trade_date='20140307', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140307_tscode_list = list() 
for ts_code_sh in weekly_20140307['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140307_tscode_list.append(stock_name)
weekly_20140307_addname_dataframe=pd.DataFrame()
weekly_20140307_addname_dataframe['cname'] = weekly_20140307_tscode_list
for table_name in weekly_20140307.columns.values.tolist():
    weekly_20140307_addname_dataframe[table_name] = weekly_20140307[table_name]
print("周线行情-时间为序  weekly_20140307 10_20140307 返回数据 row 行数 = "+str(weekly_20140307.shape[0]))
weekly_20140307_addname_dataframe.to_excel(weekly_2014_excel_writer,'10_20140307',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140314")       ###  更新 周线记录日期
weekly_20140314 = pro.weekly(trade_date='20140314', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140314_tscode_list = list() 
for ts_code_sh in weekly_20140314['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140314_tscode_list.append(stock_name)
weekly_20140314_addname_dataframe=pd.DataFrame()
weekly_20140314_addname_dataframe['cname'] = weekly_20140314_tscode_list
for table_name in weekly_20140314.columns.values.tolist():
    weekly_20140314_addname_dataframe[table_name] = weekly_20140314[table_name]
print("周线行情-时间为序  weekly_20140314 11_20140314 返回数据 row 行数 = "+str(weekly_20140314.shape[0]))
weekly_20140314_addname_dataframe.to_excel(weekly_2014_excel_writer,'11_20140314',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140321")       ###  更新 周线记录日期
weekly_20140321 = pro.weekly(trade_date='20140321', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140321_tscode_list = list() 
for ts_code_sh in weekly_20140321['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140321_tscode_list.append(stock_name)
weekly_20140321_addname_dataframe=pd.DataFrame()
weekly_20140321_addname_dataframe['cname'] = weekly_20140321_tscode_list
for table_name in weekly_20140321.columns.values.tolist():
    weekly_20140321_addname_dataframe[table_name] = weekly_20140321[table_name]
print("周线行情-时间为序  weekly_20140321 12_20140321 返回数据 row 行数 = "+str(weekly_20140321.shape[0]))
weekly_20140321_addname_dataframe.to_excel(weekly_2014_excel_writer,'12_20140321',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140328")       ###  更新 周线记录日期
weekly_20140328 = pro.weekly(trade_date='20140328', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140328_tscode_list = list() 
for ts_code_sh in weekly_20140328['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140328_tscode_list.append(stock_name)
weekly_20140328_addname_dataframe=pd.DataFrame()
weekly_20140328_addname_dataframe['cname'] = weekly_20140328_tscode_list
for table_name in weekly_20140328.columns.values.tolist():
    weekly_20140328_addname_dataframe[table_name] = weekly_20140328[table_name]
print("周线行情-时间为序  weekly_20140328 13_20140328 返回数据 row 行数 = "+str(weekly_20140328.shape[0]))
weekly_20140328_addname_dataframe.to_excel(weekly_2014_excel_writer,'13_20140328',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140404")       ###  更新 周线记录日期
weekly_20140404 = pro.weekly(trade_date='20140404', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140404_tscode_list = list() 
for ts_code_sh in weekly_20140404['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140404_tscode_list.append(stock_name)
weekly_20140404_addname_dataframe=pd.DataFrame()
weekly_20140404_addname_dataframe['cname'] = weekly_20140404_tscode_list
for table_name in weekly_20140404.columns.values.tolist():
    weekly_20140404_addname_dataframe[table_name] = weekly_20140404[table_name]
print("周线行情-时间为序  weekly_20140404 14_20140404 返回数据 row 行数 = "+str(weekly_20140404.shape[0]))
weekly_20140404_addname_dataframe.to_excel(weekly_2014_excel_writer,'14_20140404',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140411")       ###  更新 周线记录日期
weekly_20140411 = pro.weekly(trade_date='20140411', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140411_tscode_list = list() 
for ts_code_sh in weekly_20140411['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140411_tscode_list.append(stock_name)
weekly_20140411_addname_dataframe=pd.DataFrame()
weekly_20140411_addname_dataframe['cname'] = weekly_20140411_tscode_list
for table_name in weekly_20140411.columns.values.tolist():
    weekly_20140411_addname_dataframe[table_name] = weekly_20140411[table_name]
print("周线行情-时间为序  weekly_20140411 15_20140411 返回数据 row 行数 = "+str(weekly_20140411.shape[0]))
weekly_20140411_addname_dataframe.to_excel(weekly_2014_excel_writer,'15_20140411',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140418")       ###  更新 周线记录日期
weekly_20140418 = pro.weekly(trade_date='20140418', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140418_tscode_list = list() 
for ts_code_sh in weekly_20140418['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140418_tscode_list.append(stock_name)
weekly_20140418_addname_dataframe=pd.DataFrame()
weekly_20140418_addname_dataframe['cname'] = weekly_20140418_tscode_list
for table_name in weekly_20140418.columns.values.tolist():
    weekly_20140418_addname_dataframe[table_name] = weekly_20140418[table_name]
print("周线行情-时间为序  weekly_20140418 16_20140418 返回数据 row 行数 = "+str(weekly_20140418.shape[0]))
weekly_20140418_addname_dataframe.to_excel(weekly_2014_excel_writer,'16_20140418',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140425")       ###  更新 周线记录日期
weekly_20140425 = pro.weekly(trade_date='20140425', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140425_tscode_list = list() 
for ts_code_sh in weekly_20140425['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140425_tscode_list.append(stock_name)
weekly_20140425_addname_dataframe=pd.DataFrame()
weekly_20140425_addname_dataframe['cname'] = weekly_20140425_tscode_list
for table_name in weekly_20140425.columns.values.tolist():
    weekly_20140425_addname_dataframe[table_name] = weekly_20140425[table_name]
print("周线行情-时间为序  weekly_20140425 17_20140425 返回数据 row 行数 = "+str(weekly_20140425.shape[0]))
weekly_20140425_addname_dataframe.to_excel(weekly_2014_excel_writer,'17_20140425',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140430")       ###  更新 周线记录日期
weekly_20140430 = pro.weekly(trade_date='20140430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140430_tscode_list = list() 
for ts_code_sh in weekly_20140430['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140430_tscode_list.append(stock_name)
weekly_20140430_addname_dataframe=pd.DataFrame()
weekly_20140430_addname_dataframe['cname'] = weekly_20140430_tscode_list
for table_name in weekly_20140430.columns.values.tolist():
    weekly_20140430_addname_dataframe[table_name] = weekly_20140430[table_name]
print("周线行情-时间为序  weekly_20140430 18_20140430 返回数据 row 行数 = "+str(weekly_20140430.shape[0]))
weekly_20140430_addname_dataframe.to_excel(weekly_2014_excel_writer,'18_20140430',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140509")       ###  更新 周线记录日期
weekly_20140509 = pro.weekly(trade_date='20140509', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140509_tscode_list = list() 
for ts_code_sh in weekly_20140509['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140509_tscode_list.append(stock_name)
weekly_20140509_addname_dataframe=pd.DataFrame()
weekly_20140509_addname_dataframe['cname'] = weekly_20140509_tscode_list
for table_name in weekly_20140509.columns.values.tolist():
    weekly_20140509_addname_dataframe[table_name] = weekly_20140509[table_name]
print("周线行情-时间为序  weekly_20140509 19_20140509 返回数据 row 行数 = "+str(weekly_20140509.shape[0]))
weekly_20140509_addname_dataframe.to_excel(weekly_2014_excel_writer,'19_20140509',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140516")       ###  更新 周线记录日期
weekly_20140516 = pro.weekly(trade_date='20140516', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140516_tscode_list = list() 
for ts_code_sh in weekly_20140516['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140516_tscode_list.append(stock_name)
weekly_20140516_addname_dataframe=pd.DataFrame()
weekly_20140516_addname_dataframe['cname'] = weekly_20140516_tscode_list
for table_name in weekly_20140516.columns.values.tolist():
    weekly_20140516_addname_dataframe[table_name] = weekly_20140516[table_name]
print("周线行情-时间为序  weekly_20140516 20_20140516 返回数据 row 行数 = "+str(weekly_20140516.shape[0]))
weekly_20140516_addname_dataframe.to_excel(weekly_2014_excel_writer,'20_20140516',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140523")       ###  更新 周线记录日期
weekly_20140523 = pro.weekly(trade_date='20140523', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140523_tscode_list = list() 
for ts_code_sh in weekly_20140523['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140523_tscode_list.append(stock_name)
weekly_20140523_addname_dataframe=pd.DataFrame()
weekly_20140523_addname_dataframe['cname'] = weekly_20140523_tscode_list
for table_name in weekly_20140523.columns.values.tolist():
    weekly_20140523_addname_dataframe[table_name] = weekly_20140523[table_name]
print("周线行情-时间为序  weekly_20140523 21_20140523 返回数据 row 行数 = "+str(weekly_20140523.shape[0]))
weekly_20140523_addname_dataframe.to_excel(weekly_2014_excel_writer,'21_20140523',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140530")       ###  更新 周线记录日期
weekly_20140530 = pro.weekly(trade_date='20140530', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140530_tscode_list = list() 
for ts_code_sh in weekly_20140530['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140530_tscode_list.append(stock_name)
weekly_20140530_addname_dataframe=pd.DataFrame()
weekly_20140530_addname_dataframe['cname'] = weekly_20140530_tscode_list
for table_name in weekly_20140530.columns.values.tolist():
    weekly_20140530_addname_dataframe[table_name] = weekly_20140530[table_name]
print("周线行情-时间为序  weekly_20140530 22_20140530 返回数据 row 行数 = "+str(weekly_20140530.shape[0]))
weekly_20140530_addname_dataframe.to_excel(weekly_2014_excel_writer,'22_20140530',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140606")       ###  更新 周线记录日期
weekly_20140606 = pro.weekly(trade_date='20140606', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140606_tscode_list = list() 
for ts_code_sh in weekly_20140606['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140606_tscode_list.append(stock_name)
weekly_20140606_addname_dataframe=pd.DataFrame()
weekly_20140606_addname_dataframe['cname'] = weekly_20140606_tscode_list
for table_name in weekly_20140606.columns.values.tolist():
    weekly_20140606_addname_dataframe[table_name] = weekly_20140606[table_name]
print("周线行情-时间为序  weekly_20140606 23_20140606 返回数据 row 行数 = "+str(weekly_20140606.shape[0]))
weekly_20140606_addname_dataframe.to_excel(weekly_2014_excel_writer,'23_20140606',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140613")       ###  更新 周线记录日期
weekly_20140613 = pro.weekly(trade_date='20140613', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140613_tscode_list = list() 
for ts_code_sh in weekly_20140613['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140613_tscode_list.append(stock_name)
weekly_20140613_addname_dataframe=pd.DataFrame()
weekly_20140613_addname_dataframe['cname'] = weekly_20140613_tscode_list
for table_name in weekly_20140613.columns.values.tolist():
    weekly_20140613_addname_dataframe[table_name] = weekly_20140613[table_name]
print("周线行情-时间为序  weekly_20140613 24_20140613 返回数据 row 行数 = "+str(weekly_20140613.shape[0]))
weekly_20140613_addname_dataframe.to_excel(weekly_2014_excel_writer,'24_20140613',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140620")       ###  更新 周线记录日期
weekly_20140620 = pro.weekly(trade_date='20140620', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140620_tscode_list = list() 
for ts_code_sh in weekly_20140620['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140620_tscode_list.append(stock_name)
weekly_20140620_addname_dataframe=pd.DataFrame()
weekly_20140620_addname_dataframe['cname'] = weekly_20140620_tscode_list
for table_name in weekly_20140620.columns.values.tolist():
    weekly_20140620_addname_dataframe[table_name] = weekly_20140620[table_name]
print("周线行情-时间为序  weekly_20140620 25_20140620 返回数据 row 行数 = "+str(weekly_20140620.shape[0]))
weekly_20140620_addname_dataframe.to_excel(weekly_2014_excel_writer,'25_20140620',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140627")       ###  更新 周线记录日期
weekly_20140627 = pro.weekly(trade_date='20140627', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140627_tscode_list = list() 
for ts_code_sh in weekly_20140627['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140627_tscode_list.append(stock_name)
weekly_20140627_addname_dataframe=pd.DataFrame()
weekly_20140627_addname_dataframe['cname'] = weekly_20140627_tscode_list
for table_name in weekly_20140627.columns.values.tolist():
    weekly_20140627_addname_dataframe[table_name] = weekly_20140627[table_name]
print("周线行情-时间为序  weekly_20140627 26_20140627 返回数据 row 行数 = "+str(weekly_20140627.shape[0]))
weekly_20140627_addname_dataframe.to_excel(weekly_2014_excel_writer,'26_20140627',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140704")       ###  更新 周线记录日期
weekly_20140704 = pro.weekly(trade_date='20140704', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140704_tscode_list = list() 
for ts_code_sh in weekly_20140704['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140704_tscode_list.append(stock_name)
weekly_20140704_addname_dataframe=pd.DataFrame()
weekly_20140704_addname_dataframe['cname'] = weekly_20140704_tscode_list
for table_name in weekly_20140704.columns.values.tolist():
    weekly_20140704_addname_dataframe[table_name] = weekly_20140704[table_name]
print("周线行情-时间为序  weekly_20140704 27_20140704 返回数据 row 行数 = "+str(weekly_20140704.shape[0]))
weekly_20140704_addname_dataframe.to_excel(weekly_2014_excel_writer,'27_20140704',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140711")       ###  更新 周线记录日期
weekly_20140711 = pro.weekly(trade_date='20140711', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140711_tscode_list = list() 
for ts_code_sh in weekly_20140711['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140711_tscode_list.append(stock_name)
weekly_20140711_addname_dataframe=pd.DataFrame()
weekly_20140711_addname_dataframe['cname'] = weekly_20140711_tscode_list
for table_name in weekly_20140711.columns.values.tolist():
    weekly_20140711_addname_dataframe[table_name] = weekly_20140711[table_name]
print("周线行情-时间为序  weekly_20140711 28_20140711 返回数据 row 行数 = "+str(weekly_20140711.shape[0]))
weekly_20140711_addname_dataframe.to_excel(weekly_2014_excel_writer,'28_20140711',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140718")       ###  更新 周线记录日期
weekly_20140718 = pro.weekly(trade_date='20140718', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140718_tscode_list = list() 
for ts_code_sh in weekly_20140718['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140718_tscode_list.append(stock_name)
weekly_20140718_addname_dataframe=pd.DataFrame()
weekly_20140718_addname_dataframe['cname'] = weekly_20140718_tscode_list
for table_name in weekly_20140718.columns.values.tolist():
    weekly_20140718_addname_dataframe[table_name] = weekly_20140718[table_name]
print("周线行情-时间为序  weekly_20140718 29_20140718 返回数据 row 行数 = "+str(weekly_20140718.shape[0]))
weekly_20140718_addname_dataframe.to_excel(weekly_2014_excel_writer,'29_20140718',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140725")       ###  更新 周线记录日期
weekly_20140725 = pro.weekly(trade_date='20140725', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140725_tscode_list = list() 
for ts_code_sh in weekly_20140725['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140725_tscode_list.append(stock_name)
weekly_20140725_addname_dataframe=pd.DataFrame()
weekly_20140725_addname_dataframe['cname'] = weekly_20140725_tscode_list
for table_name in weekly_20140725.columns.values.tolist():
    weekly_20140725_addname_dataframe[table_name] = weekly_20140725[table_name]
print("周线行情-时间为序  weekly_20140725 30_20140725 返回数据 row 行数 = "+str(weekly_20140725.shape[0]))
weekly_20140725_addname_dataframe.to_excel(weekly_2014_excel_writer,'30_20140725',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140801")       ###  更新 周线记录日期
weekly_20140801 = pro.weekly(trade_date='20140801', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140801_tscode_list = list() 
for ts_code_sh in weekly_20140801['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140801_tscode_list.append(stock_name)
weekly_20140801_addname_dataframe=pd.DataFrame()
weekly_20140801_addname_dataframe['cname'] = weekly_20140801_tscode_list
for table_name in weekly_20140801.columns.values.tolist():
    weekly_20140801_addname_dataframe[table_name] = weekly_20140801[table_name]
print("周线行情-时间为序  weekly_20140801 31_20140801 返回数据 row 行数 = "+str(weekly_20140801.shape[0]))
weekly_20140801_addname_dataframe.to_excel(weekly_2014_excel_writer,'31_20140801',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140808")       ###  更新 周线记录日期
weekly_20140808 = pro.weekly(trade_date='20140808', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140808_tscode_list = list() 
for ts_code_sh in weekly_20140808['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140808_tscode_list.append(stock_name)
weekly_20140808_addname_dataframe=pd.DataFrame()
weekly_20140808_addname_dataframe['cname'] = weekly_20140808_tscode_list
for table_name in weekly_20140808.columns.values.tolist():
    weekly_20140808_addname_dataframe[table_name] = weekly_20140808[table_name]
print("周线行情-时间为序  weekly_20140808 32_20140808 返回数据 row 行数 = "+str(weekly_20140808.shape[0]))
weekly_20140808_addname_dataframe.to_excel(weekly_2014_excel_writer,'32_20140808',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140815")       ###  更新 周线记录日期
weekly_20140815 = pro.weekly(trade_date='20140815', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140815_tscode_list = list() 
for ts_code_sh in weekly_20140815['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140815_tscode_list.append(stock_name)
weekly_20140815_addname_dataframe=pd.DataFrame()
weekly_20140815_addname_dataframe['cname'] = weekly_20140815_tscode_list
for table_name in weekly_20140815.columns.values.tolist():
    weekly_20140815_addname_dataframe[table_name] = weekly_20140815[table_name]
print("周线行情-时间为序  weekly_20140815 33_20140815 返回数据 row 行数 = "+str(weekly_20140815.shape[0]))
weekly_20140815_addname_dataframe.to_excel(weekly_2014_excel_writer,'33_20140815',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140822")       ###  更新 周线记录日期
weekly_20140822 = pro.weekly(trade_date='20140822', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140822_tscode_list = list() 
for ts_code_sh in weekly_20140822['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140822_tscode_list.append(stock_name)
weekly_20140822_addname_dataframe=pd.DataFrame()
weekly_20140822_addname_dataframe['cname'] = weekly_20140822_tscode_list
for table_name in weekly_20140822.columns.values.tolist():
    weekly_20140822_addname_dataframe[table_name] = weekly_20140822[table_name]
print("周线行情-时间为序  weekly_20140822 34_20140822 返回数据 row 行数 = "+str(weekly_20140822.shape[0]))
weekly_20140822_addname_dataframe.to_excel(weekly_2014_excel_writer,'34_20140822',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140829")       ###  更新 周线记录日期
weekly_20140829 = pro.weekly(trade_date='20140829', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140829_tscode_list = list() 
for ts_code_sh in weekly_20140829['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140829_tscode_list.append(stock_name)
weekly_20140829_addname_dataframe=pd.DataFrame()
weekly_20140829_addname_dataframe['cname'] = weekly_20140829_tscode_list
for table_name in weekly_20140829.columns.values.tolist():
    weekly_20140829_addname_dataframe[table_name] = weekly_20140829[table_name]
print("周线行情-时间为序  weekly_20140829 35_20140829 返回数据 row 行数 = "+str(weekly_20140829.shape[0]))
weekly_20140829_addname_dataframe.to_excel(weekly_2014_excel_writer,'35_20140829',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140905")       ###  更新 周线记录日期
weekly_20140905 = pro.weekly(trade_date='20140905', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140905_tscode_list = list() 
for ts_code_sh in weekly_20140905['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140905_tscode_list.append(stock_name)
weekly_20140905_addname_dataframe=pd.DataFrame()
weekly_20140905_addname_dataframe['cname'] = weekly_20140905_tscode_list
for table_name in weekly_20140905.columns.values.tolist():
    weekly_20140905_addname_dataframe[table_name] = weekly_20140905[table_name]
print("周线行情-时间为序  weekly_20140905 36_20140905 返回数据 row 行数 = "+str(weekly_20140905.shape[0]))
weekly_20140905_addname_dataframe.to_excel(weekly_2014_excel_writer,'36_20140905',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140912")       ###  更新 周线记录日期
weekly_20140912 = pro.weekly(trade_date='20140912', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140912_tscode_list = list() 
for ts_code_sh in weekly_20140912['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140912_tscode_list.append(stock_name)
weekly_20140912_addname_dataframe=pd.DataFrame()
weekly_20140912_addname_dataframe['cname'] = weekly_20140912_tscode_list
for table_name in weekly_20140912.columns.values.tolist():
    weekly_20140912_addname_dataframe[table_name] = weekly_20140912[table_name]
print("周线行情-时间为序  weekly_20140912 37_20140912 返回数据 row 行数 = "+str(weekly_20140912.shape[0]))
weekly_20140912_addname_dataframe.to_excel(weekly_2014_excel_writer,'37_20140912',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140919")       ###  更新 周线记录日期
weekly_20140919 = pro.weekly(trade_date='20140919', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140919_tscode_list = list() 
for ts_code_sh in weekly_20140919['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140919_tscode_list.append(stock_name)
weekly_20140919_addname_dataframe=pd.DataFrame()
weekly_20140919_addname_dataframe['cname'] = weekly_20140919_tscode_list
for table_name in weekly_20140919.columns.values.tolist():
    weekly_20140919_addname_dataframe[table_name] = weekly_20140919[table_name]
print("周线行情-时间为序  weekly_20140919 38_20140919 返回数据 row 行数 = "+str(weekly_20140919.shape[0]))
weekly_20140919_addname_dataframe.to_excel(weekly_2014_excel_writer,'38_20140919',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140926")       ###  更新 周线记录日期
weekly_20140926 = pro.weekly(trade_date='20140926', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140926_tscode_list = list() 
for ts_code_sh in weekly_20140926['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140926_tscode_list.append(stock_name)
weekly_20140926_addname_dataframe=pd.DataFrame()
weekly_20140926_addname_dataframe['cname'] = weekly_20140926_tscode_list
for table_name in weekly_20140926.columns.values.tolist():
    weekly_20140926_addname_dataframe[table_name] = weekly_20140926[table_name]
print("周线行情-时间为序  weekly_20140926 39_20140926 返回数据 row 行数 = "+str(weekly_20140926.shape[0]))
weekly_20140926_addname_dataframe.to_excel(weekly_2014_excel_writer,'39_20140926',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20140930")       ###  更新 周线记录日期
weekly_20140930 = pro.weekly(trade_date='20140930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20140930_tscode_list = list() 
for ts_code_sh in weekly_20140930['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20140930_tscode_list.append(stock_name)
weekly_20140930_addname_dataframe=pd.DataFrame()
weekly_20140930_addname_dataframe['cname'] = weekly_20140930_tscode_list
for table_name in weekly_20140930.columns.values.tolist():
    weekly_20140930_addname_dataframe[table_name] = weekly_20140930[table_name]
print("周线行情-时间为序  weekly_20140930 39_20140930 返回数据 row 行数 = "+str(weekly_20140930.shape[0]))
weekly_20140930_addname_dataframe.to_excel(weekly_2014_excel_writer,'39_20140930',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141010")       ###  更新 周线记录日期
weekly_20141010 = pro.weekly(trade_date='20141010', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141010_tscode_list = list() 
for ts_code_sh in weekly_20141010['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141010_tscode_list.append(stock_name)
weekly_20141010_addname_dataframe=pd.DataFrame()
weekly_20141010_addname_dataframe['cname'] = weekly_20141010_tscode_list
for table_name in weekly_20141010.columns.values.tolist():
    weekly_20141010_addname_dataframe[table_name] = weekly_20141010[table_name]
print("周线行情-时间为序  weekly_20141010 41_20141010 返回数据 row 行数 = "+str(weekly_20141010.shape[0]))
weekly_20141010_addname_dataframe.to_excel(weekly_2014_excel_writer,'41_20141010',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141017")       ###  更新 周线记录日期
weekly_20141017 = pro.weekly(trade_date='20141017', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141017_tscode_list = list() 
for ts_code_sh in weekly_20141017['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141017_tscode_list.append(stock_name)
weekly_20141017_addname_dataframe=pd.DataFrame()
weekly_20141017_addname_dataframe['cname'] = weekly_20141017_tscode_list
for table_name in weekly_20141017.columns.values.tolist():
    weekly_20141017_addname_dataframe[table_name] = weekly_20141017[table_name]
print("周线行情-时间为序  weekly_20141017 42_20141017 返回数据 row 行数 = "+str(weekly_20141017.shape[0]))
weekly_20141017_addname_dataframe.to_excel(weekly_2014_excel_writer,'42_20141017',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141024")       ###  更新 周线记录日期
weekly_20141024 = pro.weekly(trade_date='20141024', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141024_tscode_list = list() 
for ts_code_sh in weekly_20141024['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141024_tscode_list.append(stock_name)
weekly_20141024_addname_dataframe=pd.DataFrame()
weekly_20141024_addname_dataframe['cname'] = weekly_20141024_tscode_list
for table_name in weekly_20141024.columns.values.tolist():
    weekly_20141024_addname_dataframe[table_name] = weekly_20141024[table_name]
print("周线行情-时间为序  weekly_20141024 43_20141024 返回数据 row 行数 = "+str(weekly_20141024.shape[0]))
weekly_20141024_addname_dataframe.to_excel(weekly_2014_excel_writer,'43_20141024',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141031")       ###  更新 周线记录日期
weekly_20141031 = pro.weekly(trade_date='20141031', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141031_tscode_list = list() 
for ts_code_sh in weekly_20141031['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141031_tscode_list.append(stock_name)
weekly_20141031_addname_dataframe=pd.DataFrame()
weekly_20141031_addname_dataframe['cname'] = weekly_20141031_tscode_list
for table_name in weekly_20141031.columns.values.tolist():
    weekly_20141031_addname_dataframe[table_name] = weekly_20141031[table_name]
print("周线行情-时间为序  weekly_20141031 44_20141031 返回数据 row 行数 = "+str(weekly_20141031.shape[0]))
weekly_20141031_addname_dataframe.to_excel(weekly_2014_excel_writer,'44_20141031',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141107")       ###  更新 周线记录日期
weekly_20141107 = pro.weekly(trade_date='20141107', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141107_tscode_list = list() 
for ts_code_sh in weekly_20141107['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141107_tscode_list.append(stock_name)
weekly_20141107_addname_dataframe=pd.DataFrame()
weekly_20141107_addname_dataframe['cname'] = weekly_20141107_tscode_list
for table_name in weekly_20141107.columns.values.tolist():
    weekly_20141107_addname_dataframe[table_name] = weekly_20141107[table_name]
print("周线行情-时间为序  weekly_20141107 45_20141107 返回数据 row 行数 = "+str(weekly_20141107.shape[0]))
weekly_20141107_addname_dataframe.to_excel(weekly_2014_excel_writer,'45_20141107',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141114")       ###  更新 周线记录日期
weekly_20141114 = pro.weekly(trade_date='20141114', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141114_tscode_list = list() 
for ts_code_sh in weekly_20141114['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141114_tscode_list.append(stock_name)
weekly_20141114_addname_dataframe=pd.DataFrame()
weekly_20141114_addname_dataframe['cname'] = weekly_20141114_tscode_list
for table_name in weekly_20141114.columns.values.tolist():
    weekly_20141114_addname_dataframe[table_name] = weekly_20141114[table_name]
print("周线行情-时间为序  weekly_20141114 46_20141114 返回数据 row 行数 = "+str(weekly_20141114.shape[0]))
weekly_20141114_addname_dataframe.to_excel(weekly_2014_excel_writer,'46_20141114',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141121")       ###  更新 周线记录日期
weekly_20141121 = pro.weekly(trade_date='20141121', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141121_tscode_list = list() 
for ts_code_sh in weekly_20141121['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141121_tscode_list.append(stock_name)
weekly_20141121_addname_dataframe=pd.DataFrame()
weekly_20141121_addname_dataframe['cname'] = weekly_20141121_tscode_list
for table_name in weekly_20141121.columns.values.tolist():
    weekly_20141121_addname_dataframe[table_name] = weekly_20141121[table_name]
print("周线行情-时间为序  weekly_20141121 47_20141121 返回数据 row 行数 = "+str(weekly_20141121.shape[0]))
weekly_20141121_addname_dataframe.to_excel(weekly_2014_excel_writer,'47_20141121',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141128")       ###  更新 周线记录日期
weekly_20141128 = pro.weekly(trade_date='20141128', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141128_tscode_list = list() 
for ts_code_sh in weekly_20141128['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141128_tscode_list.append(stock_name)
weekly_20141128_addname_dataframe=pd.DataFrame()
weekly_20141128_addname_dataframe['cname'] = weekly_20141128_tscode_list
for table_name in weekly_20141128.columns.values.tolist():
    weekly_20141128_addname_dataframe[table_name] = weekly_20141128[table_name]
print("周线行情-时间为序  weekly_20141128 48_20141128 返回数据 row 行数 = "+str(weekly_20141128.shape[0]))
weekly_20141128_addname_dataframe.to_excel(weekly_2014_excel_writer,'48_20141128',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141205")       ###  更新 周线记录日期
weekly_20141205 = pro.weekly(trade_date='20141205', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141205_tscode_list = list() 
for ts_code_sh in weekly_20141205['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141205_tscode_list.append(stock_name)
weekly_20141205_addname_dataframe=pd.DataFrame()
weekly_20141205_addname_dataframe['cname'] = weekly_20141205_tscode_list
for table_name in weekly_20141205.columns.values.tolist():
    weekly_20141205_addname_dataframe[table_name] = weekly_20141205[table_name]
print("周线行情-时间为序  weekly_20141205 49_20141205 返回数据 row 行数 = "+str(weekly_20141205.shape[0]))
weekly_20141205_addname_dataframe.to_excel(weekly_2014_excel_writer,'49_20141205',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141212")       ###  更新 周线记录日期
weekly_20141212 = pro.weekly(trade_date='20141212', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141212_tscode_list = list() 
for ts_code_sh in weekly_20141212['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141212_tscode_list.append(stock_name)
weekly_20141212_addname_dataframe=pd.DataFrame()
weekly_20141212_addname_dataframe['cname'] = weekly_20141212_tscode_list
for table_name in weekly_20141212.columns.values.tolist():
    weekly_20141212_addname_dataframe[table_name] = weekly_20141212[table_name]
print("周线行情-时间为序  weekly_20141212 50_20141212 返回数据 row 行数 = "+str(weekly_20141212.shape[0]))
weekly_20141212_addname_dataframe.to_excel(weekly_2014_excel_writer,'50_20141212',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141219")       ###  更新 周线记录日期
weekly_20141219 = pro.weekly(trade_date='20141219', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141219_tscode_list = list() 
for ts_code_sh in weekly_20141219['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141219_tscode_list.append(stock_name)
weekly_20141219_addname_dataframe=pd.DataFrame()
weekly_20141219_addname_dataframe['cname'] = weekly_20141219_tscode_list
for table_name in weekly_20141219.columns.values.tolist():
    weekly_20141219_addname_dataframe[table_name] = weekly_20141219[table_name]
print("周线行情-时间为序  weekly_20141219 51_20141219 返回数据 row 行数 = "+str(weekly_20141219.shape[0]))
weekly_20141219_addname_dataframe.to_excel(weekly_2014_excel_writer,'51_20141219',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141226")       ###  更新 周线记录日期
weekly_20141226 = pro.weekly(trade_date='20141226', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141226_tscode_list = list() 
for ts_code_sh in weekly_20141226['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141226_tscode_list.append(stock_name)
weekly_20141226_addname_dataframe=pd.DataFrame()
weekly_20141226_addname_dataframe['cname'] = weekly_20141226_tscode_list
for table_name in weekly_20141226.columns.values.tolist():
    weekly_20141226_addname_dataframe[table_name] = weekly_20141226[table_name]
print("周线行情-时间为序  weekly_20141226 52_20141226 返回数据 row 行数 = "+str(weekly_20141226.shape[0]))
weekly_20141226_addname_dataframe.to_excel(weekly_2014_excel_writer,'52_20141226',index=False)
weekly_2014_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20141231")       ###  更新 周线记录日期
weekly_20141231 = pro.weekly(trade_date='20141231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20141231_tscode_list = list() 
for ts_code_sh in weekly_20141231['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20141231_tscode_list.append(stock_name)
weekly_20141231_addname_dataframe=pd.DataFrame()
weekly_20141231_addname_dataframe['cname'] = weekly_20141231_tscode_list
for table_name in weekly_20141231.columns.values.tolist():
    weekly_20141231_addname_dataframe[table_name] = weekly_20141231[table_name]
print("周线行情-时间为序  weekly_20141231 53_20141231 返回数据 row 行数 = "+str(weekly_20141231.shape[0]))
weekly_20141231_addname_dataframe.to_excel(weekly_2014_excel_writer,'53_20141231',index=False)
weekly_2014_excel_writer.save()
createexcel('weekly_2015.xlsx')
weekly_2015_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2015.xlsx')
weekly_2015_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2015.xlsx', engine='openpyxl')
weekly_2015_excel_writer.book = weekly_2015_book
weekly_2015_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2015_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20150109")       ###  更新 周线记录日期
weekly_20150109 = pro.weekly(trade_date='20150109', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150109_tscode_list = list() 
for ts_code_sh in weekly_20150109['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150109_tscode_list.append(stock_name)
weekly_20150109_addname_dataframe=pd.DataFrame()
weekly_20150109_addname_dataframe['cname'] = weekly_20150109_tscode_list
for table_name in weekly_20150109.columns.values.tolist():
    weekly_20150109_addname_dataframe[table_name] = weekly_20150109[table_name]
print("周线行情-时间为序  weekly_20150109 2_20150109 返回数据 row 行数 = "+str(weekly_20150109.shape[0]))
weekly_20150109_addname_dataframe.to_excel(weekly_2015_excel_writer,'2_20150109',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150116")       ###  更新 周线记录日期
weekly_20150116 = pro.weekly(trade_date='20150116', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150116_tscode_list = list() 
for ts_code_sh in weekly_20150116['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150116_tscode_list.append(stock_name)
weekly_20150116_addname_dataframe=pd.DataFrame()
weekly_20150116_addname_dataframe['cname'] = weekly_20150116_tscode_list
for table_name in weekly_20150116.columns.values.tolist():
    weekly_20150116_addname_dataframe[table_name] = weekly_20150116[table_name]
print("周线行情-时间为序  weekly_20150116 3_20150116 返回数据 row 行数 = "+str(weekly_20150116.shape[0]))
weekly_20150116_addname_dataframe.to_excel(weekly_2015_excel_writer,'3_20150116',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150123")       ###  更新 周线记录日期
weekly_20150123 = pro.weekly(trade_date='20150123', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150123_tscode_list = list() 
for ts_code_sh in weekly_20150123['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150123_tscode_list.append(stock_name)
weekly_20150123_addname_dataframe=pd.DataFrame()
weekly_20150123_addname_dataframe['cname'] = weekly_20150123_tscode_list
for table_name in weekly_20150123.columns.values.tolist():
    weekly_20150123_addname_dataframe[table_name] = weekly_20150123[table_name]
print("周线行情-时间为序  weekly_20150123 4_20150123 返回数据 row 行数 = "+str(weekly_20150123.shape[0]))
weekly_20150123_addname_dataframe.to_excel(weekly_2015_excel_writer,'4_20150123',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150130")       ###  更新 周线记录日期
weekly_20150130 = pro.weekly(trade_date='20150130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150130_tscode_list = list() 
for ts_code_sh in weekly_20150130['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150130_tscode_list.append(stock_name)
weekly_20150130_addname_dataframe=pd.DataFrame()
weekly_20150130_addname_dataframe['cname'] = weekly_20150130_tscode_list
for table_name in weekly_20150130.columns.values.tolist():
    weekly_20150130_addname_dataframe[table_name] = weekly_20150130[table_name]
print("周线行情-时间为序  weekly_20150130 5_20150130 返回数据 row 行数 = "+str(weekly_20150130.shape[0]))
weekly_20150130_addname_dataframe.to_excel(weekly_2015_excel_writer,'5_20150130',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150206")       ###  更新 周线记录日期
weekly_20150206 = pro.weekly(trade_date='20150206', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150206_tscode_list = list() 
for ts_code_sh in weekly_20150206['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150206_tscode_list.append(stock_name)
weekly_20150206_addname_dataframe=pd.DataFrame()
weekly_20150206_addname_dataframe['cname'] = weekly_20150206_tscode_list
for table_name in weekly_20150206.columns.values.tolist():
    weekly_20150206_addname_dataframe[table_name] = weekly_20150206[table_name]
print("周线行情-时间为序  weekly_20150206 6_20150206 返回数据 row 行数 = "+str(weekly_20150206.shape[0]))
weekly_20150206_addname_dataframe.to_excel(weekly_2015_excel_writer,'6_20150206',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150213")       ###  更新 周线记录日期
weekly_20150213 = pro.weekly(trade_date='20150213', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150213_tscode_list = list() 
for ts_code_sh in weekly_20150213['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150213_tscode_list.append(stock_name)
weekly_20150213_addname_dataframe=pd.DataFrame()
weekly_20150213_addname_dataframe['cname'] = weekly_20150213_tscode_list
for table_name in weekly_20150213.columns.values.tolist():
    weekly_20150213_addname_dataframe[table_name] = weekly_20150213[table_name]
print("周线行情-时间为序  weekly_20150213 7_20150213 返回数据 row 行数 = "+str(weekly_20150213.shape[0]))
weekly_20150213_addname_dataframe.to_excel(weekly_2015_excel_writer,'7_20150213',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150217")       ###  更新 周线记录日期
weekly_20150217 = pro.weekly(trade_date='20150217', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150217_tscode_list = list() 
for ts_code_sh in weekly_20150217['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150217_tscode_list.append(stock_name)
weekly_20150217_addname_dataframe=pd.DataFrame()
weekly_20150217_addname_dataframe['cname'] = weekly_20150217_tscode_list
for table_name in weekly_20150217.columns.values.tolist():
    weekly_20150217_addname_dataframe[table_name] = weekly_20150217[table_name]
print("周线行情-时间为序  weekly_20150217 7_20150217 返回数据 row 行数 = "+str(weekly_20150217.shape[0]))
weekly_20150217_addname_dataframe.to_excel(weekly_2015_excel_writer,'7_20150217',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150227")       ###  更新 周线记录日期
weekly_20150227 = pro.weekly(trade_date='20150227', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150227_tscode_list = list() 
for ts_code_sh in weekly_20150227['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150227_tscode_list.append(stock_name)
weekly_20150227_addname_dataframe=pd.DataFrame()
weekly_20150227_addname_dataframe['cname'] = weekly_20150227_tscode_list
for table_name in weekly_20150227.columns.values.tolist():
    weekly_20150227_addname_dataframe[table_name] = weekly_20150227[table_name]
print("周线行情-时间为序  weekly_20150227 9_20150227 返回数据 row 行数 = "+str(weekly_20150227.shape[0]))
weekly_20150227_addname_dataframe.to_excel(weekly_2015_excel_writer,'9_20150227',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150306")       ###  更新 周线记录日期
weekly_20150306 = pro.weekly(trade_date='20150306', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150306_tscode_list = list() 
for ts_code_sh in weekly_20150306['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150306_tscode_list.append(stock_name)
weekly_20150306_addname_dataframe=pd.DataFrame()
weekly_20150306_addname_dataframe['cname'] = weekly_20150306_tscode_list
for table_name in weekly_20150306.columns.values.tolist():
    weekly_20150306_addname_dataframe[table_name] = weekly_20150306[table_name]
print("周线行情-时间为序  weekly_20150306 10_20150306 返回数据 row 行数 = "+str(weekly_20150306.shape[0]))
weekly_20150306_addname_dataframe.to_excel(weekly_2015_excel_writer,'10_20150306',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150313")       ###  更新 周线记录日期
weekly_20150313 = pro.weekly(trade_date='20150313', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150313_tscode_list = list() 
for ts_code_sh in weekly_20150313['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150313_tscode_list.append(stock_name)
weekly_20150313_addname_dataframe=pd.DataFrame()
weekly_20150313_addname_dataframe['cname'] = weekly_20150313_tscode_list
for table_name in weekly_20150313.columns.values.tolist():
    weekly_20150313_addname_dataframe[table_name] = weekly_20150313[table_name]
print("周线行情-时间为序  weekly_20150313 11_20150313 返回数据 row 行数 = "+str(weekly_20150313.shape[0]))
weekly_20150313_addname_dataframe.to_excel(weekly_2015_excel_writer,'11_20150313',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150320")       ###  更新 周线记录日期
weekly_20150320 = pro.weekly(trade_date='20150320', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150320_tscode_list = list() 
for ts_code_sh in weekly_20150320['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150320_tscode_list.append(stock_name)
weekly_20150320_addname_dataframe=pd.DataFrame()
weekly_20150320_addname_dataframe['cname'] = weekly_20150320_tscode_list
for table_name in weekly_20150320.columns.values.tolist():
    weekly_20150320_addname_dataframe[table_name] = weekly_20150320[table_name]
print("周线行情-时间为序  weekly_20150320 12_20150320 返回数据 row 行数 = "+str(weekly_20150320.shape[0]))
weekly_20150320_addname_dataframe.to_excel(weekly_2015_excel_writer,'12_20150320',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150327")       ###  更新 周线记录日期
weekly_20150327 = pro.weekly(trade_date='20150327', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150327_tscode_list = list() 
for ts_code_sh in weekly_20150327['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150327_tscode_list.append(stock_name)
weekly_20150327_addname_dataframe=pd.DataFrame()
weekly_20150327_addname_dataframe['cname'] = weekly_20150327_tscode_list
for table_name in weekly_20150327.columns.values.tolist():
    weekly_20150327_addname_dataframe[table_name] = weekly_20150327[table_name]
print("周线行情-时间为序  weekly_20150327 13_20150327 返回数据 row 行数 = "+str(weekly_20150327.shape[0]))
weekly_20150327_addname_dataframe.to_excel(weekly_2015_excel_writer,'13_20150327',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150403")       ###  更新 周线记录日期
weekly_20150403 = pro.weekly(trade_date='20150403', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150403_tscode_list = list() 
for ts_code_sh in weekly_20150403['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150403_tscode_list.append(stock_name)
weekly_20150403_addname_dataframe=pd.DataFrame()
weekly_20150403_addname_dataframe['cname'] = weekly_20150403_tscode_list
for table_name in weekly_20150403.columns.values.tolist():
    weekly_20150403_addname_dataframe[table_name] = weekly_20150403[table_name]
print("周线行情-时间为序  weekly_20150403 14_20150403 返回数据 row 行数 = "+str(weekly_20150403.shape[0]))
weekly_20150403_addname_dataframe.to_excel(weekly_2015_excel_writer,'14_20150403',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150410")       ###  更新 周线记录日期
weekly_20150410 = pro.weekly(trade_date='20150410', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150410_tscode_list = list() 
for ts_code_sh in weekly_20150410['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150410_tscode_list.append(stock_name)
weekly_20150410_addname_dataframe=pd.DataFrame()
weekly_20150410_addname_dataframe['cname'] = weekly_20150410_tscode_list
for table_name in weekly_20150410.columns.values.tolist():
    weekly_20150410_addname_dataframe[table_name] = weekly_20150410[table_name]
print("周线行情-时间为序  weekly_20150410 15_20150410 返回数据 row 行数 = "+str(weekly_20150410.shape[0]))
weekly_20150410_addname_dataframe.to_excel(weekly_2015_excel_writer,'15_20150410',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150417")       ###  更新 周线记录日期
weekly_20150417 = pro.weekly(trade_date='20150417', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150417_tscode_list = list() 
for ts_code_sh in weekly_20150417['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150417_tscode_list.append(stock_name)
weekly_20150417_addname_dataframe=pd.DataFrame()
weekly_20150417_addname_dataframe['cname'] = weekly_20150417_tscode_list
for table_name in weekly_20150417.columns.values.tolist():
    weekly_20150417_addname_dataframe[table_name] = weekly_20150417[table_name]
print("周线行情-时间为序  weekly_20150417 16_20150417 返回数据 row 行数 = "+str(weekly_20150417.shape[0]))
weekly_20150417_addname_dataframe.to_excel(weekly_2015_excel_writer,'16_20150417',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150424")       ###  更新 周线记录日期
weekly_20150424 = pro.weekly(trade_date='20150424', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150424_tscode_list = list() 
for ts_code_sh in weekly_20150424['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150424_tscode_list.append(stock_name)
weekly_20150424_addname_dataframe=pd.DataFrame()
weekly_20150424_addname_dataframe['cname'] = weekly_20150424_tscode_list
for table_name in weekly_20150424.columns.values.tolist():
    weekly_20150424_addname_dataframe[table_name] = weekly_20150424[table_name]
print("周线行情-时间为序  weekly_20150424 17_20150424 返回数据 row 行数 = "+str(weekly_20150424.shape[0]))
weekly_20150424_addname_dataframe.to_excel(weekly_2015_excel_writer,'17_20150424',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150430")       ###  更新 周线记录日期
weekly_20150430 = pro.weekly(trade_date='20150430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150430_tscode_list = list() 
for ts_code_sh in weekly_20150430['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150430_tscode_list.append(stock_name)
weekly_20150430_addname_dataframe=pd.DataFrame()
weekly_20150430_addname_dataframe['cname'] = weekly_20150430_tscode_list
for table_name in weekly_20150430.columns.values.tolist():
    weekly_20150430_addname_dataframe[table_name] = weekly_20150430[table_name]
print("周线行情-时间为序  weekly_20150430 18_20150430 返回数据 row 行数 = "+str(weekly_20150430.shape[0]))
weekly_20150430_addname_dataframe.to_excel(weekly_2015_excel_writer,'18_20150430',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150508")       ###  更新 周线记录日期
weekly_20150508 = pro.weekly(trade_date='20150508', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150508_tscode_list = list() 
for ts_code_sh in weekly_20150508['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150508_tscode_list.append(stock_name)
weekly_20150508_addname_dataframe=pd.DataFrame()
weekly_20150508_addname_dataframe['cname'] = weekly_20150508_tscode_list
for table_name in weekly_20150508.columns.values.tolist():
    weekly_20150508_addname_dataframe[table_name] = weekly_20150508[table_name]
print("周线行情-时间为序  weekly_20150508 19_20150508 返回数据 row 行数 = "+str(weekly_20150508.shape[0]))
weekly_20150508_addname_dataframe.to_excel(weekly_2015_excel_writer,'19_20150508',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150515")       ###  更新 周线记录日期
weekly_20150515 = pro.weekly(trade_date='20150515', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150515_tscode_list = list() 
for ts_code_sh in weekly_20150515['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150515_tscode_list.append(stock_name)
weekly_20150515_addname_dataframe=pd.DataFrame()
weekly_20150515_addname_dataframe['cname'] = weekly_20150515_tscode_list
for table_name in weekly_20150515.columns.values.tolist():
    weekly_20150515_addname_dataframe[table_name] = weekly_20150515[table_name]
print("周线行情-时间为序  weekly_20150515 20_20150515 返回数据 row 行数 = "+str(weekly_20150515.shape[0]))
weekly_20150515_addname_dataframe.to_excel(weekly_2015_excel_writer,'20_20150515',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150522")       ###  更新 周线记录日期
weekly_20150522 = pro.weekly(trade_date='20150522', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150522_tscode_list = list() 
for ts_code_sh in weekly_20150522['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150522_tscode_list.append(stock_name)
weekly_20150522_addname_dataframe=pd.DataFrame()
weekly_20150522_addname_dataframe['cname'] = weekly_20150522_tscode_list
for table_name in weekly_20150522.columns.values.tolist():
    weekly_20150522_addname_dataframe[table_name] = weekly_20150522[table_name]
print("周线行情-时间为序  weekly_20150522 21_20150522 返回数据 row 行数 = "+str(weekly_20150522.shape[0]))
weekly_20150522_addname_dataframe.to_excel(weekly_2015_excel_writer,'21_20150522',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150529")       ###  更新 周线记录日期
weekly_20150529 = pro.weekly(trade_date='20150529', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150529_tscode_list = list() 
for ts_code_sh in weekly_20150529['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150529_tscode_list.append(stock_name)
weekly_20150529_addname_dataframe=pd.DataFrame()
weekly_20150529_addname_dataframe['cname'] = weekly_20150529_tscode_list
for table_name in weekly_20150529.columns.values.tolist():
    weekly_20150529_addname_dataframe[table_name] = weekly_20150529[table_name]
print("周线行情-时间为序  weekly_20150529 22_20150529 返回数据 row 行数 = "+str(weekly_20150529.shape[0]))
weekly_20150529_addname_dataframe.to_excel(weekly_2015_excel_writer,'22_20150529',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150605")       ###  更新 周线记录日期
weekly_20150605 = pro.weekly(trade_date='20150605', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150605_tscode_list = list() 
for ts_code_sh in weekly_20150605['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150605_tscode_list.append(stock_name)
weekly_20150605_addname_dataframe=pd.DataFrame()
weekly_20150605_addname_dataframe['cname'] = weekly_20150605_tscode_list
for table_name in weekly_20150605.columns.values.tolist():
    weekly_20150605_addname_dataframe[table_name] = weekly_20150605[table_name]
print("周线行情-时间为序  weekly_20150605 23_20150605 返回数据 row 行数 = "+str(weekly_20150605.shape[0]))
weekly_20150605_addname_dataframe.to_excel(weekly_2015_excel_writer,'23_20150605',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150612")       ###  更新 周线记录日期
weekly_20150612 = pro.weekly(trade_date='20150612', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150612_tscode_list = list() 
for ts_code_sh in weekly_20150612['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150612_tscode_list.append(stock_name)
weekly_20150612_addname_dataframe=pd.DataFrame()
weekly_20150612_addname_dataframe['cname'] = weekly_20150612_tscode_list
for table_name in weekly_20150612.columns.values.tolist():
    weekly_20150612_addname_dataframe[table_name] = weekly_20150612[table_name]
print("周线行情-时间为序  weekly_20150612 24_20150612 返回数据 row 行数 = "+str(weekly_20150612.shape[0]))
weekly_20150612_addname_dataframe.to_excel(weekly_2015_excel_writer,'24_20150612',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150619")       ###  更新 周线记录日期
weekly_20150619 = pro.weekly(trade_date='20150619', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150619_tscode_list = list() 
for ts_code_sh in weekly_20150619['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150619_tscode_list.append(stock_name)
weekly_20150619_addname_dataframe=pd.DataFrame()
weekly_20150619_addname_dataframe['cname'] = weekly_20150619_tscode_list
for table_name in weekly_20150619.columns.values.tolist():
    weekly_20150619_addname_dataframe[table_name] = weekly_20150619[table_name]
print("周线行情-时间为序  weekly_20150619 25_20150619 返回数据 row 行数 = "+str(weekly_20150619.shape[0]))
weekly_20150619_addname_dataframe.to_excel(weekly_2015_excel_writer,'25_20150619',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150626")       ###  更新 周线记录日期
weekly_20150626 = pro.weekly(trade_date='20150626', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150626_tscode_list = list() 
for ts_code_sh in weekly_20150626['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150626_tscode_list.append(stock_name)
weekly_20150626_addname_dataframe=pd.DataFrame()
weekly_20150626_addname_dataframe['cname'] = weekly_20150626_tscode_list
for table_name in weekly_20150626.columns.values.tolist():
    weekly_20150626_addname_dataframe[table_name] = weekly_20150626[table_name]
print("周线行情-时间为序  weekly_20150626 26_20150626 返回数据 row 行数 = "+str(weekly_20150626.shape[0]))
weekly_20150626_addname_dataframe.to_excel(weekly_2015_excel_writer,'26_20150626',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150703")       ###  更新 周线记录日期
weekly_20150703 = pro.weekly(trade_date='20150703', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150703_tscode_list = list() 
for ts_code_sh in weekly_20150703['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150703_tscode_list.append(stock_name)
weekly_20150703_addname_dataframe=pd.DataFrame()
weekly_20150703_addname_dataframe['cname'] = weekly_20150703_tscode_list
for table_name in weekly_20150703.columns.values.tolist():
    weekly_20150703_addname_dataframe[table_name] = weekly_20150703[table_name]
print("周线行情-时间为序  weekly_20150703 27_20150703 返回数据 row 行数 = "+str(weekly_20150703.shape[0]))
weekly_20150703_addname_dataframe.to_excel(weekly_2015_excel_writer,'27_20150703',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150710")       ###  更新 周线记录日期
weekly_20150710 = pro.weekly(trade_date='20150710', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150710_tscode_list = list() 
for ts_code_sh in weekly_20150710['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150710_tscode_list.append(stock_name)
weekly_20150710_addname_dataframe=pd.DataFrame()
weekly_20150710_addname_dataframe['cname'] = weekly_20150710_tscode_list
for table_name in weekly_20150710.columns.values.tolist():
    weekly_20150710_addname_dataframe[table_name] = weekly_20150710[table_name]
print("周线行情-时间为序  weekly_20150710 28_20150710 返回数据 row 行数 = "+str(weekly_20150710.shape[0]))
weekly_20150710_addname_dataframe.to_excel(weekly_2015_excel_writer,'28_20150710',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150717")       ###  更新 周线记录日期
weekly_20150717 = pro.weekly(trade_date='20150717', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150717_tscode_list = list() 
for ts_code_sh in weekly_20150717['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150717_tscode_list.append(stock_name)
weekly_20150717_addname_dataframe=pd.DataFrame()
weekly_20150717_addname_dataframe['cname'] = weekly_20150717_tscode_list
for table_name in weekly_20150717.columns.values.tolist():
    weekly_20150717_addname_dataframe[table_name] = weekly_20150717[table_name]
print("周线行情-时间为序  weekly_20150717 29_20150717 返回数据 row 行数 = "+str(weekly_20150717.shape[0]))
weekly_20150717_addname_dataframe.to_excel(weekly_2015_excel_writer,'29_20150717',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150724")       ###  更新 周线记录日期
weekly_20150724 = pro.weekly(trade_date='20150724', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150724_tscode_list = list() 
for ts_code_sh in weekly_20150724['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150724_tscode_list.append(stock_name)
weekly_20150724_addname_dataframe=pd.DataFrame()
weekly_20150724_addname_dataframe['cname'] = weekly_20150724_tscode_list
for table_name in weekly_20150724.columns.values.tolist():
    weekly_20150724_addname_dataframe[table_name] = weekly_20150724[table_name]
print("周线行情-时间为序  weekly_20150724 30_20150724 返回数据 row 行数 = "+str(weekly_20150724.shape[0]))
weekly_20150724_addname_dataframe.to_excel(weekly_2015_excel_writer,'30_20150724',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150731")       ###  更新 周线记录日期
weekly_20150731 = pro.weekly(trade_date='20150731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150731_tscode_list = list() 
for ts_code_sh in weekly_20150731['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150731_tscode_list.append(stock_name)
weekly_20150731_addname_dataframe=pd.DataFrame()
weekly_20150731_addname_dataframe['cname'] = weekly_20150731_tscode_list
for table_name in weekly_20150731.columns.values.tolist():
    weekly_20150731_addname_dataframe[table_name] = weekly_20150731[table_name]
print("周线行情-时间为序  weekly_20150731 31_20150731 返回数据 row 行数 = "+str(weekly_20150731.shape[0]))
weekly_20150731_addname_dataframe.to_excel(weekly_2015_excel_writer,'31_20150731',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150807")       ###  更新 周线记录日期
weekly_20150807 = pro.weekly(trade_date='20150807', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150807_tscode_list = list() 
for ts_code_sh in weekly_20150807['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150807_tscode_list.append(stock_name)
weekly_20150807_addname_dataframe=pd.DataFrame()
weekly_20150807_addname_dataframe['cname'] = weekly_20150807_tscode_list
for table_name in weekly_20150807.columns.values.tolist():
    weekly_20150807_addname_dataframe[table_name] = weekly_20150807[table_name]
print("周线行情-时间为序  weekly_20150807 32_20150807 返回数据 row 行数 = "+str(weekly_20150807.shape[0]))
weekly_20150807_addname_dataframe.to_excel(weekly_2015_excel_writer,'32_20150807',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150814")       ###  更新 周线记录日期
weekly_20150814 = pro.weekly(trade_date='20150814', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150814_tscode_list = list() 
for ts_code_sh in weekly_20150814['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150814_tscode_list.append(stock_name)
weekly_20150814_addname_dataframe=pd.DataFrame()
weekly_20150814_addname_dataframe['cname'] = weekly_20150814_tscode_list
for table_name in weekly_20150814.columns.values.tolist():
    weekly_20150814_addname_dataframe[table_name] = weekly_20150814[table_name]
print("周线行情-时间为序  weekly_20150814 33_20150814 返回数据 row 行数 = "+str(weekly_20150814.shape[0]))
weekly_20150814_addname_dataframe.to_excel(weekly_2015_excel_writer,'33_20150814',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150821")       ###  更新 周线记录日期
weekly_20150821 = pro.weekly(trade_date='20150821', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150821_tscode_list = list() 
for ts_code_sh in weekly_20150821['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150821_tscode_list.append(stock_name)
weekly_20150821_addname_dataframe=pd.DataFrame()
weekly_20150821_addname_dataframe['cname'] = weekly_20150821_tscode_list
for table_name in weekly_20150821.columns.values.tolist():
    weekly_20150821_addname_dataframe[table_name] = weekly_20150821[table_name]
print("周线行情-时间为序  weekly_20150821 34_20150821 返回数据 row 行数 = "+str(weekly_20150821.shape[0]))
weekly_20150821_addname_dataframe.to_excel(weekly_2015_excel_writer,'34_20150821',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150828")       ###  更新 周线记录日期
weekly_20150828 = pro.weekly(trade_date='20150828', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150828_tscode_list = list() 
for ts_code_sh in weekly_20150828['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150828_tscode_list.append(stock_name)
weekly_20150828_addname_dataframe=pd.DataFrame()
weekly_20150828_addname_dataframe['cname'] = weekly_20150828_tscode_list
for table_name in weekly_20150828.columns.values.tolist():
    weekly_20150828_addname_dataframe[table_name] = weekly_20150828[table_name]
print("周线行情-时间为序  weekly_20150828 35_20150828 返回数据 row 行数 = "+str(weekly_20150828.shape[0]))
weekly_20150828_addname_dataframe.to_excel(weekly_2015_excel_writer,'35_20150828',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150902")       ###  更新 周线记录日期
weekly_20150902 = pro.weekly(trade_date='20150902', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150902_tscode_list = list() 
for ts_code_sh in weekly_20150902['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150902_tscode_list.append(stock_name)
weekly_20150902_addname_dataframe=pd.DataFrame()
weekly_20150902_addname_dataframe['cname'] = weekly_20150902_tscode_list
for table_name in weekly_20150902.columns.values.tolist():
    weekly_20150902_addname_dataframe[table_name] = weekly_20150902[table_name]
print("周线行情-时间为序  weekly_20150902 35_20150902 返回数据 row 行数 = "+str(weekly_20150902.shape[0]))
weekly_20150902_addname_dataframe.to_excel(weekly_2015_excel_writer,'35_20150902',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150911")       ###  更新 周线记录日期
weekly_20150911 = pro.weekly(trade_date='20150911', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150911_tscode_list = list() 
for ts_code_sh in weekly_20150911['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150911_tscode_list.append(stock_name)
weekly_20150911_addname_dataframe=pd.DataFrame()
weekly_20150911_addname_dataframe['cname'] = weekly_20150911_tscode_list
for table_name in weekly_20150911.columns.values.tolist():
    weekly_20150911_addname_dataframe[table_name] = weekly_20150911[table_name]
print("周线行情-时间为序  weekly_20150911 37_20150911 返回数据 row 行数 = "+str(weekly_20150911.shape[0]))
weekly_20150911_addname_dataframe.to_excel(weekly_2015_excel_writer,'37_20150911',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150918")       ###  更新 周线记录日期
weekly_20150918 = pro.weekly(trade_date='20150918', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150918_tscode_list = list() 
for ts_code_sh in weekly_20150918['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150918_tscode_list.append(stock_name)
weekly_20150918_addname_dataframe=pd.DataFrame()
weekly_20150918_addname_dataframe['cname'] = weekly_20150918_tscode_list
for table_name in weekly_20150918.columns.values.tolist():
    weekly_20150918_addname_dataframe[table_name] = weekly_20150918[table_name]
print("周线行情-时间为序  weekly_20150918 38_20150918 返回数据 row 行数 = "+str(weekly_20150918.shape[0]))
weekly_20150918_addname_dataframe.to_excel(weekly_2015_excel_writer,'38_20150918',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150925")       ###  更新 周线记录日期
weekly_20150925 = pro.weekly(trade_date='20150925', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150925_tscode_list = list() 
for ts_code_sh in weekly_20150925['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150925_tscode_list.append(stock_name)
weekly_20150925_addname_dataframe=pd.DataFrame()
weekly_20150925_addname_dataframe['cname'] = weekly_20150925_tscode_list
for table_name in weekly_20150925.columns.values.tolist():
    weekly_20150925_addname_dataframe[table_name] = weekly_20150925[table_name]
print("周线行情-时间为序  weekly_20150925 39_20150925 返回数据 row 行数 = "+str(weekly_20150925.shape[0]))
weekly_20150925_addname_dataframe.to_excel(weekly_2015_excel_writer,'39_20150925',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20150930")       ###  更新 周线记录日期
weekly_20150930 = pro.weekly(trade_date='20150930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20150930_tscode_list = list() 
for ts_code_sh in weekly_20150930['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20150930_tscode_list.append(stock_name)
weekly_20150930_addname_dataframe=pd.DataFrame()
weekly_20150930_addname_dataframe['cname'] = weekly_20150930_tscode_list
for table_name in weekly_20150930.columns.values.tolist():
    weekly_20150930_addname_dataframe[table_name] = weekly_20150930[table_name]
print("周线行情-时间为序  weekly_20150930 39_20150930 返回数据 row 行数 = "+str(weekly_20150930.shape[0]))
weekly_20150930_addname_dataframe.to_excel(weekly_2015_excel_writer,'39_20150930',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151009")       ###  更新 周线记录日期
weekly_20151009 = pro.weekly(trade_date='20151009', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151009_tscode_list = list() 
for ts_code_sh in weekly_20151009['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151009_tscode_list.append(stock_name)
weekly_20151009_addname_dataframe=pd.DataFrame()
weekly_20151009_addname_dataframe['cname'] = weekly_20151009_tscode_list
for table_name in weekly_20151009.columns.values.tolist():
    weekly_20151009_addname_dataframe[table_name] = weekly_20151009[table_name]
print("周线行情-时间为序  weekly_20151009 41_20151009 返回数据 row 行数 = "+str(weekly_20151009.shape[0]))
weekly_20151009_addname_dataframe.to_excel(weekly_2015_excel_writer,'41_20151009',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151016")       ###  更新 周线记录日期
weekly_20151016 = pro.weekly(trade_date='20151016', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151016_tscode_list = list() 
for ts_code_sh in weekly_20151016['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151016_tscode_list.append(stock_name)
weekly_20151016_addname_dataframe=pd.DataFrame()
weekly_20151016_addname_dataframe['cname'] = weekly_20151016_tscode_list
for table_name in weekly_20151016.columns.values.tolist():
    weekly_20151016_addname_dataframe[table_name] = weekly_20151016[table_name]
print("周线行情-时间为序  weekly_20151016 42_20151016 返回数据 row 行数 = "+str(weekly_20151016.shape[0]))
weekly_20151016_addname_dataframe.to_excel(weekly_2015_excel_writer,'42_20151016',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151023")       ###  更新 周线记录日期
weekly_20151023 = pro.weekly(trade_date='20151023', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151023_tscode_list = list() 
for ts_code_sh in weekly_20151023['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151023_tscode_list.append(stock_name)
weekly_20151023_addname_dataframe=pd.DataFrame()
weekly_20151023_addname_dataframe['cname'] = weekly_20151023_tscode_list
for table_name in weekly_20151023.columns.values.tolist():
    weekly_20151023_addname_dataframe[table_name] = weekly_20151023[table_name]
print("周线行情-时间为序  weekly_20151023 43_20151023 返回数据 row 行数 = "+str(weekly_20151023.shape[0]))
weekly_20151023_addname_dataframe.to_excel(weekly_2015_excel_writer,'43_20151023',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151030")       ###  更新 周线记录日期
weekly_20151030 = pro.weekly(trade_date='20151030', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151030_tscode_list = list() 
for ts_code_sh in weekly_20151030['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151030_tscode_list.append(stock_name)
weekly_20151030_addname_dataframe=pd.DataFrame()
weekly_20151030_addname_dataframe['cname'] = weekly_20151030_tscode_list
for table_name in weekly_20151030.columns.values.tolist():
    weekly_20151030_addname_dataframe[table_name] = weekly_20151030[table_name]
print("周线行情-时间为序  weekly_20151030 44_20151030 返回数据 row 行数 = "+str(weekly_20151030.shape[0]))
weekly_20151030_addname_dataframe.to_excel(weekly_2015_excel_writer,'44_20151030',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151106")       ###  更新 周线记录日期
weekly_20151106 = pro.weekly(trade_date='20151106', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151106_tscode_list = list() 
for ts_code_sh in weekly_20151106['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151106_tscode_list.append(stock_name)
weekly_20151106_addname_dataframe=pd.DataFrame()
weekly_20151106_addname_dataframe['cname'] = weekly_20151106_tscode_list
for table_name in weekly_20151106.columns.values.tolist():
    weekly_20151106_addname_dataframe[table_name] = weekly_20151106[table_name]
print("周线行情-时间为序  weekly_20151106 45_20151106 返回数据 row 行数 = "+str(weekly_20151106.shape[0]))
weekly_20151106_addname_dataframe.to_excel(weekly_2015_excel_writer,'45_20151106',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151113")       ###  更新 周线记录日期
weekly_20151113 = pro.weekly(trade_date='20151113', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151113_tscode_list = list() 
for ts_code_sh in weekly_20151113['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151113_tscode_list.append(stock_name)
weekly_20151113_addname_dataframe=pd.DataFrame()
weekly_20151113_addname_dataframe['cname'] = weekly_20151113_tscode_list
for table_name in weekly_20151113.columns.values.tolist():
    weekly_20151113_addname_dataframe[table_name] = weekly_20151113[table_name]
print("周线行情-时间为序  weekly_20151113 46_20151113 返回数据 row 行数 = "+str(weekly_20151113.shape[0]))
weekly_20151113_addname_dataframe.to_excel(weekly_2015_excel_writer,'46_20151113',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151120")       ###  更新 周线记录日期
weekly_20151120 = pro.weekly(trade_date='20151120', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151120_tscode_list = list() 
for ts_code_sh in weekly_20151120['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151120_tscode_list.append(stock_name)
weekly_20151120_addname_dataframe=pd.DataFrame()
weekly_20151120_addname_dataframe['cname'] = weekly_20151120_tscode_list
for table_name in weekly_20151120.columns.values.tolist():
    weekly_20151120_addname_dataframe[table_name] = weekly_20151120[table_name]
print("周线行情-时间为序  weekly_20151120 47_20151120 返回数据 row 行数 = "+str(weekly_20151120.shape[0]))
weekly_20151120_addname_dataframe.to_excel(weekly_2015_excel_writer,'47_20151120',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151127")       ###  更新 周线记录日期
weekly_20151127 = pro.weekly(trade_date='20151127', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151127_tscode_list = list() 
for ts_code_sh in weekly_20151127['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151127_tscode_list.append(stock_name)
weekly_20151127_addname_dataframe=pd.DataFrame()
weekly_20151127_addname_dataframe['cname'] = weekly_20151127_tscode_list
for table_name in weekly_20151127.columns.values.tolist():
    weekly_20151127_addname_dataframe[table_name] = weekly_20151127[table_name]
print("周线行情-时间为序  weekly_20151127 48_20151127 返回数据 row 行数 = "+str(weekly_20151127.shape[0]))
weekly_20151127_addname_dataframe.to_excel(weekly_2015_excel_writer,'48_20151127',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151204")       ###  更新 周线记录日期
weekly_20151204 = pro.weekly(trade_date='20151204', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151204_tscode_list = list() 
for ts_code_sh in weekly_20151204['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151204_tscode_list.append(stock_name)
weekly_20151204_addname_dataframe=pd.DataFrame()
weekly_20151204_addname_dataframe['cname'] = weekly_20151204_tscode_list
for table_name in weekly_20151204.columns.values.tolist():
    weekly_20151204_addname_dataframe[table_name] = weekly_20151204[table_name]
print("周线行情-时间为序  weekly_20151204 49_20151204 返回数据 row 行数 = "+str(weekly_20151204.shape[0]))
weekly_20151204_addname_dataframe.to_excel(weekly_2015_excel_writer,'49_20151204',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151211")       ###  更新 周线记录日期
weekly_20151211 = pro.weekly(trade_date='20151211', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151211_tscode_list = list() 
for ts_code_sh in weekly_20151211['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151211_tscode_list.append(stock_name)
weekly_20151211_addname_dataframe=pd.DataFrame()
weekly_20151211_addname_dataframe['cname'] = weekly_20151211_tscode_list
for table_name in weekly_20151211.columns.values.tolist():
    weekly_20151211_addname_dataframe[table_name] = weekly_20151211[table_name]
print("周线行情-时间为序  weekly_20151211 50_20151211 返回数据 row 行数 = "+str(weekly_20151211.shape[0]))
weekly_20151211_addname_dataframe.to_excel(weekly_2015_excel_writer,'50_20151211',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151218")       ###  更新 周线记录日期
weekly_20151218 = pro.weekly(trade_date='20151218', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151218_tscode_list = list() 
for ts_code_sh in weekly_20151218['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151218_tscode_list.append(stock_name)
weekly_20151218_addname_dataframe=pd.DataFrame()
weekly_20151218_addname_dataframe['cname'] = weekly_20151218_tscode_list
for table_name in weekly_20151218.columns.values.tolist():
    weekly_20151218_addname_dataframe[table_name] = weekly_20151218[table_name]
print("周线行情-时间为序  weekly_20151218 51_20151218 返回数据 row 行数 = "+str(weekly_20151218.shape[0]))
weekly_20151218_addname_dataframe.to_excel(weekly_2015_excel_writer,'51_20151218',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151225")       ###  更新 周线记录日期
weekly_20151225 = pro.weekly(trade_date='20151225', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151225_tscode_list = list() 
for ts_code_sh in weekly_20151225['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151225_tscode_list.append(stock_name)
weekly_20151225_addname_dataframe=pd.DataFrame()
weekly_20151225_addname_dataframe['cname'] = weekly_20151225_tscode_list
for table_name in weekly_20151225.columns.values.tolist():
    weekly_20151225_addname_dataframe[table_name] = weekly_20151225[table_name]
print("周线行情-时间为序  weekly_20151225 52_20151225 返回数据 row 行数 = "+str(weekly_20151225.shape[0]))
weekly_20151225_addname_dataframe.to_excel(weekly_2015_excel_writer,'52_20151225',index=False)
weekly_2015_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20151231")       ###  更新 周线记录日期
weekly_20151231 = pro.weekly(trade_date='20151231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20151231_tscode_list = list() 
for ts_code_sh in weekly_20151231['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20151231_tscode_list.append(stock_name)
weekly_20151231_addname_dataframe=pd.DataFrame()
weekly_20151231_addname_dataframe['cname'] = weekly_20151231_tscode_list
for table_name in weekly_20151231.columns.values.tolist():
    weekly_20151231_addname_dataframe[table_name] = weekly_20151231[table_name]
print("周线行情-时间为序  weekly_20151231 53_20151231 返回数据 row 行数 = "+str(weekly_20151231.shape[0]))
weekly_20151231_addname_dataframe.to_excel(weekly_2015_excel_writer,'53_20151231',index=False)
weekly_2015_excel_writer.save()
createexcel('weekly_2016.xlsx')
weekly_2016_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2016.xlsx')
weekly_2016_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2016.xlsx', engine='openpyxl')
weekly_2016_excel_writer.book = weekly_2016_book
weekly_2016_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2016_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20160108")       ###  更新 周线记录日期
weekly_20160108 = pro.weekly(trade_date='20160108', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160108_tscode_list = list() 
for ts_code_sh in weekly_20160108['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160108_tscode_list.append(stock_name)
weekly_20160108_addname_dataframe=pd.DataFrame()
weekly_20160108_addname_dataframe['cname'] = weekly_20160108_tscode_list
for table_name in weekly_20160108.columns.values.tolist():
    weekly_20160108_addname_dataframe[table_name] = weekly_20160108[table_name]
print("周线行情-时间为序  weekly_20160108 2_20160108 返回数据 row 行数 = "+str(weekly_20160108.shape[0]))
weekly_20160108_addname_dataframe.to_excel(weekly_2016_excel_writer,'2_20160108',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160115")       ###  更新 周线记录日期
weekly_20160115 = pro.weekly(trade_date='20160115', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160115_tscode_list = list() 
for ts_code_sh in weekly_20160115['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160115_tscode_list.append(stock_name)
weekly_20160115_addname_dataframe=pd.DataFrame()
weekly_20160115_addname_dataframe['cname'] = weekly_20160115_tscode_list
for table_name in weekly_20160115.columns.values.tolist():
    weekly_20160115_addname_dataframe[table_name] = weekly_20160115[table_name]
print("周线行情-时间为序  weekly_20160115 3_20160115 返回数据 row 行数 = "+str(weekly_20160115.shape[0]))
weekly_20160115_addname_dataframe.to_excel(weekly_2016_excel_writer,'3_20160115',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160122")       ###  更新 周线记录日期
weekly_20160122 = pro.weekly(trade_date='20160122', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160122_tscode_list = list() 
for ts_code_sh in weekly_20160122['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160122_tscode_list.append(stock_name)
weekly_20160122_addname_dataframe=pd.DataFrame()
weekly_20160122_addname_dataframe['cname'] = weekly_20160122_tscode_list
for table_name in weekly_20160122.columns.values.tolist():
    weekly_20160122_addname_dataframe[table_name] = weekly_20160122[table_name]
print("周线行情-时间为序  weekly_20160122 4_20160122 返回数据 row 行数 = "+str(weekly_20160122.shape[0]))
weekly_20160122_addname_dataframe.to_excel(weekly_2016_excel_writer,'4_20160122',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160129")       ###  更新 周线记录日期
weekly_20160129 = pro.weekly(trade_date='20160129', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160129_tscode_list = list() 
for ts_code_sh in weekly_20160129['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160129_tscode_list.append(stock_name)
weekly_20160129_addname_dataframe=pd.DataFrame()
weekly_20160129_addname_dataframe['cname'] = weekly_20160129_tscode_list
for table_name in weekly_20160129.columns.values.tolist():
    weekly_20160129_addname_dataframe[table_name] = weekly_20160129[table_name]
print("周线行情-时间为序  weekly_20160129 5_20160129 返回数据 row 行数 = "+str(weekly_20160129.shape[0]))
weekly_20160129_addname_dataframe.to_excel(weekly_2016_excel_writer,'5_20160129',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160205")       ###  更新 周线记录日期
weekly_20160205 = pro.weekly(trade_date='20160205', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160205_tscode_list = list() 
for ts_code_sh in weekly_20160205['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160205_tscode_list.append(stock_name)
weekly_20160205_addname_dataframe=pd.DataFrame()
weekly_20160205_addname_dataframe['cname'] = weekly_20160205_tscode_list
for table_name in weekly_20160205.columns.values.tolist():
    weekly_20160205_addname_dataframe[table_name] = weekly_20160205[table_name]
print("周线行情-时间为序  weekly_20160205 6_20160205 返回数据 row 行数 = "+str(weekly_20160205.shape[0]))
weekly_20160205_addname_dataframe.to_excel(weekly_2016_excel_writer,'6_20160205',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160219")       ###  更新 周线记录日期
weekly_20160219 = pro.weekly(trade_date='20160219', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160219_tscode_list = list() 
for ts_code_sh in weekly_20160219['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160219_tscode_list.append(stock_name)
weekly_20160219_addname_dataframe=pd.DataFrame()
weekly_20160219_addname_dataframe['cname'] = weekly_20160219_tscode_list
for table_name in weekly_20160219.columns.values.tolist():
    weekly_20160219_addname_dataframe[table_name] = weekly_20160219[table_name]
print("周线行情-时间为序  weekly_20160219 8_20160219 返回数据 row 行数 = "+str(weekly_20160219.shape[0]))
weekly_20160219_addname_dataframe.to_excel(weekly_2016_excel_writer,'8_20160219',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160226")       ###  更新 周线记录日期
weekly_20160226 = pro.weekly(trade_date='20160226', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160226_tscode_list = list() 
for ts_code_sh in weekly_20160226['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160226_tscode_list.append(stock_name)
weekly_20160226_addname_dataframe=pd.DataFrame()
weekly_20160226_addname_dataframe['cname'] = weekly_20160226_tscode_list
for table_name in weekly_20160226.columns.values.tolist():
    weekly_20160226_addname_dataframe[table_name] = weekly_20160226[table_name]
print("周线行情-时间为序  weekly_20160226 9_20160226 返回数据 row 行数 = "+str(weekly_20160226.shape[0]))
weekly_20160226_addname_dataframe.to_excel(weekly_2016_excel_writer,'9_20160226',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160304")       ###  更新 周线记录日期
weekly_20160304 = pro.weekly(trade_date='20160304', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160304_tscode_list = list() 
for ts_code_sh in weekly_20160304['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160304_tscode_list.append(stock_name)
weekly_20160304_addname_dataframe=pd.DataFrame()
weekly_20160304_addname_dataframe['cname'] = weekly_20160304_tscode_list
for table_name in weekly_20160304.columns.values.tolist():
    weekly_20160304_addname_dataframe[table_name] = weekly_20160304[table_name]
print("周线行情-时间为序  weekly_20160304 10_20160304 返回数据 row 行数 = "+str(weekly_20160304.shape[0]))
weekly_20160304_addname_dataframe.to_excel(weekly_2016_excel_writer,'10_20160304',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160311")       ###  更新 周线记录日期
weekly_20160311 = pro.weekly(trade_date='20160311', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160311_tscode_list = list() 
for ts_code_sh in weekly_20160311['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160311_tscode_list.append(stock_name)
weekly_20160311_addname_dataframe=pd.DataFrame()
weekly_20160311_addname_dataframe['cname'] = weekly_20160311_tscode_list
for table_name in weekly_20160311.columns.values.tolist():
    weekly_20160311_addname_dataframe[table_name] = weekly_20160311[table_name]
print("周线行情-时间为序  weekly_20160311 11_20160311 返回数据 row 行数 = "+str(weekly_20160311.shape[0]))
weekly_20160311_addname_dataframe.to_excel(weekly_2016_excel_writer,'11_20160311',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160318")       ###  更新 周线记录日期
weekly_20160318 = pro.weekly(trade_date='20160318', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160318_tscode_list = list() 
for ts_code_sh in weekly_20160318['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160318_tscode_list.append(stock_name)
weekly_20160318_addname_dataframe=pd.DataFrame()
weekly_20160318_addname_dataframe['cname'] = weekly_20160318_tscode_list
for table_name in weekly_20160318.columns.values.tolist():
    weekly_20160318_addname_dataframe[table_name] = weekly_20160318[table_name]
print("周线行情-时间为序  weekly_20160318 12_20160318 返回数据 row 行数 = "+str(weekly_20160318.shape[0]))
weekly_20160318_addname_dataframe.to_excel(weekly_2016_excel_writer,'12_20160318',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160325")       ###  更新 周线记录日期
weekly_20160325 = pro.weekly(trade_date='20160325', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160325_tscode_list = list() 
for ts_code_sh in weekly_20160325['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160325_tscode_list.append(stock_name)
weekly_20160325_addname_dataframe=pd.DataFrame()
weekly_20160325_addname_dataframe['cname'] = weekly_20160325_tscode_list
for table_name in weekly_20160325.columns.values.tolist():
    weekly_20160325_addname_dataframe[table_name] = weekly_20160325[table_name]
print("周线行情-时间为序  weekly_20160325 13_20160325 返回数据 row 行数 = "+str(weekly_20160325.shape[0]))
weekly_20160325_addname_dataframe.to_excel(weekly_2016_excel_writer,'13_20160325',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160401")       ###  更新 周线记录日期
weekly_20160401 = pro.weekly(trade_date='20160401', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160401_tscode_list = list() 
for ts_code_sh in weekly_20160401['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160401_tscode_list.append(stock_name)
weekly_20160401_addname_dataframe=pd.DataFrame()
weekly_20160401_addname_dataframe['cname'] = weekly_20160401_tscode_list
for table_name in weekly_20160401.columns.values.tolist():
    weekly_20160401_addname_dataframe[table_name] = weekly_20160401[table_name]
print("周线行情-时间为序  weekly_20160401 14_20160401 返回数据 row 行数 = "+str(weekly_20160401.shape[0]))
weekly_20160401_addname_dataframe.to_excel(weekly_2016_excel_writer,'14_20160401',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160408")       ###  更新 周线记录日期
weekly_20160408 = pro.weekly(trade_date='20160408', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160408_tscode_list = list() 
for ts_code_sh in weekly_20160408['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160408_tscode_list.append(stock_name)
weekly_20160408_addname_dataframe=pd.DataFrame()
weekly_20160408_addname_dataframe['cname'] = weekly_20160408_tscode_list
for table_name in weekly_20160408.columns.values.tolist():
    weekly_20160408_addname_dataframe[table_name] = weekly_20160408[table_name]
print("周线行情-时间为序  weekly_20160408 15_20160408 返回数据 row 行数 = "+str(weekly_20160408.shape[0]))
weekly_20160408_addname_dataframe.to_excel(weekly_2016_excel_writer,'15_20160408',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160415")       ###  更新 周线记录日期
weekly_20160415 = pro.weekly(trade_date='20160415', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160415_tscode_list = list() 
for ts_code_sh in weekly_20160415['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160415_tscode_list.append(stock_name)
weekly_20160415_addname_dataframe=pd.DataFrame()
weekly_20160415_addname_dataframe['cname'] = weekly_20160415_tscode_list
for table_name in weekly_20160415.columns.values.tolist():
    weekly_20160415_addname_dataframe[table_name] = weekly_20160415[table_name]
print("周线行情-时间为序  weekly_20160415 16_20160415 返回数据 row 行数 = "+str(weekly_20160415.shape[0]))
weekly_20160415_addname_dataframe.to_excel(weekly_2016_excel_writer,'16_20160415',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160422")       ###  更新 周线记录日期
weekly_20160422 = pro.weekly(trade_date='20160422', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160422_tscode_list = list() 
for ts_code_sh in weekly_20160422['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160422_tscode_list.append(stock_name)
weekly_20160422_addname_dataframe=pd.DataFrame()
weekly_20160422_addname_dataframe['cname'] = weekly_20160422_tscode_list
for table_name in weekly_20160422.columns.values.tolist():
    weekly_20160422_addname_dataframe[table_name] = weekly_20160422[table_name]
print("周线行情-时间为序  weekly_20160422 17_20160422 返回数据 row 行数 = "+str(weekly_20160422.shape[0]))
weekly_20160422_addname_dataframe.to_excel(weekly_2016_excel_writer,'17_20160422',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160429")       ###  更新 周线记录日期
weekly_20160429 = pro.weekly(trade_date='20160429', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160429_tscode_list = list() 
for ts_code_sh in weekly_20160429['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160429_tscode_list.append(stock_name)
weekly_20160429_addname_dataframe=pd.DataFrame()
weekly_20160429_addname_dataframe['cname'] = weekly_20160429_tscode_list
for table_name in weekly_20160429.columns.values.tolist():
    weekly_20160429_addname_dataframe[table_name] = weekly_20160429[table_name]
print("周线行情-时间为序  weekly_20160429 18_20160429 返回数据 row 行数 = "+str(weekly_20160429.shape[0]))
weekly_20160429_addname_dataframe.to_excel(weekly_2016_excel_writer,'18_20160429',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160506")       ###  更新 周线记录日期
weekly_20160506 = pro.weekly(trade_date='20160506', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160506_tscode_list = list() 
for ts_code_sh in weekly_20160506['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160506_tscode_list.append(stock_name)
weekly_20160506_addname_dataframe=pd.DataFrame()
weekly_20160506_addname_dataframe['cname'] = weekly_20160506_tscode_list
for table_name in weekly_20160506.columns.values.tolist():
    weekly_20160506_addname_dataframe[table_name] = weekly_20160506[table_name]
print("周线行情-时间为序  weekly_20160506 19_20160506 返回数据 row 行数 = "+str(weekly_20160506.shape[0]))
weekly_20160506_addname_dataframe.to_excel(weekly_2016_excel_writer,'19_20160506',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160513")       ###  更新 周线记录日期
weekly_20160513 = pro.weekly(trade_date='20160513', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160513_tscode_list = list() 
for ts_code_sh in weekly_20160513['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160513_tscode_list.append(stock_name)
weekly_20160513_addname_dataframe=pd.DataFrame()
weekly_20160513_addname_dataframe['cname'] = weekly_20160513_tscode_list
for table_name in weekly_20160513.columns.values.tolist():
    weekly_20160513_addname_dataframe[table_name] = weekly_20160513[table_name]
print("周线行情-时间为序  weekly_20160513 20_20160513 返回数据 row 行数 = "+str(weekly_20160513.shape[0]))
weekly_20160513_addname_dataframe.to_excel(weekly_2016_excel_writer,'20_20160513',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160520")       ###  更新 周线记录日期
weekly_20160520 = pro.weekly(trade_date='20160520', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160520_tscode_list = list() 
for ts_code_sh in weekly_20160520['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160520_tscode_list.append(stock_name)
weekly_20160520_addname_dataframe=pd.DataFrame()
weekly_20160520_addname_dataframe['cname'] = weekly_20160520_tscode_list
for table_name in weekly_20160520.columns.values.tolist():
    weekly_20160520_addname_dataframe[table_name] = weekly_20160520[table_name]
print("周线行情-时间为序  weekly_20160520 21_20160520 返回数据 row 行数 = "+str(weekly_20160520.shape[0]))
weekly_20160520_addname_dataframe.to_excel(weekly_2016_excel_writer,'21_20160520',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160527")       ###  更新 周线记录日期
weekly_20160527 = pro.weekly(trade_date='20160527', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160527_tscode_list = list() 
for ts_code_sh in weekly_20160527['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160527_tscode_list.append(stock_name)
weekly_20160527_addname_dataframe=pd.DataFrame()
weekly_20160527_addname_dataframe['cname'] = weekly_20160527_tscode_list
for table_name in weekly_20160527.columns.values.tolist():
    weekly_20160527_addname_dataframe[table_name] = weekly_20160527[table_name]
print("周线行情-时间为序  weekly_20160527 22_20160527 返回数据 row 行数 = "+str(weekly_20160527.shape[0]))
weekly_20160527_addname_dataframe.to_excel(weekly_2016_excel_writer,'22_20160527',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160603")       ###  更新 周线记录日期
weekly_20160603 = pro.weekly(trade_date='20160603', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160603_tscode_list = list() 
for ts_code_sh in weekly_20160603['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160603_tscode_list.append(stock_name)
weekly_20160603_addname_dataframe=pd.DataFrame()
weekly_20160603_addname_dataframe['cname'] = weekly_20160603_tscode_list
for table_name in weekly_20160603.columns.values.tolist():
    weekly_20160603_addname_dataframe[table_name] = weekly_20160603[table_name]
print("周线行情-时间为序  weekly_20160603 23_20160603 返回数据 row 行数 = "+str(weekly_20160603.shape[0]))
weekly_20160603_addname_dataframe.to_excel(weekly_2016_excel_writer,'23_20160603',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160608")       ###  更新 周线记录日期
weekly_20160608 = pro.weekly(trade_date='20160608', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160608_tscode_list = list() 
for ts_code_sh in weekly_20160608['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160608_tscode_list.append(stock_name)
weekly_20160608_addname_dataframe=pd.DataFrame()
weekly_20160608_addname_dataframe['cname'] = weekly_20160608_tscode_list
for table_name in weekly_20160608.columns.values.tolist():
    weekly_20160608_addname_dataframe[table_name] = weekly_20160608[table_name]
print("周线行情-时间为序  weekly_20160608 23_20160608 返回数据 row 行数 = "+str(weekly_20160608.shape[0]))
weekly_20160608_addname_dataframe.to_excel(weekly_2016_excel_writer,'23_20160608',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160617")       ###  更新 周线记录日期
weekly_20160617 = pro.weekly(trade_date='20160617', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160617_tscode_list = list() 
for ts_code_sh in weekly_20160617['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160617_tscode_list.append(stock_name)
weekly_20160617_addname_dataframe=pd.DataFrame()
weekly_20160617_addname_dataframe['cname'] = weekly_20160617_tscode_list
for table_name in weekly_20160617.columns.values.tolist():
    weekly_20160617_addname_dataframe[table_name] = weekly_20160617[table_name]
print("周线行情-时间为序  weekly_20160617 25_20160617 返回数据 row 行数 = "+str(weekly_20160617.shape[0]))
weekly_20160617_addname_dataframe.to_excel(weekly_2016_excel_writer,'25_20160617',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160624")       ###  更新 周线记录日期
weekly_20160624 = pro.weekly(trade_date='20160624', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160624_tscode_list = list() 
for ts_code_sh in weekly_20160624['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160624_tscode_list.append(stock_name)
weekly_20160624_addname_dataframe=pd.DataFrame()
weekly_20160624_addname_dataframe['cname'] = weekly_20160624_tscode_list
for table_name in weekly_20160624.columns.values.tolist():
    weekly_20160624_addname_dataframe[table_name] = weekly_20160624[table_name]
print("周线行情-时间为序  weekly_20160624 26_20160624 返回数据 row 行数 = "+str(weekly_20160624.shape[0]))
weekly_20160624_addname_dataframe.to_excel(weekly_2016_excel_writer,'26_20160624',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160701")       ###  更新 周线记录日期
weekly_20160701 = pro.weekly(trade_date='20160701', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160701_tscode_list = list() 
for ts_code_sh in weekly_20160701['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160701_tscode_list.append(stock_name)
weekly_20160701_addname_dataframe=pd.DataFrame()
weekly_20160701_addname_dataframe['cname'] = weekly_20160701_tscode_list
for table_name in weekly_20160701.columns.values.tolist():
    weekly_20160701_addname_dataframe[table_name] = weekly_20160701[table_name]
print("周线行情-时间为序  weekly_20160701 27_20160701 返回数据 row 行数 = "+str(weekly_20160701.shape[0]))
weekly_20160701_addname_dataframe.to_excel(weekly_2016_excel_writer,'27_20160701',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160708")       ###  更新 周线记录日期
weekly_20160708 = pro.weekly(trade_date='20160708', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160708_tscode_list = list() 
for ts_code_sh in weekly_20160708['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160708_tscode_list.append(stock_name)
weekly_20160708_addname_dataframe=pd.DataFrame()
weekly_20160708_addname_dataframe['cname'] = weekly_20160708_tscode_list
for table_name in weekly_20160708.columns.values.tolist():
    weekly_20160708_addname_dataframe[table_name] = weekly_20160708[table_name]
print("周线行情-时间为序  weekly_20160708 28_20160708 返回数据 row 行数 = "+str(weekly_20160708.shape[0]))
weekly_20160708_addname_dataframe.to_excel(weekly_2016_excel_writer,'28_20160708',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160715")       ###  更新 周线记录日期
weekly_20160715 = pro.weekly(trade_date='20160715', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160715_tscode_list = list() 
for ts_code_sh in weekly_20160715['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160715_tscode_list.append(stock_name)
weekly_20160715_addname_dataframe=pd.DataFrame()
weekly_20160715_addname_dataframe['cname'] = weekly_20160715_tscode_list
for table_name in weekly_20160715.columns.values.tolist():
    weekly_20160715_addname_dataframe[table_name] = weekly_20160715[table_name]
print("周线行情-时间为序  weekly_20160715 29_20160715 返回数据 row 行数 = "+str(weekly_20160715.shape[0]))
weekly_20160715_addname_dataframe.to_excel(weekly_2016_excel_writer,'29_20160715',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160722")       ###  更新 周线记录日期
weekly_20160722 = pro.weekly(trade_date='20160722', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160722_tscode_list = list() 
for ts_code_sh in weekly_20160722['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160722_tscode_list.append(stock_name)
weekly_20160722_addname_dataframe=pd.DataFrame()
weekly_20160722_addname_dataframe['cname'] = weekly_20160722_tscode_list
for table_name in weekly_20160722.columns.values.tolist():
    weekly_20160722_addname_dataframe[table_name] = weekly_20160722[table_name]
print("周线行情-时间为序  weekly_20160722 30_20160722 返回数据 row 行数 = "+str(weekly_20160722.shape[0]))
weekly_20160722_addname_dataframe.to_excel(weekly_2016_excel_writer,'30_20160722',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160729")       ###  更新 周线记录日期
weekly_20160729 = pro.weekly(trade_date='20160729', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160729_tscode_list = list() 
for ts_code_sh in weekly_20160729['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160729_tscode_list.append(stock_name)
weekly_20160729_addname_dataframe=pd.DataFrame()
weekly_20160729_addname_dataframe['cname'] = weekly_20160729_tscode_list
for table_name in weekly_20160729.columns.values.tolist():
    weekly_20160729_addname_dataframe[table_name] = weekly_20160729[table_name]
print("周线行情-时间为序  weekly_20160729 31_20160729 返回数据 row 行数 = "+str(weekly_20160729.shape[0]))
weekly_20160729_addname_dataframe.to_excel(weekly_2016_excel_writer,'31_20160729',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160805")       ###  更新 周线记录日期
weekly_20160805 = pro.weekly(trade_date='20160805', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160805_tscode_list = list() 
for ts_code_sh in weekly_20160805['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160805_tscode_list.append(stock_name)
weekly_20160805_addname_dataframe=pd.DataFrame()
weekly_20160805_addname_dataframe['cname'] = weekly_20160805_tscode_list
for table_name in weekly_20160805.columns.values.tolist():
    weekly_20160805_addname_dataframe[table_name] = weekly_20160805[table_name]
print("周线行情-时间为序  weekly_20160805 32_20160805 返回数据 row 行数 = "+str(weekly_20160805.shape[0]))
weekly_20160805_addname_dataframe.to_excel(weekly_2016_excel_writer,'32_20160805',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160812")       ###  更新 周线记录日期
weekly_20160812 = pro.weekly(trade_date='20160812', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160812_tscode_list = list() 
for ts_code_sh in weekly_20160812['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160812_tscode_list.append(stock_name)
weekly_20160812_addname_dataframe=pd.DataFrame()
weekly_20160812_addname_dataframe['cname'] = weekly_20160812_tscode_list
for table_name in weekly_20160812.columns.values.tolist():
    weekly_20160812_addname_dataframe[table_name] = weekly_20160812[table_name]
print("周线行情-时间为序  weekly_20160812 33_20160812 返回数据 row 行数 = "+str(weekly_20160812.shape[0]))
weekly_20160812_addname_dataframe.to_excel(weekly_2016_excel_writer,'33_20160812',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160819")       ###  更新 周线记录日期
weekly_20160819 = pro.weekly(trade_date='20160819', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160819_tscode_list = list() 
for ts_code_sh in weekly_20160819['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160819_tscode_list.append(stock_name)
weekly_20160819_addname_dataframe=pd.DataFrame()
weekly_20160819_addname_dataframe['cname'] = weekly_20160819_tscode_list
for table_name in weekly_20160819.columns.values.tolist():
    weekly_20160819_addname_dataframe[table_name] = weekly_20160819[table_name]
print("周线行情-时间为序  weekly_20160819 34_20160819 返回数据 row 行数 = "+str(weekly_20160819.shape[0]))
weekly_20160819_addname_dataframe.to_excel(weekly_2016_excel_writer,'34_20160819',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160826")       ###  更新 周线记录日期
weekly_20160826 = pro.weekly(trade_date='20160826', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160826_tscode_list = list() 
for ts_code_sh in weekly_20160826['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160826_tscode_list.append(stock_name)
weekly_20160826_addname_dataframe=pd.DataFrame()
weekly_20160826_addname_dataframe['cname'] = weekly_20160826_tscode_list
for table_name in weekly_20160826.columns.values.tolist():
    weekly_20160826_addname_dataframe[table_name] = weekly_20160826[table_name]
print("周线行情-时间为序  weekly_20160826 35_20160826 返回数据 row 行数 = "+str(weekly_20160826.shape[0]))
weekly_20160826_addname_dataframe.to_excel(weekly_2016_excel_writer,'35_20160826',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160902")       ###  更新 周线记录日期
weekly_20160902 = pro.weekly(trade_date='20160902', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160902_tscode_list = list() 
for ts_code_sh in weekly_20160902['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160902_tscode_list.append(stock_name)
weekly_20160902_addname_dataframe=pd.DataFrame()
weekly_20160902_addname_dataframe['cname'] = weekly_20160902_tscode_list
for table_name in weekly_20160902.columns.values.tolist():
    weekly_20160902_addname_dataframe[table_name] = weekly_20160902[table_name]
print("周线行情-时间为序  weekly_20160902 36_20160902 返回数据 row 行数 = "+str(weekly_20160902.shape[0]))
weekly_20160902_addname_dataframe.to_excel(weekly_2016_excel_writer,'36_20160902',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160909")       ###  更新 周线记录日期
weekly_20160909 = pro.weekly(trade_date='20160909', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160909_tscode_list = list() 
for ts_code_sh in weekly_20160909['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160909_tscode_list.append(stock_name)
weekly_20160909_addname_dataframe=pd.DataFrame()
weekly_20160909_addname_dataframe['cname'] = weekly_20160909_tscode_list
for table_name in weekly_20160909.columns.values.tolist():
    weekly_20160909_addname_dataframe[table_name] = weekly_20160909[table_name]
print("周线行情-时间为序  weekly_20160909 37_20160909 返回数据 row 行数 = "+str(weekly_20160909.shape[0]))
weekly_20160909_addname_dataframe.to_excel(weekly_2016_excel_writer,'37_20160909',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160914")       ###  更新 周线记录日期
weekly_20160914 = pro.weekly(trade_date='20160914', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160914_tscode_list = list() 
for ts_code_sh in weekly_20160914['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160914_tscode_list.append(stock_name)
weekly_20160914_addname_dataframe=pd.DataFrame()
weekly_20160914_addname_dataframe['cname'] = weekly_20160914_tscode_list
for table_name in weekly_20160914.columns.values.tolist():
    weekly_20160914_addname_dataframe[table_name] = weekly_20160914[table_name]
print("周线行情-时间为序  weekly_20160914 37_20160914 返回数据 row 行数 = "+str(weekly_20160914.shape[0]))
weekly_20160914_addname_dataframe.to_excel(weekly_2016_excel_writer,'37_20160914',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160923")       ###  更新 周线记录日期
weekly_20160923 = pro.weekly(trade_date='20160923', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160923_tscode_list = list() 
for ts_code_sh in weekly_20160923['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160923_tscode_list.append(stock_name)
weekly_20160923_addname_dataframe=pd.DataFrame()
weekly_20160923_addname_dataframe['cname'] = weekly_20160923_tscode_list
for table_name in weekly_20160923.columns.values.tolist():
    weekly_20160923_addname_dataframe[table_name] = weekly_20160923[table_name]
print("周线行情-时间为序  weekly_20160923 39_20160923 返回数据 row 行数 = "+str(weekly_20160923.shape[0]))
weekly_20160923_addname_dataframe.to_excel(weekly_2016_excel_writer,'39_20160923',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20160930")       ###  更新 周线记录日期
weekly_20160930 = pro.weekly(trade_date='20160930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20160930_tscode_list = list() 
for ts_code_sh in weekly_20160930['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20160930_tscode_list.append(stock_name)
weekly_20160930_addname_dataframe=pd.DataFrame()
weekly_20160930_addname_dataframe['cname'] = weekly_20160930_tscode_list
for table_name in weekly_20160930.columns.values.tolist():
    weekly_20160930_addname_dataframe[table_name] = weekly_20160930[table_name]
print("周线行情-时间为序  weekly_20160930 40_20160930 返回数据 row 行数 = "+str(weekly_20160930.shape[0]))
weekly_20160930_addname_dataframe.to_excel(weekly_2016_excel_writer,'40_20160930',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161014")       ###  更新 周线记录日期
weekly_20161014 = pro.weekly(trade_date='20161014', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161014_tscode_list = list() 
for ts_code_sh in weekly_20161014['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161014_tscode_list.append(stock_name)
weekly_20161014_addname_dataframe=pd.DataFrame()
weekly_20161014_addname_dataframe['cname'] = weekly_20161014_tscode_list
for table_name in weekly_20161014.columns.values.tolist():
    weekly_20161014_addname_dataframe[table_name] = weekly_20161014[table_name]
print("周线行情-时间为序  weekly_20161014 42_20161014 返回数据 row 行数 = "+str(weekly_20161014.shape[0]))
weekly_20161014_addname_dataframe.to_excel(weekly_2016_excel_writer,'42_20161014',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161021")       ###  更新 周线记录日期
weekly_20161021 = pro.weekly(trade_date='20161021', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161021_tscode_list = list() 
for ts_code_sh in weekly_20161021['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161021_tscode_list.append(stock_name)
weekly_20161021_addname_dataframe=pd.DataFrame()
weekly_20161021_addname_dataframe['cname'] = weekly_20161021_tscode_list
for table_name in weekly_20161021.columns.values.tolist():
    weekly_20161021_addname_dataframe[table_name] = weekly_20161021[table_name]
print("周线行情-时间为序  weekly_20161021 43_20161021 返回数据 row 行数 = "+str(weekly_20161021.shape[0]))
weekly_20161021_addname_dataframe.to_excel(weekly_2016_excel_writer,'43_20161021',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161028")       ###  更新 周线记录日期
weekly_20161028 = pro.weekly(trade_date='20161028', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161028_tscode_list = list() 
for ts_code_sh in weekly_20161028['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161028_tscode_list.append(stock_name)
weekly_20161028_addname_dataframe=pd.DataFrame()
weekly_20161028_addname_dataframe['cname'] = weekly_20161028_tscode_list
for table_name in weekly_20161028.columns.values.tolist():
    weekly_20161028_addname_dataframe[table_name] = weekly_20161028[table_name]
print("周线行情-时间为序  weekly_20161028 44_20161028 返回数据 row 行数 = "+str(weekly_20161028.shape[0]))
weekly_20161028_addname_dataframe.to_excel(weekly_2016_excel_writer,'44_20161028',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161104")       ###  更新 周线记录日期
weekly_20161104 = pro.weekly(trade_date='20161104', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161104_tscode_list = list() 
for ts_code_sh in weekly_20161104['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161104_tscode_list.append(stock_name)
weekly_20161104_addname_dataframe=pd.DataFrame()
weekly_20161104_addname_dataframe['cname'] = weekly_20161104_tscode_list
for table_name in weekly_20161104.columns.values.tolist():
    weekly_20161104_addname_dataframe[table_name] = weekly_20161104[table_name]
print("周线行情-时间为序  weekly_20161104 45_20161104 返回数据 row 行数 = "+str(weekly_20161104.shape[0]))
weekly_20161104_addname_dataframe.to_excel(weekly_2016_excel_writer,'45_20161104',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161111")       ###  更新 周线记录日期
weekly_20161111 = pro.weekly(trade_date='20161111', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161111_tscode_list = list() 
for ts_code_sh in weekly_20161111['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161111_tscode_list.append(stock_name)
weekly_20161111_addname_dataframe=pd.DataFrame()
weekly_20161111_addname_dataframe['cname'] = weekly_20161111_tscode_list
for table_name in weekly_20161111.columns.values.tolist():
    weekly_20161111_addname_dataframe[table_name] = weekly_20161111[table_name]
print("周线行情-时间为序  weekly_20161111 46_20161111 返回数据 row 行数 = "+str(weekly_20161111.shape[0]))
weekly_20161111_addname_dataframe.to_excel(weekly_2016_excel_writer,'46_20161111',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161118")       ###  更新 周线记录日期
weekly_20161118 = pro.weekly(trade_date='20161118', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161118_tscode_list = list() 
for ts_code_sh in weekly_20161118['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161118_tscode_list.append(stock_name)
weekly_20161118_addname_dataframe=pd.DataFrame()
weekly_20161118_addname_dataframe['cname'] = weekly_20161118_tscode_list
for table_name in weekly_20161118.columns.values.tolist():
    weekly_20161118_addname_dataframe[table_name] = weekly_20161118[table_name]
print("周线行情-时间为序  weekly_20161118 47_20161118 返回数据 row 行数 = "+str(weekly_20161118.shape[0]))
weekly_20161118_addname_dataframe.to_excel(weekly_2016_excel_writer,'47_20161118',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161125")       ###  更新 周线记录日期
weekly_20161125 = pro.weekly(trade_date='20161125', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161125_tscode_list = list() 
for ts_code_sh in weekly_20161125['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161125_tscode_list.append(stock_name)
weekly_20161125_addname_dataframe=pd.DataFrame()
weekly_20161125_addname_dataframe['cname'] = weekly_20161125_tscode_list
for table_name in weekly_20161125.columns.values.tolist():
    weekly_20161125_addname_dataframe[table_name] = weekly_20161125[table_name]
print("周线行情-时间为序  weekly_20161125 48_20161125 返回数据 row 行数 = "+str(weekly_20161125.shape[0]))
weekly_20161125_addname_dataframe.to_excel(weekly_2016_excel_writer,'48_20161125',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161202")       ###  更新 周线记录日期
weekly_20161202 = pro.weekly(trade_date='20161202', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161202_tscode_list = list() 
for ts_code_sh in weekly_20161202['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161202_tscode_list.append(stock_name)
weekly_20161202_addname_dataframe=pd.DataFrame()
weekly_20161202_addname_dataframe['cname'] = weekly_20161202_tscode_list
for table_name in weekly_20161202.columns.values.tolist():
    weekly_20161202_addname_dataframe[table_name] = weekly_20161202[table_name]
print("周线行情-时间为序  weekly_20161202 49_20161202 返回数据 row 行数 = "+str(weekly_20161202.shape[0]))
weekly_20161202_addname_dataframe.to_excel(weekly_2016_excel_writer,'49_20161202',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161209")       ###  更新 周线记录日期
weekly_20161209 = pro.weekly(trade_date='20161209', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161209_tscode_list = list() 
for ts_code_sh in weekly_20161209['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161209_tscode_list.append(stock_name)
weekly_20161209_addname_dataframe=pd.DataFrame()
weekly_20161209_addname_dataframe['cname'] = weekly_20161209_tscode_list
for table_name in weekly_20161209.columns.values.tolist():
    weekly_20161209_addname_dataframe[table_name] = weekly_20161209[table_name]
print("周线行情-时间为序  weekly_20161209 50_20161209 返回数据 row 行数 = "+str(weekly_20161209.shape[0]))
weekly_20161209_addname_dataframe.to_excel(weekly_2016_excel_writer,'50_20161209',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161216")       ###  更新 周线记录日期
weekly_20161216 = pro.weekly(trade_date='20161216', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161216_tscode_list = list() 
for ts_code_sh in weekly_20161216['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161216_tscode_list.append(stock_name)
weekly_20161216_addname_dataframe=pd.DataFrame()
weekly_20161216_addname_dataframe['cname'] = weekly_20161216_tscode_list
for table_name in weekly_20161216.columns.values.tolist():
    weekly_20161216_addname_dataframe[table_name] = weekly_20161216[table_name]
print("周线行情-时间为序  weekly_20161216 51_20161216 返回数据 row 行数 = "+str(weekly_20161216.shape[0]))
weekly_20161216_addname_dataframe.to_excel(weekly_2016_excel_writer,'51_20161216',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161223")       ###  更新 周线记录日期
weekly_20161223 = pro.weekly(trade_date='20161223', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161223_tscode_list = list() 
for ts_code_sh in weekly_20161223['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161223_tscode_list.append(stock_name)
weekly_20161223_addname_dataframe=pd.DataFrame()
weekly_20161223_addname_dataframe['cname'] = weekly_20161223_tscode_list
for table_name in weekly_20161223.columns.values.tolist():
    weekly_20161223_addname_dataframe[table_name] = weekly_20161223[table_name]
print("周线行情-时间为序  weekly_20161223 52_20161223 返回数据 row 行数 = "+str(weekly_20161223.shape[0]))
weekly_20161223_addname_dataframe.to_excel(weekly_2016_excel_writer,'52_20161223',index=False)
weekly_2016_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20161230")       ###  更新 周线记录日期
weekly_20161230 = pro.weekly(trade_date='20161230', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20161230_tscode_list = list() 
for ts_code_sh in weekly_20161230['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20161230_tscode_list.append(stock_name)
weekly_20161230_addname_dataframe=pd.DataFrame()
weekly_20161230_addname_dataframe['cname'] = weekly_20161230_tscode_list
for table_name in weekly_20161230.columns.values.tolist():
    weekly_20161230_addname_dataframe[table_name] = weekly_20161230[table_name]
print("周线行情-时间为序  weekly_20161230 53_20161230 返回数据 row 行数 = "+str(weekly_20161230.shape[0]))
weekly_20161230_addname_dataframe.to_excel(weekly_2016_excel_writer,'53_20161230',index=False)
weekly_2016_excel_writer.save()
createexcel('weekly_2017.xlsx')
weekly_2017_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2017.xlsx')
weekly_2017_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2017.xlsx', engine='openpyxl')
weekly_2017_excel_writer.book = weekly_2017_book
weekly_2017_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2017_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20170106")       ###  更新 周线记录日期
weekly_20170106 = pro.weekly(trade_date='20170106', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170106_tscode_list = list() 
for ts_code_sh in weekly_20170106['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170106_tscode_list.append(stock_name)
weekly_20170106_addname_dataframe=pd.DataFrame()
weekly_20170106_addname_dataframe['cname'] = weekly_20170106_tscode_list
for table_name in weekly_20170106.columns.values.tolist():
    weekly_20170106_addname_dataframe[table_name] = weekly_20170106[table_name]
print("周线行情-时间为序  weekly_20170106 1_20170106 返回数据 row 行数 = "+str(weekly_20170106.shape[0]))
weekly_20170106_addname_dataframe.to_excel(weekly_2017_excel_writer,'1_20170106',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170113")       ###  更新 周线记录日期
weekly_20170113 = pro.weekly(trade_date='20170113', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170113_tscode_list = list() 
for ts_code_sh in weekly_20170113['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170113_tscode_list.append(stock_name)
weekly_20170113_addname_dataframe=pd.DataFrame()
weekly_20170113_addname_dataframe['cname'] = weekly_20170113_tscode_list
for table_name in weekly_20170113.columns.values.tolist():
    weekly_20170113_addname_dataframe[table_name] = weekly_20170113[table_name]
print("周线行情-时间为序  weekly_20170113 2_20170113 返回数据 row 行数 = "+str(weekly_20170113.shape[0]))
weekly_20170113_addname_dataframe.to_excel(weekly_2017_excel_writer,'2_20170113',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170120")       ###  更新 周线记录日期
weekly_20170120 = pro.weekly(trade_date='20170120', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170120_tscode_list = list() 
for ts_code_sh in weekly_20170120['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170120_tscode_list.append(stock_name)
weekly_20170120_addname_dataframe=pd.DataFrame()
weekly_20170120_addname_dataframe['cname'] = weekly_20170120_tscode_list
for table_name in weekly_20170120.columns.values.tolist():
    weekly_20170120_addname_dataframe[table_name] = weekly_20170120[table_name]
print("周线行情-时间为序  weekly_20170120 3_20170120 返回数据 row 行数 = "+str(weekly_20170120.shape[0]))
weekly_20170120_addname_dataframe.to_excel(weekly_2017_excel_writer,'3_20170120',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170126")       ###  更新 周线记录日期
weekly_20170126 = pro.weekly(trade_date='20170126', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170126_tscode_list = list() 
for ts_code_sh in weekly_20170126['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170126_tscode_list.append(stock_name)
weekly_20170126_addname_dataframe=pd.DataFrame()
weekly_20170126_addname_dataframe['cname'] = weekly_20170126_tscode_list
for table_name in weekly_20170126.columns.values.tolist():
    weekly_20170126_addname_dataframe[table_name] = weekly_20170126[table_name]
print("周线行情-时间为序  weekly_20170126 4_20170126 返回数据 row 行数 = "+str(weekly_20170126.shape[0]))
weekly_20170126_addname_dataframe.to_excel(weekly_2017_excel_writer,'4_20170126',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170203")       ###  更新 周线记录日期
weekly_20170203 = pro.weekly(trade_date='20170203', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170203_tscode_list = list() 
for ts_code_sh in weekly_20170203['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170203_tscode_list.append(stock_name)
weekly_20170203_addname_dataframe=pd.DataFrame()
weekly_20170203_addname_dataframe['cname'] = weekly_20170203_tscode_list
for table_name in weekly_20170203.columns.values.tolist():
    weekly_20170203_addname_dataframe[table_name] = weekly_20170203[table_name]
print("周线行情-时间为序  weekly_20170203 5_20170203 返回数据 row 行数 = "+str(weekly_20170203.shape[0]))
weekly_20170203_addname_dataframe.to_excel(weekly_2017_excel_writer,'5_20170203',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170210")       ###  更新 周线记录日期
weekly_20170210 = pro.weekly(trade_date='20170210', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170210_tscode_list = list() 
for ts_code_sh in weekly_20170210['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170210_tscode_list.append(stock_name)
weekly_20170210_addname_dataframe=pd.DataFrame()
weekly_20170210_addname_dataframe['cname'] = weekly_20170210_tscode_list
for table_name in weekly_20170210.columns.values.tolist():
    weekly_20170210_addname_dataframe[table_name] = weekly_20170210[table_name]
print("周线行情-时间为序  weekly_20170210 6_20170210 返回数据 row 行数 = "+str(weekly_20170210.shape[0]))
weekly_20170210_addname_dataframe.to_excel(weekly_2017_excel_writer,'6_20170210',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170217")       ###  更新 周线记录日期
weekly_20170217 = pro.weekly(trade_date='20170217', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170217_tscode_list = list() 
for ts_code_sh in weekly_20170217['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170217_tscode_list.append(stock_name)
weekly_20170217_addname_dataframe=pd.DataFrame()
weekly_20170217_addname_dataframe['cname'] = weekly_20170217_tscode_list
for table_name in weekly_20170217.columns.values.tolist():
    weekly_20170217_addname_dataframe[table_name] = weekly_20170217[table_name]
print("周线行情-时间为序  weekly_20170217 7_20170217 返回数据 row 行数 = "+str(weekly_20170217.shape[0]))
weekly_20170217_addname_dataframe.to_excel(weekly_2017_excel_writer,'7_20170217',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170224")       ###  更新 周线记录日期
weekly_20170224 = pro.weekly(trade_date='20170224', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170224_tscode_list = list() 
for ts_code_sh in weekly_20170224['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170224_tscode_list.append(stock_name)
weekly_20170224_addname_dataframe=pd.DataFrame()
weekly_20170224_addname_dataframe['cname'] = weekly_20170224_tscode_list
for table_name in weekly_20170224.columns.values.tolist():
    weekly_20170224_addname_dataframe[table_name] = weekly_20170224[table_name]
print("周线行情-时间为序  weekly_20170224 8_20170224 返回数据 row 行数 = "+str(weekly_20170224.shape[0]))
weekly_20170224_addname_dataframe.to_excel(weekly_2017_excel_writer,'8_20170224',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170303")       ###  更新 周线记录日期
weekly_20170303 = pro.weekly(trade_date='20170303', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170303_tscode_list = list() 
for ts_code_sh in weekly_20170303['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170303_tscode_list.append(stock_name)
weekly_20170303_addname_dataframe=pd.DataFrame()
weekly_20170303_addname_dataframe['cname'] = weekly_20170303_tscode_list
for table_name in weekly_20170303.columns.values.tolist():
    weekly_20170303_addname_dataframe[table_name] = weekly_20170303[table_name]
print("周线行情-时间为序  weekly_20170303 9_20170303 返回数据 row 行数 = "+str(weekly_20170303.shape[0]))
weekly_20170303_addname_dataframe.to_excel(weekly_2017_excel_writer,'9_20170303',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170310")       ###  更新 周线记录日期
weekly_20170310 = pro.weekly(trade_date='20170310', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170310_tscode_list = list() 
for ts_code_sh in weekly_20170310['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170310_tscode_list.append(stock_name)
weekly_20170310_addname_dataframe=pd.DataFrame()
weekly_20170310_addname_dataframe['cname'] = weekly_20170310_tscode_list
for table_name in weekly_20170310.columns.values.tolist():
    weekly_20170310_addname_dataframe[table_name] = weekly_20170310[table_name]
print("周线行情-时间为序  weekly_20170310 10_20170310 返回数据 row 行数 = "+str(weekly_20170310.shape[0]))
weekly_20170310_addname_dataframe.to_excel(weekly_2017_excel_writer,'10_20170310',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170317")       ###  更新 周线记录日期
weekly_20170317 = pro.weekly(trade_date='20170317', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170317_tscode_list = list() 
for ts_code_sh in weekly_20170317['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170317_tscode_list.append(stock_name)
weekly_20170317_addname_dataframe=pd.DataFrame()
weekly_20170317_addname_dataframe['cname'] = weekly_20170317_tscode_list
for table_name in weekly_20170317.columns.values.tolist():
    weekly_20170317_addname_dataframe[table_name] = weekly_20170317[table_name]
print("周线行情-时间为序  weekly_20170317 11_20170317 返回数据 row 行数 = "+str(weekly_20170317.shape[0]))
weekly_20170317_addname_dataframe.to_excel(weekly_2017_excel_writer,'11_20170317',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170324")       ###  更新 周线记录日期
weekly_20170324 = pro.weekly(trade_date='20170324', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170324_tscode_list = list() 
for ts_code_sh in weekly_20170324['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170324_tscode_list.append(stock_name)
weekly_20170324_addname_dataframe=pd.DataFrame()
weekly_20170324_addname_dataframe['cname'] = weekly_20170324_tscode_list
for table_name in weekly_20170324.columns.values.tolist():
    weekly_20170324_addname_dataframe[table_name] = weekly_20170324[table_name]
print("周线行情-时间为序  weekly_20170324 12_20170324 返回数据 row 行数 = "+str(weekly_20170324.shape[0]))
weekly_20170324_addname_dataframe.to_excel(weekly_2017_excel_writer,'12_20170324',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170331")       ###  更新 周线记录日期
weekly_20170331 = pro.weekly(trade_date='20170331', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170331_tscode_list = list() 
for ts_code_sh in weekly_20170331['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170331_tscode_list.append(stock_name)
weekly_20170331_addname_dataframe=pd.DataFrame()
weekly_20170331_addname_dataframe['cname'] = weekly_20170331_tscode_list
for table_name in weekly_20170331.columns.values.tolist():
    weekly_20170331_addname_dataframe[table_name] = weekly_20170331[table_name]
print("周线行情-时间为序  weekly_20170331 13_20170331 返回数据 row 行数 = "+str(weekly_20170331.shape[0]))
weekly_20170331_addname_dataframe.to_excel(weekly_2017_excel_writer,'13_20170331',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170407")       ###  更新 周线记录日期
weekly_20170407 = pro.weekly(trade_date='20170407', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170407_tscode_list = list() 
for ts_code_sh in weekly_20170407['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170407_tscode_list.append(stock_name)
weekly_20170407_addname_dataframe=pd.DataFrame()
weekly_20170407_addname_dataframe['cname'] = weekly_20170407_tscode_list
for table_name in weekly_20170407.columns.values.tolist():
    weekly_20170407_addname_dataframe[table_name] = weekly_20170407[table_name]
print("周线行情-时间为序  weekly_20170407 14_20170407 返回数据 row 行数 = "+str(weekly_20170407.shape[0]))
weekly_20170407_addname_dataframe.to_excel(weekly_2017_excel_writer,'14_20170407',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170414")       ###  更新 周线记录日期
weekly_20170414 = pro.weekly(trade_date='20170414', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170414_tscode_list = list() 
for ts_code_sh in weekly_20170414['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170414_tscode_list.append(stock_name)
weekly_20170414_addname_dataframe=pd.DataFrame()
weekly_20170414_addname_dataframe['cname'] = weekly_20170414_tscode_list
for table_name in weekly_20170414.columns.values.tolist():
    weekly_20170414_addname_dataframe[table_name] = weekly_20170414[table_name]
print("周线行情-时间为序  weekly_20170414 15_20170414 返回数据 row 行数 = "+str(weekly_20170414.shape[0]))
weekly_20170414_addname_dataframe.to_excel(weekly_2017_excel_writer,'15_20170414',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170421")       ###  更新 周线记录日期
weekly_20170421 = pro.weekly(trade_date='20170421', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170421_tscode_list = list() 
for ts_code_sh in weekly_20170421['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170421_tscode_list.append(stock_name)
weekly_20170421_addname_dataframe=pd.DataFrame()
weekly_20170421_addname_dataframe['cname'] = weekly_20170421_tscode_list
for table_name in weekly_20170421.columns.values.tolist():
    weekly_20170421_addname_dataframe[table_name] = weekly_20170421[table_name]
print("周线行情-时间为序  weekly_20170421 16_20170421 返回数据 row 行数 = "+str(weekly_20170421.shape[0]))
weekly_20170421_addname_dataframe.to_excel(weekly_2017_excel_writer,'16_20170421',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170428")       ###  更新 周线记录日期
weekly_20170428 = pro.weekly(trade_date='20170428', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170428_tscode_list = list() 
for ts_code_sh in weekly_20170428['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170428_tscode_list.append(stock_name)
weekly_20170428_addname_dataframe=pd.DataFrame()
weekly_20170428_addname_dataframe['cname'] = weekly_20170428_tscode_list
for table_name in weekly_20170428.columns.values.tolist():
    weekly_20170428_addname_dataframe[table_name] = weekly_20170428[table_name]
print("周线行情-时间为序  weekly_20170428 17_20170428 返回数据 row 行数 = "+str(weekly_20170428.shape[0]))
weekly_20170428_addname_dataframe.to_excel(weekly_2017_excel_writer,'17_20170428',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170505")       ###  更新 周线记录日期
weekly_20170505 = pro.weekly(trade_date='20170505', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170505_tscode_list = list() 
for ts_code_sh in weekly_20170505['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170505_tscode_list.append(stock_name)
weekly_20170505_addname_dataframe=pd.DataFrame()
weekly_20170505_addname_dataframe['cname'] = weekly_20170505_tscode_list
for table_name in weekly_20170505.columns.values.tolist():
    weekly_20170505_addname_dataframe[table_name] = weekly_20170505[table_name]
print("周线行情-时间为序  weekly_20170505 18_20170505 返回数据 row 行数 = "+str(weekly_20170505.shape[0]))
weekly_20170505_addname_dataframe.to_excel(weekly_2017_excel_writer,'18_20170505',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170512")       ###  更新 周线记录日期
weekly_20170512 = pro.weekly(trade_date='20170512', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170512_tscode_list = list() 
for ts_code_sh in weekly_20170512['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170512_tscode_list.append(stock_name)
weekly_20170512_addname_dataframe=pd.DataFrame()
weekly_20170512_addname_dataframe['cname'] = weekly_20170512_tscode_list
for table_name in weekly_20170512.columns.values.tolist():
    weekly_20170512_addname_dataframe[table_name] = weekly_20170512[table_name]
print("周线行情-时间为序  weekly_20170512 19_20170512 返回数据 row 行数 = "+str(weekly_20170512.shape[0]))
weekly_20170512_addname_dataframe.to_excel(weekly_2017_excel_writer,'19_20170512',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170519")       ###  更新 周线记录日期
weekly_20170519 = pro.weekly(trade_date='20170519', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170519_tscode_list = list() 
for ts_code_sh in weekly_20170519['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170519_tscode_list.append(stock_name)
weekly_20170519_addname_dataframe=pd.DataFrame()
weekly_20170519_addname_dataframe['cname'] = weekly_20170519_tscode_list
for table_name in weekly_20170519.columns.values.tolist():
    weekly_20170519_addname_dataframe[table_name] = weekly_20170519[table_name]
print("周线行情-时间为序  weekly_20170519 20_20170519 返回数据 row 行数 = "+str(weekly_20170519.shape[0]))
weekly_20170519_addname_dataframe.to_excel(weekly_2017_excel_writer,'20_20170519',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170526")       ###  更新 周线记录日期
weekly_20170526 = pro.weekly(trade_date='20170526', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170526_tscode_list = list() 
for ts_code_sh in weekly_20170526['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170526_tscode_list.append(stock_name)
weekly_20170526_addname_dataframe=pd.DataFrame()
weekly_20170526_addname_dataframe['cname'] = weekly_20170526_tscode_list
for table_name in weekly_20170526.columns.values.tolist():
    weekly_20170526_addname_dataframe[table_name] = weekly_20170526[table_name]
print("周线行情-时间为序  weekly_20170526 21_20170526 返回数据 row 行数 = "+str(weekly_20170526.shape[0]))
weekly_20170526_addname_dataframe.to_excel(weekly_2017_excel_writer,'21_20170526',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170602")       ###  更新 周线记录日期
weekly_20170602 = pro.weekly(trade_date='20170602', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170602_tscode_list = list() 
for ts_code_sh in weekly_20170602['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170602_tscode_list.append(stock_name)
weekly_20170602_addname_dataframe=pd.DataFrame()
weekly_20170602_addname_dataframe['cname'] = weekly_20170602_tscode_list
for table_name in weekly_20170602.columns.values.tolist():
    weekly_20170602_addname_dataframe[table_name] = weekly_20170602[table_name]
print("周线行情-时间为序  weekly_20170602 22_20170602 返回数据 row 行数 = "+str(weekly_20170602.shape[0]))
weekly_20170602_addname_dataframe.to_excel(weekly_2017_excel_writer,'22_20170602',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170609")       ###  更新 周线记录日期
weekly_20170609 = pro.weekly(trade_date='20170609', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170609_tscode_list = list() 
for ts_code_sh in weekly_20170609['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170609_tscode_list.append(stock_name)
weekly_20170609_addname_dataframe=pd.DataFrame()
weekly_20170609_addname_dataframe['cname'] = weekly_20170609_tscode_list
for table_name in weekly_20170609.columns.values.tolist():
    weekly_20170609_addname_dataframe[table_name] = weekly_20170609[table_name]
print("周线行情-时间为序  weekly_20170609 23_20170609 返回数据 row 行数 = "+str(weekly_20170609.shape[0]))
weekly_20170609_addname_dataframe.to_excel(weekly_2017_excel_writer,'23_20170609',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170616")       ###  更新 周线记录日期
weekly_20170616 = pro.weekly(trade_date='20170616', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170616_tscode_list = list() 
for ts_code_sh in weekly_20170616['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170616_tscode_list.append(stock_name)
weekly_20170616_addname_dataframe=pd.DataFrame()
weekly_20170616_addname_dataframe['cname'] = weekly_20170616_tscode_list
for table_name in weekly_20170616.columns.values.tolist():
    weekly_20170616_addname_dataframe[table_name] = weekly_20170616[table_name]
print("周线行情-时间为序  weekly_20170616 24_20170616 返回数据 row 行数 = "+str(weekly_20170616.shape[0]))
weekly_20170616_addname_dataframe.to_excel(weekly_2017_excel_writer,'24_20170616',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170623")       ###  更新 周线记录日期
weekly_20170623 = pro.weekly(trade_date='20170623', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170623_tscode_list = list() 
for ts_code_sh in weekly_20170623['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170623_tscode_list.append(stock_name)
weekly_20170623_addname_dataframe=pd.DataFrame()
weekly_20170623_addname_dataframe['cname'] = weekly_20170623_tscode_list
for table_name in weekly_20170623.columns.values.tolist():
    weekly_20170623_addname_dataframe[table_name] = weekly_20170623[table_name]
print("周线行情-时间为序  weekly_20170623 25_20170623 返回数据 row 行数 = "+str(weekly_20170623.shape[0]))
weekly_20170623_addname_dataframe.to_excel(weekly_2017_excel_writer,'25_20170623',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170630")       ###  更新 周线记录日期
weekly_20170630 = pro.weekly(trade_date='20170630', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170630_tscode_list = list() 
for ts_code_sh in weekly_20170630['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170630_tscode_list.append(stock_name)
weekly_20170630_addname_dataframe=pd.DataFrame()
weekly_20170630_addname_dataframe['cname'] = weekly_20170630_tscode_list
for table_name in weekly_20170630.columns.values.tolist():
    weekly_20170630_addname_dataframe[table_name] = weekly_20170630[table_name]
print("周线行情-时间为序  weekly_20170630 26_20170630 返回数据 row 行数 = "+str(weekly_20170630.shape[0]))
weekly_20170630_addname_dataframe.to_excel(weekly_2017_excel_writer,'26_20170630',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170707")       ###  更新 周线记录日期
weekly_20170707 = pro.weekly(trade_date='20170707', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170707_tscode_list = list() 
for ts_code_sh in weekly_20170707['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170707_tscode_list.append(stock_name)
weekly_20170707_addname_dataframe=pd.DataFrame()
weekly_20170707_addname_dataframe['cname'] = weekly_20170707_tscode_list
for table_name in weekly_20170707.columns.values.tolist():
    weekly_20170707_addname_dataframe[table_name] = weekly_20170707[table_name]
print("周线行情-时间为序  weekly_20170707 27_20170707 返回数据 row 行数 = "+str(weekly_20170707.shape[0]))
weekly_20170707_addname_dataframe.to_excel(weekly_2017_excel_writer,'27_20170707',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170714")       ###  更新 周线记录日期
weekly_20170714 = pro.weekly(trade_date='20170714', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170714_tscode_list = list() 
for ts_code_sh in weekly_20170714['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170714_tscode_list.append(stock_name)
weekly_20170714_addname_dataframe=pd.DataFrame()
weekly_20170714_addname_dataframe['cname'] = weekly_20170714_tscode_list
for table_name in weekly_20170714.columns.values.tolist():
    weekly_20170714_addname_dataframe[table_name] = weekly_20170714[table_name]
print("周线行情-时间为序  weekly_20170714 28_20170714 返回数据 row 行数 = "+str(weekly_20170714.shape[0]))
weekly_20170714_addname_dataframe.to_excel(weekly_2017_excel_writer,'28_20170714',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170721")       ###  更新 周线记录日期
weekly_20170721 = pro.weekly(trade_date='20170721', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170721_tscode_list = list() 
for ts_code_sh in weekly_20170721['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170721_tscode_list.append(stock_name)
weekly_20170721_addname_dataframe=pd.DataFrame()
weekly_20170721_addname_dataframe['cname'] = weekly_20170721_tscode_list
for table_name in weekly_20170721.columns.values.tolist():
    weekly_20170721_addname_dataframe[table_name] = weekly_20170721[table_name]
print("周线行情-时间为序  weekly_20170721 29_20170721 返回数据 row 行数 = "+str(weekly_20170721.shape[0]))
weekly_20170721_addname_dataframe.to_excel(weekly_2017_excel_writer,'29_20170721',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170728")       ###  更新 周线记录日期
weekly_20170728 = pro.weekly(trade_date='20170728', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170728_tscode_list = list() 
for ts_code_sh in weekly_20170728['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170728_tscode_list.append(stock_name)
weekly_20170728_addname_dataframe=pd.DataFrame()
weekly_20170728_addname_dataframe['cname'] = weekly_20170728_tscode_list
for table_name in weekly_20170728.columns.values.tolist():
    weekly_20170728_addname_dataframe[table_name] = weekly_20170728[table_name]
print("周线行情-时间为序  weekly_20170728 30_20170728 返回数据 row 行数 = "+str(weekly_20170728.shape[0]))
weekly_20170728_addname_dataframe.to_excel(weekly_2017_excel_writer,'30_20170728',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170804")       ###  更新 周线记录日期
weekly_20170804 = pro.weekly(trade_date='20170804', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170804_tscode_list = list() 
for ts_code_sh in weekly_20170804['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170804_tscode_list.append(stock_name)
weekly_20170804_addname_dataframe=pd.DataFrame()
weekly_20170804_addname_dataframe['cname'] = weekly_20170804_tscode_list
for table_name in weekly_20170804.columns.values.tolist():
    weekly_20170804_addname_dataframe[table_name] = weekly_20170804[table_name]
print("周线行情-时间为序  weekly_20170804 31_20170804 返回数据 row 行数 = "+str(weekly_20170804.shape[0]))
weekly_20170804_addname_dataframe.to_excel(weekly_2017_excel_writer,'31_20170804',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170811")       ###  更新 周线记录日期
weekly_20170811 = pro.weekly(trade_date='20170811', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170811_tscode_list = list() 
for ts_code_sh in weekly_20170811['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170811_tscode_list.append(stock_name)
weekly_20170811_addname_dataframe=pd.DataFrame()
weekly_20170811_addname_dataframe['cname'] = weekly_20170811_tscode_list
for table_name in weekly_20170811.columns.values.tolist():
    weekly_20170811_addname_dataframe[table_name] = weekly_20170811[table_name]
print("周线行情-时间为序  weekly_20170811 32_20170811 返回数据 row 行数 = "+str(weekly_20170811.shape[0]))
weekly_20170811_addname_dataframe.to_excel(weekly_2017_excel_writer,'32_20170811',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170818")       ###  更新 周线记录日期
weekly_20170818 = pro.weekly(trade_date='20170818', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170818_tscode_list = list() 
for ts_code_sh in weekly_20170818['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170818_tscode_list.append(stock_name)
weekly_20170818_addname_dataframe=pd.DataFrame()
weekly_20170818_addname_dataframe['cname'] = weekly_20170818_tscode_list
for table_name in weekly_20170818.columns.values.tolist():
    weekly_20170818_addname_dataframe[table_name] = weekly_20170818[table_name]
print("周线行情-时间为序  weekly_20170818 33_20170818 返回数据 row 行数 = "+str(weekly_20170818.shape[0]))
weekly_20170818_addname_dataframe.to_excel(weekly_2017_excel_writer,'33_20170818',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170825")       ###  更新 周线记录日期
weekly_20170825 = pro.weekly(trade_date='20170825', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170825_tscode_list = list() 
for ts_code_sh in weekly_20170825['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170825_tscode_list.append(stock_name)
weekly_20170825_addname_dataframe=pd.DataFrame()
weekly_20170825_addname_dataframe['cname'] = weekly_20170825_tscode_list
for table_name in weekly_20170825.columns.values.tolist():
    weekly_20170825_addname_dataframe[table_name] = weekly_20170825[table_name]
print("周线行情-时间为序  weekly_20170825 34_20170825 返回数据 row 行数 = "+str(weekly_20170825.shape[0]))
weekly_20170825_addname_dataframe.to_excel(weekly_2017_excel_writer,'34_20170825',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170901")       ###  更新 周线记录日期
weekly_20170901 = pro.weekly(trade_date='20170901', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170901_tscode_list = list() 
for ts_code_sh in weekly_20170901['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170901_tscode_list.append(stock_name)
weekly_20170901_addname_dataframe=pd.DataFrame()
weekly_20170901_addname_dataframe['cname'] = weekly_20170901_tscode_list
for table_name in weekly_20170901.columns.values.tolist():
    weekly_20170901_addname_dataframe[table_name] = weekly_20170901[table_name]
print("周线行情-时间为序  weekly_20170901 35_20170901 返回数据 row 行数 = "+str(weekly_20170901.shape[0]))
weekly_20170901_addname_dataframe.to_excel(weekly_2017_excel_writer,'35_20170901',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170908")       ###  更新 周线记录日期
weekly_20170908 = pro.weekly(trade_date='20170908', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170908_tscode_list = list() 
for ts_code_sh in weekly_20170908['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170908_tscode_list.append(stock_name)
weekly_20170908_addname_dataframe=pd.DataFrame()
weekly_20170908_addname_dataframe['cname'] = weekly_20170908_tscode_list
for table_name in weekly_20170908.columns.values.tolist():
    weekly_20170908_addname_dataframe[table_name] = weekly_20170908[table_name]
print("周线行情-时间为序  weekly_20170908 36_20170908 返回数据 row 行数 = "+str(weekly_20170908.shape[0]))
weekly_20170908_addname_dataframe.to_excel(weekly_2017_excel_writer,'36_20170908',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170915")       ###  更新 周线记录日期
weekly_20170915 = pro.weekly(trade_date='20170915', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170915_tscode_list = list() 
for ts_code_sh in weekly_20170915['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170915_tscode_list.append(stock_name)
weekly_20170915_addname_dataframe=pd.DataFrame()
weekly_20170915_addname_dataframe['cname'] = weekly_20170915_tscode_list
for table_name in weekly_20170915.columns.values.tolist():
    weekly_20170915_addname_dataframe[table_name] = weekly_20170915[table_name]
print("周线行情-时间为序  weekly_20170915 37_20170915 返回数据 row 行数 = "+str(weekly_20170915.shape[0]))
weekly_20170915_addname_dataframe.to_excel(weekly_2017_excel_writer,'37_20170915',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170922")       ###  更新 周线记录日期
weekly_20170922 = pro.weekly(trade_date='20170922', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170922_tscode_list = list() 
for ts_code_sh in weekly_20170922['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170922_tscode_list.append(stock_name)
weekly_20170922_addname_dataframe=pd.DataFrame()
weekly_20170922_addname_dataframe['cname'] = weekly_20170922_tscode_list
for table_name in weekly_20170922.columns.values.tolist():
    weekly_20170922_addname_dataframe[table_name] = weekly_20170922[table_name]
print("周线行情-时间为序  weekly_20170922 38_20170922 返回数据 row 行数 = "+str(weekly_20170922.shape[0]))
weekly_20170922_addname_dataframe.to_excel(weekly_2017_excel_writer,'38_20170922',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20170929")       ###  更新 周线记录日期
weekly_20170929 = pro.weekly(trade_date='20170929', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20170929_tscode_list = list() 
for ts_code_sh in weekly_20170929['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20170929_tscode_list.append(stock_name)
weekly_20170929_addname_dataframe=pd.DataFrame()
weekly_20170929_addname_dataframe['cname'] = weekly_20170929_tscode_list
for table_name in weekly_20170929.columns.values.tolist():
    weekly_20170929_addname_dataframe[table_name] = weekly_20170929[table_name]
print("周线行情-时间为序  weekly_20170929 39_20170929 返回数据 row 行数 = "+str(weekly_20170929.shape[0]))
weekly_20170929_addname_dataframe.to_excel(weekly_2017_excel_writer,'39_20170929',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171013")       ###  更新 周线记录日期
weekly_20171013 = pro.weekly(trade_date='20171013', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171013_tscode_list = list() 
for ts_code_sh in weekly_20171013['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171013_tscode_list.append(stock_name)
weekly_20171013_addname_dataframe=pd.DataFrame()
weekly_20171013_addname_dataframe['cname'] = weekly_20171013_tscode_list
for table_name in weekly_20171013.columns.values.tolist():
    weekly_20171013_addname_dataframe[table_name] = weekly_20171013[table_name]
print("周线行情-时间为序  weekly_20171013 41_20171013 返回数据 row 行数 = "+str(weekly_20171013.shape[0]))
weekly_20171013_addname_dataframe.to_excel(weekly_2017_excel_writer,'41_20171013',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171020")       ###  更新 周线记录日期
weekly_20171020 = pro.weekly(trade_date='20171020', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171020_tscode_list = list() 
for ts_code_sh in weekly_20171020['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171020_tscode_list.append(stock_name)
weekly_20171020_addname_dataframe=pd.DataFrame()
weekly_20171020_addname_dataframe['cname'] = weekly_20171020_tscode_list
for table_name in weekly_20171020.columns.values.tolist():
    weekly_20171020_addname_dataframe[table_name] = weekly_20171020[table_name]
print("周线行情-时间为序  weekly_20171020 42_20171020 返回数据 row 行数 = "+str(weekly_20171020.shape[0]))
weekly_20171020_addname_dataframe.to_excel(weekly_2017_excel_writer,'42_20171020',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171027")       ###  更新 周线记录日期
weekly_20171027 = pro.weekly(trade_date='20171027', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171027_tscode_list = list() 
for ts_code_sh in weekly_20171027['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171027_tscode_list.append(stock_name)
weekly_20171027_addname_dataframe=pd.DataFrame()
weekly_20171027_addname_dataframe['cname'] = weekly_20171027_tscode_list
for table_name in weekly_20171027.columns.values.tolist():
    weekly_20171027_addname_dataframe[table_name] = weekly_20171027[table_name]
print("周线行情-时间为序  weekly_20171027 43_20171027 返回数据 row 行数 = "+str(weekly_20171027.shape[0]))
weekly_20171027_addname_dataframe.to_excel(weekly_2017_excel_writer,'43_20171027',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171103")       ###  更新 周线记录日期
weekly_20171103 = pro.weekly(trade_date='20171103', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171103_tscode_list = list() 
for ts_code_sh in weekly_20171103['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171103_tscode_list.append(stock_name)
weekly_20171103_addname_dataframe=pd.DataFrame()
weekly_20171103_addname_dataframe['cname'] = weekly_20171103_tscode_list
for table_name in weekly_20171103.columns.values.tolist():
    weekly_20171103_addname_dataframe[table_name] = weekly_20171103[table_name]
print("周线行情-时间为序  weekly_20171103 44_20171103 返回数据 row 行数 = "+str(weekly_20171103.shape[0]))
weekly_20171103_addname_dataframe.to_excel(weekly_2017_excel_writer,'44_20171103',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171110")       ###  更新 周线记录日期
weekly_20171110 = pro.weekly(trade_date='20171110', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171110_tscode_list = list() 
for ts_code_sh in weekly_20171110['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171110_tscode_list.append(stock_name)
weekly_20171110_addname_dataframe=pd.DataFrame()
weekly_20171110_addname_dataframe['cname'] = weekly_20171110_tscode_list
for table_name in weekly_20171110.columns.values.tolist():
    weekly_20171110_addname_dataframe[table_name] = weekly_20171110[table_name]
print("周线行情-时间为序  weekly_20171110 45_20171110 返回数据 row 行数 = "+str(weekly_20171110.shape[0]))
weekly_20171110_addname_dataframe.to_excel(weekly_2017_excel_writer,'45_20171110',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171117")       ###  更新 周线记录日期
weekly_20171117 = pro.weekly(trade_date='20171117', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171117_tscode_list = list() 
for ts_code_sh in weekly_20171117['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171117_tscode_list.append(stock_name)
weekly_20171117_addname_dataframe=pd.DataFrame()
weekly_20171117_addname_dataframe['cname'] = weekly_20171117_tscode_list
for table_name in weekly_20171117.columns.values.tolist():
    weekly_20171117_addname_dataframe[table_name] = weekly_20171117[table_name]
print("周线行情-时间为序  weekly_20171117 46_20171117 返回数据 row 行数 = "+str(weekly_20171117.shape[0]))
weekly_20171117_addname_dataframe.to_excel(weekly_2017_excel_writer,'46_20171117',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171124")       ###  更新 周线记录日期
weekly_20171124 = pro.weekly(trade_date='20171124', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171124_tscode_list = list() 
for ts_code_sh in weekly_20171124['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171124_tscode_list.append(stock_name)
weekly_20171124_addname_dataframe=pd.DataFrame()
weekly_20171124_addname_dataframe['cname'] = weekly_20171124_tscode_list
for table_name in weekly_20171124.columns.values.tolist():
    weekly_20171124_addname_dataframe[table_name] = weekly_20171124[table_name]
print("周线行情-时间为序  weekly_20171124 47_20171124 返回数据 row 行数 = "+str(weekly_20171124.shape[0]))
weekly_20171124_addname_dataframe.to_excel(weekly_2017_excel_writer,'47_20171124',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171201")       ###  更新 周线记录日期
weekly_20171201 = pro.weekly(trade_date='20171201', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171201_tscode_list = list() 
for ts_code_sh in weekly_20171201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171201_tscode_list.append(stock_name)
weekly_20171201_addname_dataframe=pd.DataFrame()
weekly_20171201_addname_dataframe['cname'] = weekly_20171201_tscode_list
for table_name in weekly_20171201.columns.values.tolist():
    weekly_20171201_addname_dataframe[table_name] = weekly_20171201[table_name]
print("周线行情-时间为序  weekly_20171201 48_20171201 返回数据 row 行数 = "+str(weekly_20171201.shape[0]))
weekly_20171201_addname_dataframe.to_excel(weekly_2017_excel_writer,'48_20171201',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171208")       ###  更新 周线记录日期
weekly_20171208 = pro.weekly(trade_date='20171208', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171208_tscode_list = list() 
for ts_code_sh in weekly_20171208['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171208_tscode_list.append(stock_name)
weekly_20171208_addname_dataframe=pd.DataFrame()
weekly_20171208_addname_dataframe['cname'] = weekly_20171208_tscode_list
for table_name in weekly_20171208.columns.values.tolist():
    weekly_20171208_addname_dataframe[table_name] = weekly_20171208[table_name]
print("周线行情-时间为序  weekly_20171208 49_20171208 返回数据 row 行数 = "+str(weekly_20171208.shape[0]))
weekly_20171208_addname_dataframe.to_excel(weekly_2017_excel_writer,'49_20171208',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171215")       ###  更新 周线记录日期
weekly_20171215 = pro.weekly(trade_date='20171215', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171215_tscode_list = list() 
for ts_code_sh in weekly_20171215['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171215_tscode_list.append(stock_name)
weekly_20171215_addname_dataframe=pd.DataFrame()
weekly_20171215_addname_dataframe['cname'] = weekly_20171215_tscode_list
for table_name in weekly_20171215.columns.values.tolist():
    weekly_20171215_addname_dataframe[table_name] = weekly_20171215[table_name]
print("周线行情-时间为序  weekly_20171215 50_20171215 返回数据 row 行数 = "+str(weekly_20171215.shape[0]))
weekly_20171215_addname_dataframe.to_excel(weekly_2017_excel_writer,'50_20171215',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171222")       ###  更新 周线记录日期
weekly_20171222 = pro.weekly(trade_date='20171222', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171222_tscode_list = list() 
for ts_code_sh in weekly_20171222['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171222_tscode_list.append(stock_name)
weekly_20171222_addname_dataframe=pd.DataFrame()
weekly_20171222_addname_dataframe['cname'] = weekly_20171222_tscode_list
for table_name in weekly_20171222.columns.values.tolist():
    weekly_20171222_addname_dataframe[table_name] = weekly_20171222[table_name]
print("周线行情-时间为序  weekly_20171222 51_20171222 返回数据 row 行数 = "+str(weekly_20171222.shape[0]))
weekly_20171222_addname_dataframe.to_excel(weekly_2017_excel_writer,'51_20171222',index=False)
weekly_2017_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20171229")       ###  更新 周线记录日期
weekly_20171229 = pro.weekly(trade_date='20171229', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20171229_tscode_list = list() 
for ts_code_sh in weekly_20171229['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20171229_tscode_list.append(stock_name)
weekly_20171229_addname_dataframe=pd.DataFrame()
weekly_20171229_addname_dataframe['cname'] = weekly_20171229_tscode_list
for table_name in weekly_20171229.columns.values.tolist():
    weekly_20171229_addname_dataframe[table_name] = weekly_20171229[table_name]
print("周线行情-时间为序  weekly_20171229 52_20171229 返回数据 row 行数 = "+str(weekly_20171229.shape[0]))
weekly_20171229_addname_dataframe.to_excel(weekly_2017_excel_writer,'52_20171229',index=False)
weekly_2017_excel_writer.save()
createexcel('weekly_2018.xlsx')
weekly_2018_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2018.xlsx')
weekly_2018_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2018.xlsx', engine='openpyxl')
weekly_2018_excel_writer.book = weekly_2018_book
weekly_2018_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2018_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20180105")       ###  更新 周线记录日期
weekly_20180105 = pro.weekly(trade_date='20180105', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180105_tscode_list = list() 
for ts_code_sh in weekly_20180105['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180105_tscode_list.append(stock_name)
weekly_20180105_addname_dataframe=pd.DataFrame()
weekly_20180105_addname_dataframe['cname'] = weekly_20180105_tscode_list
for table_name in weekly_20180105.columns.values.tolist():
    weekly_20180105_addname_dataframe[table_name] = weekly_20180105[table_name]
print("周线行情-时间为序  weekly_20180105 1_20180105 返回数据 row 行数 = "+str(weekly_20180105.shape[0]))
weekly_20180105_addname_dataframe.to_excel(weekly_2018_excel_writer,'1_20180105',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180112")       ###  更新 周线记录日期
weekly_20180112 = pro.weekly(trade_date='20180112', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180112_tscode_list = list() 
for ts_code_sh in weekly_20180112['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180112_tscode_list.append(stock_name)
weekly_20180112_addname_dataframe=pd.DataFrame()
weekly_20180112_addname_dataframe['cname'] = weekly_20180112_tscode_list
for table_name in weekly_20180112.columns.values.tolist():
    weekly_20180112_addname_dataframe[table_name] = weekly_20180112[table_name]
print("周线行情-时间为序  weekly_20180112 2_20180112 返回数据 row 行数 = "+str(weekly_20180112.shape[0]))
weekly_20180112_addname_dataframe.to_excel(weekly_2018_excel_writer,'2_20180112',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180119")       ###  更新 周线记录日期
weekly_20180119 = pro.weekly(trade_date='20180119', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180119_tscode_list = list() 
for ts_code_sh in weekly_20180119['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180119_tscode_list.append(stock_name)
weekly_20180119_addname_dataframe=pd.DataFrame()
weekly_20180119_addname_dataframe['cname'] = weekly_20180119_tscode_list
for table_name in weekly_20180119.columns.values.tolist():
    weekly_20180119_addname_dataframe[table_name] = weekly_20180119[table_name]
print("周线行情-时间为序  weekly_20180119 3_20180119 返回数据 row 行数 = "+str(weekly_20180119.shape[0]))
weekly_20180119_addname_dataframe.to_excel(weekly_2018_excel_writer,'3_20180119',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180126")       ###  更新 周线记录日期
weekly_20180126 = pro.weekly(trade_date='20180126', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180126_tscode_list = list() 
for ts_code_sh in weekly_20180126['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180126_tscode_list.append(stock_name)
weekly_20180126_addname_dataframe=pd.DataFrame()
weekly_20180126_addname_dataframe['cname'] = weekly_20180126_tscode_list
for table_name in weekly_20180126.columns.values.tolist():
    weekly_20180126_addname_dataframe[table_name] = weekly_20180126[table_name]
print("周线行情-时间为序  weekly_20180126 4_20180126 返回数据 row 行数 = "+str(weekly_20180126.shape[0]))
weekly_20180126_addname_dataframe.to_excel(weekly_2018_excel_writer,'4_20180126',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180202")       ###  更新 周线记录日期
weekly_20180202 = pro.weekly(trade_date='20180202', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180202_tscode_list = list() 
for ts_code_sh in weekly_20180202['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180202_tscode_list.append(stock_name)
weekly_20180202_addname_dataframe=pd.DataFrame()
weekly_20180202_addname_dataframe['cname'] = weekly_20180202_tscode_list
for table_name in weekly_20180202.columns.values.tolist():
    weekly_20180202_addname_dataframe[table_name] = weekly_20180202[table_name]
print("周线行情-时间为序  weekly_20180202 5_20180202 返回数据 row 行数 = "+str(weekly_20180202.shape[0]))
weekly_20180202_addname_dataframe.to_excel(weekly_2018_excel_writer,'5_20180202',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180209")       ###  更新 周线记录日期
weekly_20180209 = pro.weekly(trade_date='20180209', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180209_tscode_list = list() 
for ts_code_sh in weekly_20180209['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180209_tscode_list.append(stock_name)
weekly_20180209_addname_dataframe=pd.DataFrame()
weekly_20180209_addname_dataframe['cname'] = weekly_20180209_tscode_list
for table_name in weekly_20180209.columns.values.tolist():
    weekly_20180209_addname_dataframe[table_name] = weekly_20180209[table_name]
print("周线行情-时间为序  weekly_20180209 6_20180209 返回数据 row 行数 = "+str(weekly_20180209.shape[0]))
weekly_20180209_addname_dataframe.to_excel(weekly_2018_excel_writer,'6_20180209',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180214")       ###  更新 周线记录日期
weekly_20180214 = pro.weekly(trade_date='20180214', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180214_tscode_list = list() 
for ts_code_sh in weekly_20180214['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180214_tscode_list.append(stock_name)
weekly_20180214_addname_dataframe=pd.DataFrame()
weekly_20180214_addname_dataframe['cname'] = weekly_20180214_tscode_list
for table_name in weekly_20180214.columns.values.tolist():
    weekly_20180214_addname_dataframe[table_name] = weekly_20180214[table_name]
print("周线行情-时间为序  weekly_20180214 7_20180214 返回数据 row 行数 = "+str(weekly_20180214.shape[0]))
weekly_20180214_addname_dataframe.to_excel(weekly_2018_excel_writer,'7_20180214',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180223")       ###  更新 周线记录日期
weekly_20180223 = pro.weekly(trade_date='20180223', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180223_tscode_list = list() 
for ts_code_sh in weekly_20180223['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180223_tscode_list.append(stock_name)
weekly_20180223_addname_dataframe=pd.DataFrame()
weekly_20180223_addname_dataframe['cname'] = weekly_20180223_tscode_list
for table_name in weekly_20180223.columns.values.tolist():
    weekly_20180223_addname_dataframe[table_name] = weekly_20180223[table_name]
print("周线行情-时间为序  weekly_20180223 8_20180223 返回数据 row 行数 = "+str(weekly_20180223.shape[0]))
weekly_20180223_addname_dataframe.to_excel(weekly_2018_excel_writer,'8_20180223',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180302")       ###  更新 周线记录日期
weekly_20180302 = pro.weekly(trade_date='20180302', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180302_tscode_list = list() 
for ts_code_sh in weekly_20180302['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180302_tscode_list.append(stock_name)
weekly_20180302_addname_dataframe=pd.DataFrame()
weekly_20180302_addname_dataframe['cname'] = weekly_20180302_tscode_list
for table_name in weekly_20180302.columns.values.tolist():
    weekly_20180302_addname_dataframe[table_name] = weekly_20180302[table_name]
print("周线行情-时间为序  weekly_20180302 9_20180302 返回数据 row 行数 = "+str(weekly_20180302.shape[0]))
weekly_20180302_addname_dataframe.to_excel(weekly_2018_excel_writer,'9_20180302',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180309")       ###  更新 周线记录日期
weekly_20180309 = pro.weekly(trade_date='20180309', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180309_tscode_list = list() 
for ts_code_sh in weekly_20180309['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180309_tscode_list.append(stock_name)
weekly_20180309_addname_dataframe=pd.DataFrame()
weekly_20180309_addname_dataframe['cname'] = weekly_20180309_tscode_list
for table_name in weekly_20180309.columns.values.tolist():
    weekly_20180309_addname_dataframe[table_name] = weekly_20180309[table_name]
print("周线行情-时间为序  weekly_20180309 10_20180309 返回数据 row 行数 = "+str(weekly_20180309.shape[0]))
weekly_20180309_addname_dataframe.to_excel(weekly_2018_excel_writer,'10_20180309',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180316")       ###  更新 周线记录日期
weekly_20180316 = pro.weekly(trade_date='20180316', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180316_tscode_list = list() 
for ts_code_sh in weekly_20180316['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180316_tscode_list.append(stock_name)
weekly_20180316_addname_dataframe=pd.DataFrame()
weekly_20180316_addname_dataframe['cname'] = weekly_20180316_tscode_list
for table_name in weekly_20180316.columns.values.tolist():
    weekly_20180316_addname_dataframe[table_name] = weekly_20180316[table_name]
print("周线行情-时间为序  weekly_20180316 11_20180316 返回数据 row 行数 = "+str(weekly_20180316.shape[0]))
weekly_20180316_addname_dataframe.to_excel(weekly_2018_excel_writer,'11_20180316',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180323")       ###  更新 周线记录日期
weekly_20180323 = pro.weekly(trade_date='20180323', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180323_tscode_list = list() 
for ts_code_sh in weekly_20180323['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180323_tscode_list.append(stock_name)
weekly_20180323_addname_dataframe=pd.DataFrame()
weekly_20180323_addname_dataframe['cname'] = weekly_20180323_tscode_list
for table_name in weekly_20180323.columns.values.tolist():
    weekly_20180323_addname_dataframe[table_name] = weekly_20180323[table_name]
print("周线行情-时间为序  weekly_20180323 12_20180323 返回数据 row 行数 = "+str(weekly_20180323.shape[0]))
weekly_20180323_addname_dataframe.to_excel(weekly_2018_excel_writer,'12_20180323',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180330")       ###  更新 周线记录日期
weekly_20180330 = pro.weekly(trade_date='20180330', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180330_tscode_list = list() 
for ts_code_sh in weekly_20180330['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180330_tscode_list.append(stock_name)
weekly_20180330_addname_dataframe=pd.DataFrame()
weekly_20180330_addname_dataframe['cname'] = weekly_20180330_tscode_list
for table_name in weekly_20180330.columns.values.tolist():
    weekly_20180330_addname_dataframe[table_name] = weekly_20180330[table_name]
print("周线行情-时间为序  weekly_20180330 13_20180330 返回数据 row 行数 = "+str(weekly_20180330.shape[0]))
weekly_20180330_addname_dataframe.to_excel(weekly_2018_excel_writer,'13_20180330',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180404")       ###  更新 周线记录日期
weekly_20180404 = pro.weekly(trade_date='20180404', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180404_tscode_list = list() 
for ts_code_sh in weekly_20180404['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180404_tscode_list.append(stock_name)
weekly_20180404_addname_dataframe=pd.DataFrame()
weekly_20180404_addname_dataframe['cname'] = weekly_20180404_tscode_list
for table_name in weekly_20180404.columns.values.tolist():
    weekly_20180404_addname_dataframe[table_name] = weekly_20180404[table_name]
print("周线行情-时间为序  weekly_20180404 14_20180404 返回数据 row 行数 = "+str(weekly_20180404.shape[0]))
weekly_20180404_addname_dataframe.to_excel(weekly_2018_excel_writer,'14_20180404',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180413")       ###  更新 周线记录日期
weekly_20180413 = pro.weekly(trade_date='20180413', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180413_tscode_list = list() 
for ts_code_sh in weekly_20180413['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180413_tscode_list.append(stock_name)
weekly_20180413_addname_dataframe=pd.DataFrame()
weekly_20180413_addname_dataframe['cname'] = weekly_20180413_tscode_list
for table_name in weekly_20180413.columns.values.tolist():
    weekly_20180413_addname_dataframe[table_name] = weekly_20180413[table_name]
print("周线行情-时间为序  weekly_20180413 15_20180413 返回数据 row 行数 = "+str(weekly_20180413.shape[0]))
weekly_20180413_addname_dataframe.to_excel(weekly_2018_excel_writer,'15_20180413',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180420")       ###  更新 周线记录日期
weekly_20180420 = pro.weekly(trade_date='20180420', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180420_tscode_list = list() 
for ts_code_sh in weekly_20180420['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180420_tscode_list.append(stock_name)
weekly_20180420_addname_dataframe=pd.DataFrame()
weekly_20180420_addname_dataframe['cname'] = weekly_20180420_tscode_list
for table_name in weekly_20180420.columns.values.tolist():
    weekly_20180420_addname_dataframe[table_name] = weekly_20180420[table_name]
print("周线行情-时间为序  weekly_20180420 16_20180420 返回数据 row 行数 = "+str(weekly_20180420.shape[0]))
weekly_20180420_addname_dataframe.to_excel(weekly_2018_excel_writer,'16_20180420',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180427")       ###  更新 周线记录日期
weekly_20180427 = pro.weekly(trade_date='20180427', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180427_tscode_list = list() 
for ts_code_sh in weekly_20180427['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180427_tscode_list.append(stock_name)
weekly_20180427_addname_dataframe=pd.DataFrame()
weekly_20180427_addname_dataframe['cname'] = weekly_20180427_tscode_list
for table_name in weekly_20180427.columns.values.tolist():
    weekly_20180427_addname_dataframe[table_name] = weekly_20180427[table_name]
print("周线行情-时间为序  weekly_20180427 17_20180427 返回数据 row 行数 = "+str(weekly_20180427.shape[0]))
weekly_20180427_addname_dataframe.to_excel(weekly_2018_excel_writer,'17_20180427',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180504")       ###  更新 周线记录日期
weekly_20180504 = pro.weekly(trade_date='20180504', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180504_tscode_list = list() 
for ts_code_sh in weekly_20180504['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180504_tscode_list.append(stock_name)
weekly_20180504_addname_dataframe=pd.DataFrame()
weekly_20180504_addname_dataframe['cname'] = weekly_20180504_tscode_list
for table_name in weekly_20180504.columns.values.tolist():
    weekly_20180504_addname_dataframe[table_name] = weekly_20180504[table_name]
print("周线行情-时间为序  weekly_20180504 18_20180504 返回数据 row 行数 = "+str(weekly_20180504.shape[0]))
weekly_20180504_addname_dataframe.to_excel(weekly_2018_excel_writer,'18_20180504',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180511")       ###  更新 周线记录日期
weekly_20180511 = pro.weekly(trade_date='20180511', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180511_tscode_list = list() 
for ts_code_sh in weekly_20180511['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180511_tscode_list.append(stock_name)
weekly_20180511_addname_dataframe=pd.DataFrame()
weekly_20180511_addname_dataframe['cname'] = weekly_20180511_tscode_list
for table_name in weekly_20180511.columns.values.tolist():
    weekly_20180511_addname_dataframe[table_name] = weekly_20180511[table_name]
print("周线行情-时间为序  weekly_20180511 19_20180511 返回数据 row 行数 = "+str(weekly_20180511.shape[0]))
weekly_20180511_addname_dataframe.to_excel(weekly_2018_excel_writer,'19_20180511',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180518")       ###  更新 周线记录日期
weekly_20180518 = pro.weekly(trade_date='20180518', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180518_tscode_list = list() 
for ts_code_sh in weekly_20180518['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180518_tscode_list.append(stock_name)
weekly_20180518_addname_dataframe=pd.DataFrame()
weekly_20180518_addname_dataframe['cname'] = weekly_20180518_tscode_list
for table_name in weekly_20180518.columns.values.tolist():
    weekly_20180518_addname_dataframe[table_name] = weekly_20180518[table_name]
print("周线行情-时间为序  weekly_20180518 20_20180518 返回数据 row 行数 = "+str(weekly_20180518.shape[0]))
weekly_20180518_addname_dataframe.to_excel(weekly_2018_excel_writer,'20_20180518',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180525")       ###  更新 周线记录日期
weekly_20180525 = pro.weekly(trade_date='20180525', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180525_tscode_list = list() 
for ts_code_sh in weekly_20180525['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180525_tscode_list.append(stock_name)
weekly_20180525_addname_dataframe=pd.DataFrame()
weekly_20180525_addname_dataframe['cname'] = weekly_20180525_tscode_list
for table_name in weekly_20180525.columns.values.tolist():
    weekly_20180525_addname_dataframe[table_name] = weekly_20180525[table_name]
print("周线行情-时间为序  weekly_20180525 21_20180525 返回数据 row 行数 = "+str(weekly_20180525.shape[0]))
weekly_20180525_addname_dataframe.to_excel(weekly_2018_excel_writer,'21_20180525',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180601")       ###  更新 周线记录日期
weekly_20180601 = pro.weekly(trade_date='20180601', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180601_tscode_list = list() 
for ts_code_sh in weekly_20180601['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180601_tscode_list.append(stock_name)
weekly_20180601_addname_dataframe=pd.DataFrame()
weekly_20180601_addname_dataframe['cname'] = weekly_20180601_tscode_list
for table_name in weekly_20180601.columns.values.tolist():
    weekly_20180601_addname_dataframe[table_name] = weekly_20180601[table_name]
print("周线行情-时间为序  weekly_20180601 22_20180601 返回数据 row 行数 = "+str(weekly_20180601.shape[0]))
weekly_20180601_addname_dataframe.to_excel(weekly_2018_excel_writer,'22_20180601',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180608")       ###  更新 周线记录日期
weekly_20180608 = pro.weekly(trade_date='20180608', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180608_tscode_list = list() 
for ts_code_sh in weekly_20180608['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180608_tscode_list.append(stock_name)
weekly_20180608_addname_dataframe=pd.DataFrame()
weekly_20180608_addname_dataframe['cname'] = weekly_20180608_tscode_list
for table_name in weekly_20180608.columns.values.tolist():
    weekly_20180608_addname_dataframe[table_name] = weekly_20180608[table_name]
print("周线行情-时间为序  weekly_20180608 23_20180608 返回数据 row 行数 = "+str(weekly_20180608.shape[0]))
weekly_20180608_addname_dataframe.to_excel(weekly_2018_excel_writer,'23_20180608',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180615")       ###  更新 周线记录日期
weekly_20180615 = pro.weekly(trade_date='20180615', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180615_tscode_list = list() 
for ts_code_sh in weekly_20180615['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180615_tscode_list.append(stock_name)
weekly_20180615_addname_dataframe=pd.DataFrame()
weekly_20180615_addname_dataframe['cname'] = weekly_20180615_tscode_list
for table_name in weekly_20180615.columns.values.tolist():
    weekly_20180615_addname_dataframe[table_name] = weekly_20180615[table_name]
print("周线行情-时间为序  weekly_20180615 24_20180615 返回数据 row 行数 = "+str(weekly_20180615.shape[0]))
weekly_20180615_addname_dataframe.to_excel(weekly_2018_excel_writer,'24_20180615',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180622")       ###  更新 周线记录日期
weekly_20180622 = pro.weekly(trade_date='20180622', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180622_tscode_list = list() 
for ts_code_sh in weekly_20180622['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180622_tscode_list.append(stock_name)
weekly_20180622_addname_dataframe=pd.DataFrame()
weekly_20180622_addname_dataframe['cname'] = weekly_20180622_tscode_list
for table_name in weekly_20180622.columns.values.tolist():
    weekly_20180622_addname_dataframe[table_name] = weekly_20180622[table_name]
print("周线行情-时间为序  weekly_20180622 25_20180622 返回数据 row 行数 = "+str(weekly_20180622.shape[0]))
weekly_20180622_addname_dataframe.to_excel(weekly_2018_excel_writer,'25_20180622',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180629")       ###  更新 周线记录日期
weekly_20180629 = pro.weekly(trade_date='20180629', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180629_tscode_list = list() 
for ts_code_sh in weekly_20180629['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180629_tscode_list.append(stock_name)
weekly_20180629_addname_dataframe=pd.DataFrame()
weekly_20180629_addname_dataframe['cname'] = weekly_20180629_tscode_list
for table_name in weekly_20180629.columns.values.tolist():
    weekly_20180629_addname_dataframe[table_name] = weekly_20180629[table_name]
print("周线行情-时间为序  weekly_20180629 26_20180629 返回数据 row 行数 = "+str(weekly_20180629.shape[0]))
weekly_20180629_addname_dataframe.to_excel(weekly_2018_excel_writer,'26_20180629',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180706")       ###  更新 周线记录日期
weekly_20180706 = pro.weekly(trade_date='20180706', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180706_tscode_list = list() 
for ts_code_sh in weekly_20180706['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180706_tscode_list.append(stock_name)
weekly_20180706_addname_dataframe=pd.DataFrame()
weekly_20180706_addname_dataframe['cname'] = weekly_20180706_tscode_list
for table_name in weekly_20180706.columns.values.tolist():
    weekly_20180706_addname_dataframe[table_name] = weekly_20180706[table_name]
print("周线行情-时间为序  weekly_20180706 27_20180706 返回数据 row 行数 = "+str(weekly_20180706.shape[0]))
weekly_20180706_addname_dataframe.to_excel(weekly_2018_excel_writer,'27_20180706',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180713")       ###  更新 周线记录日期
weekly_20180713 = pro.weekly(trade_date='20180713', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180713_tscode_list = list() 
for ts_code_sh in weekly_20180713['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180713_tscode_list.append(stock_name)
weekly_20180713_addname_dataframe=pd.DataFrame()
weekly_20180713_addname_dataframe['cname'] = weekly_20180713_tscode_list
for table_name in weekly_20180713.columns.values.tolist():
    weekly_20180713_addname_dataframe[table_name] = weekly_20180713[table_name]
print("周线行情-时间为序  weekly_20180713 28_20180713 返回数据 row 行数 = "+str(weekly_20180713.shape[0]))
weekly_20180713_addname_dataframe.to_excel(weekly_2018_excel_writer,'28_20180713',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180720")       ###  更新 周线记录日期
weekly_20180720 = pro.weekly(trade_date='20180720', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180720_tscode_list = list() 
for ts_code_sh in weekly_20180720['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180720_tscode_list.append(stock_name)
weekly_20180720_addname_dataframe=pd.DataFrame()
weekly_20180720_addname_dataframe['cname'] = weekly_20180720_tscode_list
for table_name in weekly_20180720.columns.values.tolist():
    weekly_20180720_addname_dataframe[table_name] = weekly_20180720[table_name]
print("周线行情-时间为序  weekly_20180720 29_20180720 返回数据 row 行数 = "+str(weekly_20180720.shape[0]))
weekly_20180720_addname_dataframe.to_excel(weekly_2018_excel_writer,'29_20180720',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180727")       ###  更新 周线记录日期
weekly_20180727 = pro.weekly(trade_date='20180727', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180727_tscode_list = list() 
for ts_code_sh in weekly_20180727['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180727_tscode_list.append(stock_name)
weekly_20180727_addname_dataframe=pd.DataFrame()
weekly_20180727_addname_dataframe['cname'] = weekly_20180727_tscode_list
for table_name in weekly_20180727.columns.values.tolist():
    weekly_20180727_addname_dataframe[table_name] = weekly_20180727[table_name]
print("周线行情-时间为序  weekly_20180727 30_20180727 返回数据 row 行数 = "+str(weekly_20180727.shape[0]))
weekly_20180727_addname_dataframe.to_excel(weekly_2018_excel_writer,'30_20180727',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180803")       ###  更新 周线记录日期
weekly_20180803 = pro.weekly(trade_date='20180803', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180803_tscode_list = list() 
for ts_code_sh in weekly_20180803['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180803_tscode_list.append(stock_name)
weekly_20180803_addname_dataframe=pd.DataFrame()
weekly_20180803_addname_dataframe['cname'] = weekly_20180803_tscode_list
for table_name in weekly_20180803.columns.values.tolist():
    weekly_20180803_addname_dataframe[table_name] = weekly_20180803[table_name]
print("周线行情-时间为序  weekly_20180803 31_20180803 返回数据 row 行数 = "+str(weekly_20180803.shape[0]))
weekly_20180803_addname_dataframe.to_excel(weekly_2018_excel_writer,'31_20180803',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180810")       ###  更新 周线记录日期
weekly_20180810 = pro.weekly(trade_date='20180810', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180810_tscode_list = list() 
for ts_code_sh in weekly_20180810['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180810_tscode_list.append(stock_name)
weekly_20180810_addname_dataframe=pd.DataFrame()
weekly_20180810_addname_dataframe['cname'] = weekly_20180810_tscode_list
for table_name in weekly_20180810.columns.values.tolist():
    weekly_20180810_addname_dataframe[table_name] = weekly_20180810[table_name]
print("周线行情-时间为序  weekly_20180810 32_20180810 返回数据 row 行数 = "+str(weekly_20180810.shape[0]))
weekly_20180810_addname_dataframe.to_excel(weekly_2018_excel_writer,'32_20180810',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180817")       ###  更新 周线记录日期
weekly_20180817 = pro.weekly(trade_date='20180817', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180817_tscode_list = list() 
for ts_code_sh in weekly_20180817['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180817_tscode_list.append(stock_name)
weekly_20180817_addname_dataframe=pd.DataFrame()
weekly_20180817_addname_dataframe['cname'] = weekly_20180817_tscode_list
for table_name in weekly_20180817.columns.values.tolist():
    weekly_20180817_addname_dataframe[table_name] = weekly_20180817[table_name]
print("周线行情-时间为序  weekly_20180817 33_20180817 返回数据 row 行数 = "+str(weekly_20180817.shape[0]))
weekly_20180817_addname_dataframe.to_excel(weekly_2018_excel_writer,'33_20180817',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180824")       ###  更新 周线记录日期
weekly_20180824 = pro.weekly(trade_date='20180824', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180824_tscode_list = list() 
for ts_code_sh in weekly_20180824['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180824_tscode_list.append(stock_name)
weekly_20180824_addname_dataframe=pd.DataFrame()
weekly_20180824_addname_dataframe['cname'] = weekly_20180824_tscode_list
for table_name in weekly_20180824.columns.values.tolist():
    weekly_20180824_addname_dataframe[table_name] = weekly_20180824[table_name]
print("周线行情-时间为序  weekly_20180824 34_20180824 返回数据 row 行数 = "+str(weekly_20180824.shape[0]))
weekly_20180824_addname_dataframe.to_excel(weekly_2018_excel_writer,'34_20180824',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180831")       ###  更新 周线记录日期
weekly_20180831 = pro.weekly(trade_date='20180831', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180831_tscode_list = list() 
for ts_code_sh in weekly_20180831['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180831_tscode_list.append(stock_name)
weekly_20180831_addname_dataframe=pd.DataFrame()
weekly_20180831_addname_dataframe['cname'] = weekly_20180831_tscode_list
for table_name in weekly_20180831.columns.values.tolist():
    weekly_20180831_addname_dataframe[table_name] = weekly_20180831[table_name]
print("周线行情-时间为序  weekly_20180831 35_20180831 返回数据 row 行数 = "+str(weekly_20180831.shape[0]))
weekly_20180831_addname_dataframe.to_excel(weekly_2018_excel_writer,'35_20180831',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180907")       ###  更新 周线记录日期
weekly_20180907 = pro.weekly(trade_date='20180907', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180907_tscode_list = list() 
for ts_code_sh in weekly_20180907['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180907_tscode_list.append(stock_name)
weekly_20180907_addname_dataframe=pd.DataFrame()
weekly_20180907_addname_dataframe['cname'] = weekly_20180907_tscode_list
for table_name in weekly_20180907.columns.values.tolist():
    weekly_20180907_addname_dataframe[table_name] = weekly_20180907[table_name]
print("周线行情-时间为序  weekly_20180907 36_20180907 返回数据 row 行数 = "+str(weekly_20180907.shape[0]))
weekly_20180907_addname_dataframe.to_excel(weekly_2018_excel_writer,'36_20180907',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180914")       ###  更新 周线记录日期
weekly_20180914 = pro.weekly(trade_date='20180914', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180914_tscode_list = list() 
for ts_code_sh in weekly_20180914['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180914_tscode_list.append(stock_name)
weekly_20180914_addname_dataframe=pd.DataFrame()
weekly_20180914_addname_dataframe['cname'] = weekly_20180914_tscode_list
for table_name in weekly_20180914.columns.values.tolist():
    weekly_20180914_addname_dataframe[table_name] = weekly_20180914[table_name]
print("周线行情-时间为序  weekly_20180914 37_20180914 返回数据 row 行数 = "+str(weekly_20180914.shape[0]))
weekly_20180914_addname_dataframe.to_excel(weekly_2018_excel_writer,'37_20180914',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180921")       ###  更新 周线记录日期
weekly_20180921 = pro.weekly(trade_date='20180921', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180921_tscode_list = list() 
for ts_code_sh in weekly_20180921['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180921_tscode_list.append(stock_name)
weekly_20180921_addname_dataframe=pd.DataFrame()
weekly_20180921_addname_dataframe['cname'] = weekly_20180921_tscode_list
for table_name in weekly_20180921.columns.values.tolist():
    weekly_20180921_addname_dataframe[table_name] = weekly_20180921[table_name]
print("周线行情-时间为序  weekly_20180921 38_20180921 返回数据 row 行数 = "+str(weekly_20180921.shape[0]))
weekly_20180921_addname_dataframe.to_excel(weekly_2018_excel_writer,'38_20180921',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20180928")       ###  更新 周线记录日期
weekly_20180928 = pro.weekly(trade_date='20180928', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20180928_tscode_list = list() 
for ts_code_sh in weekly_20180928['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20180928_tscode_list.append(stock_name)
weekly_20180928_addname_dataframe=pd.DataFrame()
weekly_20180928_addname_dataframe['cname'] = weekly_20180928_tscode_list
for table_name in weekly_20180928.columns.values.tolist():
    weekly_20180928_addname_dataframe[table_name] = weekly_20180928[table_name]
print("周线行情-时间为序  weekly_20180928 39_20180928 返回数据 row 行数 = "+str(weekly_20180928.shape[0]))
weekly_20180928_addname_dataframe.to_excel(weekly_2018_excel_writer,'39_20180928',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181012")       ###  更新 周线记录日期
weekly_20181012 = pro.weekly(trade_date='20181012', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181012_tscode_list = list() 
for ts_code_sh in weekly_20181012['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181012_tscode_list.append(stock_name)
weekly_20181012_addname_dataframe=pd.DataFrame()
weekly_20181012_addname_dataframe['cname'] = weekly_20181012_tscode_list
for table_name in weekly_20181012.columns.values.tolist():
    weekly_20181012_addname_dataframe[table_name] = weekly_20181012[table_name]
print("周线行情-时间为序  weekly_20181012 41_20181012 返回数据 row 行数 = "+str(weekly_20181012.shape[0]))
weekly_20181012_addname_dataframe.to_excel(weekly_2018_excel_writer,'41_20181012',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181019")       ###  更新 周线记录日期
weekly_20181019 = pro.weekly(trade_date='20181019', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181019_tscode_list = list() 
for ts_code_sh in weekly_20181019['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181019_tscode_list.append(stock_name)
weekly_20181019_addname_dataframe=pd.DataFrame()
weekly_20181019_addname_dataframe['cname'] = weekly_20181019_tscode_list
for table_name in weekly_20181019.columns.values.tolist():
    weekly_20181019_addname_dataframe[table_name] = weekly_20181019[table_name]
print("周线行情-时间为序  weekly_20181019 42_20181019 返回数据 row 行数 = "+str(weekly_20181019.shape[0]))
weekly_20181019_addname_dataframe.to_excel(weekly_2018_excel_writer,'42_20181019',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181026")       ###  更新 周线记录日期
weekly_20181026 = pro.weekly(trade_date='20181026', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181026_tscode_list = list() 
for ts_code_sh in weekly_20181026['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181026_tscode_list.append(stock_name)
weekly_20181026_addname_dataframe=pd.DataFrame()
weekly_20181026_addname_dataframe['cname'] = weekly_20181026_tscode_list
for table_name in weekly_20181026.columns.values.tolist():
    weekly_20181026_addname_dataframe[table_name] = weekly_20181026[table_name]
print("周线行情-时间为序  weekly_20181026 43_20181026 返回数据 row 行数 = "+str(weekly_20181026.shape[0]))
weekly_20181026_addname_dataframe.to_excel(weekly_2018_excel_writer,'43_20181026',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181102")       ###  更新 周线记录日期
weekly_20181102 = pro.weekly(trade_date='20181102', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181102_tscode_list = list() 
for ts_code_sh in weekly_20181102['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181102_tscode_list.append(stock_name)
weekly_20181102_addname_dataframe=pd.DataFrame()
weekly_20181102_addname_dataframe['cname'] = weekly_20181102_tscode_list
for table_name in weekly_20181102.columns.values.tolist():
    weekly_20181102_addname_dataframe[table_name] = weekly_20181102[table_name]
print("周线行情-时间为序  weekly_20181102 44_20181102 返回数据 row 行数 = "+str(weekly_20181102.shape[0]))
weekly_20181102_addname_dataframe.to_excel(weekly_2018_excel_writer,'44_20181102',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181109")       ###  更新 周线记录日期
weekly_20181109 = pro.weekly(trade_date='20181109', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181109_tscode_list = list() 
for ts_code_sh in weekly_20181109['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181109_tscode_list.append(stock_name)
weekly_20181109_addname_dataframe=pd.DataFrame()
weekly_20181109_addname_dataframe['cname'] = weekly_20181109_tscode_list
for table_name in weekly_20181109.columns.values.tolist():
    weekly_20181109_addname_dataframe[table_name] = weekly_20181109[table_name]
print("周线行情-时间为序  weekly_20181109 45_20181109 返回数据 row 行数 = "+str(weekly_20181109.shape[0]))
weekly_20181109_addname_dataframe.to_excel(weekly_2018_excel_writer,'45_20181109',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181116")       ###  更新 周线记录日期
weekly_20181116 = pro.weekly(trade_date='20181116', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181116_tscode_list = list() 
for ts_code_sh in weekly_20181116['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181116_tscode_list.append(stock_name)
weekly_20181116_addname_dataframe=pd.DataFrame()
weekly_20181116_addname_dataframe['cname'] = weekly_20181116_tscode_list
for table_name in weekly_20181116.columns.values.tolist():
    weekly_20181116_addname_dataframe[table_name] = weekly_20181116[table_name]
print("周线行情-时间为序  weekly_20181116 46_20181116 返回数据 row 行数 = "+str(weekly_20181116.shape[0]))
weekly_20181116_addname_dataframe.to_excel(weekly_2018_excel_writer,'46_20181116',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181123")       ###  更新 周线记录日期
weekly_20181123 = pro.weekly(trade_date='20181123', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181123_tscode_list = list() 
for ts_code_sh in weekly_20181123['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181123_tscode_list.append(stock_name)
weekly_20181123_addname_dataframe=pd.DataFrame()
weekly_20181123_addname_dataframe['cname'] = weekly_20181123_tscode_list
for table_name in weekly_20181123.columns.values.tolist():
    weekly_20181123_addname_dataframe[table_name] = weekly_20181123[table_name]
print("周线行情-时间为序  weekly_20181123 47_20181123 返回数据 row 行数 = "+str(weekly_20181123.shape[0]))
weekly_20181123_addname_dataframe.to_excel(weekly_2018_excel_writer,'47_20181123',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181130")       ###  更新 周线记录日期
weekly_20181130 = pro.weekly(trade_date='20181130', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181130_tscode_list = list() 
for ts_code_sh in weekly_20181130['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181130_tscode_list.append(stock_name)
weekly_20181130_addname_dataframe=pd.DataFrame()
weekly_20181130_addname_dataframe['cname'] = weekly_20181130_tscode_list
for table_name in weekly_20181130.columns.values.tolist():
    weekly_20181130_addname_dataframe[table_name] = weekly_20181130[table_name]
print("周线行情-时间为序  weekly_20181130 48_20181130 返回数据 row 行数 = "+str(weekly_20181130.shape[0]))
weekly_20181130_addname_dataframe.to_excel(weekly_2018_excel_writer,'48_20181130',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181207")       ###  更新 周线记录日期
weekly_20181207 = pro.weekly(trade_date='20181207', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181207_tscode_list = list() 
for ts_code_sh in weekly_20181207['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181207_tscode_list.append(stock_name)
weekly_20181207_addname_dataframe=pd.DataFrame()
weekly_20181207_addname_dataframe['cname'] = weekly_20181207_tscode_list
for table_name in weekly_20181207.columns.values.tolist():
    weekly_20181207_addname_dataframe[table_name] = weekly_20181207[table_name]
print("周线行情-时间为序  weekly_20181207 49_20181207 返回数据 row 行数 = "+str(weekly_20181207.shape[0]))
weekly_20181207_addname_dataframe.to_excel(weekly_2018_excel_writer,'49_20181207',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181214")       ###  更新 周线记录日期
weekly_20181214 = pro.weekly(trade_date='20181214', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181214_tscode_list = list() 
for ts_code_sh in weekly_20181214['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181214_tscode_list.append(stock_name)
weekly_20181214_addname_dataframe=pd.DataFrame()
weekly_20181214_addname_dataframe['cname'] = weekly_20181214_tscode_list
for table_name in weekly_20181214.columns.values.tolist():
    weekly_20181214_addname_dataframe[table_name] = weekly_20181214[table_name]
print("周线行情-时间为序  weekly_20181214 50_20181214 返回数据 row 行数 = "+str(weekly_20181214.shape[0]))
weekly_20181214_addname_dataframe.to_excel(weekly_2018_excel_writer,'50_20181214',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181221")       ###  更新 周线记录日期
weekly_20181221 = pro.weekly(trade_date='20181221', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181221_tscode_list = list() 
for ts_code_sh in weekly_20181221['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181221_tscode_list.append(stock_name)
weekly_20181221_addname_dataframe=pd.DataFrame()
weekly_20181221_addname_dataframe['cname'] = weekly_20181221_tscode_list
for table_name in weekly_20181221.columns.values.tolist():
    weekly_20181221_addname_dataframe[table_name] = weekly_20181221[table_name]
print("周线行情-时间为序  weekly_20181221 51_20181221 返回数据 row 行数 = "+str(weekly_20181221.shape[0]))
weekly_20181221_addname_dataframe.to_excel(weekly_2018_excel_writer,'51_20181221',index=False)
weekly_2018_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20181228")       ###  更新 周线记录日期
weekly_20181228 = pro.weekly(trade_date='20181228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20181228_tscode_list = list() 
for ts_code_sh in weekly_20181228['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20181228_tscode_list.append(stock_name)
weekly_20181228_addname_dataframe=pd.DataFrame()
weekly_20181228_addname_dataframe['cname'] = weekly_20181228_tscode_list
for table_name in weekly_20181228.columns.values.tolist():
    weekly_20181228_addname_dataframe[table_name] = weekly_20181228[table_name]
print("周线行情-时间为序  weekly_20181228 52_20181228 返回数据 row 行数 = "+str(weekly_20181228.shape[0]))
weekly_20181228_addname_dataframe.to_excel(weekly_2018_excel_writer,'52_20181228',index=False)
weekly_2018_excel_writer.save()
createexcel('weekly_2019.xlsx')
weekly_2019_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2019.xlsx')
weekly_2019_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2019.xlsx', engine='openpyxl')
weekly_2019_excel_writer.book = weekly_2019_book
weekly_2019_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2019_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20190104")       ###  更新 周线记录日期
weekly_20190104 = pro.weekly(trade_date='20190104', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190104_tscode_list = list() 
for ts_code_sh in weekly_20190104['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190104_tscode_list.append(stock_name)
weekly_20190104_addname_dataframe=pd.DataFrame()
weekly_20190104_addname_dataframe['cname'] = weekly_20190104_tscode_list
for table_name in weekly_20190104.columns.values.tolist():
    weekly_20190104_addname_dataframe[table_name] = weekly_20190104[table_name]
print("周线行情-时间为序  weekly_20190104 1_20190104 返回数据 row 行数 = "+str(weekly_20190104.shape[0]))
weekly_20190104_addname_dataframe.to_excel(weekly_2019_excel_writer,'1_20190104',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190111")       ###  更新 周线记录日期
weekly_20190111 = pro.weekly(trade_date='20190111', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190111_tscode_list = list() 
for ts_code_sh in weekly_20190111['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190111_tscode_list.append(stock_name)
weekly_20190111_addname_dataframe=pd.DataFrame()
weekly_20190111_addname_dataframe['cname'] = weekly_20190111_tscode_list
for table_name in weekly_20190111.columns.values.tolist():
    weekly_20190111_addname_dataframe[table_name] = weekly_20190111[table_name]
print("周线行情-时间为序  weekly_20190111 2_20190111 返回数据 row 行数 = "+str(weekly_20190111.shape[0]))
weekly_20190111_addname_dataframe.to_excel(weekly_2019_excel_writer,'2_20190111',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190118")       ###  更新 周线记录日期
weekly_20190118 = pro.weekly(trade_date='20190118', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190118_tscode_list = list() 
for ts_code_sh in weekly_20190118['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190118_tscode_list.append(stock_name)
weekly_20190118_addname_dataframe=pd.DataFrame()
weekly_20190118_addname_dataframe['cname'] = weekly_20190118_tscode_list
for table_name in weekly_20190118.columns.values.tolist():
    weekly_20190118_addname_dataframe[table_name] = weekly_20190118[table_name]
print("周线行情-时间为序  weekly_20190118 3_20190118 返回数据 row 行数 = "+str(weekly_20190118.shape[0]))
weekly_20190118_addname_dataframe.to_excel(weekly_2019_excel_writer,'3_20190118',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190125")       ###  更新 周线记录日期
weekly_20190125 = pro.weekly(trade_date='20190125', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190125_tscode_list = list() 
for ts_code_sh in weekly_20190125['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190125_tscode_list.append(stock_name)
weekly_20190125_addname_dataframe=pd.DataFrame()
weekly_20190125_addname_dataframe['cname'] = weekly_20190125_tscode_list
for table_name in weekly_20190125.columns.values.tolist():
    weekly_20190125_addname_dataframe[table_name] = weekly_20190125[table_name]
print("周线行情-时间为序  weekly_20190125 4_20190125 返回数据 row 行数 = "+str(weekly_20190125.shape[0]))
weekly_20190125_addname_dataframe.to_excel(weekly_2019_excel_writer,'4_20190125',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190201")       ###  更新 周线记录日期
weekly_20190201 = pro.weekly(trade_date='20190201', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190201_tscode_list = list() 
for ts_code_sh in weekly_20190201['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190201_tscode_list.append(stock_name)
weekly_20190201_addname_dataframe=pd.DataFrame()
weekly_20190201_addname_dataframe['cname'] = weekly_20190201_tscode_list
for table_name in weekly_20190201.columns.values.tolist():
    weekly_20190201_addname_dataframe[table_name] = weekly_20190201[table_name]
print("周线行情-时间为序  weekly_20190201 5_20190201 返回数据 row 行数 = "+str(weekly_20190201.shape[0]))
weekly_20190201_addname_dataframe.to_excel(weekly_2019_excel_writer,'5_20190201',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190215")       ###  更新 周线记录日期
weekly_20190215 = pro.weekly(trade_date='20190215', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190215_tscode_list = list() 
for ts_code_sh in weekly_20190215['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190215_tscode_list.append(stock_name)
weekly_20190215_addname_dataframe=pd.DataFrame()
weekly_20190215_addname_dataframe['cname'] = weekly_20190215_tscode_list
for table_name in weekly_20190215.columns.values.tolist():
    weekly_20190215_addname_dataframe[table_name] = weekly_20190215[table_name]
print("周线行情-时间为序  weekly_20190215 7_20190215 返回数据 row 行数 = "+str(weekly_20190215.shape[0]))
weekly_20190215_addname_dataframe.to_excel(weekly_2019_excel_writer,'7_20190215',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190222")       ###  更新 周线记录日期
weekly_20190222 = pro.weekly(trade_date='20190222', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190222_tscode_list = list() 
for ts_code_sh in weekly_20190222['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190222_tscode_list.append(stock_name)
weekly_20190222_addname_dataframe=pd.DataFrame()
weekly_20190222_addname_dataframe['cname'] = weekly_20190222_tscode_list
for table_name in weekly_20190222.columns.values.tolist():
    weekly_20190222_addname_dataframe[table_name] = weekly_20190222[table_name]
print("周线行情-时间为序  weekly_20190222 8_20190222 返回数据 row 行数 = "+str(weekly_20190222.shape[0]))
weekly_20190222_addname_dataframe.to_excel(weekly_2019_excel_writer,'8_20190222',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190301")       ###  更新 周线记录日期
weekly_20190301 = pro.weekly(trade_date='20190301', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190301_tscode_list = list() 
for ts_code_sh in weekly_20190301['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190301_tscode_list.append(stock_name)
weekly_20190301_addname_dataframe=pd.DataFrame()
weekly_20190301_addname_dataframe['cname'] = weekly_20190301_tscode_list
for table_name in weekly_20190301.columns.values.tolist():
    weekly_20190301_addname_dataframe[table_name] = weekly_20190301[table_name]
print("周线行情-时间为序  weekly_20190301 9_20190301 返回数据 row 行数 = "+str(weekly_20190301.shape[0]))
weekly_20190301_addname_dataframe.to_excel(weekly_2019_excel_writer,'9_20190301',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190308")       ###  更新 周线记录日期
weekly_20190308 = pro.weekly(trade_date='20190308', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190308_tscode_list = list() 
for ts_code_sh in weekly_20190308['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190308_tscode_list.append(stock_name)
weekly_20190308_addname_dataframe=pd.DataFrame()
weekly_20190308_addname_dataframe['cname'] = weekly_20190308_tscode_list
for table_name in weekly_20190308.columns.values.tolist():
    weekly_20190308_addname_dataframe[table_name] = weekly_20190308[table_name]
print("周线行情-时间为序  weekly_20190308 10_20190308 返回数据 row 行数 = "+str(weekly_20190308.shape[0]))
weekly_20190308_addname_dataframe.to_excel(weekly_2019_excel_writer,'10_20190308',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190315")       ###  更新 周线记录日期
weekly_20190315 = pro.weekly(trade_date='20190315', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190315_tscode_list = list() 
for ts_code_sh in weekly_20190315['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190315_tscode_list.append(stock_name)
weekly_20190315_addname_dataframe=pd.DataFrame()
weekly_20190315_addname_dataframe['cname'] = weekly_20190315_tscode_list
for table_name in weekly_20190315.columns.values.tolist():
    weekly_20190315_addname_dataframe[table_name] = weekly_20190315[table_name]
print("周线行情-时间为序  weekly_20190315 11_20190315 返回数据 row 行数 = "+str(weekly_20190315.shape[0]))
weekly_20190315_addname_dataframe.to_excel(weekly_2019_excel_writer,'11_20190315',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190322")       ###  更新 周线记录日期
weekly_20190322 = pro.weekly(trade_date='20190322', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190322_tscode_list = list() 
for ts_code_sh in weekly_20190322['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190322_tscode_list.append(stock_name)
weekly_20190322_addname_dataframe=pd.DataFrame()
weekly_20190322_addname_dataframe['cname'] = weekly_20190322_tscode_list
for table_name in weekly_20190322.columns.values.tolist():
    weekly_20190322_addname_dataframe[table_name] = weekly_20190322[table_name]
print("周线行情-时间为序  weekly_20190322 12_20190322 返回数据 row 行数 = "+str(weekly_20190322.shape[0]))
weekly_20190322_addname_dataframe.to_excel(weekly_2019_excel_writer,'12_20190322',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190329")       ###  更新 周线记录日期
weekly_20190329 = pro.weekly(trade_date='20190329', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190329_tscode_list = list() 
for ts_code_sh in weekly_20190329['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190329_tscode_list.append(stock_name)
weekly_20190329_addname_dataframe=pd.DataFrame()
weekly_20190329_addname_dataframe['cname'] = weekly_20190329_tscode_list
for table_name in weekly_20190329.columns.values.tolist():
    weekly_20190329_addname_dataframe[table_name] = weekly_20190329[table_name]
print("周线行情-时间为序  weekly_20190329 13_20190329 返回数据 row 行数 = "+str(weekly_20190329.shape[0]))
weekly_20190329_addname_dataframe.to_excel(weekly_2019_excel_writer,'13_20190329',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190404")       ###  更新 周线记录日期
weekly_20190404 = pro.weekly(trade_date='20190404', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190404_tscode_list = list() 
for ts_code_sh in weekly_20190404['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190404_tscode_list.append(stock_name)
weekly_20190404_addname_dataframe=pd.DataFrame()
weekly_20190404_addname_dataframe['cname'] = weekly_20190404_tscode_list
for table_name in weekly_20190404.columns.values.tolist():
    weekly_20190404_addname_dataframe[table_name] = weekly_20190404[table_name]
print("周线行情-时间为序  weekly_20190404 14_20190404 返回数据 row 行数 = "+str(weekly_20190404.shape[0]))
weekly_20190404_addname_dataframe.to_excel(weekly_2019_excel_writer,'14_20190404',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190412")       ###  更新 周线记录日期
weekly_20190412 = pro.weekly(trade_date='20190412', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190412_tscode_list = list() 
for ts_code_sh in weekly_20190412['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190412_tscode_list.append(stock_name)
weekly_20190412_addname_dataframe=pd.DataFrame()
weekly_20190412_addname_dataframe['cname'] = weekly_20190412_tscode_list
for table_name in weekly_20190412.columns.values.tolist():
    weekly_20190412_addname_dataframe[table_name] = weekly_20190412[table_name]
print("周线行情-时间为序  weekly_20190412 15_20190412 返回数据 row 行数 = "+str(weekly_20190412.shape[0]))
weekly_20190412_addname_dataframe.to_excel(weekly_2019_excel_writer,'15_20190412',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190419")       ###  更新 周线记录日期
weekly_20190419 = pro.weekly(trade_date='20190419', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190419_tscode_list = list() 
for ts_code_sh in weekly_20190419['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190419_tscode_list.append(stock_name)
weekly_20190419_addname_dataframe=pd.DataFrame()
weekly_20190419_addname_dataframe['cname'] = weekly_20190419_tscode_list
for table_name in weekly_20190419.columns.values.tolist():
    weekly_20190419_addname_dataframe[table_name] = weekly_20190419[table_name]
print("周线行情-时间为序  weekly_20190419 16_20190419 返回数据 row 行数 = "+str(weekly_20190419.shape[0]))
weekly_20190419_addname_dataframe.to_excel(weekly_2019_excel_writer,'16_20190419',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190426")       ###  更新 周线记录日期
weekly_20190426 = pro.weekly(trade_date='20190426', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190426_tscode_list = list() 
for ts_code_sh in weekly_20190426['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190426_tscode_list.append(stock_name)
weekly_20190426_addname_dataframe=pd.DataFrame()
weekly_20190426_addname_dataframe['cname'] = weekly_20190426_tscode_list
for table_name in weekly_20190426.columns.values.tolist():
    weekly_20190426_addname_dataframe[table_name] = weekly_20190426[table_name]
print("周线行情-时间为序  weekly_20190426 17_20190426 返回数据 row 行数 = "+str(weekly_20190426.shape[0]))
weekly_20190426_addname_dataframe.to_excel(weekly_2019_excel_writer,'17_20190426',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190430")       ###  更新 周线记录日期
weekly_20190430 = pro.weekly(trade_date='20190430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190430_tscode_list = list() 
for ts_code_sh in weekly_20190430['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190430_tscode_list.append(stock_name)
weekly_20190430_addname_dataframe=pd.DataFrame()
weekly_20190430_addname_dataframe['cname'] = weekly_20190430_tscode_list
for table_name in weekly_20190430.columns.values.tolist():
    weekly_20190430_addname_dataframe[table_name] = weekly_20190430[table_name]
print("周线行情-时间为序  weekly_20190430 18_20190430 返回数据 row 行数 = "+str(weekly_20190430.shape[0]))
weekly_20190430_addname_dataframe.to_excel(weekly_2019_excel_writer,'18_20190430',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190510")       ###  更新 周线记录日期
weekly_20190510 = pro.weekly(trade_date='20190510', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190510_tscode_list = list() 
for ts_code_sh in weekly_20190510['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190510_tscode_list.append(stock_name)
weekly_20190510_addname_dataframe=pd.DataFrame()
weekly_20190510_addname_dataframe['cname'] = weekly_20190510_tscode_list
for table_name in weekly_20190510.columns.values.tolist():
    weekly_20190510_addname_dataframe[table_name] = weekly_20190510[table_name]
print("周线行情-时间为序  weekly_20190510 19_20190510 返回数据 row 行数 = "+str(weekly_20190510.shape[0]))
weekly_20190510_addname_dataframe.to_excel(weekly_2019_excel_writer,'19_20190510',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190517")       ###  更新 周线记录日期
weekly_20190517 = pro.weekly(trade_date='20190517', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190517_tscode_list = list() 
for ts_code_sh in weekly_20190517['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190517_tscode_list.append(stock_name)
weekly_20190517_addname_dataframe=pd.DataFrame()
weekly_20190517_addname_dataframe['cname'] = weekly_20190517_tscode_list
for table_name in weekly_20190517.columns.values.tolist():
    weekly_20190517_addname_dataframe[table_name] = weekly_20190517[table_name]
print("周线行情-时间为序  weekly_20190517 20_20190517 返回数据 row 行数 = "+str(weekly_20190517.shape[0]))
weekly_20190517_addname_dataframe.to_excel(weekly_2019_excel_writer,'20_20190517',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190524")       ###  更新 周线记录日期
weekly_20190524 = pro.weekly(trade_date='20190524', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190524_tscode_list = list() 
for ts_code_sh in weekly_20190524['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190524_tscode_list.append(stock_name)
weekly_20190524_addname_dataframe=pd.DataFrame()
weekly_20190524_addname_dataframe['cname'] = weekly_20190524_tscode_list
for table_name in weekly_20190524.columns.values.tolist():
    weekly_20190524_addname_dataframe[table_name] = weekly_20190524[table_name]
print("周线行情-时间为序  weekly_20190524 21_20190524 返回数据 row 行数 = "+str(weekly_20190524.shape[0]))
weekly_20190524_addname_dataframe.to_excel(weekly_2019_excel_writer,'21_20190524',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190531")       ###  更新 周线记录日期
weekly_20190531 = pro.weekly(trade_date='20190531', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190531_tscode_list = list() 
for ts_code_sh in weekly_20190531['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190531_tscode_list.append(stock_name)
weekly_20190531_addname_dataframe=pd.DataFrame()
weekly_20190531_addname_dataframe['cname'] = weekly_20190531_tscode_list
for table_name in weekly_20190531.columns.values.tolist():
    weekly_20190531_addname_dataframe[table_name] = weekly_20190531[table_name]
print("周线行情-时间为序  weekly_20190531 22_20190531 返回数据 row 行数 = "+str(weekly_20190531.shape[0]))
weekly_20190531_addname_dataframe.to_excel(weekly_2019_excel_writer,'22_20190531',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190606")       ###  更新 周线记录日期
weekly_20190606 = pro.weekly(trade_date='20190606', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190606_tscode_list = list() 
for ts_code_sh in weekly_20190606['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190606_tscode_list.append(stock_name)
weekly_20190606_addname_dataframe=pd.DataFrame()
weekly_20190606_addname_dataframe['cname'] = weekly_20190606_tscode_list
for table_name in weekly_20190606.columns.values.tolist():
    weekly_20190606_addname_dataframe[table_name] = weekly_20190606[table_name]
print("周线行情-时间为序  weekly_20190606 23_20190606 返回数据 row 行数 = "+str(weekly_20190606.shape[0]))
weekly_20190606_addname_dataframe.to_excel(weekly_2019_excel_writer,'23_20190606',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190614")       ###  更新 周线记录日期
weekly_20190614 = pro.weekly(trade_date='20190614', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190614_tscode_list = list() 
for ts_code_sh in weekly_20190614['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190614_tscode_list.append(stock_name)
weekly_20190614_addname_dataframe=pd.DataFrame()
weekly_20190614_addname_dataframe['cname'] = weekly_20190614_tscode_list
for table_name in weekly_20190614.columns.values.tolist():
    weekly_20190614_addname_dataframe[table_name] = weekly_20190614[table_name]
print("周线行情-时间为序  weekly_20190614 24_20190614 返回数据 row 行数 = "+str(weekly_20190614.shape[0]))
weekly_20190614_addname_dataframe.to_excel(weekly_2019_excel_writer,'24_20190614',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190621")       ###  更新 周线记录日期
weekly_20190621 = pro.weekly(trade_date='20190621', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190621_tscode_list = list() 
for ts_code_sh in weekly_20190621['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190621_tscode_list.append(stock_name)
weekly_20190621_addname_dataframe=pd.DataFrame()
weekly_20190621_addname_dataframe['cname'] = weekly_20190621_tscode_list
for table_name in weekly_20190621.columns.values.tolist():
    weekly_20190621_addname_dataframe[table_name] = weekly_20190621[table_name]
print("周线行情-时间为序  weekly_20190621 25_20190621 返回数据 row 行数 = "+str(weekly_20190621.shape[0]))
weekly_20190621_addname_dataframe.to_excel(weekly_2019_excel_writer,'25_20190621',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190628")       ###  更新 周线记录日期
weekly_20190628 = pro.weekly(trade_date='20190628', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190628_tscode_list = list() 
for ts_code_sh in weekly_20190628['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190628_tscode_list.append(stock_name)
weekly_20190628_addname_dataframe=pd.DataFrame()
weekly_20190628_addname_dataframe['cname'] = weekly_20190628_tscode_list
for table_name in weekly_20190628.columns.values.tolist():
    weekly_20190628_addname_dataframe[table_name] = weekly_20190628[table_name]
print("周线行情-时间为序  weekly_20190628 26_20190628 返回数据 row 行数 = "+str(weekly_20190628.shape[0]))
weekly_20190628_addname_dataframe.to_excel(weekly_2019_excel_writer,'26_20190628',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190705")       ###  更新 周线记录日期
weekly_20190705 = pro.weekly(trade_date='20190705', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190705_tscode_list = list() 
for ts_code_sh in weekly_20190705['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190705_tscode_list.append(stock_name)
weekly_20190705_addname_dataframe=pd.DataFrame()
weekly_20190705_addname_dataframe['cname'] = weekly_20190705_tscode_list
for table_name in weekly_20190705.columns.values.tolist():
    weekly_20190705_addname_dataframe[table_name] = weekly_20190705[table_name]
print("周线行情-时间为序  weekly_20190705 27_20190705 返回数据 row 行数 = "+str(weekly_20190705.shape[0]))
weekly_20190705_addname_dataframe.to_excel(weekly_2019_excel_writer,'27_20190705',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190712")       ###  更新 周线记录日期
weekly_20190712 = pro.weekly(trade_date='20190712', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190712_tscode_list = list() 
for ts_code_sh in weekly_20190712['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190712_tscode_list.append(stock_name)
weekly_20190712_addname_dataframe=pd.DataFrame()
weekly_20190712_addname_dataframe['cname'] = weekly_20190712_tscode_list
for table_name in weekly_20190712.columns.values.tolist():
    weekly_20190712_addname_dataframe[table_name] = weekly_20190712[table_name]
print("周线行情-时间为序  weekly_20190712 28_20190712 返回数据 row 行数 = "+str(weekly_20190712.shape[0]))
weekly_20190712_addname_dataframe.to_excel(weekly_2019_excel_writer,'28_20190712',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190719")       ###  更新 周线记录日期
weekly_20190719 = pro.weekly(trade_date='20190719', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190719_tscode_list = list() 
for ts_code_sh in weekly_20190719['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190719_tscode_list.append(stock_name)
weekly_20190719_addname_dataframe=pd.DataFrame()
weekly_20190719_addname_dataframe['cname'] = weekly_20190719_tscode_list
for table_name in weekly_20190719.columns.values.tolist():
    weekly_20190719_addname_dataframe[table_name] = weekly_20190719[table_name]
print("周线行情-时间为序  weekly_20190719 29_20190719 返回数据 row 行数 = "+str(weekly_20190719.shape[0]))
weekly_20190719_addname_dataframe.to_excel(weekly_2019_excel_writer,'29_20190719',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190726")       ###  更新 周线记录日期
weekly_20190726 = pro.weekly(trade_date='20190726', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190726_tscode_list = list() 
for ts_code_sh in weekly_20190726['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190726_tscode_list.append(stock_name)
weekly_20190726_addname_dataframe=pd.DataFrame()
weekly_20190726_addname_dataframe['cname'] = weekly_20190726_tscode_list
for table_name in weekly_20190726.columns.values.tolist():
    weekly_20190726_addname_dataframe[table_name] = weekly_20190726[table_name]
print("周线行情-时间为序  weekly_20190726 30_20190726 返回数据 row 行数 = "+str(weekly_20190726.shape[0]))
weekly_20190726_addname_dataframe.to_excel(weekly_2019_excel_writer,'30_20190726',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190802")       ###  更新 周线记录日期
weekly_20190802 = pro.weekly(trade_date='20190802', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190802_tscode_list = list() 
for ts_code_sh in weekly_20190802['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190802_tscode_list.append(stock_name)
weekly_20190802_addname_dataframe=pd.DataFrame()
weekly_20190802_addname_dataframe['cname'] = weekly_20190802_tscode_list
for table_name in weekly_20190802.columns.values.tolist():
    weekly_20190802_addname_dataframe[table_name] = weekly_20190802[table_name]
print("周线行情-时间为序  weekly_20190802 31_20190802 返回数据 row 行数 = "+str(weekly_20190802.shape[0]))
weekly_20190802_addname_dataframe.to_excel(weekly_2019_excel_writer,'31_20190802',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190809")       ###  更新 周线记录日期
weekly_20190809 = pro.weekly(trade_date='20190809', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190809_tscode_list = list() 
for ts_code_sh in weekly_20190809['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190809_tscode_list.append(stock_name)
weekly_20190809_addname_dataframe=pd.DataFrame()
weekly_20190809_addname_dataframe['cname'] = weekly_20190809_tscode_list
for table_name in weekly_20190809.columns.values.tolist():
    weekly_20190809_addname_dataframe[table_name] = weekly_20190809[table_name]
print("周线行情-时间为序  weekly_20190809 32_20190809 返回数据 row 行数 = "+str(weekly_20190809.shape[0]))
weekly_20190809_addname_dataframe.to_excel(weekly_2019_excel_writer,'32_20190809',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190816")       ###  更新 周线记录日期
weekly_20190816 = pro.weekly(trade_date='20190816', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190816_tscode_list = list() 
for ts_code_sh in weekly_20190816['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190816_tscode_list.append(stock_name)
weekly_20190816_addname_dataframe=pd.DataFrame()
weekly_20190816_addname_dataframe['cname'] = weekly_20190816_tscode_list
for table_name in weekly_20190816.columns.values.tolist():
    weekly_20190816_addname_dataframe[table_name] = weekly_20190816[table_name]
print("周线行情-时间为序  weekly_20190816 33_20190816 返回数据 row 行数 = "+str(weekly_20190816.shape[0]))
weekly_20190816_addname_dataframe.to_excel(weekly_2019_excel_writer,'33_20190816',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190823")       ###  更新 周线记录日期
weekly_20190823 = pro.weekly(trade_date='20190823', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190823_tscode_list = list() 
for ts_code_sh in weekly_20190823['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190823_tscode_list.append(stock_name)
weekly_20190823_addname_dataframe=pd.DataFrame()
weekly_20190823_addname_dataframe['cname'] = weekly_20190823_tscode_list
for table_name in weekly_20190823.columns.values.tolist():
    weekly_20190823_addname_dataframe[table_name] = weekly_20190823[table_name]
print("周线行情-时间为序  weekly_20190823 34_20190823 返回数据 row 行数 = "+str(weekly_20190823.shape[0]))
weekly_20190823_addname_dataframe.to_excel(weekly_2019_excel_writer,'34_20190823',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190830")       ###  更新 周线记录日期
weekly_20190830 = pro.weekly(trade_date='20190830', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190830_tscode_list = list() 
for ts_code_sh in weekly_20190830['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190830_tscode_list.append(stock_name)
weekly_20190830_addname_dataframe=pd.DataFrame()
weekly_20190830_addname_dataframe['cname'] = weekly_20190830_tscode_list
for table_name in weekly_20190830.columns.values.tolist():
    weekly_20190830_addname_dataframe[table_name] = weekly_20190830[table_name]
print("周线行情-时间为序  weekly_20190830 35_20190830 返回数据 row 行数 = "+str(weekly_20190830.shape[0]))
weekly_20190830_addname_dataframe.to_excel(weekly_2019_excel_writer,'35_20190830',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190906")       ###  更新 周线记录日期
weekly_20190906 = pro.weekly(trade_date='20190906', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190906_tscode_list = list() 
for ts_code_sh in weekly_20190906['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190906_tscode_list.append(stock_name)
weekly_20190906_addname_dataframe=pd.DataFrame()
weekly_20190906_addname_dataframe['cname'] = weekly_20190906_tscode_list
for table_name in weekly_20190906.columns.values.tolist():
    weekly_20190906_addname_dataframe[table_name] = weekly_20190906[table_name]
print("周线行情-时间为序  weekly_20190906 36_20190906 返回数据 row 行数 = "+str(weekly_20190906.shape[0]))
weekly_20190906_addname_dataframe.to_excel(weekly_2019_excel_writer,'36_20190906',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190912")       ###  更新 周线记录日期
weekly_20190912 = pro.weekly(trade_date='20190912', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190912_tscode_list = list() 
for ts_code_sh in weekly_20190912['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190912_tscode_list.append(stock_name)
weekly_20190912_addname_dataframe=pd.DataFrame()
weekly_20190912_addname_dataframe['cname'] = weekly_20190912_tscode_list
for table_name in weekly_20190912.columns.values.tolist():
    weekly_20190912_addname_dataframe[table_name] = weekly_20190912[table_name]
print("周线行情-时间为序  weekly_20190912 37_20190912 返回数据 row 行数 = "+str(weekly_20190912.shape[0]))
weekly_20190912_addname_dataframe.to_excel(weekly_2019_excel_writer,'37_20190912',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190920")       ###  更新 周线记录日期
weekly_20190920 = pro.weekly(trade_date='20190920', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190920_tscode_list = list() 
for ts_code_sh in weekly_20190920['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190920_tscode_list.append(stock_name)
weekly_20190920_addname_dataframe=pd.DataFrame()
weekly_20190920_addname_dataframe['cname'] = weekly_20190920_tscode_list
for table_name in weekly_20190920.columns.values.tolist():
    weekly_20190920_addname_dataframe[table_name] = weekly_20190920[table_name]
print("周线行情-时间为序  weekly_20190920 38_20190920 返回数据 row 行数 = "+str(weekly_20190920.shape[0]))
weekly_20190920_addname_dataframe.to_excel(weekly_2019_excel_writer,'38_20190920',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190927")       ###  更新 周线记录日期
weekly_20190927 = pro.weekly(trade_date='20190927', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190927_tscode_list = list() 
for ts_code_sh in weekly_20190927['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190927_tscode_list.append(stock_name)
weekly_20190927_addname_dataframe=pd.DataFrame()
weekly_20190927_addname_dataframe['cname'] = weekly_20190927_tscode_list
for table_name in weekly_20190927.columns.values.tolist():
    weekly_20190927_addname_dataframe[table_name] = weekly_20190927[table_name]
print("周线行情-时间为序  weekly_20190927 39_20190927 返回数据 row 行数 = "+str(weekly_20190927.shape[0]))
weekly_20190927_addname_dataframe.to_excel(weekly_2019_excel_writer,'39_20190927',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20190930")       ###  更新 周线记录日期
weekly_20190930 = pro.weekly(trade_date='20190930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20190930_tscode_list = list() 
for ts_code_sh in weekly_20190930['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20190930_tscode_list.append(stock_name)
weekly_20190930_addname_dataframe=pd.DataFrame()
weekly_20190930_addname_dataframe['cname'] = weekly_20190930_tscode_list
for table_name in weekly_20190930.columns.values.tolist():
    weekly_20190930_addname_dataframe[table_name] = weekly_20190930[table_name]
print("周线行情-时间为序  weekly_20190930 39_20190930 返回数据 row 行数 = "+str(weekly_20190930.shape[0]))
weekly_20190930_addname_dataframe.to_excel(weekly_2019_excel_writer,'39_20190930',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191011")       ###  更新 周线记录日期
weekly_20191011 = pro.weekly(trade_date='20191011', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191011_tscode_list = list() 
for ts_code_sh in weekly_20191011['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191011_tscode_list.append(stock_name)
weekly_20191011_addname_dataframe=pd.DataFrame()
weekly_20191011_addname_dataframe['cname'] = weekly_20191011_tscode_list
for table_name in weekly_20191011.columns.values.tolist():
    weekly_20191011_addname_dataframe[table_name] = weekly_20191011[table_name]
print("周线行情-时间为序  weekly_20191011 41_20191011 返回数据 row 行数 = "+str(weekly_20191011.shape[0]))
weekly_20191011_addname_dataframe.to_excel(weekly_2019_excel_writer,'41_20191011',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191018")       ###  更新 周线记录日期
weekly_20191018 = pro.weekly(trade_date='20191018', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191018_tscode_list = list() 
for ts_code_sh in weekly_20191018['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191018_tscode_list.append(stock_name)
weekly_20191018_addname_dataframe=pd.DataFrame()
weekly_20191018_addname_dataframe['cname'] = weekly_20191018_tscode_list
for table_name in weekly_20191018.columns.values.tolist():
    weekly_20191018_addname_dataframe[table_name] = weekly_20191018[table_name]
print("周线行情-时间为序  weekly_20191018 42_20191018 返回数据 row 行数 = "+str(weekly_20191018.shape[0]))
weekly_20191018_addname_dataframe.to_excel(weekly_2019_excel_writer,'42_20191018',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191025")       ###  更新 周线记录日期
weekly_20191025 = pro.weekly(trade_date='20191025', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191025_tscode_list = list() 
for ts_code_sh in weekly_20191025['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191025_tscode_list.append(stock_name)
weekly_20191025_addname_dataframe=pd.DataFrame()
weekly_20191025_addname_dataframe['cname'] = weekly_20191025_tscode_list
for table_name in weekly_20191025.columns.values.tolist():
    weekly_20191025_addname_dataframe[table_name] = weekly_20191025[table_name]
print("周线行情-时间为序  weekly_20191025 43_20191025 返回数据 row 行数 = "+str(weekly_20191025.shape[0]))
weekly_20191025_addname_dataframe.to_excel(weekly_2019_excel_writer,'43_20191025',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191101")       ###  更新 周线记录日期
weekly_20191101 = pro.weekly(trade_date='20191101', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191101_tscode_list = list() 
for ts_code_sh in weekly_20191101['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191101_tscode_list.append(stock_name)
weekly_20191101_addname_dataframe=pd.DataFrame()
weekly_20191101_addname_dataframe['cname'] = weekly_20191101_tscode_list
for table_name in weekly_20191101.columns.values.tolist():
    weekly_20191101_addname_dataframe[table_name] = weekly_20191101[table_name]
print("周线行情-时间为序  weekly_20191101 44_20191101 返回数据 row 行数 = "+str(weekly_20191101.shape[0]))
weekly_20191101_addname_dataframe.to_excel(weekly_2019_excel_writer,'44_20191101',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191108")       ###  更新 周线记录日期
weekly_20191108 = pro.weekly(trade_date='20191108', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191108_tscode_list = list() 
for ts_code_sh in weekly_20191108['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191108_tscode_list.append(stock_name)
weekly_20191108_addname_dataframe=pd.DataFrame()
weekly_20191108_addname_dataframe['cname'] = weekly_20191108_tscode_list
for table_name in weekly_20191108.columns.values.tolist():
    weekly_20191108_addname_dataframe[table_name] = weekly_20191108[table_name]
print("周线行情-时间为序  weekly_20191108 45_20191108 返回数据 row 行数 = "+str(weekly_20191108.shape[0]))
weekly_20191108_addname_dataframe.to_excel(weekly_2019_excel_writer,'45_20191108',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191115")       ###  更新 周线记录日期
weekly_20191115 = pro.weekly(trade_date='20191115', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191115_tscode_list = list() 
for ts_code_sh in weekly_20191115['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191115_tscode_list.append(stock_name)
weekly_20191115_addname_dataframe=pd.DataFrame()
weekly_20191115_addname_dataframe['cname'] = weekly_20191115_tscode_list
for table_name in weekly_20191115.columns.values.tolist():
    weekly_20191115_addname_dataframe[table_name] = weekly_20191115[table_name]
print("周线行情-时间为序  weekly_20191115 46_20191115 返回数据 row 行数 = "+str(weekly_20191115.shape[0]))
weekly_20191115_addname_dataframe.to_excel(weekly_2019_excel_writer,'46_20191115',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191122")       ###  更新 周线记录日期
weekly_20191122 = pro.weekly(trade_date='20191122', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191122_tscode_list = list() 
for ts_code_sh in weekly_20191122['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191122_tscode_list.append(stock_name)
weekly_20191122_addname_dataframe=pd.DataFrame()
weekly_20191122_addname_dataframe['cname'] = weekly_20191122_tscode_list
for table_name in weekly_20191122.columns.values.tolist():
    weekly_20191122_addname_dataframe[table_name] = weekly_20191122[table_name]
print("周线行情-时间为序  weekly_20191122 47_20191122 返回数据 row 行数 = "+str(weekly_20191122.shape[0]))
weekly_20191122_addname_dataframe.to_excel(weekly_2019_excel_writer,'47_20191122',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191129")       ###  更新 周线记录日期
weekly_20191129 = pro.weekly(trade_date='20191129', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191129_tscode_list = list() 
for ts_code_sh in weekly_20191129['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191129_tscode_list.append(stock_name)
weekly_20191129_addname_dataframe=pd.DataFrame()
weekly_20191129_addname_dataframe['cname'] = weekly_20191129_tscode_list
for table_name in weekly_20191129.columns.values.tolist():
    weekly_20191129_addname_dataframe[table_name] = weekly_20191129[table_name]
print("周线行情-时间为序  weekly_20191129 48_20191129 返回数据 row 行数 = "+str(weekly_20191129.shape[0]))
weekly_20191129_addname_dataframe.to_excel(weekly_2019_excel_writer,'48_20191129',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191206")       ###  更新 周线记录日期
weekly_20191206 = pro.weekly(trade_date='20191206', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191206_tscode_list = list() 
for ts_code_sh in weekly_20191206['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191206_tscode_list.append(stock_name)
weekly_20191206_addname_dataframe=pd.DataFrame()
weekly_20191206_addname_dataframe['cname'] = weekly_20191206_tscode_list
for table_name in weekly_20191206.columns.values.tolist():
    weekly_20191206_addname_dataframe[table_name] = weekly_20191206[table_name]
print("周线行情-时间为序  weekly_20191206 49_20191206 返回数据 row 行数 = "+str(weekly_20191206.shape[0]))
weekly_20191206_addname_dataframe.to_excel(weekly_2019_excel_writer,'49_20191206',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191213")       ###  更新 周线记录日期
weekly_20191213 = pro.weekly(trade_date='20191213', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191213_tscode_list = list() 
for ts_code_sh in weekly_20191213['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191213_tscode_list.append(stock_name)
weekly_20191213_addname_dataframe=pd.DataFrame()
weekly_20191213_addname_dataframe['cname'] = weekly_20191213_tscode_list
for table_name in weekly_20191213.columns.values.tolist():
    weekly_20191213_addname_dataframe[table_name] = weekly_20191213[table_name]
print("周线行情-时间为序  weekly_20191213 50_20191213 返回数据 row 行数 = "+str(weekly_20191213.shape[0]))
weekly_20191213_addname_dataframe.to_excel(weekly_2019_excel_writer,'50_20191213',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191220")       ###  更新 周线记录日期
weekly_20191220 = pro.weekly(trade_date='20191220', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191220_tscode_list = list() 
for ts_code_sh in weekly_20191220['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191220_tscode_list.append(stock_name)
weekly_20191220_addname_dataframe=pd.DataFrame()
weekly_20191220_addname_dataframe['cname'] = weekly_20191220_tscode_list
for table_name in weekly_20191220.columns.values.tolist():
    weekly_20191220_addname_dataframe[table_name] = weekly_20191220[table_name]
print("周线行情-时间为序  weekly_20191220 51_20191220 返回数据 row 行数 = "+str(weekly_20191220.shape[0]))
weekly_20191220_addname_dataframe.to_excel(weekly_2019_excel_writer,'51_20191220',index=False)
weekly_2019_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20191227")       ###  更新 周线记录日期
weekly_20191227 = pro.weekly(trade_date='20191227', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20191227_tscode_list = list() 
for ts_code_sh in weekly_20191227['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20191227_tscode_list.append(stock_name)
weekly_20191227_addname_dataframe=pd.DataFrame()
weekly_20191227_addname_dataframe['cname'] = weekly_20191227_tscode_list
for table_name in weekly_20191227.columns.values.tolist():
    weekly_20191227_addname_dataframe[table_name] = weekly_20191227[table_name]
print("周线行情-时间为序  weekly_20191227 52_20191227 返回数据 row 行数 = "+str(weekly_20191227.shape[0]))
weekly_20191227_addname_dataframe.to_excel(weekly_2019_excel_writer,'52_20191227',index=False)
weekly_2019_excel_writer.save()
createexcel('weekly_2020.xlsx')
weekly_2020_book = load_workbook('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2020.xlsx')
weekly_2020_excel_writer = pd.ExcelWriter('C:\\Users\\zhuzj5\\Desktop\\zbin\\J0_Data\\weekly_2020.xlsx', engine='openpyxl')
weekly_2020_excel_writer.book = weekly_2020_book
weekly_2020_excel_writer.sheets = dict((ws.title, ws) for ws in weekly_2020_book.worksheets)
J0_PROPS.put(tree_node_name+"record_date", "20200103")       ###  更新 周线记录日期
weekly_20200103 = pro.weekly(trade_date='20200103', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200103_tscode_list = list() 
for ts_code_sh in weekly_20200103['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200103_tscode_list.append(stock_name)
weekly_20200103_addname_dataframe=pd.DataFrame()
weekly_20200103_addname_dataframe['cname'] = weekly_20200103_tscode_list
for table_name in weekly_20200103.columns.values.tolist():
    weekly_20200103_addname_dataframe[table_name] = weekly_20200103[table_name]
print("周线行情-时间为序  weekly_20200103 1_20200103 返回数据 row 行数 = "+str(weekly_20200103.shape[0]))
weekly_20200103_addname_dataframe.to_excel(weekly_2020_excel_writer,'1_20200103',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200110")       ###  更新 周线记录日期
weekly_20200110 = pro.weekly(trade_date='20200110', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200110_tscode_list = list() 
for ts_code_sh in weekly_20200110['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200110_tscode_list.append(stock_name)
weekly_20200110_addname_dataframe=pd.DataFrame()
weekly_20200110_addname_dataframe['cname'] = weekly_20200110_tscode_list
for table_name in weekly_20200110.columns.values.tolist():
    weekly_20200110_addname_dataframe[table_name] = weekly_20200110[table_name]
print("周线行情-时间为序  weekly_20200110 2_20200110 返回数据 row 行数 = "+str(weekly_20200110.shape[0]))
weekly_20200110_addname_dataframe.to_excel(weekly_2020_excel_writer,'2_20200110',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200117")       ###  更新 周线记录日期
weekly_20200117 = pro.weekly(trade_date='20200117', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200117_tscode_list = list() 
for ts_code_sh in weekly_20200117['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200117_tscode_list.append(stock_name)
weekly_20200117_addname_dataframe=pd.DataFrame()
weekly_20200117_addname_dataframe['cname'] = weekly_20200117_tscode_list
for table_name in weekly_20200117.columns.values.tolist():
    weekly_20200117_addname_dataframe[table_name] = weekly_20200117[table_name]
print("周线行情-时间为序  weekly_20200117 3_20200117 返回数据 row 行数 = "+str(weekly_20200117.shape[0]))
weekly_20200117_addname_dataframe.to_excel(weekly_2020_excel_writer,'3_20200117',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200123")       ###  更新 周线记录日期
weekly_20200123 = pro.weekly(trade_date='20200123', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200123_tscode_list = list() 
for ts_code_sh in weekly_20200123['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200123_tscode_list.append(stock_name)
weekly_20200123_addname_dataframe=pd.DataFrame()
weekly_20200123_addname_dataframe['cname'] = weekly_20200123_tscode_list
for table_name in weekly_20200123.columns.values.tolist():
    weekly_20200123_addname_dataframe[table_name] = weekly_20200123[table_name]
print("周线行情-时间为序  weekly_20200123 4_20200123 返回数据 row 行数 = "+str(weekly_20200123.shape[0]))
weekly_20200123_addname_dataframe.to_excel(weekly_2020_excel_writer,'4_20200123',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200207")       ###  更新 周线记录日期
weekly_20200207 = pro.weekly(trade_date='20200207', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200207_tscode_list = list() 
for ts_code_sh in weekly_20200207['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200207_tscode_list.append(stock_name)
weekly_20200207_addname_dataframe=pd.DataFrame()
weekly_20200207_addname_dataframe['cname'] = weekly_20200207_tscode_list
for table_name in weekly_20200207.columns.values.tolist():
    weekly_20200207_addname_dataframe[table_name] = weekly_20200207[table_name]
print("周线行情-时间为序  weekly_20200207 6_20200207 返回数据 row 行数 = "+str(weekly_20200207.shape[0]))
weekly_20200207_addname_dataframe.to_excel(weekly_2020_excel_writer,'6_20200207',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200214")       ###  更新 周线记录日期
weekly_20200214 = pro.weekly(trade_date='20200214', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200214_tscode_list = list() 
for ts_code_sh in weekly_20200214['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200214_tscode_list.append(stock_name)
weekly_20200214_addname_dataframe=pd.DataFrame()
weekly_20200214_addname_dataframe['cname'] = weekly_20200214_tscode_list
for table_name in weekly_20200214.columns.values.tolist():
    weekly_20200214_addname_dataframe[table_name] = weekly_20200214[table_name]
print("周线行情-时间为序  weekly_20200214 7_20200214 返回数据 row 行数 = "+str(weekly_20200214.shape[0]))
weekly_20200214_addname_dataframe.to_excel(weekly_2020_excel_writer,'7_20200214',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200221")       ###  更新 周线记录日期
weekly_20200221 = pro.weekly(trade_date='20200221', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200221_tscode_list = list() 
for ts_code_sh in weekly_20200221['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200221_tscode_list.append(stock_name)
weekly_20200221_addname_dataframe=pd.DataFrame()
weekly_20200221_addname_dataframe['cname'] = weekly_20200221_tscode_list
for table_name in weekly_20200221.columns.values.tolist():
    weekly_20200221_addname_dataframe[table_name] = weekly_20200221[table_name]
print("周线行情-时间为序  weekly_20200221 8_20200221 返回数据 row 行数 = "+str(weekly_20200221.shape[0]))
weekly_20200221_addname_dataframe.to_excel(weekly_2020_excel_writer,'8_20200221',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200228")       ###  更新 周线记录日期
weekly_20200228 = pro.weekly(trade_date='20200228', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200228_tscode_list = list() 
for ts_code_sh in weekly_20200228['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200228_tscode_list.append(stock_name)
weekly_20200228_addname_dataframe=pd.DataFrame()
weekly_20200228_addname_dataframe['cname'] = weekly_20200228_tscode_list
for table_name in weekly_20200228.columns.values.tolist():
    weekly_20200228_addname_dataframe[table_name] = weekly_20200228[table_name]
print("周线行情-时间为序  weekly_20200228 9_20200228 返回数据 row 行数 = "+str(weekly_20200228.shape[0]))
weekly_20200228_addname_dataframe.to_excel(weekly_2020_excel_writer,'9_20200228',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200306")       ###  更新 周线记录日期
weekly_20200306 = pro.weekly(trade_date='20200306', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200306_tscode_list = list() 
for ts_code_sh in weekly_20200306['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200306_tscode_list.append(stock_name)
weekly_20200306_addname_dataframe=pd.DataFrame()
weekly_20200306_addname_dataframe['cname'] = weekly_20200306_tscode_list
for table_name in weekly_20200306.columns.values.tolist():
    weekly_20200306_addname_dataframe[table_name] = weekly_20200306[table_name]
print("周线行情-时间为序  weekly_20200306 10_20200306 返回数据 row 行数 = "+str(weekly_20200306.shape[0]))
weekly_20200306_addname_dataframe.to_excel(weekly_2020_excel_writer,'10_20200306',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200313")       ###  更新 周线记录日期
weekly_20200313 = pro.weekly(trade_date='20200313', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200313_tscode_list = list() 
for ts_code_sh in weekly_20200313['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200313_tscode_list.append(stock_name)
weekly_20200313_addname_dataframe=pd.DataFrame()
weekly_20200313_addname_dataframe['cname'] = weekly_20200313_tscode_list
for table_name in weekly_20200313.columns.values.tolist():
    weekly_20200313_addname_dataframe[table_name] = weekly_20200313[table_name]
print("周线行情-时间为序  weekly_20200313 11_20200313 返回数据 row 行数 = "+str(weekly_20200313.shape[0]))
weekly_20200313_addname_dataframe.to_excel(weekly_2020_excel_writer,'11_20200313',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200320")       ###  更新 周线记录日期
weekly_20200320 = pro.weekly(trade_date='20200320', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200320_tscode_list = list() 
for ts_code_sh in weekly_20200320['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200320_tscode_list.append(stock_name)
weekly_20200320_addname_dataframe=pd.DataFrame()
weekly_20200320_addname_dataframe['cname'] = weekly_20200320_tscode_list
for table_name in weekly_20200320.columns.values.tolist():
    weekly_20200320_addname_dataframe[table_name] = weekly_20200320[table_name]
print("周线行情-时间为序  weekly_20200320 12_20200320 返回数据 row 行数 = "+str(weekly_20200320.shape[0]))
weekly_20200320_addname_dataframe.to_excel(weekly_2020_excel_writer,'12_20200320',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200327")       ###  更新 周线记录日期
weekly_20200327 = pro.weekly(trade_date='20200327', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200327_tscode_list = list() 
for ts_code_sh in weekly_20200327['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200327_tscode_list.append(stock_name)
weekly_20200327_addname_dataframe=pd.DataFrame()
weekly_20200327_addname_dataframe['cname'] = weekly_20200327_tscode_list
for table_name in weekly_20200327.columns.values.tolist():
    weekly_20200327_addname_dataframe[table_name] = weekly_20200327[table_name]
print("周线行情-时间为序  weekly_20200327 13_20200327 返回数据 row 行数 = "+str(weekly_20200327.shape[0]))
weekly_20200327_addname_dataframe.to_excel(weekly_2020_excel_writer,'13_20200327',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200403")       ###  更新 周线记录日期
weekly_20200403 = pro.weekly(trade_date='20200403', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200403_tscode_list = list() 
for ts_code_sh in weekly_20200403['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200403_tscode_list.append(stock_name)
weekly_20200403_addname_dataframe=pd.DataFrame()
weekly_20200403_addname_dataframe['cname'] = weekly_20200403_tscode_list
for table_name in weekly_20200403.columns.values.tolist():
    weekly_20200403_addname_dataframe[table_name] = weekly_20200403[table_name]
print("周线行情-时间为序  weekly_20200403 14_20200403 返回数据 row 行数 = "+str(weekly_20200403.shape[0]))
weekly_20200403_addname_dataframe.to_excel(weekly_2020_excel_writer,'14_20200403',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200410")       ###  更新 周线记录日期
weekly_20200410 = pro.weekly(trade_date='20200410', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200410_tscode_list = list() 
for ts_code_sh in weekly_20200410['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200410_tscode_list.append(stock_name)
weekly_20200410_addname_dataframe=pd.DataFrame()
weekly_20200410_addname_dataframe['cname'] = weekly_20200410_tscode_list
for table_name in weekly_20200410.columns.values.tolist():
    weekly_20200410_addname_dataframe[table_name] = weekly_20200410[table_name]
print("周线行情-时间为序  weekly_20200410 15_20200410 返回数据 row 行数 = "+str(weekly_20200410.shape[0]))
weekly_20200410_addname_dataframe.to_excel(weekly_2020_excel_writer,'15_20200410',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200417")       ###  更新 周线记录日期
weekly_20200417 = pro.weekly(trade_date='20200417', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200417_tscode_list = list() 
for ts_code_sh in weekly_20200417['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200417_tscode_list.append(stock_name)
weekly_20200417_addname_dataframe=pd.DataFrame()
weekly_20200417_addname_dataframe['cname'] = weekly_20200417_tscode_list
for table_name in weekly_20200417.columns.values.tolist():
    weekly_20200417_addname_dataframe[table_name] = weekly_20200417[table_name]
print("周线行情-时间为序  weekly_20200417 16_20200417 返回数据 row 行数 = "+str(weekly_20200417.shape[0]))
weekly_20200417_addname_dataframe.to_excel(weekly_2020_excel_writer,'16_20200417',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200424")       ###  更新 周线记录日期
weekly_20200424 = pro.weekly(trade_date='20200424', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200424_tscode_list = list() 
for ts_code_sh in weekly_20200424['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200424_tscode_list.append(stock_name)
weekly_20200424_addname_dataframe=pd.DataFrame()
weekly_20200424_addname_dataframe['cname'] = weekly_20200424_tscode_list
for table_name in weekly_20200424.columns.values.tolist():
    weekly_20200424_addname_dataframe[table_name] = weekly_20200424[table_name]
print("周线行情-时间为序  weekly_20200424 17_20200424 返回数据 row 行数 = "+str(weekly_20200424.shape[0]))
weekly_20200424_addname_dataframe.to_excel(weekly_2020_excel_writer,'17_20200424',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200430")       ###  更新 周线记录日期
weekly_20200430 = pro.weekly(trade_date='20200430', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200430_tscode_list = list() 
for ts_code_sh in weekly_20200430['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200430_tscode_list.append(stock_name)
weekly_20200430_addname_dataframe=pd.DataFrame()
weekly_20200430_addname_dataframe['cname'] = weekly_20200430_tscode_list
for table_name in weekly_20200430.columns.values.tolist():
    weekly_20200430_addname_dataframe[table_name] = weekly_20200430[table_name]
print("周线行情-时间为序  weekly_20200430 18_20200430 返回数据 row 行数 = "+str(weekly_20200430.shape[0]))
weekly_20200430_addname_dataframe.to_excel(weekly_2020_excel_writer,'18_20200430',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200508")       ###  更新 周线记录日期
weekly_20200508 = pro.weekly(trade_date='20200508', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200508_tscode_list = list() 
for ts_code_sh in weekly_20200508['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200508_tscode_list.append(stock_name)
weekly_20200508_addname_dataframe=pd.DataFrame()
weekly_20200508_addname_dataframe['cname'] = weekly_20200508_tscode_list
for table_name in weekly_20200508.columns.values.tolist():
    weekly_20200508_addname_dataframe[table_name] = weekly_20200508[table_name]
print("周线行情-时间为序  weekly_20200508 19_20200508 返回数据 row 行数 = "+str(weekly_20200508.shape[0]))
weekly_20200508_addname_dataframe.to_excel(weekly_2020_excel_writer,'19_20200508',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200515")       ###  更新 周线记录日期
weekly_20200515 = pro.weekly(trade_date='20200515', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200515_tscode_list = list() 
for ts_code_sh in weekly_20200515['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200515_tscode_list.append(stock_name)
weekly_20200515_addname_dataframe=pd.DataFrame()
weekly_20200515_addname_dataframe['cname'] = weekly_20200515_tscode_list
for table_name in weekly_20200515.columns.values.tolist():
    weekly_20200515_addname_dataframe[table_name] = weekly_20200515[table_name]
print("周线行情-时间为序  weekly_20200515 20_20200515 返回数据 row 行数 = "+str(weekly_20200515.shape[0]))
weekly_20200515_addname_dataframe.to_excel(weekly_2020_excel_writer,'20_20200515',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200522")       ###  更新 周线记录日期
weekly_20200522 = pro.weekly(trade_date='20200522', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200522_tscode_list = list() 
for ts_code_sh in weekly_20200522['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200522_tscode_list.append(stock_name)
weekly_20200522_addname_dataframe=pd.DataFrame()
weekly_20200522_addname_dataframe['cname'] = weekly_20200522_tscode_list
for table_name in weekly_20200522.columns.values.tolist():
    weekly_20200522_addname_dataframe[table_name] = weekly_20200522[table_name]
print("周线行情-时间为序  weekly_20200522 21_20200522 返回数据 row 行数 = "+str(weekly_20200522.shape[0]))
weekly_20200522_addname_dataframe.to_excel(weekly_2020_excel_writer,'21_20200522',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200529")       ###  更新 周线记录日期
weekly_20200529 = pro.weekly(trade_date='20200529', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200529_tscode_list = list() 
for ts_code_sh in weekly_20200529['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200529_tscode_list.append(stock_name)
weekly_20200529_addname_dataframe=pd.DataFrame()
weekly_20200529_addname_dataframe['cname'] = weekly_20200529_tscode_list
for table_name in weekly_20200529.columns.values.tolist():
    weekly_20200529_addname_dataframe[table_name] = weekly_20200529[table_name]
print("周线行情-时间为序  weekly_20200529 22_20200529 返回数据 row 行数 = "+str(weekly_20200529.shape[0]))
weekly_20200529_addname_dataframe.to_excel(weekly_2020_excel_writer,'22_20200529',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200605")       ###  更新 周线记录日期
weekly_20200605 = pro.weekly(trade_date='20200605', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200605_tscode_list = list() 
for ts_code_sh in weekly_20200605['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200605_tscode_list.append(stock_name)
weekly_20200605_addname_dataframe=pd.DataFrame()
weekly_20200605_addname_dataframe['cname'] = weekly_20200605_tscode_list
for table_name in weekly_20200605.columns.values.tolist():
    weekly_20200605_addname_dataframe[table_name] = weekly_20200605[table_name]
print("周线行情-时间为序  weekly_20200605 23_20200605 返回数据 row 行数 = "+str(weekly_20200605.shape[0]))
weekly_20200605_addname_dataframe.to_excel(weekly_2020_excel_writer,'23_20200605',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200612")       ###  更新 周线记录日期
weekly_20200612 = pro.weekly(trade_date='20200612', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200612_tscode_list = list() 
for ts_code_sh in weekly_20200612['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200612_tscode_list.append(stock_name)
weekly_20200612_addname_dataframe=pd.DataFrame()
weekly_20200612_addname_dataframe['cname'] = weekly_20200612_tscode_list
for table_name in weekly_20200612.columns.values.tolist():
    weekly_20200612_addname_dataframe[table_name] = weekly_20200612[table_name]
print("周线行情-时间为序  weekly_20200612 24_20200612 返回数据 row 行数 = "+str(weekly_20200612.shape[0]))
weekly_20200612_addname_dataframe.to_excel(weekly_2020_excel_writer,'24_20200612',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200619")       ###  更新 周线记录日期
weekly_20200619 = pro.weekly(trade_date='20200619', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200619_tscode_list = list() 
for ts_code_sh in weekly_20200619['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200619_tscode_list.append(stock_name)
weekly_20200619_addname_dataframe=pd.DataFrame()
weekly_20200619_addname_dataframe['cname'] = weekly_20200619_tscode_list
for table_name in weekly_20200619.columns.values.tolist():
    weekly_20200619_addname_dataframe[table_name] = weekly_20200619[table_name]
print("周线行情-时间为序  weekly_20200619 25_20200619 返回数据 row 行数 = "+str(weekly_20200619.shape[0]))
weekly_20200619_addname_dataframe.to_excel(weekly_2020_excel_writer,'25_20200619',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200624")       ###  更新 周线记录日期
weekly_20200624 = pro.weekly(trade_date='20200624', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200624_tscode_list = list() 
for ts_code_sh in weekly_20200624['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200624_tscode_list.append(stock_name)
weekly_20200624_addname_dataframe=pd.DataFrame()
weekly_20200624_addname_dataframe['cname'] = weekly_20200624_tscode_list
for table_name in weekly_20200624.columns.values.tolist():
    weekly_20200624_addname_dataframe[table_name] = weekly_20200624[table_name]
print("周线行情-时间为序  weekly_20200624 26_20200624 返回数据 row 行数 = "+str(weekly_20200624.shape[0]))
weekly_20200624_addname_dataframe.to_excel(weekly_2020_excel_writer,'26_20200624',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200703")       ###  更新 周线记录日期
weekly_20200703 = pro.weekly(trade_date='20200703', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200703_tscode_list = list() 
for ts_code_sh in weekly_20200703['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200703_tscode_list.append(stock_name)
weekly_20200703_addname_dataframe=pd.DataFrame()
weekly_20200703_addname_dataframe['cname'] = weekly_20200703_tscode_list
for table_name in weekly_20200703.columns.values.tolist():
    weekly_20200703_addname_dataframe[table_name] = weekly_20200703[table_name]
print("周线行情-时间为序  weekly_20200703 27_20200703 返回数据 row 行数 = "+str(weekly_20200703.shape[0]))
weekly_20200703_addname_dataframe.to_excel(weekly_2020_excel_writer,'27_20200703',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200710")       ###  更新 周线记录日期
weekly_20200710 = pro.weekly(trade_date='20200710', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200710_tscode_list = list() 
for ts_code_sh in weekly_20200710['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200710_tscode_list.append(stock_name)
weekly_20200710_addname_dataframe=pd.DataFrame()
weekly_20200710_addname_dataframe['cname'] = weekly_20200710_tscode_list
for table_name in weekly_20200710.columns.values.tolist():
    weekly_20200710_addname_dataframe[table_name] = weekly_20200710[table_name]
print("周线行情-时间为序  weekly_20200710 28_20200710 返回数据 row 行数 = "+str(weekly_20200710.shape[0]))
weekly_20200710_addname_dataframe.to_excel(weekly_2020_excel_writer,'28_20200710',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200717")       ###  更新 周线记录日期
weekly_20200717 = pro.weekly(trade_date='20200717', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200717_tscode_list = list() 
for ts_code_sh in weekly_20200717['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200717_tscode_list.append(stock_name)
weekly_20200717_addname_dataframe=pd.DataFrame()
weekly_20200717_addname_dataframe['cname'] = weekly_20200717_tscode_list
for table_name in weekly_20200717.columns.values.tolist():
    weekly_20200717_addname_dataframe[table_name] = weekly_20200717[table_name]
print("周线行情-时间为序  weekly_20200717 29_20200717 返回数据 row 行数 = "+str(weekly_20200717.shape[0]))
weekly_20200717_addname_dataframe.to_excel(weekly_2020_excel_writer,'29_20200717',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200724")       ###  更新 周线记录日期
weekly_20200724 = pro.weekly(trade_date='20200724', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200724_tscode_list = list() 
for ts_code_sh in weekly_20200724['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200724_tscode_list.append(stock_name)
weekly_20200724_addname_dataframe=pd.DataFrame()
weekly_20200724_addname_dataframe['cname'] = weekly_20200724_tscode_list
for table_name in weekly_20200724.columns.values.tolist():
    weekly_20200724_addname_dataframe[table_name] = weekly_20200724[table_name]
print("周线行情-时间为序  weekly_20200724 30_20200724 返回数据 row 行数 = "+str(weekly_20200724.shape[0]))
weekly_20200724_addname_dataframe.to_excel(weekly_2020_excel_writer,'30_20200724',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200731")       ###  更新 周线记录日期
weekly_20200731 = pro.weekly(trade_date='20200731', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200731_tscode_list = list() 
for ts_code_sh in weekly_20200731['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200731_tscode_list.append(stock_name)
weekly_20200731_addname_dataframe=pd.DataFrame()
weekly_20200731_addname_dataframe['cname'] = weekly_20200731_tscode_list
for table_name in weekly_20200731.columns.values.tolist():
    weekly_20200731_addname_dataframe[table_name] = weekly_20200731[table_name]
print("周线行情-时间为序  weekly_20200731 31_20200731 返回数据 row 行数 = "+str(weekly_20200731.shape[0]))
weekly_20200731_addname_dataframe.to_excel(weekly_2020_excel_writer,'31_20200731',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200807")       ###  更新 周线记录日期
weekly_20200807 = pro.weekly(trade_date='20200807', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200807_tscode_list = list() 
for ts_code_sh in weekly_20200807['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200807_tscode_list.append(stock_name)
weekly_20200807_addname_dataframe=pd.DataFrame()
weekly_20200807_addname_dataframe['cname'] = weekly_20200807_tscode_list
for table_name in weekly_20200807.columns.values.tolist():
    weekly_20200807_addname_dataframe[table_name] = weekly_20200807[table_name]
print("周线行情-时间为序  weekly_20200807 32_20200807 返回数据 row 行数 = "+str(weekly_20200807.shape[0]))
weekly_20200807_addname_dataframe.to_excel(weekly_2020_excel_writer,'32_20200807',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200814")       ###  更新 周线记录日期
weekly_20200814 = pro.weekly(trade_date='20200814', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200814_tscode_list = list() 
for ts_code_sh in weekly_20200814['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200814_tscode_list.append(stock_name)
weekly_20200814_addname_dataframe=pd.DataFrame()
weekly_20200814_addname_dataframe['cname'] = weekly_20200814_tscode_list
for table_name in weekly_20200814.columns.values.tolist():
    weekly_20200814_addname_dataframe[table_name] = weekly_20200814[table_name]
print("周线行情-时间为序  weekly_20200814 33_20200814 返回数据 row 行数 = "+str(weekly_20200814.shape[0]))
weekly_20200814_addname_dataframe.to_excel(weekly_2020_excel_writer,'33_20200814',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200821")       ###  更新 周线记录日期
weekly_20200821 = pro.weekly(trade_date='20200821', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200821_tscode_list = list() 
for ts_code_sh in weekly_20200821['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200821_tscode_list.append(stock_name)
weekly_20200821_addname_dataframe=pd.DataFrame()
weekly_20200821_addname_dataframe['cname'] = weekly_20200821_tscode_list
for table_name in weekly_20200821.columns.values.tolist():
    weekly_20200821_addname_dataframe[table_name] = weekly_20200821[table_name]
print("周线行情-时间为序  weekly_20200821 34_20200821 返回数据 row 行数 = "+str(weekly_20200821.shape[0]))
weekly_20200821_addname_dataframe.to_excel(weekly_2020_excel_writer,'34_20200821',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200828")       ###  更新 周线记录日期
weekly_20200828 = pro.weekly(trade_date='20200828', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200828_tscode_list = list() 
for ts_code_sh in weekly_20200828['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200828_tscode_list.append(stock_name)
weekly_20200828_addname_dataframe=pd.DataFrame()
weekly_20200828_addname_dataframe['cname'] = weekly_20200828_tscode_list
for table_name in weekly_20200828.columns.values.tolist():
    weekly_20200828_addname_dataframe[table_name] = weekly_20200828[table_name]
print("周线行情-时间为序  weekly_20200828 35_20200828 返回数据 row 行数 = "+str(weekly_20200828.shape[0]))
weekly_20200828_addname_dataframe.to_excel(weekly_2020_excel_writer,'35_20200828',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200904")       ###  更新 周线记录日期
weekly_20200904 = pro.weekly(trade_date='20200904', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200904_tscode_list = list() 
for ts_code_sh in weekly_20200904['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200904_tscode_list.append(stock_name)
weekly_20200904_addname_dataframe=pd.DataFrame()
weekly_20200904_addname_dataframe['cname'] = weekly_20200904_tscode_list
for table_name in weekly_20200904.columns.values.tolist():
    weekly_20200904_addname_dataframe[table_name] = weekly_20200904[table_name]
print("周线行情-时间为序  weekly_20200904 36_20200904 返回数据 row 行数 = "+str(weekly_20200904.shape[0]))
weekly_20200904_addname_dataframe.to_excel(weekly_2020_excel_writer,'36_20200904',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200911")       ###  更新 周线记录日期
weekly_20200911 = pro.weekly(trade_date='20200911', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200911_tscode_list = list() 
for ts_code_sh in weekly_20200911['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200911_tscode_list.append(stock_name)
weekly_20200911_addname_dataframe=pd.DataFrame()
weekly_20200911_addname_dataframe['cname'] = weekly_20200911_tscode_list
for table_name in weekly_20200911.columns.values.tolist():
    weekly_20200911_addname_dataframe[table_name] = weekly_20200911[table_name]
print("周线行情-时间为序  weekly_20200911 37_20200911 返回数据 row 行数 = "+str(weekly_20200911.shape[0]))
weekly_20200911_addname_dataframe.to_excel(weekly_2020_excel_writer,'37_20200911',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200918")       ###  更新 周线记录日期
weekly_20200918 = pro.weekly(trade_date='20200918', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200918_tscode_list = list() 
for ts_code_sh in weekly_20200918['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200918_tscode_list.append(stock_name)
weekly_20200918_addname_dataframe=pd.DataFrame()
weekly_20200918_addname_dataframe['cname'] = weekly_20200918_tscode_list
for table_name in weekly_20200918.columns.values.tolist():
    weekly_20200918_addname_dataframe[table_name] = weekly_20200918[table_name]
print("周线行情-时间为序  weekly_20200918 38_20200918 返回数据 row 行数 = "+str(weekly_20200918.shape[0]))
weekly_20200918_addname_dataframe.to_excel(weekly_2020_excel_writer,'38_20200918',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200925")       ###  更新 周线记录日期
weekly_20200925 = pro.weekly(trade_date='20200925', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200925_tscode_list = list() 
for ts_code_sh in weekly_20200925['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200925_tscode_list.append(stock_name)
weekly_20200925_addname_dataframe=pd.DataFrame()
weekly_20200925_addname_dataframe['cname'] = weekly_20200925_tscode_list
for table_name in weekly_20200925.columns.values.tolist():
    weekly_20200925_addname_dataframe[table_name] = weekly_20200925[table_name]
print("周线行情-时间为序  weekly_20200925 39_20200925 返回数据 row 行数 = "+str(weekly_20200925.shape[0]))
weekly_20200925_addname_dataframe.to_excel(weekly_2020_excel_writer,'39_20200925',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20200930")       ###  更新 周线记录日期
weekly_20200930 = pro.weekly(trade_date='20200930', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20200930_tscode_list = list() 
for ts_code_sh in weekly_20200930['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20200930_tscode_list.append(stock_name)
weekly_20200930_addname_dataframe=pd.DataFrame()
weekly_20200930_addname_dataframe['cname'] = weekly_20200930_tscode_list
for table_name in weekly_20200930.columns.values.tolist():
    weekly_20200930_addname_dataframe[table_name] = weekly_20200930[table_name]
print("周线行情-时间为序  weekly_20200930 40_20200930 返回数据 row 行数 = "+str(weekly_20200930.shape[0]))
weekly_20200930_addname_dataframe.to_excel(weekly_2020_excel_writer,'40_20200930',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201009")       ###  更新 周线记录日期
weekly_20201009 = pro.weekly(trade_date='20201009', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201009_tscode_list = list() 
for ts_code_sh in weekly_20201009['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201009_tscode_list.append(stock_name)
weekly_20201009_addname_dataframe=pd.DataFrame()
weekly_20201009_addname_dataframe['cname'] = weekly_20201009_tscode_list
for table_name in weekly_20201009.columns.values.tolist():
    weekly_20201009_addname_dataframe[table_name] = weekly_20201009[table_name]
print("周线行情-时间为序  weekly_20201009 41_20201009 返回数据 row 行数 = "+str(weekly_20201009.shape[0]))
weekly_20201009_addname_dataframe.to_excel(weekly_2020_excel_writer,'41_20201009',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201016")       ###  更新 周线记录日期
weekly_20201016 = pro.weekly(trade_date='20201016', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201016_tscode_list = list() 
for ts_code_sh in weekly_20201016['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201016_tscode_list.append(stock_name)
weekly_20201016_addname_dataframe=pd.DataFrame()
weekly_20201016_addname_dataframe['cname'] = weekly_20201016_tscode_list
for table_name in weekly_20201016.columns.values.tolist():
    weekly_20201016_addname_dataframe[table_name] = weekly_20201016[table_name]
print("周线行情-时间为序  weekly_20201016 42_20201016 返回数据 row 行数 = "+str(weekly_20201016.shape[0]))
weekly_20201016_addname_dataframe.to_excel(weekly_2020_excel_writer,'42_20201016',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201023")       ###  更新 周线记录日期
weekly_20201023 = pro.weekly(trade_date='20201023', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201023_tscode_list = list() 
for ts_code_sh in weekly_20201023['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201023_tscode_list.append(stock_name)
weekly_20201023_addname_dataframe=pd.DataFrame()
weekly_20201023_addname_dataframe['cname'] = weekly_20201023_tscode_list
for table_name in weekly_20201023.columns.values.tolist():
    weekly_20201023_addname_dataframe[table_name] = weekly_20201023[table_name]
print("周线行情-时间为序  weekly_20201023 43_20201023 返回数据 row 行数 = "+str(weekly_20201023.shape[0]))
weekly_20201023_addname_dataframe.to_excel(weekly_2020_excel_writer,'43_20201023',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201030")       ###  更新 周线记录日期
weekly_20201030 = pro.weekly(trade_date='20201030', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201030_tscode_list = list() 
for ts_code_sh in weekly_20201030['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201030_tscode_list.append(stock_name)
weekly_20201030_addname_dataframe=pd.DataFrame()
weekly_20201030_addname_dataframe['cname'] = weekly_20201030_tscode_list
for table_name in weekly_20201030.columns.values.tolist():
    weekly_20201030_addname_dataframe[table_name] = weekly_20201030[table_name]
print("周线行情-时间为序  weekly_20201030 44_20201030 返回数据 row 行数 = "+str(weekly_20201030.shape[0]))
weekly_20201030_addname_dataframe.to_excel(weekly_2020_excel_writer,'44_20201030',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201106")       ###  更新 周线记录日期
weekly_20201106 = pro.weekly(trade_date='20201106', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201106_tscode_list = list() 
for ts_code_sh in weekly_20201106['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201106_tscode_list.append(stock_name)
weekly_20201106_addname_dataframe=pd.DataFrame()
weekly_20201106_addname_dataframe['cname'] = weekly_20201106_tscode_list
for table_name in weekly_20201106.columns.values.tolist():
    weekly_20201106_addname_dataframe[table_name] = weekly_20201106[table_name]
print("周线行情-时间为序  weekly_20201106 45_20201106 返回数据 row 行数 = "+str(weekly_20201106.shape[0]))
weekly_20201106_addname_dataframe.to_excel(weekly_2020_excel_writer,'45_20201106',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201113")       ###  更新 周线记录日期
weekly_20201113 = pro.weekly(trade_date='20201113', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201113_tscode_list = list() 
for ts_code_sh in weekly_20201113['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201113_tscode_list.append(stock_name)
weekly_20201113_addname_dataframe=pd.DataFrame()
weekly_20201113_addname_dataframe['cname'] = weekly_20201113_tscode_list
for table_name in weekly_20201113.columns.values.tolist():
    weekly_20201113_addname_dataframe[table_name] = weekly_20201113[table_name]
print("周线行情-时间为序  weekly_20201113 46_20201113 返回数据 row 行数 = "+str(weekly_20201113.shape[0]))
weekly_20201113_addname_dataframe.to_excel(weekly_2020_excel_writer,'46_20201113',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201120")       ###  更新 周线记录日期
weekly_20201120 = pro.weekly(trade_date='20201120', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201120_tscode_list = list() 
for ts_code_sh in weekly_20201120['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201120_tscode_list.append(stock_name)
weekly_20201120_addname_dataframe=pd.DataFrame()
weekly_20201120_addname_dataframe['cname'] = weekly_20201120_tscode_list
for table_name in weekly_20201120.columns.values.tolist():
    weekly_20201120_addname_dataframe[table_name] = weekly_20201120[table_name]
print("周线行情-时间为序  weekly_20201120 47_20201120 返回数据 row 行数 = "+str(weekly_20201120.shape[0]))
weekly_20201120_addname_dataframe.to_excel(weekly_2020_excel_writer,'47_20201120',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201127")       ###  更新 周线记录日期
weekly_20201127 = pro.weekly(trade_date='20201127', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201127_tscode_list = list() 
for ts_code_sh in weekly_20201127['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201127_tscode_list.append(stock_name)
weekly_20201127_addname_dataframe=pd.DataFrame()
weekly_20201127_addname_dataframe['cname'] = weekly_20201127_tscode_list
for table_name in weekly_20201127.columns.values.tolist():
    weekly_20201127_addname_dataframe[table_name] = weekly_20201127[table_name]
print("周线行情-时间为序  weekly_20201127 48_20201127 返回数据 row 行数 = "+str(weekly_20201127.shape[0]))
weekly_20201127_addname_dataframe.to_excel(weekly_2020_excel_writer,'48_20201127',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201204")       ###  更新 周线记录日期
weekly_20201204 = pro.weekly(trade_date='20201204', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201204_tscode_list = list() 
for ts_code_sh in weekly_20201204['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201204_tscode_list.append(stock_name)
weekly_20201204_addname_dataframe=pd.DataFrame()
weekly_20201204_addname_dataframe['cname'] = weekly_20201204_tscode_list
for table_name in weekly_20201204.columns.values.tolist():
    weekly_20201204_addname_dataframe[table_name] = weekly_20201204[table_name]
print("周线行情-时间为序  weekly_20201204 49_20201204 返回数据 row 行数 = "+str(weekly_20201204.shape[0]))
weekly_20201204_addname_dataframe.to_excel(weekly_2020_excel_writer,'49_20201204',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201211")       ###  更新 周线记录日期
weekly_20201211 = pro.weekly(trade_date='20201211', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201211_tscode_list = list() 
for ts_code_sh in weekly_20201211['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201211_tscode_list.append(stock_name)
weekly_20201211_addname_dataframe=pd.DataFrame()
weekly_20201211_addname_dataframe['cname'] = weekly_20201211_tscode_list
for table_name in weekly_20201211.columns.values.tolist():
    weekly_20201211_addname_dataframe[table_name] = weekly_20201211[table_name]
print("周线行情-时间为序  weekly_20201211 50_20201211 返回数据 row 行数 = "+str(weekly_20201211.shape[0]))
weekly_20201211_addname_dataframe.to_excel(weekly_2020_excel_writer,'50_20201211',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201218")       ###  更新 周线记录日期
weekly_20201218 = pro.weekly(trade_date='20201218', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201218_tscode_list = list() 
for ts_code_sh in weekly_20201218['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201218_tscode_list.append(stock_name)
weekly_20201218_addname_dataframe=pd.DataFrame()
weekly_20201218_addname_dataframe['cname'] = weekly_20201218_tscode_list
for table_name in weekly_20201218.columns.values.tolist():
    weekly_20201218_addname_dataframe[table_name] = weekly_20201218[table_name]
print("周线行情-时间为序  weekly_20201218 51_20201218 返回数据 row 行数 = "+str(weekly_20201218.shape[0]))
weekly_20201218_addname_dataframe.to_excel(weekly_2020_excel_writer,'51_20201218',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201225")       ###  更新 周线记录日期
weekly_20201225 = pro.weekly(trade_date='20201225', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201225_tscode_list = list() 
for ts_code_sh in weekly_20201225['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201225_tscode_list.append(stock_name)
weekly_20201225_addname_dataframe=pd.DataFrame()
weekly_20201225_addname_dataframe['cname'] = weekly_20201225_tscode_list
for table_name in weekly_20201225.columns.values.tolist():
    weekly_20201225_addname_dataframe[table_name] = weekly_20201225[table_name]
print("周线行情-时间为序  weekly_20201225 52_20201225 返回数据 row 行数 = "+str(weekly_20201225.shape[0]))
weekly_20201225_addname_dataframe.to_excel(weekly_2020_excel_writer,'52_20201225',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "20201231")       ###  更新 周线记录日期
weekly_20201231 = pro.weekly(trade_date='20201231', fields='ts_code,trade_date,close,open,high,low,pre_close,change,pct_chg,vol,amount')
weekly_20201231_tscode_list = list() 
for ts_code_sh in weekly_20201231['ts_code']:
    stock_name = tscode_name_dict.get(ts_code_sh)
    if stock_name is None:
        stock_name = 'null'
    weekly_20201231_tscode_list.append(stock_name)
weekly_20201231_addname_dataframe=pd.DataFrame()
weekly_20201231_addname_dataframe['cname'] = weekly_20201231_tscode_list
for table_name in weekly_20201231.columns.values.tolist():
    weekly_20201231_addname_dataframe[table_name] = weekly_20201231[table_name]
print("周线行情-时间为序  weekly_20201231 53_20201231 返回数据 row 行数 = "+str(weekly_20201231.shape[0]))
weekly_20201231_addname_dataframe.to_excel(weekly_2020_excel_writer,'53_20201231',index=False)
weekly_2020_excel_writer.save()
J0_PROPS.put(tree_node_name+"record_date", "19901219")       ###  更新 记录日期
